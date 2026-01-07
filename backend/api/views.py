from django.shortcuts import render
from django.contrib.auth.models import User as DjangoUser
from django.shortcuts import get_object_or_404
from django.db import models
from django.db import IntegrityError
from rest_framework import generics, status
from .models import Contact
from .models import Note, NoteCategory
from .models import UserDetails
from .models import Team
from .models import Event
from .models import TeamMember
from .models import Log
from .models import Role, Permission, PermissionRole, Status, Source, Platform, Document, SMTPConfig, Email, EmailSignature, ChatRoom, Message, Notification, NotificationPreference, FosseSettings, OTP, Transaction, RIB, ContactView
from .serializer import (
    UserSerializer, ContactSerializer, ContactMigrationSerializer, NoteSerializer, NoteCategorySerializer,
    TeamSerializer, TeamDetailSerializer, UserDetailsSerializer, EventSerializer, TeamMemberSerializer,
    RoleSerializer, PermissionSerializer, PermissionRoleSerializer, StatusSerializer, SourceSerializer, PlatformSerializer, LogSerializer, DocumentSerializer,
    SMTPConfigSerializer, EmailSerializer, EmailSignatureSerializer, ChatRoomSerializer, MessageSerializer, NotificationSerializer,
    NotificationPreferenceSerializer, FosseSettingsSerializer, TransactionSerializer, RIBSerializer, ContactViewSerializer
)
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
import uuid
from datetime import datetime, date, timedelta
from django.utils import timezone
from django.db.models import Count, Q, Sum, F, Case, When, IntegerField, Value
from django.db.models.functions import Cast, MD5, Substr, Coalesce, Concat
from django.db.models import CharField, Value
import boto3
from botocore.exceptions import ClientError
import os
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.conf import settings
from django.http import StreamingHttpResponse, HttpResponse
from io import BytesIO
import csv
import io
import smtplib
import imaplib
import email
import email.utils
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.header import decode_header
import re
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync


def send_event_notification(event, notification_type='assigned', minutes_before=None):
    """
    Send event notification via WebSocket to the assigned user.
    
    Args:
        event: Event instance
        notification_type: 'assigned', '30min_before', or '5min_before'
        minutes_before: Number of minutes before event (for reminder notifications)
    """
    if not event.userId:
        return
    
    try:
        channel_layer = get_channel_layer()
        if not channel_layer:
            return
        
        # Format event datetime - convert to Europe/Paris timezone for display
        event_datetime = event.datetime
        if event_datetime:
            try:
                import pytz
                paris_tz = pytz.timezone('Europe/Paris')
                
                # Convert to Europe/Paris timezone if datetime is timezone-aware
                if timezone.is_aware(event_datetime):
                    event_datetime_local = event_datetime.astimezone(paris_tz)
                else:
                    # If naive, assume it's UTC and convert
                    utc_tz = pytz.UTC
                    event_datetime_local = utc_tz.localize(event_datetime).astimezone(paris_tz)
                
                event_datetime_str = event_datetime_local.strftime('%d/%m/%Y à %H:%M')
            except Exception as e:
                # Fallback: use UTC if conversion fails
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Failed to convert datetime to Paris timezone: {e}, using UTC")
                if timezone.is_aware(event_datetime):
                    event_datetime_str = event_datetime.strftime('%d/%m/%Y à %H:%M')
                else:
                    event_datetime_str = event_datetime.strftime('%d/%m/%Y à %H:%M')
        else:
            event_datetime_str = 'Date non définie'
        
        # Build notification message based on type
        if notification_type == 'assigned':
            title = 'Nouvel événement assigné'
            message = f"Vous avez été assigné à un événement le {event_datetime_str}"
        elif notification_type == '30min_before':
            title = 'Rappel événement'
            message = f"Votre événement commence dans 30 minutes ({event_datetime_str})"
        elif notification_type == '10min_before':
            title = 'Rappel événement'
            message = f"Votre événement commence dans 10 minutes ({event_datetime_str})"
        elif notification_type == '5min_before':
            # Keep for backward compatibility
            title = 'Rappel événement'
            message = f"Votre événement commence dans 5 minutes ({event_datetime_str})"
        else:
            title = 'Notification événement'
            message = f"Événement le {event_datetime_str}"
        
        # Add contact info if available
        if event.contactId:
            contact_name = f"{event.contactId.fname} {event.contactId.lname}".strip()
            if contact_name:
                message += f" - {contact_name}"
        
        # Add comment if available
        if event.comment:
            message += f"\n{event.comment[:100]}"  # Limit comment length
        
        # Create notification data
        notification_data = {
            'type': 'event_notification',
            'event': {
                'id': event.id,
                'datetime': event.datetime.isoformat() if event.datetime else None,
                'contactId': event.contactId.id if event.contactId else None,
                'contactName': f"{event.contactId.fname} {event.contactId.lname}".strip() if event.contactId else None,
                'comment': event.comment or '',
            },
            'notification_type': notification_type,
            'title': title,
            'message': message,
            'minutes_before': minutes_before,
        }
        
        # Send via WebSocket
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[send_event_notification] Sending {notification_type} notification for event {event.id} to user {event.userId.id}")
        
        async_to_sync(channel_layer.group_send)(
            f'notifications_{event.userId.id}',
            {
                'type': 'event_notification',
                'notification': notification_data,
            }
        )
        
        logger.info(f"[send_event_notification] WebSocket event sent to group notifications_{event.userId.id}")
        
        # Also create a database notification for persistence
        try:
            notification_id = uuid.uuid4().hex[:12]
            while Notification.objects.filter(id=notification_id).exists():
                notification_id = uuid.uuid4().hex[:12]
            
            Notification.objects.create(
                id=notification_id,
                user=event.userId,
                type='event',
                title=title,
                message=message,
                event_id=event.id,
                is_read=False,
                data={
                    'notification_type': notification_type,
                    'minutes_before': minutes_before,
                }
            )
        except Exception as e:
            # Log but don't fail - WebSocket notification is more important
            import traceback
            print(f"Error creating database notification for event {event.id}: {str(e)}")
            print(traceback.format_exc())
            
    except Exception as e:
        import traceback
        print(f"Error sending event notification: {str(e)}")
        print(traceback.format_exc())


def send_contact_notification(contact, notification_type='nouveau_client'):
    """
    Send contact notification via WebSocket to users who can see the contact.
    
    Args:
        contact: Contact instance
        notification_type: Type of notification (e.g., 'nouveau_client')
    """
    try:
        # Get all users who can see this contact based on data_access permissions
        users_to_notify = []
        
        # Get contact's assigned users
        contact_users = []
        if contact.teleoperator:
            contact_users.append(contact.teleoperator)
        if contact.confirmateur:
            contact_users.append(contact.confirmateur)
        if contact.creator:
            contact_users.append(contact.creator)
        
        # Get all users with their roles and data_access
        all_user_details = UserDetails.objects.select_related('role_id', 'django_user').all()
        
        for user_details in all_user_details:
            user = user_details.django_user
            if not user_details.role:
                continue
                
            data_access = user_details.role.data_access
            
            # Check if user can see this contact
            can_see_contact = False
            
            if data_access == 'all':
                # User can see all contacts
                can_see_contact = True
            elif data_access == 'own_only':
                # User can only see contacts where they are teleoperator or confirmateur
                if user in contact_users:
                    can_see_contact = True
            elif data_access == 'team_only':
                # User can see contacts from their team
                team_member = user_details.team_memberships.select_related('team').first()
                if team_member:
                    team = team_member.team
                    # Get all users in the same team
                    team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                    # Check if contact's teleoperator/confirmateur/creator is in the same team
                    if (contact.teleoperator and contact.teleoperator.id in team_user_ids) or \
                       (contact.confirmateur and contact.confirmateur.id in team_user_ids) or \
                       (contact.creator and contact.creator.id in team_user_ids) or \
                       (user in contact_users):
                        can_see_contact = True
                else:
                    # User has no team, fall back to own_only behavior
                    if user in contact_users:
                        can_see_contact = True
            
            if can_see_contact:
                users_to_notify.append(user)
        
        # Remove duplicates
        users_to_notify = list(set(users_to_notify))
        
        if not users_to_notify:
            return  # No users to notify
        
        # Build notification message
        contact_name = f"{contact.fname} {contact.lname}".strip() or "Contact"
        title = 'Nouveau client'
        message = f"Le contact {contact_name} est devenu un nouveau client"
        
        # Create database notifications for each user
        # The signal handler will automatically send them via WebSocket
        for user in users_to_notify:
            try:
                notification_id = uuid.uuid4().hex[:12]
                while Notification.objects.filter(id=notification_id).exists():
                    notification_id = uuid.uuid4().hex[:12]
                
                Notification.objects.create(
                    id=notification_id,
                    user=user,
                    type='contact',
                    title=title,
                    message=message,
                    contact_id=contact.id,
                    is_read=False,
                    data={
                        'notification_type': notification_type,
                    }
                )
            except Exception as e:
                import traceback
                print(f"Error creating notification for contact {contact.id} to user {user.id}: {str(e)}")
                print(traceback.format_exc())
                
    except Exception as e:
        import traceback
        print(f"Error sending contact notification: {str(e)}")
        print(traceback.format_exc())


def send_transaction_update_notification(contact, transaction):
    """
    Send notification when a transaction with type 'Ouverture' is updated.
    Notifies all users who can see the contact.
    
    Args:
        contact: Contact instance
        transaction: Transaction instance (should be type 'Ouverture')
    """
    try:
        # Only send notification for 'Ouverture' transactions
        if transaction.type != 'Ouverture':
            return
        
        # Get all users who can see this contact based on data_access permissions
        users_to_notify = []
        
        # Get contact's assigned users
        contact_users = []
        if contact.teleoperator:
            contact_users.append(contact.teleoperator)
        if contact.confirmateur:
            contact_users.append(contact.confirmateur)
        if contact.creator:
            contact_users.append(contact.creator)
        
        # Get all users with their roles and data_access
        all_user_details = UserDetails.objects.select_related('role_id', 'django_user').all()
        
        for user_details in all_user_details:
            user = user_details.django_user
            if not user_details.role:
                continue
                
            data_access = user_details.role.data_access
            
            # Check if user can see this contact
            can_see_contact = False
            
            if data_access == 'all':
                # User can see all contacts
                can_see_contact = True
            elif data_access == 'own_only':
                # User can only see contacts where they are teleoperator or confirmateur
                if user in contact_users:
                    can_see_contact = True
            elif data_access == 'team_only':
                # User can see contacts from their team
                team_member = user_details.team_memberships.select_related('team').first()
                if team_member:
                    team = team_member.team
                    # Get all users in the same team
                    team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                    # Check if contact's teleoperator/confirmateur/creator is in the same team
                    if (contact.teleoperator and contact.teleoperator.id in team_user_ids) or \
                       (contact.confirmateur and contact.confirmateur.id in team_user_ids) or \
                       (contact.creator and contact.creator.id in team_user_ids) or \
                       (user in contact_users):
                        can_see_contact = True
                else:
                    # User has no team, fall back to own_only behavior
                    if user in contact_users:
                        can_see_contact = True
            
            if can_see_contact:
                users_to_notify.append(user)
        
        # Remove duplicates
        users_to_notify = list(set(users_to_notify))
        
        if not users_to_notify:
            return  # No users to notify
        
        # Build notification message
        contact_name = f"{contact.fname} {contact.lname}".strip() or "Contact"
        title = 'Transaction modifiée'
        message = f"La transaction d'ouverture a été modifiée pour le contact {contact_name}"
        
        # Create database notifications for each user
        # The signal handler will automatically send them via WebSocket
        for user in users_to_notify:
            try:
                notification_id = uuid.uuid4().hex[:12]
                while Notification.objects.filter(id=notification_id).exists():
                    notification_id = uuid.uuid4().hex[:12]
                
                Notification.objects.create(
                    id=notification_id,
                    user=user,
                    type='contact',
                    title=title,
                    message=message,
                    contact_id=contact.id,
                    is_read=False,
                    data={
                        'notification_type': 'transaction_updated',
                        'transaction_id': transaction.id,
                    }
                )
            except Exception as e:
                import traceback
                print(f"Error creating transaction update notification for contact {contact.id} to user {user.id}: {str(e)}")
                print(traceback.format_exc())
                
    except Exception as e:
        import traceback
        print(f"Error sending transaction update notification: {str(e)}")
        print(traceback.format_exc())


def send_confirmateur_assignment_notification(contact, confirmateur_user):
    """
    Send notification to confirmateur when they are assigned to a contact with client_default=True status.
    
    Args:
        contact: Contact instance
        confirmateur_user: Django User instance of the confirmateur
    """
    try:
        # Check if contact has a status with client_default=True
        if not contact.status or not contact.status.client_default:
            return  # Contact doesn't have client_default status, no notification needed
        
        # Build notification message
        contact_name = f"{contact.fname} {contact.lname}".strip() or "Contact"
        title = 'Nouveau client'
        message = f"Nouveau client vous a été assigné: {contact_name}"
        
        # Create database notification
        # The signal handler will automatically send it via WebSocket
        try:
            notification_id = uuid.uuid4().hex[:12]
            while Notification.objects.filter(id=notification_id).exists():
                notification_id = uuid.uuid4().hex[:12]
            
            Notification.objects.create(
                id=notification_id,
                user=confirmateur_user,
                type='contact',
                title=title,
                message=message,
                contact_id=contact.id,
                is_read=False,
                data={
                    'notification_type': 'confirmateur_assigned',
                }
            )
        except Exception as e:
            import traceback
            print(f"Error creating notification for confirmateur {confirmateur_user.id} for contact {contact.id}: {str(e)}")
            print(traceback.format_exc())
            
    except Exception as e:
        import traceback
        print(f"Error sending confirmateur assignment notification: {str(e)}")
        print(traceback.format_exc())


def get_client_ip(request):
    """Extract client IP address from request, checking multiple headers"""
    # Check various headers that might contain the real client IP
    # X-Forwarded-For can contain multiple IPs (client, proxy1, proxy2)
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        # Get the first IP (original client) and strip whitespace
        ip = x_forwarded_for.split(',')[0].strip()
        if ip:
            # Remove port if present (e.g., "192.168.1.1:12345" -> "192.168.1.1")
            if ':' in ip and not ip.startswith('['):  # IPv6 addresses are in brackets
                ip = ip.split(':')[0]
            return ip.strip()
    
    # Check X-Real-IP header (used by some proxies)
    x_real_ip = request.META.get('HTTP_X_REAL_IP')
    if x_real_ip:
        ip = x_real_ip.strip()
        if ip:
            # Remove port if present
            if ':' in ip and not ip.startswith('['):
                ip = ip.split(':')[0]
            return ip.strip()
    
    # Check CF-Connecting-IP (Cloudflare)
    cf_connecting_ip = request.META.get('HTTP_CF_CONNECTING_IP')
    if cf_connecting_ip:
        ip = cf_connecting_ip.strip()
        if ip:
            # Remove port if present
            if ':' in ip and not ip.startswith('['):
                ip = ip.split(':')[0]
            return ip.strip()
    
    # Fallback to REMOTE_ADDR
    ip = request.META.get('REMOTE_ADDR', '')
    if ip:
        # Remove port if present
        if ':' in ip and not ip.startswith('['):
            ip = ip.split(':')[0]
    return ip.strip() if ip else 'Unknown'


def normalize_ip(ip_string):
    """Normalize IP address by removing port, whitespace, and converting to lowercase"""
    if not ip_string:
        return None
    ip = ip_string.strip().lower()
    # Remove port if present (e.g., "192.168.1.1:12345" -> "192.168.1.1")
    # But preserve IPv6 format (e.g., "[::1]:12345" -> "::1")
    if ':' in ip and not ip.startswith('['):
        # IPv4 with port
        ip = ip.split(':')[0]
    elif ip.startswith('[') and ']:' in ip:
        # IPv6 with port
        ip = ip.split(']:')[0][1:]  # Remove brackets
    return ip

def is_ip_allowed(client_ip, ip_whitelist):
    """Check if client IP is in the whitelist"""
    import logging
    logger = logging.getLogger(__name__)
    
    # If whitelist is empty or invalid, deny access (whitelist is enabled, so empty = deny all)
    if not ip_whitelist or not isinstance(ip_whitelist, list) or len(ip_whitelist) == 0:
        logger.warning(f"IP whitelist check: Whitelist is empty or invalid, denying access")
        return False, "IP whitelist is empty"
    
    # If client IP is unknown or empty, deny access when whitelist is enabled
    if not client_ip or client_ip.strip().lower() == 'unknown':
        logger.warning(f"IP whitelist check: Client IP is unknown or empty")
        return False, "Client IP could not be determined"
    
    # Normalize IPs for comparison
    normalized_client_ip = normalize_ip(client_ip)
    normalized_whitelist = [normalize_ip(ip) for ip in ip_whitelist if ip]
    
    # Remove None values from whitelist
    normalized_whitelist = [ip for ip in normalized_whitelist if ip]
    
    if not normalized_client_ip:
        logger.warning(f"IP whitelist check: Could not normalize client IP '{client_ip}'")
        return False, f"Could not normalize client IP: {client_ip}"
    
    logger.info(f"IP whitelist check: Client IP='{normalized_client_ip}' (original: '{client_ip}'), Whitelist={normalized_whitelist}")
    
    # Check exact match first
    if normalized_client_ip in normalized_whitelist:
        logger.info(f"IP whitelist check: Exact match found for {normalized_client_ip}")
        return True, None
    
    # Check CIDR notation (e.g., 192.168.1.0/24)
    import ipaddress
    try:
        client_ip_obj = ipaddress.ip_address(normalized_client_ip)
        for whitelist_ip in normalized_whitelist:
            try:
                # Try to parse as CIDR (works for both CIDR and single IPs)
                network = ipaddress.ip_network(whitelist_ip, strict=False)
                if client_ip_obj in network:
                    logger.info(f"IP whitelist check: CIDR match found - {normalized_client_ip} in {whitelist_ip}")
                    return True, None
            except ValueError as e:
                # Not a valid IP/CIDR format, skip
                logger.warning(f"IP whitelist check: Invalid IP format '{whitelist_ip}': {e}")
                continue
    except ValueError as e:
        # Invalid client IP format, deny access
        logger.warning(f"IP whitelist check: Invalid client IP format '{normalized_client_ip}': {e}")
        return False, f"Invalid client IP format: {normalized_client_ip}"
    
    logger.warning(f"IP whitelist check: Access denied - '{normalized_client_ip}' not in whitelist {normalized_whitelist}")
    return False, f"IP '{normalized_client_ip}' is not in the whitelist"


def get_browser_info(request):
    """Extract browser information from request headers"""
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    # Parse browser info from user agent
    browser_info = {
        'user_agent': user_agent,
    }
    
    # Try to extract browser name and version
    if user_agent:
        user_agent_lower = user_agent.lower()
        
        # Detect browser
        if 'chrome' in user_agent_lower and 'edg' not in user_agent_lower:
            browser_info['browser'] = 'Chrome'
            # Extract Chrome version
            try:
                chrome_index = user_agent_lower.find('chrome/')
                if chrome_index != -1:
                    version_part = user_agent[chrome_index + 7:chrome_index + 20]
                    version = version_part.split()[0].split('.')[0]
                    browser_info['browser_version'] = version
            except:
                pass
        elif 'firefox' in user_agent_lower:
            browser_info['browser'] = 'Firefox'
            try:
                firefox_index = user_agent_lower.find('firefox/')
                if firefox_index != -1:
                    version_part = user_agent[firefox_index + 8:firefox_index + 20]
                    version = version_part.split()[0].split('.')[0]
                    browser_info['browser_version'] = version
            except:
                pass
        elif 'safari' in user_agent_lower and 'chrome' not in user_agent_lower:
            browser_info['browser'] = 'Safari'
        elif 'edg' in user_agent_lower:
            browser_info['browser'] = 'Edge'
        elif 'opera' in user_agent_lower or 'opr' in user_agent_lower:
            browser_info['browser'] = 'Opera'
        
        # Detect OS
        if 'windows' in user_agent_lower:
            browser_info['os'] = 'Windows'
            if 'windows nt 10.0' in user_agent_lower:
                browser_info['os_version'] = '10'
            elif 'windows nt 11.0' in user_agent_lower:
                browser_info['os_version'] = '11'
        elif 'mac' in user_agent_lower or 'macintosh' in user_agent_lower:
            browser_info['os'] = 'macOS'
        elif 'linux' in user_agent_lower:
            browser_info['os'] = 'Linux'
        elif 'android' in user_agent_lower:
            browser_info['os'] = 'Android'
        elif 'ios' in user_agent_lower or 'iphone' in user_agent_lower or 'ipad' in user_agent_lower:
            browser_info['os'] = 'iOS'
    
    return browser_info


def get_user_data_for_log(django_user, user_details=None):
    """Helper function to extract user data for logging"""
    user_data = {
        'id': str(django_user.id),
        'username': django_user.username,
        'email': django_user.email or '',
        'first_name': django_user.first_name or '',
        'last_name': django_user.last_name or '',
    }
    
    # Get UserDetails if not provided
    if user_details is None:
        try:
            user_details = UserDetails.objects.get(django_user=django_user)
        except UserDetails.DoesNotExist:
            return user_data
    
    if user_details:
        user_data['user_details_id'] = user_details.id
        user_data['role'] = user_details.role.id if user_details.role else None
        user_data['roleName'] = user_details.role.name if user_details.role else None
        if user_details.phone:
            user_data['phone'] = user_details.phone
        
        # Get team ID if user is in a team
        team_member = user_details.team_memberships.first()
        if team_member:
            user_data['teamId'] = team_member.team.id
    
    return user_data


def get_team_data_for_log(team):
    """Helper function to extract team data for logging"""
    team_data = {
        'id': team.id,
        'name': team.name,
    }
    
    # Get team members count
    team_members_count = team.team_members.count()
    if team_members_count > 0:
        team_data['members_count'] = team_members_count
    
    return team_data


def serialize_for_json(obj):
    """Convert datetime and date objects to strings for JSON serialization"""
    if isinstance(obj, dict):
        return {key: serialize_for_json(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [serialize_for_json(item) for item in obj]
    elif isinstance(obj, (datetime, date)):
        return obj.isoformat()
    else:
        return obj

def clean_contact_data_for_log(contact_data, include_created_at=False):
    """Clean contact data for log storage - keep only relevant fields for users
    Always includes all fields even if empty to maintain consistent structure
    """
    if not isinstance(contact_data, dict):
        return contact_data
    
    # Fields to keep in specific order (only camelCase, user-friendly fields)
    # Excluded: teleoperatorId, statusColor, fullName, sourceId, statusId
    # Using statusName instead of statusId
    field_order = [
        'firstName',
        'lastName',
        'mobile',
        'source',  # Show source name instead of sourceId
        'statusName',  # Show status name instead of statusId
        'teleoperatorName',
        'creatorName',
        'confirmateurName',
        'civility',
        'email',
        'phone',
        'birthDate',
        'birthPlace',
        'nationality',
        'address',
        'addressComplement',
        'postalCode',
        'city',
        'campaign',
    ]
    
    # Add createdAt only if requested (for old_value, not for new_value)
    if include_created_at:
        field_order.append('createdAt')
    
    cleaned_data = {}
    for field in field_order:
        # Always include field, even if empty, to maintain full structure
        if field in contact_data:
            value = contact_data[field]
            # Convert None to empty string for consistency
            cleaned_data[field] = value if value is not None else ''
        else:
            # Field not present in data, set to empty string
            cleaned_data[field] = ''
    
    return cleaned_data

def compute_changed_fields(old_value, new_value):
    """Compute only the fields that changed between old_value and new_value
    Returns a dictionary with only changed fields, showing old and new values
    """
    if not isinstance(old_value, dict) or not isinstance(new_value, dict):
        return {}
    
    changes = {}
    
    # Get all unique keys from both dictionaries
    all_keys = set(old_value.keys()) | set(new_value.keys())
    
    for key in all_keys:
        old_val = old_value.get(key, '')
        new_val = new_value.get(key, '')
        
        # Normalize values for comparison (handle None, empty strings, etc.)
        old_val_normalized = old_val if old_val is not None else ''
        new_val_normalized = new_val if new_val is not None else ''
        
        # Compare normalized values
        if str(old_val_normalized) != str(new_val_normalized):
            changes[key] = {
                'old': old_val_normalized,
                'new': new_val_normalized
            }
    
    return changes

def create_log_entry(event_type, user_id, request, old_value=None, new_value=None, contact_id=None):
    """Create a log entry for an activity"""
    try:
        print(f"[LOG ENTRY] Creating log entry: event_type={event_type}, contact_id={contact_id.id if contact_id else None}")
        # Generate log ID
        log_id = uuid.uuid4().hex[:12]
        while Log.objects.filter(id=log_id).exists():
            log_id = uuid.uuid4().hex[:12]
        
        # Extract details from request
        details = {
            'ip_address': get_client_ip(request),
            'browser': get_browser_info(request),
        }
        
        # Serialize old_value and new_value to handle datetime objects
        serialized_old_value = serialize_for_json(old_value) if old_value else {}
        serialized_new_value = serialize_for_json(new_value) if new_value else {}
        
        print(f"[LOG ENTRY] Serialized data - old_value keys: {list(serialized_old_value.keys()) if serialized_old_value else []}, new_value keys: {list(serialized_new_value.keys()) if serialized_new_value else []}")
        
        # Create log entry
        log = Log.objects.create(
            id=log_id,
            event_type=event_type,
            user_id=user_id if user_id else None,
            contact_id=contact_id if contact_id else None,
            details=details,
            old_value=serialized_old_value,
            new_value=serialized_new_value
        )
        print(f"[LOG ENTRY] Log entry created successfully: id={log.id}, event_type={log.event_type}, contact_id={log.contact_id.id if log.contact_id else None}")
    except Exception as e:
        # Log the error but don't fail the request
        import traceback
        error_details = traceback.format_exc()
        print(f"[LOG ENTRY] ERROR creating log entry for {event_type}: {str(e)}")
        print(f"[LOG ENTRY] Error details: {error_details}")


class UserCreateView(generics.CreateAPIView):
    queryset = DjangoUser.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]
    
    def perform_create(self, serializer):
        # Save the user (this will trigger the serializer's create method)
        user = serializer.save()
        
        # Get the user who created this (if authenticated, otherwise None)
        created_by_user = self.request.user if self.request.user.is_authenticated else None
        
        # Prepare new_value with user data using helper function
        new_value = get_user_data_for_log(user)
        
        # Create log entry
        create_log_entry(
            event_type='createUser',
            user_id=created_by_user,
            request=self.request,
            old_value={},  # No old value for creation
            new_value=new_value
        )

class NoteListCreateView(generics.ListCreateAPIView):
    serializer_class = NoteSerializer
    permission_classes = [IsAuthenticated]
    
    # Request-level cache for permissions (cleared after each request)
    _request_cache = {}

    def get_user_accessible_category_ids(self, user):
        """Get list of note category IDs the user has view permission for"""
        # Use request-level cache to avoid re-querying permissions on every request
        cache_key = f"accessible_categories_{user.id}"
        if hasattr(self, '_request_cache') and cache_key in self._request_cache:
            return self._request_cache[cache_key]
        
        try:
            # Optimize query with select_related and prefetch_related
            user_details = UserDetails.objects.select_related('role_id').prefetch_related(
                'role_id__permission_roles__permission'
            ).get(django_user=user)
            
            if not user_details.role_id:
                # No role - return empty list (no access)
                result = []
            else:
                role = user_details.role_id
                
                # Get all permission roles for this role - already prefetched, so no additional query
                permission_roles = role.permission_roles.all()
                
                # Build sets of permission IDs for faster lookup
                category_permission_field_names = set()
                
                for perm_role in permission_roles:
                    # Permission is already prefetched, so this won't cause additional queries
                    perm = perm_role.permission
                    if (perm.component == 'note_categories' and 
                        perm.action == 'view' and 
                        perm.field_name is None and 
                        perm.status is None):
                        # User has general permission - can see all categories
                        result = None
                        break
                    elif (perm.component == 'note_categories' and 
                          perm.action == 'view' and 
                          perm.field_name is not None):
                        # Specific category permission
                        category_permission_field_names.add(perm.field_name)
                
                # If we have specific category permissions, return them
                if category_permission_field_names:
                    # Ensure all category IDs are strings and strip whitespace (NoteCategory.id is CharField)
                    result = [str(cat_id).strip() for cat_id in category_permission_field_names if cat_id]
                    # Log for debugging - verify all categories are included
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.debug(f"User {user.username} has access to {len(result)} note categories: {result}")
                else:
                    # No permissions found - return empty list
                    result = []
        except UserDetails.DoesNotExist:
            # User has no UserDetails - return empty list (no access)
            result = []
        except Exception as e:
            # If there's any error (e.g., NoteCategory doesn't exist yet), allow all notes
            # This prevents errors when the database hasn't been migrated yet
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Error checking note category permissions: {e}")
            result = None  # Allow all notes if there's an error
        
        # Cache the result for this request
        if not hasattr(self, '_request_cache'):
            self._request_cache = {}
        self._request_cache[cache_key] = result
        return result

    def get_queryset(self):
        # Allow filtering by contactId if provided as query parameter
        contact_id = self.request.query_params.get('contactId', None)
        user = self.request.user
        
        try:
            # Optimize query with select_related to avoid N+1 queries
            queryset = Note.objects.select_related('userId', 'categ_id', 'contactId').order_by('-created_at')
        except Exception as e:
            # If select_related fails (e.g., categ_id field doesn't exist yet), use basic queryset
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Error with select_related on categ_id: {e}")
            try:
                queryset = Note.objects.select_related('userId', 'contactId').order_by('-created_at')
            except:
                queryset = Note.objects.select_related('userId').order_by('-created_at')
        
        if contact_id:
            queryset = queryset.filter(contactId=contact_id)
        else:
            # If no contactId, return current user's notes (for backward compatibility)
            queryset = queryset.filter(userId=user)
        
        # No permission filtering - return all notes
        # Permissions are enforced at the tab level in the frontend
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """
        Override list to return all notes without pagination when contactId is provided.
        This ensures the popover shows all notes for a contact.
        """
        contact_id = request.query_params.get('contactId', None)
        
        # If contactId is provided, return all notes without pagination
        if contact_id:
            queryset = self.get_queryset()
            # filter_queryset() applies any additional filter backends (if any)
            queryset = self.filter_queryset(queryset)
            
            # With CONN_MAX_AGE: 0, we need to evaluate the queryset before serialization
            # to ensure all related fields (select_related) are loaded before connection closes
            # Convert to list to evaluate queryset immediately while connection is still open
            notes_list = list(queryset)
            
            # Now serialize the evaluated list (no database access needed)
            serializer = self.get_serializer(notes_list, many=True)
            return Response(serializer.data)
        
        # Otherwise, use default pagination behavior
        return super().list(request, *args, **kwargs)

    def perform_create(self, serializer):
        # serializer is already validated at this point
        # contactId can be null if not provided - preserve it from validated_data
        # If not provided, it will be None/null which is allowed
        validated_data = serializer.validated_data
        note_id = validated_data.get('id')
        # Generate a unique ID if not provided
        if not note_id:
            # Generate a 12-character unique ID
            note_id = uuid.uuid4().hex[:12]
            # Ensure uniqueness (small chance of collision, but unlikely)
            while Note.objects.filter(id=note_id).exists():
                note_id = uuid.uuid4().hex[:12]
        serializer.save(
            id=note_id,
            userId=self.request.user, 
            contactId=validated_data.get('contactId'),  # Can be None/null
            categ_id=validated_data.get('categ_id')  # Can be None/null
        )

class NoteUpdateView(generics.UpdateAPIView):
    serializer_class = NoteSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    lookup_url_kwarg = 'pk'
    
    def get_queryset(self):
        user = self.request.user
        return Note.objects.filter(userId=user)
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}
        
        return Response(serializer.data)

class NoteDeleteView(generics.DestroyAPIView):
    serializer_class = NoteSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    lookup_url_kwarg = 'pk'
    
    def get_queryset(self):
        user = self.request.user
        return Note.objects.filter(userId=user)

class ContactView(generics.ListAPIView):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [IsAuthenticated]  # Explicitly set permission
    
    def get_queryset(self):
        # If we have a filtered queryset stored, use it (for pagination counting)
        # This is set in list() method after applying filters
        if hasattr(self, '_filtered_queryset') and self._filtered_queryset is not None:
            return self._filtered_queryset
        
        # Check if all_contacts=true parameter is provided (for admin views)
        all_contacts = self.request.query_params.get('all_contacts', 'false').lower() == 'true'
        if all_contacts:
            # Bypass permission filtering and return all contacts
            queryset = Contact.objects.all()
        else:
            """
            Filter contacts based on user's role data_access level:
            - own_only: 
                - If user is teleoperateur: Only contacts where user is teleoperator
                - If user is confirmateur: Only contacts where user is confirmateur
                - Otherwise: Contacts where user is teleoperator, confirmateur, or creator
            - team_only: Contacts where user is assigned OR contacts from users in the same team
            - all: All contacts (no filtering)
            """
            queryset = Contact.objects.all()
            user = self.request.user
            
            # Get user's role and data_access level
            try:
                user_details = UserDetails.objects.select_related('role_id').get(django_user=user)
                if user_details.role:
                    data_access = user_details.role.data_access
                    
                    if data_access == 'own_only':
                        # Only show contacts where user is linked as teleoperateur OR confirmateur
                        # Do not include creator field - only teleoperateur and confirmateur
                        queryset = queryset.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user)
                        )
                    elif data_access == 'team_only':
                        # Get user's team members
                        team_member = user_details.team_memberships.select_related('team').first()
                        if team_member:
                            team = team_member.team
                            # Get all users in the same team
                            team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                            # Show contacts where:
                            # - User is teleoperator, confirmateur, or creator
                            # - OR contact's teleoperator/confirmateur/creator is in the same team
                            queryset = queryset.filter(
                                models.Q(teleoperator=user) |
                                models.Q(confirmateur=user) |
                                models.Q(creator=user) |
                                models.Q(teleoperator__id__in=team_user_ids) |
                                models.Q(confirmateur__id__in=team_user_ids) |
                                models.Q(creator__id__in=team_user_ids)
                            )
                        else:
                            # User has no team, fall back to own_only behavior
                            queryset = queryset.filter(
                                models.Q(teleoperator=user) |
                                models.Q(confirmateur=user) |
                                models.Q(creator=user)
                            )
                    # If data_access is 'all', show all contacts (no filtering)
            except UserDetails.DoesNotExist:
                # If user has no UserDetails, show no contacts (safety default)
                queryset = Contact.objects.none()
        
        # Optimize queries with select_related for ForeignKey relationships
        # This prevents N+1 queries when accessing related objects
        queryset = queryset.select_related(
            'status',
            'source',
            'teleoperator',
            'teleoperator__user_details',  # Select user_details to avoid query when accessing it
            'confirmateur',
            'creator'
        )
        
        # Prefetch related team_memberships for teleoperator's user_details
        # This optimizes the serializer's access to managerTeamId and managerTeamName
        from django.db.models import Prefetch, Subquery, OuterRef
        from .models import Log
        
        # Annotate with most recent log date to avoid N+1 queries
        # Explicitly set output_field to ensure proper datetime type for sorting
        latest_log = Log.objects.filter(
            contact_id=OuterRef('pk')
        ).order_by('-created_at').values('created_at')[:1]
        
        queryset = queryset.annotate(
            last_log_date=Subquery(latest_log, output_field=models.DateTimeField())
        )
        
        queryset = queryset.prefetch_related(
            Prefetch(
                'teleoperator__user_details__team_memberships',
                queryset=TeamMember.objects.select_related('team')
            ),
            Prefetch(
                'contact_notes',
                queryset=Note.objects.select_related('categ_id').order_by('-created_at')
            )
        )
        
        # Order by created_at descending (most recent first)
        queryset = queryset.order_by('-created_at')
        
        return queryset
    
    def _apply_filters(self, queryset, request):
        """Helper method to apply all filters to a queryset"""
        # Apply search filter
        search = request.query_params.get('search', '').strip()
        if search:
            # Remove spaces from search value for phone number matching
            phone_search = ''.join(str(search).split())
            
            # Search in combined full name, email, and phone numbers
            queryset = queryset.annotate(
                full_name_search=Concat(
                    Coalesce('fname', Value('')), 
                    Value(' '), 
                    Coalesce('lname', Value(''))
                ),
                phone_str=Cast('phone', CharField()),
                mobile_str=Cast('mobile', CharField())
            ).filter(
                models.Q(full_name_search__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(phone_str__contains=phone_search) |
                models.Q(mobile_str__contains=phone_search)
            )
        
        # Apply team filter (filter by teleoperator's team)
        team_id = request.query_params.get('team')
        if team_id and team_id != 'all':
            # Get team members
            team_user_ids = TeamMember.objects.filter(team_id=team_id).values_list('user__django_user__id', flat=True)
            queryset = queryset.filter(
                models.Q(teleoperator__id__in=team_user_ids) |
                models.Q(confirmateur__id__in=team_user_ids) |
                models.Q(creator__id__in=team_user_ids)
            )
        
        # Apply status type filter
        status_type = request.query_params.get('status_type')
        if status_type and status_type != 'all':
            queryset = queryset.filter(status__type=status_type)
        
        # Apply column filters
        # First, collect date range filters separately
        date_range_filters = {}
        # Collect multi-select filters (same key can appear multiple times)
        multi_select_filters = {}
        
        # Debug: Print all filter parameters
        filter_params = {k: request.query_params.getlist(k) if k.startswith('filter_') and k.replace('filter_', '') in ['status', 'creator', 'teleoperator', 'confirmateur', 'source', 'postalCode', 'nationality', 'campaign', 'civility', 'managerTeam'] else request.query_params.get(k) for k in request.query_params.keys() if k.startswith('filter_')}
        if filter_params:
            print(f"[DEBUG] Received filter parameters: {filter_params}")
            print(f"[DEBUG] Full query string: {request.META.get('QUERY_STRING', '')}")
        
        # Process all filter parameters
        for key in request.query_params.keys():
            if key.startswith('filter_'):
                if key.endswith('_from') or key.endswith('_to'):
                    # This is a date range filter
                    base_column = key.replace('filter_', '').replace('_from', '').replace('_to', '')
                    value = request.query_params.get(key)
                    if value:
                        if base_column not in date_range_filters:
                            date_range_filters[base_column] = {}
                        if key.endswith('_from'):
                            date_range_filters[base_column]['from'] = value
                        elif key.endswith('_to'):
                            date_range_filters[base_column]['to'] = value
                else:
                    # Regular filter - check if it's a multi-select column
                    column_id = key.replace('filter_', '')
                    # Multi-select columns: status, creator, teleoperator, confirmateur, source, postalCode, nationality, campaign, civility, managerTeam
                    if column_id in ['status', 'creator', 'teleoperator', 'confirmateur', 'source', 'postalCode', 'nationality', 'campaign', 'civility', 'managerTeam']:
                        # Use getlist to get all values for this key
                        values = request.query_params.getlist(key)
                        if values:
                            print(f"[DEBUG] Found multi-select filter for '{column_id}': {values}")
                            multi_select_filters[column_id] = values
                    else:
                        # Single value filter
                        value = request.query_params.get(key)
                        if value:
                            value = value.strip()  # Strip whitespace
                            if not value:  # Skip empty strings after stripping
                                continue
                            
                            count_before = queryset.count()
                            
                            if column_id == 'email':
                                queryset = queryset.filter(email__icontains=value)
                            elif column_id == 'phone':
                                # Remove spaces from search value and convert to string
                                phone_search = ''.join(str(value).split())
                                if phone_search:
                                    # Convert integer phone fields to strings for partial matching
                                    queryset = queryset.annotate(
                                        phone_str=Cast('phone', CharField()),
                                        mobile_str=Cast('mobile', CharField())
                                    ).filter(
                                        models.Q(phone_str__contains=phone_search) | 
                                        models.Q(mobile_str__contains=phone_search)
                                    )
                            elif column_id == 'mobile':
                                # Remove spaces from search value and convert to string
                                mobile_search = ''.join(str(value).split())
                                if mobile_search:
                                    # Convert integer mobile field to string for partial matching
                                    queryset = queryset.annotate(
                                        mobile_str=Cast('mobile', CharField())
                                    ).filter(
                                        models.Q(mobile_str__contains=mobile_search)
                                    )
                            elif column_id == 'fullName':
                                # Search in the combined full name string (first name + space + last name)
                                # Handle NULL values with Coalesce to avoid NULL results
                                # Use unique annotation name to avoid conflicts
                                queryset = queryset.annotate(
                                    full_name_search=Concat(
                                        Coalesce('fname', Value('')), 
                                        Value(' '), 
                                        Coalesce('lname', Value(''))
                                    )
                                ).filter(full_name_search__icontains=value)
                            elif column_id == 'firstName':
                                queryset = queryset.filter(fname__icontains=value)
                            elif column_id == 'lastName':
                                queryset = queryset.filter(lname__icontains=value)
                            elif column_id == 'city':
                                queryset = queryset.filter(city__icontains=value)
                            elif column_id == 'address':
                                queryset = queryset.filter(address__icontains=value)
                            elif column_id == 'addressComplement':
                                queryset = queryset.filter(address_complement__icontains=value)
                            elif column_id == 'birthPlace':
                                queryset = queryset.filter(birth_place__icontains=value)
                            elif column_id == 'campaign':
                                queryset = queryset.filter(campaign__icontains=value)
                            elif column_id == 'oldContactId':
                                queryset = queryset.filter(old_contact_id__icontains=value)
                            else:
                                # Unknown filter column, skip it
                                continue
                            
                            count_after = queryset.count()
                            print(f"[DEBUG] Filter '{column_id}' applied with value '{value}': {count_before} -> {count_after} contacts")
                            # Add more column filters as needed
        
        # Apply multi-select filters
        for column_id, values in multi_select_filters.items():
            if values:
                # Strip and filter out empty strings
                values = [v.strip() if isinstance(v, str) else str(v).strip() for v in values if v and str(v).strip()]
                
                if not values:
                    continue
                
                # Check if empty option is selected
                has_empty = '__empty__' in values
                # Filter out empty option from values for regular filtering
                regular_values = [v for v in values if v != '__empty__']
                
                # Debug logging
                print(f"[DEBUG] Applying filter for column '{column_id}': values={values}, regular_values={regular_values}, has_empty={has_empty}")
                
                # Build Q objects for filtering
                q_objects = []
                
                # If only empty is selected (has_empty=True and regular_values is empty), we should ONLY show empty/null
                # This is a special case that needs to be handled first
                if has_empty and not regular_values:
                    print(f"[DEBUG] Only empty option selected for '{column_id}' - will filter for null/empty only")
                
                # Add regular value filters if any
                # IMPORTANT: If regular_values exist, we should filter by them
                if regular_values:
                    if column_id == 'status':
                        # Filter by current status (status_id field) - NOT previous status from logs
                        # This ensures we only show contacts with the most recent/latest status
                        q_objects.append(models.Q(status_id__in=regular_values))
                    elif column_id == 'source':
                        print(f"[DEBUG] Filtering by source_id__in={regular_values}")
                        # Check if source IDs exist in database
                        from api.models import Source
                        existing_source_ids = list(Source.objects.filter(id__in=regular_values).values_list('id', flat=True))
                        print(f"[DEBUG] Found {len(existing_source_ids)} matching sources in DB: {existing_source_ids}")
                        if existing_source_ids:
                            q_objects.append(models.Q(source_id__in=existing_source_ids))
                        else:
                            print(f"[DEBUG] WARNING: No matching sources found in database for IDs: {regular_values}")
                            # Still apply filter to return empty result
                            q_objects.append(models.Q(source_id__in=regular_values))
                    elif column_id == 'teleoperator':
                        # Convert UserDetails IDs to Django User IDs (frontend sends UserDetails IDs)
                        print(f"[DEBUG] ===== TELEOPERATOR FILTER START =====")
                        print(f"[DEBUG] Teleoperator filter - regular_values: {regular_values} (type: {[type(v) for v in regular_values]})")
                        django_user_ids = []
                        
                        # Try bulk lookup first for UserDetails IDs (more efficient)
                        try:
                            # Convert all values to strings for UserDetails lookup
                            user_details_ids_str = [str(v) for v in regular_values]
                            print(f"[DEBUG] Looking up UserDetails with IDs: {user_details_ids_str}")
                            user_details_list = list(UserDetails.objects.filter(id__in=user_details_ids_str).select_related('django_user'))
                            print(f"[DEBUG] Found {len(user_details_list)} UserDetails records")
                            
                            # Extract Django User IDs from found UserDetails
                            for user_details in user_details_list:
                                if user_details and user_details.django_user:
                                    django_user_ids.append(user_details.django_user.id)
                                    print(f"[DEBUG] Found UserDetails ID {user_details.id} -> Django User ID {user_details.django_user.id}")
                            
                            # For any values not found as UserDetails IDs, try as Django User IDs directly
                            found_user_details_ids = {ud.id for ud in user_details_list}
                            print(f"[DEBUG] Found UserDetails IDs: {found_user_details_ids}")
                            for user_details_id in regular_values:
                                if str(user_details_id) not in found_user_details_ids:
                                    # Try as Django User ID (if numeric)
                                    try:
                                        int_id = int(user_details_id)
                                        if DjangoUser.objects.filter(id=int_id).exists():
                                            django_user_ids.append(int_id)
                                            print(f"[DEBUG] Found Django User ID directly: {int_id}")
                                        else:
                                            print(f"[DEBUG] Django User ID {int_id} does not exist")
                                    except (ValueError, TypeError) as e:
                                        print(f"[DEBUG] Could not convert '{user_details_id}' to integer: {e}")
                        except Exception as e:
                            print(f"[DEBUG] Error in bulk UserDetails lookup: {e}")
                            import traceback
                            traceback.print_exc()
                            # Fallback to individual lookups
                            for user_details_id in regular_values:
                                try:
                                    # Try as UserDetails ID first (string)
                                    user_details = UserDetails.objects.filter(id=str(user_details_id)).first()
                                    if user_details and user_details.django_user:
                                        django_user_ids.append(user_details.django_user.id)
                                        print(f"[DEBUG] Found UserDetails ID {user_details.id} -> Django User ID {user_details.django_user.id} (individual lookup)")
                                    else:
                                        # Fallback: try as Django User ID (if numeric)
                                        try:
                                            int_id = int(user_details_id)
                                            if DjangoUser.objects.filter(id=int_id).exists():
                                                django_user_ids.append(int_id)
                                                print(f"[DEBUG] Found Django User ID directly: {int_id} (individual lookup)")
                                        except (ValueError, TypeError):
                                            pass
                                except Exception as e:
                                    print(f"[DEBUG] Error converting teleoperator UserDetails ID {user_details_id}: {e}")
                        
                        # Ensure all IDs are integers and remove duplicates
                        django_user_ids = list(set([int(uid) for uid in django_user_ids if uid is not None]))
                        
                        if django_user_ids:
                            q_objects.append(models.Q(teleoperator_id__in=django_user_ids))
                            print(f"[DEBUG] Teleoperator filter SUCCESS: converted {len(regular_values)} values to {len(django_user_ids)} Django User IDs: {django_user_ids}")
                        else:
                            print(f"[DEBUG] WARNING: No valid teleoperator IDs found for filter values: {regular_values}")
                            # Return empty queryset by filtering with impossible condition
                            q_objects.append(models.Q(teleoperator_id__in=[]))
                        print(f"[DEBUG] ===== TELEOPERATOR FILTER END =====")
                    elif column_id == 'confirmateur':
                        # Convert UserDetails IDs to Django User IDs (frontend sends UserDetails IDs)
                        django_user_ids = []
                        for user_details_id in regular_values:
                            try:
                                # Try as UserDetails ID first (string)
                                user_details = UserDetails.objects.filter(id=str(user_details_id)).first()
                                if user_details and user_details.django_user:
                                    django_user_ids.append(user_details.django_user.id)
                                else:
                                    # Fallback: try as Django User ID (if numeric)
                                    try:
                                        int_id = int(user_details_id)
                                        if DjangoUser.objects.filter(id=int_id).exists():
                                            django_user_ids.append(int_id)
                                    except (ValueError, TypeError):
                                        pass
                            except Exception as e:
                                print(f"[DEBUG] Error converting confirmateur UserDetails ID {user_details_id}: {e}")
                                # Fallback: try as Django User ID
                                try:
                                    int_id = int(user_details_id)
                                    if DjangoUser.objects.filter(id=int_id).exists():
                                        django_user_ids.append(int_id)
                                except (ValueError, TypeError):
                                    pass
                        
                        if django_user_ids:
                            q_objects.append(models.Q(confirmateur_id__in=django_user_ids))
                            print(f"[DEBUG] Confirmateur filter: converted {len(regular_values)} UserDetails IDs to {len(django_user_ids)} Django User IDs: {django_user_ids}")
                        else:
                            print(f"[DEBUG] WARNING: No valid confirmateur IDs found for filter values: {regular_values}")
                            # Return empty queryset by filtering with impossible condition
                            q_objects.append(models.Q(confirmateur_id__in=[]))
                    elif column_id == 'creator':
                        # Convert UserDetails IDs to Django User IDs (frontend sends UserDetails IDs)
                        django_user_ids = []
                        for user_details_id in regular_values:
                            try:
                                # Try as UserDetails ID first (string)
                                user_details = UserDetails.objects.filter(id=str(user_details_id)).first()
                                if user_details and user_details.django_user:
                                    django_user_ids.append(user_details.django_user.id)
                                else:
                                    # Fallback: try as Django User ID (if numeric)
                                    try:
                                        int_id = int(user_details_id)
                                        if DjangoUser.objects.filter(id=int_id).exists():
                                            django_user_ids.append(int_id)
                                    except (ValueError, TypeError):
                                        pass
                            except Exception as e:
                                print(f"[DEBUG] Error converting creator UserDetails ID {user_details_id}: {e}")
                                # Fallback: try as Django User ID
                                try:
                                    int_id = int(user_details_id)
                                    if DjangoUser.objects.filter(id=int_id).exists():
                                        django_user_ids.append(int_id)
                                except (ValueError, TypeError):
                                    pass
                        
                        if django_user_ids:
                            q_objects.append(models.Q(creator_id__in=django_user_ids))
                            print(f"[DEBUG] Creator filter: converted {len(regular_values)} UserDetails IDs to {len(django_user_ids)} Django User IDs: {django_user_ids}")
                        else:
                            print(f"[DEBUG] WARNING: No valid creator IDs found for filter values: {regular_values}")
                            # Return empty queryset by filtering with impossible condition
                            q_objects.append(models.Q(creator_id__in=[]))
                    elif column_id == 'postalCode':
                        q_objects.append(models.Q(postal_code__in=regular_values))
                    elif column_id == 'nationality':
                        q_objects.append(models.Q(nationality__in=regular_values))
                    elif column_id == 'campaign':
                        q_objects.append(models.Q(campaign__in=regular_values))
                    elif column_id == 'civility':
                        q_objects.append(models.Q(civility__in=regular_values))
                    elif column_id == 'managerTeam':
                        # For managerTeam, filter by team memberships
                        q_objects.append(
                            models.Q(teleoperator__user_details__team_memberships__team_id__in=regular_values) |
                            models.Q(confirmateur__user_details__team_memberships__team_id__in=regular_values)
                        )
                    elif column_id == 'previousStatus':
                        # Filter by IMMEDIATE previous status (the status right before the current one)
                        # NOT any status in the history - only the most recent previous status
                        from .models import Log
                        
                        matching_contact_ids = set()
                        print(f"[CONTACT DEBUG] previousStatus filter - regular_values: {regular_values}")
                        
                        # Get all contacts with status change logs
                        contacts_with_logs = Contact.objects.filter(
                            contact_logs__old_value__statusName__isnull=False,
                            contact_logs__new_value__statusName__isnull=False
                        ).exclude(
                            contact_logs__old_value__statusName=models.F('contact_logs__new_value__statusName')
                        ).distinct()
                        
                        print(f"[CONTACT DEBUG] Found {contacts_with_logs.count()} contacts with status change logs")
                        
                        for contact in contacts_with_logs:
                            # Get current status name for this contact
                            current_status_name = contact.status.name if contact.status else ''
                            
                            # Get the most recent status change log where new_status matches current status
                            # This ensures we get the IMMEDIATE previous status (right before current)
                            most_recent_log = Log.objects.filter(
                                contact_id=contact,
                                old_value__statusName__isnull=False,
                                new_value__statusName__isnull=False
                            ).exclude(
                                old_value__statusName=models.F('new_value__statusName')
                            ).order_by('-created_at')
                            
                            # Find the log where new_status matches the current status
                            # IMPORTANT: Match serializer logic exactly - check old_status != new_status and old_status is truthy
                            matching_log = None
                            for log in most_recent_log:
                                if log.old_value and log.new_value:
                                    old_status = log.old_value.get('statusName', '') if log.old_value else ''
                                    new_status = log.new_value.get('statusName', '') if log.new_value else ''
                                    # Match serializer logic: old_status must be truthy, different from new_status, and new_status matches current
                                    if old_status and old_status != new_status and new_status == current_status_name:
                                        matching_log = log
                                        break
                            
                            if matching_log:
                                previous_status = matching_log.old_value.get('statusName', '') if matching_log.old_value else ''
                                # Check if the immediate previous status matches any of the filter values
                                if previous_status in regular_values:
                                    matching_contact_ids.add(contact.id)
                                    print(f"[CONTACT DEBUG] Contact {contact.id} - current: '{current_status_name}', previous: '{previous_status}' - MATCH")
                            elif not current_status_name:
                                # Contact has no current status - check if empty is selected
                                pass
                        
                        print(f"[CONTACT DEBUG] previousStatus filter - total matching_contact_ids: {len(matching_contact_ids)}, has_empty: {has_empty}")
                        
                        # Match preview logic:
                        # - If regularValues.length > 0: matches = regularValues.includes(previousStatus) || (hasEmpty && !previousStatus)
                        # - So we need to add regular values filter, and empty filter will be combined with OR if has_empty is True
                        if matching_contact_ids:
                            q_objects.append(models.Q(id__in=matching_contact_ids))
                            print(f"[CONTACT DEBUG] previousStatus filter - Added Q object with {len(matching_contact_ids)} contact IDs")
                        elif not has_empty:
                            # No matches and no empty option - exclude all (strict filter)
                            print(f"[CONTACT DEBUG] previousStatus filter - No matches found and empty not selected, excluding all contacts")
                            queryset = queryset.none()
                            break
                        # If has_empty is True and matching_contact_ids is empty, don't add anything for regular values
                        # The empty filter (already added above) will be applied, showing only contacts with empty previousStatus
                        elif has_empty:
                            print(f"[CONTACT DEBUG] previousStatus filter - No matches found but empty selected, will show only empty contacts")
                    elif column_id == 'previousTeleoperator':
                        # Filter by previous teleoperator from logs
                        # Similar to previousStatus - find contacts with matching old_value.teleoperatorName
                        from .models import Log
                        from django.db.models import OuterRef, Exists
                        
                        matching_contact_ids = set()
                        for teleoperator_name in regular_values:
                            # Find logs where teleoperator changed and old_value.teleoperatorName matches
                            matching_logs = Log.objects.filter(
                                contact_id=OuterRef('pk'),
                                old_value__teleoperatorName=teleoperator_name,
                                new_value__teleoperatorName__isnull=False
                            ).exclude(
                                old_value__teleoperatorName=models.F('new_value__teleoperatorName')
                            )
                            
                            # Get contacts that have at least one matching log
                            contacts = Contact.objects.filter(
                                Exists(matching_logs)
                            ).values_list('id', flat=True)
                            
                            matching_contact_ids.update(contacts)
                        
                        if matching_contact_ids:
                            q_objects.append(models.Q(id__in=matching_contact_ids))
                        elif not has_empty:
                            # No matches and no empty option - exclude all
                            queryset = queryset.none()
                            break
                
                # Add empty/null filter if empty option is selected
                # If we have regular_values, combine with OR (regular values OR empty)
                # If we only have empty option, filter by null/empty only
                if has_empty:
                    if column_id == 'status':
                        empty_q = models.Q(status_id__isnull=True)
                    elif column_id == 'source':
                        empty_q = models.Q(source_id__isnull=True)
                        print(f"[DEBUG] Created empty_q for source: {empty_q}")
                    elif column_id == 'teleoperator':
                        empty_q = models.Q(teleoperator_id__isnull=True)
                    elif column_id == 'confirmateur':
                        empty_q = models.Q(confirmateur_id__isnull=True)
                    elif column_id == 'creator':
                        empty_q = models.Q(creator_id__isnull=True)
                    elif column_id == 'postalCode':
                        empty_q = models.Q(postal_code__isnull=True) | models.Q(postal_code='')
                    elif column_id == 'nationality':
                        empty_q = models.Q(nationality__isnull=True) | models.Q(nationality='')
                    elif column_id == 'campaign':
                        empty_q = models.Q(campaign__isnull=True) | models.Q(campaign='')
                    elif column_id == 'civility':
                        empty_q = models.Q(civility__isnull=True) | models.Q(civility='')
                    elif column_id == 'managerTeam':
                        # For managerTeam, we need to check if the contact's teleoperator/confirmateur has no team
                        empty_q = (
                            models.Q(teleoperator__isnull=True) |
                            models.Q(teleoperator__user_details__team_memberships__isnull=True) |
                            models.Q(confirmateur__isnull=True) |
                            models.Q(confirmateur__user_details__team_memberships__isnull=True)
                        )
                    elif column_id == 'previousStatus':
                        # Empty previousStatus means no previous status (contact never had a status change)
                        # This is tricky - we need to find contacts that have no status change logs
                        from .models import Log
                        # Get all contact IDs that have status change logs
                        contacts_with_status_changes = Log.objects.filter(
                            contact_id__isnull=False,
                            old_value__statusName__isnull=False,
                            new_value__statusName__isnull=False
                        ).exclude(
                            old_value__statusName=models.F('new_value__statusName')
                        ).values_list('contact_id_id', flat=True).distinct()
                        
                        # Contacts without status changes (empty previousStatus)
                        empty_q = ~models.Q(id__in=contacts_with_status_changes)
                    else:
                        empty_q = None
                        print(f"[DEBUG] WARNING: No empty_q created for column '{column_id}'")
                    
                    if empty_q:
                        if regular_values:
                            # Combine regular values with empty: (regular_values) OR (empty)
                            # This means: show contacts with selected sources OR contacts with no source
                            print(f"[DEBUG] Combining regular values with empty filter for '{column_id}'")
                            q_objects.append(empty_q)
                        else:
                            # Only empty option selected - show ONLY contacts with no source
                            # Clear any existing q_objects and only use empty filter
                            q_objects = [empty_q]
                            print(f"[DEBUG] Only empty option selected for '{column_id}', q_objects set to: {q_objects}")
                    else:
                        print(f"[DEBUG] ERROR: empty_q is None for column '{column_id}' even though has_empty=True")
                
                # Apply combined filter
                print(f"[DEBUG] Final q_objects for '{column_id}': {len(q_objects)} object(s)")
                if q_objects:
                    print(f"[DEBUG] Applying {len(q_objects)} Q object(s) for column '{column_id}': {q_objects}")
                    queryset_before_count = queryset.count()
                    # Sample a few source_ids before filtering to verify
                    if column_id == 'source':
                        sample_before = list(queryset.values_list('source_id', flat=True)[:5])
                        null_count_before = queryset.filter(source_id__isnull=True).count()
                        print(f"[DEBUG] Sample source_ids before filter: {sample_before}")
                        print(f"[DEBUG] Contacts with null source_id before filter: {null_count_before}")
                    
                    if len(q_objects) == 1:
                        # Single filter - apply directly
                        print(f"[DEBUG] Applying single Q object: {q_objects[0]}")
                        queryset = queryset.filter(q_objects[0])
                    else:
                        # Multiple filters - combine with OR (regular values OR empty)
                        combined_q = q_objects[0]
                        for q_obj in q_objects[1:]:
                            combined_q |= q_obj
                        print(f"[DEBUG] Applying combined Q object: {combined_q}")
                        queryset = queryset.filter(combined_q)
                    
                    queryset_after_count = queryset.count()
                    # Sample a few source_ids after filtering to verify
                    if column_id == 'source':
                        sample_after = list(queryset.values_list('source_id', flat=True)[:5])
                        print(f"[DEBUG] Sample source_ids after filter: {sample_after}")
                        # Verify filter correctness
                        if has_empty and not regular_values:
                            # Should only have null source_ids
                            non_null_sources = queryset.exclude(source_id__isnull=True).values_list('source_id', flat=True).distinct()[:5]
                            if non_null_sources:
                                print(f"[DEBUG] ERROR: Found contacts with non-null source_ids when filtering for empty only: {list(non_null_sources)}")
                            else:
                                print(f"[DEBUG] Verified: All filtered contacts have null source_id (empty source)")
                        elif regular_values:
                            # Should only have source_ids in regular_values
                            wrong_sources = queryset.exclude(source_id__in=regular_values).exclude(source_id__isnull=True).values_list('source_id', flat=True).distinct()[:5]
                            if wrong_sources and not has_empty:
                                print(f"[DEBUG] WARNING: Found contacts with wrong source_ids: {list(wrong_sources)}")
                            else:
                                print(f"[DEBUG] Verified: All filtered contacts have source_id in {regular_values} or null")
                    
                    print(f"[DEBUG] Filter '{column_id}' applied: {queryset_before_count} -> {queryset_after_count} contacts")
                else:
                    print(f"[DEBUG] WARNING: No q_objects to apply for column '{column_id}' - filter will NOT be applied!")
        
        # Apply date range filters
        for column_id, date_range in date_range_filters.items():
            try:
                if column_id == 'createdAt':
                    if 'from' in date_range and date_range['from']:
                        date_from = datetime.strptime(date_range['from'], '%Y-%m-%d').date()
                        # Filter contacts created on or after this date (start of day)
                        queryset = queryset.filter(created_at__date__gte=date_from)
                    if 'to' in date_range and date_range['to']:
                        date_to = datetime.strptime(date_range['to'], '%Y-%m-%d').date()
                        # Filter contacts created on or before this date (end of day)
                        # Use __lt with next day to ensure we only include contacts up to end of selected day
                        date_to_next = date_to + timedelta(days=1)
                        queryset = queryset.filter(created_at__date__lt=date_to_next)
                elif column_id == 'updatedAt':
                    if 'from' in date_range and date_range['from']:
                        date_from = datetime.strptime(date_range['from'], '%Y-%m-%d').date()
                        queryset = queryset.filter(updated_at__date__gte=date_from)
                    if 'to' in date_range and date_range['to']:
                        date_to = datetime.strptime(date_range['to'], '%Y-%m-%d').date()
                        # Use __lt with next day to ensure we only include contacts up to end of selected day
                        date_to_next = date_to + timedelta(days=1)
                        queryset = queryset.filter(updated_at__date__lt=date_to_next)
                elif column_id == 'birthDate':
                    if 'from' in date_range and date_range['from']:
                        date_from = datetime.strptime(date_range['from'], '%Y-%m-%d').date()
                        queryset = queryset.filter(birth_date__gte=date_from)
                    if 'to' in date_range and date_range['to']:
                        date_to = datetime.strptime(date_range['to'], '%Y-%m-%d').date()
                        queryset = queryset.filter(birth_date__lte=date_to)
            except (ValueError, TypeError) as e:
                # Invalid date format, skip this filter
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Invalid date format in filter {column_id}: {e}")
                pass
        
        # Apply ordering from query parameter
        order_param = request.query_params.get('order')
        if order_param:
            # Clear any existing ordering from get_queryset
            queryset = queryset.order_by()
            
            if order_param == 'created_at_asc':
                queryset = queryset.order_by('created_at')
            elif order_param == 'created_at_desc':
                queryset = queryset.order_by('-created_at')
            elif order_param == 'updated_at_asc':
                # Sort by updated_at only (modification date)
                # Put NULL values last using Case/When, use -created_at as secondary sort for consistency
                queryset = queryset.annotate(
                    updated_at_null_order=Case(
                        When(updated_at__isnull=True, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ).order_by('updated_at_null_order', 'updated_at', '-created_at')
            elif order_param == 'updated_at_desc':
                # Sort by updated_at only (modification date, newest first)
                # Put NULL values last using Case/When, use -created_at as secondary sort for consistency
                queryset = queryset.annotate(
                    updated_at_null_order=Case(
                        When(updated_at__isnull=True, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ).order_by('updated_at_null_order', '-updated_at', '-created_at')
            elif order_param == 'email_asc':
                queryset = queryset.order_by('email')
            elif order_param == 'assigned_at_asc':
                queryset = queryset.order_by('assigned_at')
            elif order_param == 'assigned_at_desc':
                queryset = queryset.order_by('-assigned_at')
            elif order_param == 'date_lead_to_client_asc':
                # Sort by date_lead_to_client ascending (oldest first)
                # Put NULL values last using Case/When
                queryset = queryset.annotate(
                    date_lead_to_client_null_order=Case(
                        When(date_lead_to_client__isnull=True, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ).order_by('date_lead_to_client_null_order', 'date_lead_to_client', '-created_at')
            elif order_param == 'date_lead_to_client_desc':
                # Sort by date_lead_to_client descending (newest first)
                # Put NULL values last using Case/When
                queryset = queryset.annotate(
                    date_lead_to_client_null_order=Case(
                        When(date_lead_to_client__isnull=True, then=Value(1)),
                        default=Value(0),
                        output_field=IntegerField()
                    )
                ).order_by('date_lead_to_client_null_order', '-date_lead_to_client', '-created_at')
            elif order_param == 'random':
                queryset = queryset.order_by('?')
            else:
                # Fallback to default ordering (creation date, most recent first)
                queryset = queryset.order_by('-created_at')
        # If no order parameter, keep the ordering from get_queryset (-created_at)
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        # Check if pagination is requested (preferred method for large datasets)
        requested_page = request.query_params.get('page')
        requested_page_size = request.query_params.get('page_size')
        
        # Use pagination if page or page_size are provided
        if requested_page or requested_page_size:
            # Use pagination for large datasets
            from rest_framework.pagination import PageNumberPagination
            
            page_size = int(requested_page_size) if requested_page_size else 100
            page = int(requested_page) if requested_page else 1
            
            # Ensure reasonable page size (max 1000 per page)
            # Note: Heroku has a 30-second HTTP timeout, so queries must complete within that time
            # The queryset is optimized with select_related/prefetch_related to minimize database queries
            MAX_PAGE_SIZE = 1000
            if page_size > MAX_PAGE_SIZE:
                page_size = MAX_PAGE_SIZE
            if page_size < 1:
                page_size = 100
            
            # Get base queryset
            queryset = self.get_queryset()
            
            # Apply all filters
            queryset = self._apply_filters(queryset, request)
            
            # CRITICAL: Store the filtered queryset so get_queryset() returns it
            # This ensures DRF pagination uses the filtered queryset for counting
            self._filtered_queryset = queryset
            
            # Use normal pagination with count
            # Create pagination class with captured page_size using closure
            def create_pagination_class(page_size_value):
                class ContactPagination(PageNumberPagination):
                    page_size = page_size_value
                    page_size_query_param = 'page_size'
                    max_page_size = MAX_PAGE_SIZE
                return ContactPagination
            
            self.pagination_class = create_pagination_class(page_size)
            
            try:
                # For medium page sizes (500-999), still optimize with defer
                if page_size >= 500:
                    queryset = queryset.defer('notes', 'logs')
                
                response = super().list(request, *args, **kwargs)
                
                # Use pagination's count directly - it's already optimized
                return Response({
                    'contacts': response.data['results'],
                    'total': response.data['count'],
                    'next': response.data.get('next'),
                    'previous': response.data.get('previous'),
                    'page': page,
                    'page_size': page_size
                })
            finally:
                # Clear the filtered queryset to avoid affecting other requests
                self._filtered_queryset = None
        
        # Check if limit is requested (for backward compatibility)
        limit = request.query_params.get('limit')
        if limit:
            try:
                limit = int(limit)
                # Ensure limit is reasonable (max 500)
                MAX_LIMIT = 500
                if limit > MAX_LIMIT:
                    limit = MAX_LIMIT
                if limit < 1:
                    limit = 1
                # Get base queryset
                queryset = self.get_queryset()
                
                # Apply all filters using helper method
                queryset = self._apply_filters(queryset, request)
                
                # Get actual total count BEFORE applying limit
                # This gives users the real total number of contacts matching their filters
                total_count = queryset.count()
                
                # CRITICAL PERFORMANCE FIX: Apply limit BEFORE serialization
                # This prevents loading thousands of contacts into memory
                queryset = queryset[:limit]
                
                serializer = self.get_serializer(queryset, many=True, context={'request': request})
                return Response({
                    'contacts': serializer.data,
                    'total': total_count,
                    'limit': limit
                })
            except (ValueError, TypeError):
                # Invalid limit, fall through to default behavior
                pass
            except Exception as e:
                # Log the error and return a proper error response
                import traceback
                error_details = traceback.format_exc()
                print(f"Error in ContactView.list with limit: {error_details}")
                return Response({'error': str(e), 'details': error_details}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Fallback: If no pagination params provided, use default pagination with reasonable limit
        # This ensures we don't load all contacts at once (performance protection)
        from rest_framework.pagination import PageNumberPagination
        
        class DefaultContactPagination(PageNumberPagination):
            page_size = 50  # Default page size matching frontend default
            page_size_query_param = 'page_size'
            max_page_size = 1000  # Max 1000 per page
        
        self.pagination_class = DefaultContactPagination
        queryset = self.get_queryset()
        queryset = self._apply_filters(queryset, request)
        
        # Store filtered queryset for proper pagination counting
        self._filtered_queryset = queryset
        
        try:
            response = super().list(request, *args, **kwargs)
            return Response({
                'contacts': response.data['results'],
                'total': response.data['count'],
                'next': response.data.get('next'),
                'previous': response.data.get('previous'),
                'page': response.data.get('page', 1),
                'page_size': response.data.get('page_size', 50)
            })
        finally:
            self._filtered_queryset = None


class FosseContactView(generics.ListAPIView):
    """
    View for "Fosse" page - shows all contacts that are not assigned to anyone
    (teleoperator is null AND confirmateur is null).
    No permission filtering - shows all unassigned contacts to authenticated users.
    """
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    permission_classes = [IsAuthenticated]
    
    def get_serializer_context(self):
        """Override to inject cached FosseSettings into serializer context"""
        context = super().get_serializer_context()
        # Cache FosseSettings lookup to avoid N+1 queries in serializer
        if not hasattr(self, '_cached_fosse_default_status'):
            from .models import UserDetails, FosseSettings, Status
            fosse_default_status_name = None
            try:
                user_details = UserDetails.objects.select_related('role_id').filter(django_user=self.request.user).first()
                if user_details and user_details.role_id:
                    fosse_setting = FosseSettings.objects.select_related('default_status').filter(role=user_details.role_id).first()
                    if fosse_setting and fosse_setting.default_status:
                        fosse_default_status_name = fosse_setting.default_status.name
                    else:
                        # Fallback to status with is_fosse_default=True
                        default_status = Status.objects.filter(is_fosse_default=True).first()
                        if default_status:
                            fosse_default_status_name = default_status.name
            except Exception:
                pass
            self._cached_fosse_default_status = fosse_default_status_name
        context['fosse_default_status_name'] = self._cached_fosse_default_status
        return context
    
    def get_queryset(self):
        """
        Filter contacts to only show unassigned contacts (teleoperator=null AND confirmateur=null).
        Bypasses all permission-based filtering - shows all contacts regardless of status.
        """
        # If we have a filtered queryset stored, use it (for pagination counting)
        # This is set in list() method after applying filters
        if hasattr(self, '_filtered_queryset') and self._filtered_queryset is not None:
            return self._filtered_queryset
        
        # Start with all contacts
        queryset = Contact.objects.all()
        
        # Filter for unassigned contacts only (teleoperator is null AND confirmateur is null)
        queryset = queryset.filter(
            teleoperator__isnull=True,
            confirmateur__isnull=True
        )
        
        # Optimize queries with select_related for ForeignKey relationships
        queryset = queryset.select_related(
            'status',
            'source',
            'teleoperator',
            'teleoperator__user_details',  # Select user_details to avoid query when accessing it
            'confirmateur',
            'creator'
        )
        
        # Annotate with most recent log date to avoid N+1 queries
        # This annotation is needed for updated_at ordering in _apply_filters_fosse
        # Explicitly set output_field to ensure proper datetime type for sorting
        from django.db.models import Prefetch, Subquery, OuterRef
        latest_log = Log.objects.filter(
            contact_id=OuterRef('pk')
        ).order_by('-created_at').values('created_at')[:1]
        
        queryset = queryset.annotate(
            last_log_date=Subquery(latest_log, output_field=models.DateTimeField())
        )
        
        # Prefetch related team_memberships for teleoperator's user_details
        # CRITICAL: Prefetch logs to avoid N+1 queries in serializer
        queryset = queryset.prefetch_related(
            Prefetch(
                'teleoperator__user_details__team_memberships',
                queryset=TeamMember.objects.select_related('team')
            ),
            Prefetch(
                'contact_notes',
                queryset=Note.objects.select_related('categ_id').order_by('-created_at')
            ),
            Prefetch(
                'contact_logs',
                queryset=Log.objects.filter(event_type='editContact').order_by('-created_at'),
                to_attr='prefetched_edit_logs'
            )
        )
        
        # Default order by created_at descending (most recent first)
        # This will be overridden by FosseSettings.default_order in _apply_filters_fosse if settings exist
        queryset = queryset.order_by('-created_at')
        
        return queryset
    
    def _apply_filters_fosse(self, queryset, request):
        """Helper method to apply all filters to a Fosse queryset"""
        # Track if full_name_search annotation has been applied to avoid double annotation
        full_name_search_annotated = False
        
        # Apply forced filters from FosseSettings (server-side enforcement)
        user = request.user
        default_order = None
        try:
            from .models import UserDetails, FosseSettings
            user_details = UserDetails.objects.select_related('role_id').get(django_user=user)
            if user_details.role_id:
                try:
                    fosse_setting = FosseSettings.objects.get(role=user_details.role_id)
                    forced_filters = fosse_setting.forced_filters or {}
                    # Get default_order from settings - preserve 'none' value, don't convert None to 'created_at_desc'
                    # This allows the order query parameter to be used when default_order is 'none' or not set
                    # Handle empty string as None
                    default_order = fosse_setting.default_order if fosse_setting.default_order and fosse_setting.default_order.strip() else None
                    
                    # Collect query parameter filters first to check which columns have explicit filters
                    # Query parameters override forced filters for those columns (for preview functionality)
                    query_param_columns = set()
                    for key in request.query_params.keys():
                        if key.startswith('filter_'):
                            if key.endswith('_from') or key.endswith('_to'):
                                base_column = key.replace('filter_', '').replace('_from', '').replace('_to', '')
                                query_param_columns.add(base_column)
                            else:
                                column_id = key.replace('filter_', '')
                                query_param_columns.add(column_id)
                    
                    # Apply forced filters server-side (both 'defined' and 'open' types)
                    # This ensures forced filters are always applied even if frontend doesn't send them
                    # BUT: Skip forced filters for columns that have query parameters (query params override forced filters)
                    for column_id, filter_config in forced_filters.items():
                        # Skip this column if query parameters are present (they override forced filters)
                        # This allows preview to show contacts with configured filters
                        if column_id in query_param_columns:
                            continue
                            
                        config = filter_config if isinstance(filter_config, dict) else {}
                        config_type = config.get('type')
                        config_values = config.get('values')
                        
                        # Handle multi-select filters (both 'defined' and 'open' types with values)
                        if (config_type == 'defined' or config_type == 'open') and config_values and len(config_values) > 0:
                            values = config_values
                            has_empty = '__empty__' in values
                            regular_values = [v for v in values if v != '__empty__']
                            
                            # Build Q objects for filtering
                            q_objects = []
                            
                            # IMPORTANT: For forced filters, only apply filters if values are explicitly provided
                            # If user forced a specific value (not empty), has_empty should be False
                            # and we should only show contacts matching that value
                            
                            # Add regular value filters if any
                            if regular_values:
                                if column_id == 'status':
                                    # Filter by current status (status_id field) - NOT previous status from logs
                                    # This ensures we only show contacts with the most recent/latest status
                                    q_objects.append(models.Q(status_id__in=regular_values))
                                elif column_id == 'source':
                                    q_objects.append(models.Q(source_id__in=regular_values))
                                elif column_id == 'creator':
                                    # Convert UserDetails IDs to Django User IDs (frontend sends UserDetails IDs)
                                    django_user_ids = []
                                    for user_details_id in regular_values:
                                        try:
                                            # Try as UserDetails ID first (string)
                                            user_details = UserDetails.objects.filter(id=str(user_details_id)).first()
                                            if user_details and user_details.django_user:
                                                django_user_ids.append(user_details.django_user.id)
                                            else:
                                                # Fallback: try as Django User ID (if numeric)
                                                try:
                                                    int_id = int(user_details_id)
                                                    if DjangoUser.objects.filter(id=int_id).exists():
                                                        django_user_ids.append(int_id)
                                                except (ValueError, TypeError):
                                                    pass
                                        except Exception as e:
                                            print(f"[DEBUG] Error converting creator UserDetails ID {user_details_id}: {e}")
                                            # Fallback: try as Django User ID
                                            try:
                                                int_id = int(user_details_id)
                                                if DjangoUser.objects.filter(id=int_id).exists():
                                                    django_user_ids.append(int_id)
                                            except (ValueError, TypeError):
                                                pass
                                    
                                    if django_user_ids:
                                        q_objects.append(models.Q(creator_id__in=django_user_ids))
                                    else:
                                        # Return empty queryset by filtering with impossible condition
                                        q_objects.append(models.Q(creator_id__in=[]))
                                elif column_id == 'postalCode':
                                    q_objects.append(models.Q(postal_code__in=regular_values))
                                elif column_id == 'nationality':
                                    q_objects.append(models.Q(nationality__in=regular_values))
                                elif column_id == 'campaign':
                                    q_objects.append(models.Q(campaign__in=regular_values))
                                elif column_id == 'civility':
                                    q_objects.append(models.Q(civility__in=regular_values))
                                elif column_id == 'previousStatus':
                                    # Filter by IMMEDIATE previous status (the status right before the current one)
                                    # PERFORMANCE: Use efficient query with proper error handling
                                    from .models import Log
                                    from django.db.models import F
                                    
                                    try:
                                        # Find matching contacts efficiently using database query
                                        # Get logs where old_status matches filter values, then verify new_status matches contact's current status
                                        # Use values() to get only needed fields for efficiency
                                        matching_logs = Log.objects.filter(
                                            event_type='editContact',
                                            old_value__statusName__in=regular_values,
                                            new_value__statusName__isnull=False
                                        ).exclude(
                                            old_value__statusName=F('new_value__statusName')
                                        ).select_related('contact_id__status').order_by('contact_id_id', '-created_at')
                                        
                                        # Process logs to find matches - track seen contacts to get most recent per contact
                                        matching_contact_ids = []
                                        seen_contacts = set()
                                        
                                        for log in matching_logs.iterator(chunk_size=1000):
                                            contact_id = log.contact_id_id if log.contact_id_id else None
                                            if not contact_id or contact_id in seen_contacts:
                                                continue  # Skip if already processed this contact
                                            
                                            seen_contacts.add(contact_id)
                                            
                                            if log.contact_id and log.contact_id.status:
                                                log_new_status = log.new_value.get('statusName', '') if log.new_value else ''
                                                current_status_name = log.contact_id.status.name
                                                if log_new_status == current_status_name:
                                                    matching_contact_ids.append(contact_id)
                                        
                                        if matching_contact_ids:
                                            q_objects.append(models.Q(id__in=matching_contact_ids))
                                        elif not has_empty:
                                            # No matches and no empty option - exclude all (strict filter)
                                            queryset = queryset.none()
                                            break
                                    except Exception as e:
                                        # If query fails, log error and skip this filter to prevent 500 error
                                        import logging
                                        logger = logging.getLogger(__name__)
                                        logger.error(f"Error in previousStatus filter: {e}")
                                        # Skip this filter rather than failing
                                        continue
                                elif column_id == 'previousTeleoperator':
                                    # Filter by previous teleoperator from logs - same logic as regular filter
                                    from .models import Log
                                    from django.db.models import OuterRef, Exists
                                    
                                    matching_contact_ids = set()
                                    for teleoperator_name in regular_values:
                                        matching_logs = Log.objects.filter(
                                            contact_id=OuterRef('pk'),
                                            old_value__teleoperatorName=teleoperator_name,
                                            new_value__teleoperatorName__isnull=False
                                        ).exclude(
                                            old_value__teleoperatorName=models.F('new_value__teleoperatorName')
                                        )
                                        # Get contacts that have matching logs - use all contacts, not just current queryset
                                        # because we need to find contacts by their log history
                                        contacts = Contact.objects.filter(
                                            Exists(matching_logs)
                                        ).values_list('id', flat=True)
                                        matching_contact_ids.update(contacts)
                                    
                                    if matching_contact_ids:
                                        q_objects.append(models.Q(id__in=matching_contact_ids))
                                    # If no matches found and has_empty is False, we'll exclude all contacts below
                                    # If has_empty is True, empty_q will be added below, but only if empty was explicitly selected
                            
                            # Add empty filter ONLY if empty option was explicitly selected (has_empty is True)
                            # For forced filters, if user selected specific values without empty, only show those
                            if has_empty:
                                if column_id == 'status':
                                    empty_q = models.Q(status_id__isnull=True)
                                elif column_id == 'source':
                                    empty_q = models.Q(source_id__isnull=True)
                                elif column_id == 'creator':
                                    empty_q = models.Q(creator_id__isnull=True)
                                elif column_id == 'postalCode':
                                    empty_q = models.Q(postal_code__isnull=True) | models.Q(postal_code='')
                                elif column_id == 'nationality':
                                    empty_q = models.Q(nationality__isnull=True) | models.Q(nationality='')
                                elif column_id == 'campaign':
                                    empty_q = models.Q(campaign__isnull=True) | models.Q(campaign='')
                                elif column_id == 'civility':
                                    empty_q = models.Q(civility__isnull=True) | models.Q(civility='')
                                elif column_id == 'previousStatus':
                                    # Empty previousStatus means no previous status (contact never had a status change)
                                    from .models import Log
                                    contacts_with_status_changes = Log.objects.filter(
                                        contact_id__isnull=False,
                                        old_value__statusName__isnull=False,
                                        new_value__statusName__isnull=False
                                    ).exclude(
                                        old_value__statusName=models.F('new_value__statusName')
                                    ).values_list('contact_id_id', flat=True).distinct()
                                    empty_q = ~models.Q(id__in=contacts_with_status_changes)
                                elif column_id == 'previousTeleoperator':
                                    # Empty previousTeleoperator means no previous teleoperator (contact never had a teleoperator change)
                                    from .models import Log
                                    contacts_with_teleoperator_changes = Log.objects.filter(
                                        contact_id__isnull=False,
                                        old_value__teleoperatorName__isnull=False,
                                        new_value__teleoperatorName__isnull=False
                                    ).exclude(
                                        old_value__teleoperatorName=models.F('new_value__teleoperatorName')
                                    ).values_list('contact_id_id', flat=True).distinct()
                                    empty_q = ~models.Q(id__in=contacts_with_teleoperator_changes)
                                else:
                                    empty_q = None
                                
                                if empty_q:
                                    if regular_values:
                                        # Regular values exist - only add empty_q if has_empty is True
                                        # This matches preview logic: if (hasEmpty && !previousStatus) matches = true;
                                        # So we combine: (regular_values) OR (empty) when has_empty is True
                                        if has_empty:
                                            q_objects.append(empty_q)
                                        # If has_empty is False, don't add empty_q - only show regular values
                                    else:
                                        # Only empty option selected - show ONLY contacts with empty/null
                                        # This matches preview logic: matches = hasEmpty && !previousStatus
                                        q_objects = [empty_q]
                            
                            # Apply combined filter - match preview table logic exactly
                            # Preview logic:
                            # - If regularValues.length === 0: matches = hasEmpty && !previousStatus
                            # - If regularValues.length > 0: matches = regularValues.includes(previousStatus) || (hasEmpty && !previousStatus)
                            if q_objects:
                                if len(q_objects) == 1:
                                    # Single filter - apply directly
                                    queryset = queryset.filter(q_objects[0])
                                else:
                                    # Multiple filters - combine with OR (regular values OR empty)
                                    # This matches preview: matches = regularValues.includes(previousStatus) || (hasEmpty && !previousStatus)
                                    combined_q = q_objects[0]
                                    for q_obj in q_objects[1:]:
                                        combined_q |= q_obj
                                    queryset = queryset.filter(combined_q)
                            elif column_id in ['previousStatus', 'previousTeleoperator']:
                                # No q_objects but we're filtering - match preview logic
                                # Preview logic: if regularValues.length === 0, matches = hasEmpty && !previousStatus
                                # So if only empty is selected and no q_objects, something went wrong
                                if not regular_values and has_empty:
                                    # empty_q should have been added above, but if not, exclude all to be safe
                                    queryset = queryset.none()
                                # Preview logic: if regularValues.length > 0 and no matches, matches = false (unless hasEmpty && !previousStatus)
                                # So if regularValues exist but no matches and hasEmpty is false, exclude all
                                elif regular_values and not has_empty:
                                    queryset = queryset.none()
                                # If regularValues exist but no matches and hasEmpty is true, empty_q should have been added
                                elif regular_values and has_empty:
                                    # This shouldn't happen - empty_q should have been added in the has_empty block above
                                    # But if it didn't, exclude all to be safe
                                    queryset = queryset.none()
                        elif config_type == 'open':
                            # Apply 'open' type filters server-side if they have text values or date ranges
                            # These are pre-filled but users can modify them
                            value = config.get('value')
                            if value and isinstance(value, str) and value.strip():
                                value = value.strip()
                                if column_id == 'email':
                                    queryset = queryset.filter(email__icontains=value)
                                elif column_id == 'fullName':
                                    # Search in the combined full name string (first name + space + last name)
                                    # Handle NULL values with Coalesce to avoid NULL results
                                    # Track that annotation was applied to avoid double annotation
                                    if not full_name_search_annotated:
                                        queryset = queryset.annotate(
                                            full_name_search=Concat(
                                                Coalesce('fname', Value('')), 
                                                Value(' '), 
                                                Coalesce('lname', Value(''))
                                            )
                                        )
                                        full_name_search_annotated = True
                                    queryset = queryset.filter(full_name_search__icontains=value)
                            date_range = config.get('dateRange')
                            if date_range and isinstance(date_range, dict):
                                if date_range.get('from') or date_range.get('to'):
                                    if column_id == 'createdAt':
                                        if date_range.get('from'):
                                            queryset = queryset.filter(created_at__gte=date_range.get('from'))
                                        if date_range.get('to'):
                                            queryset = queryset.filter(created_at__lte=date_range.get('to'))
                                    elif column_id == 'updatedAt':
                                        # Filter by updated_at (modification date)
                                        if date_range.get('from'):
                                            queryset = queryset.filter(updated_at__gte=date_range.get('from'))
                                        if date_range.get('to'):
                                            queryset = queryset.filter(updated_at__lte=date_range.get('to'))
                except FosseSettings.DoesNotExist:
                    pass
        except (UserDetails.DoesNotExist, Exception):
            pass
        
        # Check if we need full_name_search annotation (for search or fullName filter)
        # MUST be done BEFORE applying search filter
        # Only apply if not already annotated by forced filters
        search = request.query_params.get('search', '').strip()
        has_fullname_filter = any(key.startswith('filter_fullName') for key in request.query_params.keys())
        
        # Track if phone annotations are needed for search
        phone_annotations_needed = bool(search)
        
        # Apply full_name_search annotation if needed (for search or fullName filter)
        if (search or has_fullname_filter) and not full_name_search_annotated:
            # Apply annotation once for both search and fullName filter
            if search:
                # Include phone annotations when searching
                queryset = queryset.annotate(
                    full_name_search=Concat(
                        Coalesce('fname', Value('')), 
                        Value(' '), 
                        Coalesce('lname', Value(''))
                    ),
                    phone_str=Cast('phone', CharField()),
                    mobile_str=Cast('mobile', CharField())
                )
                phone_annotations_needed = False  # Already added
            else:
                queryset = queryset.annotate(
                    full_name_search=Concat(
                        Coalesce('fname', Value('')), 
                        Value(' '), 
                        Coalesce('lname', Value(''))
                    )
                )
            full_name_search_annotated = True
        
        # If search exists and phone annotations weren't added above (because full_name_search was already annotated),
        # add phone annotations separately
        if phone_annotations_needed:
            queryset = queryset.annotate(
                phone_str=Cast('phone', CharField()),
                mobile_str=Cast('mobile', CharField())
            )
        
        # Apply search filter (annotation must be applied above if needed)
        if search:
            # Remove spaces from search value for phone number matching
            phone_search = ''.join(str(search).split())
            # Search in combined full name, email, and phone numbers (same as regular ContactView)
            queryset = queryset.filter(
                models.Q(full_name_search__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(phone_str__contains=phone_search) |
                models.Q(mobile_str__contains=phone_search)
            )
        
        # Apply team filter (filter by creator's team for Fosse contacts)
        # Note: For fosse, contacts are unassigned (teleoperator and confirmateur are null),
        # so we only filter by creator's team
        team_id = request.query_params.get('team')
        if team_id and team_id != 'all':
            # Get team members
            team_user_ids = TeamMember.objects.filter(team_id=team_id).values_list('user__django_user__id', flat=True)
            # For Fosse, only filter by creator since teleoperator and confirmateur are null
            queryset = queryset.filter(creator__id__in=team_user_ids)
        
        # Apply status type filter
        status_type = request.query_params.get('status_type')
        if status_type and status_type != 'all':
            queryset = queryset.filter(status__type=status_type)
        
        # Apply column filters (similar to ContactView but with Fosse-specific handling)
        date_range_filters = {}
        multi_select_filters = {}
        
        for key in request.query_params.keys():
            if key.startswith('filter_'):
                if key.endswith('_from') or key.endswith('_to'):
                    base_column = key.replace('filter_', '').replace('_from', '').replace('_to', '')
                    value = request.query_params.get(key)
                    if value:
                        if base_column not in date_range_filters:
                            date_range_filters[base_column] = {}
                        if key.endswith('_from'):
                            date_range_filters[base_column]['from'] = value
                        elif key.endswith('_to'):
                            date_range_filters[base_column]['to'] = value
                else:
                    column_id = key.replace('filter_', '')
                    if column_id in ['status', 'creator', 'teleoperator', 'confirmateur', 'source', 'postalCode', 'nationality', 'campaign', 'civility', 'managerTeam', 'previousStatus', 'previousTeleoperator']:
                        values = request.query_params.getlist(key)
                        if values:
                            multi_select_filters[column_id] = values
                    else:
                        value = request.query_params.get(key)
                        if value:
                            if column_id == 'email':
                                queryset = queryset.filter(email__icontains=value)
                            elif column_id == 'phone':
                                # Remove spaces from search value and convert to string
                                phone_search = ''.join(str(value).split())
                                if phone_search:
                                    # Convert integer phone fields to strings for partial matching
                                    queryset = queryset.annotate(
                                        phone_str=Cast('phone', CharField()),
                                        mobile_str=Cast('mobile', CharField())
                                    ).filter(
                                        models.Q(phone_str__contains=phone_search) | 
                                        models.Q(mobile_str__contains=phone_search)
                                    )
                            elif column_id == 'fullName':
                                # Search in the combined full name string (first name + space + last name)
                                # Annotation already applied above if search or fullName filter is present
                                queryset = queryset.filter(full_name_search__icontains=value)
                            elif column_id == 'firstName':
                                queryset = queryset.filter(fname__icontains=value)
                            elif column_id == 'lastName':
                                queryset = queryset.filter(lname__icontains=value)
                            elif column_id == 'city':
                                queryset = queryset.filter(city__icontains=value)
                            elif column_id == 'address':
                                queryset = queryset.filter(address__icontains=value)
                            elif column_id == 'addressComplement':
                                queryset = queryset.filter(address_complement__icontains=value)
                            elif column_id == 'birthPlace':
                                queryset = queryset.filter(birth_place__icontains=value)
        
        # Apply multi-select filters with Fosse-specific logic
        for column_id, values in multi_select_filters.items():
            if values:
                # Strip and filter out empty strings
                values = [v.strip() if isinstance(v, str) else str(v).strip() for v in values if v and str(v).strip()]
                
                if not values:
                    continue
                
                has_empty = '__empty__' in values
                regular_values = [v for v in values if v != '__empty__']
                q_objects = []
                
                if has_empty:
                    if column_id == 'status':
                        q_objects.append(models.Q(status_id__isnull=True))
                    elif column_id == 'source':
                        q_objects.append(models.Q(source_id__isnull=True))
                    elif column_id == 'teleoperator':
                        # For Fosse, teleoperator is always null, so empty option matches all
                        pass
                    elif column_id == 'confirmateur':
                        # For Fosse, confirmateur is always null, so empty option matches all
                        pass
                    elif column_id == 'creator':
                        q_objects.append(models.Q(creator_id__isnull=True))
                    elif column_id == 'postalCode':
                        q_objects.append(models.Q(postal_code__isnull=True) | models.Q(postal_code=''))
                    elif column_id == 'nationality':
                        q_objects.append(models.Q(nationality__isnull=True) | models.Q(nationality=''))
                    elif column_id == 'campaign':
                        q_objects.append(models.Q(campaign__isnull=True) | models.Q(campaign=''))
                    elif column_id == 'civility':
                        q_objects.append(models.Q(civility__isnull=True) | models.Q(civility=''))
                    elif column_id == 'managerTeam':
                        # For Fosse, contacts have no team
                        pass
                    elif column_id == 'previousStatus':
                        # Empty previousStatus means no previous status (contact never had a status change)
                        # This is tricky - we need to find contacts that have no status change logs
                        from .models import Log
                        # Get all contact IDs that have status change logs
                        contacts_with_status_changes = Log.objects.filter(
                            contact_id__isnull=False,
                            old_value__statusName__isnull=False,
                            new_value__statusName__isnull=False
                        ).exclude(
                            old_value__statusName=models.F('new_value__statusName')
                        ).values_list('contact_id_id', flat=True).distinct()
                        
                        # Contacts without status changes (empty previousStatus)
                        # Use a subquery to handle empty case correctly
                        q_objects.append(~models.Q(id__in=contacts_with_status_changes))
                    elif column_id == 'previousTeleoperator':
                        # Empty previousTeleoperator means no previous teleoperator (contact never had a teleoperator change)
                        from .models import Log
                        # Get all contact IDs that have teleoperator change logs
                        contacts_with_teleoperator_changes = Log.objects.filter(
                            contact_id__isnull=False,
                            old_value__teleoperatorName__isnull=False,
                            new_value__teleoperatorName__isnull=False
                        ).exclude(
                            old_value__teleoperatorName=models.F('new_value__teleoperatorName')
                        ).values_list('contact_id_id', flat=True).distinct()
                        
                        # Contacts without teleoperator changes (empty previousTeleoperator)
                        # Use a subquery to handle empty case correctly
                        q_objects.append(~models.Q(id__in=contacts_with_teleoperator_changes))
                
                # Add regular value filters if any
                if regular_values:
                    if column_id == 'status':
                        # Filter by current status (status_id field) - NOT previous status from logs
                        # This ensures we only show contacts with the most recent/latest status
                        q_objects.append(models.Q(status_id__in=regular_values))
                    elif column_id == 'source':
                        q_objects.append(models.Q(source_id__in=regular_values))
                    elif column_id == 'teleoperator':
                        # For Fosse, teleoperator is always null, so regular values would exclude all
                        queryset = queryset.none()
                        break
                    elif column_id == 'confirmateur':
                        # For Fosse, confirmateur is always null, so regular values would exclude all
                        queryset = queryset.none()
                        break
                    elif column_id == 'creator':
                        # Convert UserDetails IDs to Django User IDs (frontend sends UserDetails IDs)
                        django_user_ids = []
                        for user_details_id in regular_values:
                            try:
                                # Try as UserDetails ID first (string)
                                user_details = UserDetails.objects.filter(id=str(user_details_id)).first()
                                if user_details and user_details.django_user:
                                    django_user_ids.append(user_details.django_user.id)
                                else:
                                    # Fallback: try as Django User ID (if numeric)
                                    try:
                                        int_id = int(user_details_id)
                                        if DjangoUser.objects.filter(id=int_id).exists():
                                            django_user_ids.append(int_id)
                                    except (ValueError, TypeError):
                                        pass
                            except Exception as e:
                                print(f"[DEBUG] Error converting creator UserDetails ID {user_details_id}: {e}")
                                # Fallback: try as Django User ID
                                try:
                                    int_id = int(user_details_id)
                                    if DjangoUser.objects.filter(id=int_id).exists():
                                        django_user_ids.append(int_id)
                                except (ValueError, TypeError):
                                    pass
                        
                        if django_user_ids:
                            q_objects.append(models.Q(creator_id__in=django_user_ids))
                        else:
                            # Return empty queryset by filtering with impossible condition
                            q_objects.append(models.Q(creator_id__in=[]))
                    elif column_id == 'postalCode':
                        q_objects.append(models.Q(postal_code__in=regular_values))
                    elif column_id == 'nationality':
                        q_objects.append(models.Q(nationality__in=regular_values))
                    elif column_id == 'campaign':
                        q_objects.append(models.Q(campaign__in=regular_values))
                    elif column_id == 'civility':
                        q_objects.append(models.Q(civility__in=regular_values))
                    elif column_id == 'managerTeam':
                        # For Fosse, contacts have no team, so regular values would exclude all
                        queryset = queryset.none()
                        break
                    elif column_id == 'previousStatus':
                        # Filter by IMMEDIATE previous status (the status right before the current one)
                        # PERFORMANCE: Use efficient query with proper error handling
                        from .models import Log
                        from django.db.models import F
                        
                        try:
                            # Find matching contacts efficiently using database query
                            # Get logs where old_status matches filter values, then verify new_status matches contact's current status
                            # Use values() to get only needed fields for efficiency
                            matching_logs = Log.objects.filter(
                                event_type='editContact',
                                old_value__statusName__in=regular_values,
                                new_value__statusName__isnull=False
                            ).exclude(
                                old_value__statusName=F('new_value__statusName')
                            ).select_related('contact_id__status').order_by('contact_id_id', '-created_at')
                            
                            # Process logs to find matches - track seen contacts to get most recent per contact
                            matching_contact_ids = []
                            seen_contacts = set()
                            
                            for log in matching_logs.iterator(chunk_size=1000):
                                contact_id = log.contact_id_id if log.contact_id_id else None
                                if not contact_id or contact_id in seen_contacts:
                                    continue  # Skip if already processed this contact
                                
                                seen_contacts.add(contact_id)
                                
                                if log.contact_id and log.contact_id.status:
                                    log_new_status = log.new_value.get('statusName', '') if log.new_value else ''
                                    current_status_name = log.contact_id.status.name
                                    if log_new_status == current_status_name:
                                        matching_contact_ids.append(contact_id)
                            
                            # Match preview logic:
                            # - If regularValues.length > 0: matches = regularValues.includes(previousStatus) || (hasEmpty && !previousStatus)
                            # - So we need to add regular values filter, and empty filter will be combined with OR if has_empty is True
                            if matching_contact_ids:
                                q_objects.append(models.Q(id__in=matching_contact_ids))
                            elif not has_empty:
                                # No matches and no empty option - exclude all (strict filter)
                                queryset = queryset.none()
                                break
                            # If has_empty is True and matching_contact_ids is empty, don't add anything for regular values
                            # The empty filter (already added above) will be applied, showing only contacts with empty previousStatus
                        except Exception as e:
                            # If query fails, log error and skip this filter to prevent 500 error
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.error(f"Error in previousStatus filter: {e}")
                            # Skip this filter rather than failing - don't break, just continue
                            pass
                    elif column_id == 'previousTeleoperator':
                        # Filter by previous teleoperator from logs
                        # Similar to previousStatus - find contacts with matching old_value.teleoperatorName
                        from .models import Log
                        from django.db.models import OuterRef, Exists
                        
                        matching_contact_ids = set()
                        for teleoperator_name in regular_values:
                            # Find logs where teleoperator changed and old_value.teleoperatorName matches
                            matching_logs = Log.objects.filter(
                                contact_id=OuterRef('pk'),
                                old_value__teleoperatorName=teleoperator_name,
                                new_value__teleoperatorName__isnull=False
                            ).exclude(
                                old_value__teleoperatorName=models.F('new_value__teleoperatorName')
                            )
                            
                            # Get contacts that have at least one matching log
                            contacts = Contact.objects.filter(
                                Exists(matching_logs)
                            ).values_list('id', flat=True)
                            
                            matching_contact_ids.update(contacts)
                        
                        if matching_contact_ids:
                            q_objects.append(models.Q(id__in=matching_contact_ids))
                        elif not has_empty:
                            # No matches and no empty option - exclude all
                            queryset = queryset.none()
                            break
                
                # Apply combined filter
                if q_objects:
                    if len(q_objects) == 1:
                        # Single filter - apply directly
                        queryset = queryset.filter(q_objects[0])
                    else:
                        # Multiple filters - combine with OR (empty OR regular values)
                        # For previousStatus and previousTeleoperator, ensure empty filter is only included if has_empty is True
                        combined_q = q_objects[0]
                        for q_obj in q_objects[1:]:
                            combined_q |= q_obj
                        queryset = queryset.filter(combined_q)
                else:
                    # Special handling for managerTeam in Fosse
                    if column_id == 'managerTeam' and regular_values:
                        # Filter by creator's team for Fosse contacts
                        team_user_ids = TeamMember.objects.filter(team_id__in=regular_values).values_list('user__django_user__id', flat=True)
                        queryset = queryset.filter(creator_id__in=team_user_ids)
        
        # Apply date range filters
        for column_id, date_range in date_range_filters.items():
            try:
                if column_id == 'createdAt':
                    if 'from' in date_range and date_range['from']:
                        date_from = datetime.strptime(date_range['from'], '%Y-%m-%d').date()
                        queryset = queryset.filter(created_at__date__gte=date_from)
                    if 'to' in date_range and date_range['to']:
                        date_to = datetime.strptime(date_range['to'], '%Y-%m-%d').date()
                        date_to_next = date_to + timedelta(days=1)
                        queryset = queryset.filter(created_at__date__lt=date_to_next)
                elif column_id == 'updatedAt':
                    if 'from' in date_range and date_range['from']:
                        date_from = datetime.strptime(date_range['from'], '%Y-%m-%d').date()
                        queryset = queryset.filter(updated_at__date__gte=date_from)
                    if 'to' in date_range and date_range['to']:
                        date_to = datetime.strptime(date_range['to'], '%Y-%m-%d').date()
                        date_to_next = date_to + timedelta(days=1)
                        queryset = queryset.filter(updated_at__date__lt=date_to_next)
                elif column_id == 'birthDate':
                    if 'from' in date_range and date_range['from']:
                        date_from = datetime.strptime(date_range['from'], '%Y-%m-%d').date()
                        queryset = queryset.filter(birth_date__gte=date_from)
                    if 'to' in date_range and date_range['to']:
                        date_to = datetime.strptime(date_range['to'], '%Y-%m-%d').date()
                        queryset = queryset.filter(birth_date__lte=date_to)
            except (ValueError, TypeError) as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Invalid date format in filter {column_id}: {e}")
                pass
        
        # Apply ordering: first check FosseSettings.default_order, then check order query parameter
        order_to_apply = None
        
        # Priority 1: Use forced order from FosseSettings if set and not 'none'
        if default_order and default_order != 'none':
            order_to_apply = default_order
        # Priority 2: Use order from query parameter (from select dropdown)
        elif request.query_params.get('order'):
            order_to_apply = request.query_params.get('order')
        
        # Apply the determined order
        if order_to_apply:
            try:
                # Clear any existing ordering first
                queryset = queryset.order_by()
                
                if order_to_apply == 'created_at_asc':
                    queryset = queryset.order_by('created_at')
                elif order_to_apply == 'created_at_desc':
                    queryset = queryset.order_by('-created_at')
                elif order_to_apply == 'updated_at_asc':
                    # For Fosse, use updated_at directly (simpler and more reliable)
                    queryset = queryset.order_by('updated_at')
                elif order_to_apply == 'updated_at_desc':
                    # For Fosse, use updated_at directly (simpler and more reliable)
                    queryset = queryset.order_by('-updated_at')
                elif order_to_apply == 'email_asc':
                    queryset = queryset.order_by('email')
                elif order_to_apply == 'assigned_at_asc':
                    queryset = queryset.order_by('assigned_at')
                elif order_to_apply == 'assigned_at_desc':
                    queryset = queryset.order_by('-assigned_at')
                elif order_to_apply == 'date_lead_to_client_asc':
                    # Sort by date_lead_to_client ascending (oldest first)
                    # Put NULL values last using Case/When
                    queryset = queryset.annotate(
                        date_lead_to_client_null_order=Case(
                            When(date_lead_to_client__isnull=True, then=Value(1)),
                            default=Value(0),
                            output_field=IntegerField()
                        )
                    ).order_by('date_lead_to_client_null_order', 'date_lead_to_client', '-created_at')
                elif order_to_apply == 'date_lead_to_client_desc':
                    # Sort by date_lead_to_client descending (newest first)
                    # Put NULL values last using Case/When
                    queryset = queryset.annotate(
                        date_lead_to_client_null_order=Case(
                            When(date_lead_to_client__isnull=True, then=Value(1)),
                            default=Value(0),
                            output_field=IntegerField()
                        )
                    ).order_by('date_lead_to_client_null_order', '-date_lead_to_client', '-created_at')
                elif order_to_apply == 'random':
                    queryset = queryset.order_by('?')
                else:
                    # Fallback to default ordering (creation date, most recent first)
                    queryset = queryset.order_by('-created_at')
            except Exception as e:
                # Log the error and fall back to default ordering
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error applying order '{order_to_apply}' in FosseContactView: {e}")
                # Fall back to default ordering
                queryset = queryset.order_by('-created_at')
        # If no order specified, keep the ordering from get_queryset (-created_at)
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        # Check if pagination is requested (preferred method for large datasets)
        requested_page = request.query_params.get('page')
        requested_page_size = request.query_params.get('page_size')
        
        # Use pagination if page or page_size are provided
        if requested_page or requested_page_size:
            from rest_framework.pagination import PageNumberPagination
            
            page_size = int(requested_page_size) if requested_page_size else 100
            page = int(requested_page) if requested_page else 1
            
            # Ensure reasonable page size (max 1000 per page)
            # Note: Heroku has a 30-second HTTP timeout, so queries must complete within that time
            MAX_PAGE_SIZE = 1000
            if page_size > MAX_PAGE_SIZE:
                page_size = MAX_PAGE_SIZE
            if page_size < 1:
                page_size = 100
            
            # Capture page_size for use in class definition
            pagination_page_size = page_size
            
            class FosseContactPagination(PageNumberPagination):
                page_size = pagination_page_size
                page_size_query_param = 'page_size'
                max_page_size = MAX_PAGE_SIZE
            
            queryset = self.get_queryset()
            queryset = self._apply_filters_fosse(queryset, request)
            
            # CRITICAL: Store the filtered queryset so get_queryset() returns it
            # This ensures DRF pagination uses the filtered queryset for counting
            cloned_queryset = queryset._clone()
            self._filtered_queryset = cloned_queryset
            
            # Cache FosseSettings lookup to avoid N+1 queries in serializer
            from .models import UserDetails, FosseSettings, Status
            fosse_default_status_name = None
            try:
                user_details = UserDetails.objects.select_related('role_id').filter(django_user=request.user).first()
                if user_details and user_details.role_id:
                    fosse_setting = FosseSettings.objects.select_related('default_status').filter(role=user_details.role_id).first()
                    if fosse_setting and fosse_setting.default_status:
                        fosse_default_status_name = fosse_setting.default_status.name
                    else:
                        # Fallback to status with is_fosse_default=True
                        default_status = Status.objects.filter(is_fosse_default=True).first()
                        if default_status:
                            fosse_default_status_name = default_status.name
            except Exception:
                pass
            
            self.pagination_class = FosseContactPagination
            
            # Pass cached FosseSettings to serializer context
            response = super().list(request, *args, **kwargs)
            
            # Clean up after pagination
            self._filtered_queryset = None
            return Response({
                'contacts': response.data['results'],
                'total': response.data['count'],
                'next': response.data.get('next'),
                'previous': response.data.get('previous'),
                'page': page,
                'page_size': page_size
            })
        
        # Check if limit is requested (for backward compatibility)
        limit = request.query_params.get('limit')
        if limit:
            try:
                limit = int(limit)
                # Ensure limit is reasonable (max 500)
                MAX_LIMIT = 500
                if limit > MAX_LIMIT:
                    limit = MAX_LIMIT
                if limit < 1:
                    limit = 1
                # Get base queryset
                queryset = self.get_queryset()
                
                # Apply all filters using helper method
                queryset = self._apply_filters_fosse(queryset, request)
                
                # Get actual total count BEFORE applying limit
                # This gives users the real total number of contacts matching their filters
                total_count = queryset.count()
                
                # CRITICAL PERFORMANCE FIX: Apply limit BEFORE serialization
                # This prevents loading thousands of contacts into memory
                queryset = queryset[:limit]
                
                # Cache FosseSettings lookup to avoid N+1 queries in serializer
                from .models import UserDetails, FosseSettings, Status
                fosse_default_status_name = None
                try:
                    user_details = UserDetails.objects.select_related('role_id').filter(django_user=request.user).first()
                    if user_details and user_details.role_id:
                        fosse_setting = FosseSettings.objects.select_related('default_status').filter(role=user_details.role_id).first()
                        if fosse_setting and fosse_setting.default_status:
                            fosse_default_status_name = fosse_setting.default_status.name
                        else:
                            # Fallback to status with is_fosse_default=True
                            default_status = Status.objects.filter(is_fosse_default=True).first()
                            if default_status:
                                fosse_default_status_name = default_status.name
                except Exception:
                    pass
                
                serializer = self.get_serializer(queryset, many=True, context={'request': request, 'fosse_default_status_name': fosse_default_status_name})
                return Response({
                    'contacts': serializer.data,
                    'total': total_count,
                    'limit': limit
                })
            except (ValueError, TypeError):
                # Invalid limit, fall through to default behavior
                pass
            except Exception as e:
                # Log the error and return a proper error response
                import traceback
                error_details = traceback.format_exc()
                print(f"Error in FosseContactView.list with limit: {error_details}")
                return Response({'error': str(e), 'details': error_details}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Default behavior: return contacts with a reasonable limit to prevent performance issues
        # Always apply a limit to prevent loading all contacts at once
        queryset = self.get_queryset()
        # Apply filters even in default path
        queryset = self._apply_filters_fosse(queryset, request)
        DEFAULT_LIMIT = 1000
        limited_queryset = queryset[:DEFAULT_LIMIT]
        
        # Cache FosseSettings lookup to avoid N+1 queries in serializer
        from .models import UserDetails, FosseSettings, Status
        fosse_default_status_name = None
        try:
            user_details = UserDetails.objects.select_related('role_id').filter(django_user=request.user).first()
            if user_details and user_details.role_id:
                fosse_setting = FosseSettings.objects.select_related('default_status').filter(role=user_details.role_id).first()
                if fosse_setting and fosse_setting.default_status:
                    fosse_default_status_name = fosse_setting.default_status.name
                else:
                    # Fallback to status with is_fosse_default=True
                    default_status = Status.objects.filter(is_fosse_default=True).first()
                    if default_status:
                        fosse_default_status_name = default_status.name
        except Exception:
            pass
        
        serializer = self.get_serializer(limited_queryset, many=True, context={'request': request, 'fosse_default_status_name': fosse_default_status_name})
        # Don't call count() on the full queryset - it's too slow. Use len() on limited queryset
        total_count = len(serializer.data)  # Approximate count
        return Response({
            'contacts': serializer.data,
            'total': total_count,
            'limit': DEFAULT_LIMIT
        })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def contacts_assigned_today_count(request):
    """Get count of contacts assigned to current user today with status type = lead"""
    try:
        from django.utils import timezone
        
        # Get current user
        user = request.user
        
        # Get today's date (based on day only, ignoring time)
        today = timezone.now().date()
        
        # Count contacts assigned to current user today with status type = lead
        # Filter by teleoperator = current user AND created_at date is today AND status type = lead
        # Using __date lookup compares only the date part, ignoring time
        count = Contact.objects.filter(
            teleoperator=user,
            created_at__date=today,
            status__type='lead'
        ).count()
        
        return Response({'count': count}, status=status.HTTP_200_OK)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error getting today's assigned contacts count: {e}")
        return Response({'error': str(e), 'count': 0}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def contact_create(request):
    # Validate required fields - only statusId is required
    if not request.data.get('statusId'):
        return Response({'error': 'Le statut est requis'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if email already exists (if provided)
    email = request.data.get('email', '').strip()
    if email and Contact.objects.filter(email=email).exists():
        return Response({'error': 'Un contact avec cet email existe déjà'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Verify email domain (DNS MX record check)
    email_verification_status = 'not_verified'
    if email:
        try:
            from api.utils.email_verification import verify_email_domain
            email_verification_status, _ = verify_email_domain(email)
        except Exception as e:
            # If verification fails, default to not_verified
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Email verification failed for {email}: {str(e)}")
            email_verification_status = 'not_verified'
    
    # Generate contact ID
    contact_id = uuid.uuid4().hex[:12]
    while Contact.objects.filter(id=contact_id).exists():
        contact_id = uuid.uuid4().hex[:12]
    
    # Helper function to safely get date
    def get_date(value):
        if not value or value == '':
            return None
        try:
            from datetime import datetime
            # Handle both YYYY-MM-DD and DD/MM/YYYY formats
            if '/' in str(value):
                parts = str(value).split('/')
                if len(parts) == 3:
                    day, month, year = parts
                    return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d").date()
            # Try ISO format (YYYY-MM-DD)
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None
    
    # Map frontend field names to model field names
    # Convert phone numbers to integers (remove spaces first)
    phone_value = request.data.get('phone', '') or ''
    mobile_value = request.data.get('mobile', '') or ''
    
    def phone_to_int(value):
        """Convert phone number string to integer, return None if empty"""
        if not value:
            return None
        # Remove all whitespace and convert to int
        cleaned = ''.join(str(value).split())
        if not cleaned:
            return None
        try:
            return int(cleaned)
        except (ValueError, TypeError):
            return None
    
    contact_data = {
        'id': contact_id,
        'civility': request.data.get('civility', '') or '',
        'fname': request.data.get('firstName', '') or '',
        'lname': request.data.get('lastName', '') or '',
        'phone': phone_to_int(phone_value),
        'mobile': phone_to_int(mobile_value),
        'email': request.data.get('email', '') or '',
        'email_verification_status': email_verification_status,
        'birth_date': get_date(request.data.get('birthDate')),
        'birth_place': request.data.get('birthPlace', '') or '',
        'address': request.data.get('address', '') or '',
        'address_complement': request.data.get('addressComplement', '') or '',
        'postal_code': request.data.get('postalCode', '') or '',
        'city': request.data.get('city', '') or '',
        'nationality': request.data.get('nationality', '') or '',
        'autre_informations': request.data.get('autreInformations', '') or '',
        'date_d_inscription': request.data.get('dateInscription', '') or '',
        'campaign': request.data.get('campaign', '') or '',
    }
    
    # Handle old_contact_id - convert empty strings to None
    old_contact_id_value = request.data.get('oldContactId', '') or ''
    contact_data['old_contact_id'] = old_contact_id_value.strip() if old_contact_id_value.strip() else None
    
    # Set creator to the current user
    contact_data['creator'] = request.user
    
    # Handle status - use get() for better performance
    status_id = request.data.get('statusId')
    if status_id:
        try:
            status_obj = Status.objects.filter(id=status_id).only('id').first()
            if status_obj:
                contact_data['status'] = status_obj
        except Exception:
            pass
    
    # Handle source - use get() for better performance
    source_id = request.data.get('sourceId')
    if source_id:
        try:
            source_obj = Source.objects.filter(id=source_id).only('id').first()
            if source_obj:
                contact_data['source'] = source_obj
        except Exception:
            pass
    
    # Handle teleoperator - optimized lookup
    teleoperator_id = request.data.get('teleoperatorId')
    if teleoperator_id:
        try:
            teleoperator_user = None
            # Prioritize UserDetails ID lookup first (since frontend sends UserDetails IDs)
            from api.models import UserDetails
            user_details = UserDetails.objects.filter(id=str(teleoperator_id)).select_related('django_user').only('id', 'django_user').first()
            if user_details and user_details.django_user:
                teleoperator_user = user_details.django_user
            
            # Fallback: Try Django User ID lookup only if UserDetails lookup failed
            if not teleoperator_user:
                try:
                    int_id = int(teleoperator_id)
                    teleoperator_user = DjangoUser.objects.filter(id=int_id).only('id').first()
                except (ValueError, TypeError):
                    pass
            
            if teleoperator_user:
                contact_data['teleoperator'] = teleoperator_user
                # Set assigned_at when creating contact with teleoperator
                from django.utils import timezone
                contact_data['assigned_at'] = timezone.now()
        except Exception as e:
            print(f"[DEBUG] Error setting teleoperator during contact creation: {e}")
            pass
    
    # Handle confirmateur - optimized lookup
    confirmateur_id = request.data.get('confirmateurId')
    if confirmateur_id:
        try:
            confirmateur_user = None
            # Prioritize UserDetails ID lookup first (since frontend sends UserDetails IDs)
            from api.models import UserDetails
            user_details = UserDetails.objects.filter(id=str(confirmateur_id)).select_related('django_user').only('id', 'django_user').first()
            if user_details and user_details.django_user:
                confirmateur_user = user_details.django_user
            
            # Fallback: Try Django User ID lookup only if UserDetails lookup failed
            if not confirmateur_user:
                try:
                    int_id = int(confirmateur_id)
                    confirmateur_user = DjangoUser.objects.filter(id=int_id).only('id').first()
                except (ValueError, TypeError):
                    pass
            
            if confirmateur_user:
                contact_data['confirmateur'] = confirmateur_user
        except Exception as e:
            print(f"[DEBUG] Error setting confirmateur during contact creation: {e}")
            pass
    
    # Handle confirmateur email and telephone
    if request.data.get('confirmateurEmail'):
        contact_data['confirmateur_email'] = request.data.get('confirmateurEmail', '').strip()
    if request.data.get('confirmateurTelephone'):
        contact_data['confirmateur_telephone'] = request.data.get('confirmateurTelephone', '').strip()
    
    # Handle platform - use get() for better performance
    platform_id = request.data.get('platformId')
    if platform_id:
        try:
            platform_obj = Platform.objects.filter(id=platform_id).only('id').first()
            if platform_obj:
                contact_data['platform'] = platform_obj
        except Exception:
            pass
    
    # Handle datetime fields from migration/CSV
    def parse_datetime(datetime_str):
        """Parse datetime string to timezone-aware datetime object"""
        if not datetime_str or str(datetime_str).strip() == '':
            return None
        from datetime import datetime
        from django.utils import timezone
        import pytz
        utc = pytz.UTC
        
        # If already a datetime object, make it timezone-aware and return
        if isinstance(datetime_str, datetime):
            if timezone.is_aware(datetime_str):
                return datetime_str
            # For naive datetime, assume UTC to avoid timezone offset issues
            return utc.localize(datetime_str)
        datetime_str = str(datetime_str).strip()
        
        # Try parsing ISO format with timezone first (e.g., "2024-01-01T10:00:00Z" or "2024-01-01T10:00:00+00:00")
        try:
            from dateutil import parser
            parsed = parser.parse(datetime_str)
            if timezone.is_aware(parsed):
                return parsed
            # If parsed as naive, assume UTC
            return utc.localize(parsed)
        except (ValueError, ImportError):
            pass
        
        # Try common datetime formats (assume UTC for naive datetimes)
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%S.%f',
            '%Y-%m-%dT%H:%M',
            '%Y-%m-%d',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y',
        ]
        for fmt in formats:
            try:
                parsed = datetime.strptime(datetime_str, fmt)
                # Assume UTC instead of server local timezone to avoid offset issues
                return utc.localize(parsed)
            except ValueError:
                continue
        return None
    
    # Set created_at from CSV if provided (for migration)
    if request.data.get('createdAt'):
        parsed_dt = parse_datetime(request.data.get('createdAt'))
        if parsed_dt:
            contact_data['created_at'] = parsed_dt
    
    # Handle updatedAt and assignedAt
    if request.data.get('updatedAt'):
        parsed_dt = parse_datetime(request.data.get('updatedAt'))
        if parsed_dt:
            contact_data['updated_at'] = parsed_dt
    if request.data.get('assignedAt'):
        parsed_dt = parse_datetime(request.data.get('assignedAt'))
        if parsed_dt:
            contact_data['assigned_at'] = parsed_dt
    
    try:
        # If created_at is set, we need to create the object and then update it
        # because auto_now_add might override it during create()
        if 'created_at' in contact_data:
            # Create without created_at first, then update it
            created_at_value = contact_data.pop('created_at')
            contact = Contact.objects.create(**contact_data)
            # Update created_at after creation to override auto_now_add
            Contact.objects.filter(id=contact.id).update(created_at=created_at_value)
            # Refresh the contact object to get the updated created_at
            contact.refresh_from_db()
        else:
            contact = Contact.objects.create(**contact_data)
        
        # Create log entry for contact creation
        serializer = ContactSerializer(contact, context={'request': request})
        contact_data_raw = serializer.data
        contact_data_for_log = clean_contact_data_for_log(contact_data_raw, include_created_at=False)
        
        create_log_entry(
            event_type='addContact',
            user_id=request.user if request.user.is_authenticated else None,
            request=request,
            old_value={},  # No old value for creation
            new_value=contact_data_for_log,
            contact_id=contact
        )
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error creating contact: {error_details}")
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def contacts_bulk_create(request):
    """Bulk create contacts - optimized for migration/import"""
    try:
        # Ensure request.data is a list
        contacts_data = request.data
        if isinstance(contacts_data, dict):
            # If it's a dict, try to get the 'contacts' key or convert to list
            if 'contacts' in contacts_data:
                contacts_data = contacts_data['contacts']
            elif isinstance(contacts_data.get('data'), list):
                contacts_data = contacts_data['data']
            else:
                return Response({'error': 'Expected a list of contacts'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not isinstance(contacts_data, list):
            return Response({'error': 'Expected a list of contacts'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Convert to list if it's a tuple (can happen with some parsers)
        if isinstance(contacts_data, tuple):
            contacts_data = list(contacts_data)
        if not contacts_data:
            return Response({'error': 'No contacts provided'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        # Handle database connection errors during request processing
        from django.db import OperationalError
        import traceback
        error_details = traceback.format_exc()
        print(f"Error in contacts_bulk_create (request processing): {error_details}")
        
        if isinstance(e, OperationalError) or 'timeout' in str(e).lower() or 'connection' in str(e).lower():
            return Response({
                'error': 'Database connection timeout. Please try again in a moment.',
                'results': [],
                'total': 0,
                'success': 0,
                'failed': 0
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        
        return Response({
            'error': 'An error occurred while processing the request',
            'results': [],
            'total': 0,
            'success': 0,
            'failed': 0
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # Limit the number of contacts per request to avoid timeouts
    MAX_CONTACTS_PER_REQUEST = 500
    if len(contacts_data) > MAX_CONTACTS_PER_REQUEST:
        return Response({
            'error': f'Too many contacts. Maximum {MAX_CONTACTS_PER_REQUEST} contacts per request. Please split into smaller batches.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Helper functions
    def parse_date(date_str):
        """Parse date string to date object"""
        if not date_str or str(date_str).strip() == '':
            return None
        # If already a date object, return it
        if isinstance(date_str, date) and not isinstance(date_str, datetime):
            return date_str
        # If datetime object, extract date
        if isinstance(date_str, datetime):
            return date_str.date()
        date_str = str(date_str).strip()
        # Try common date formats
        formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None
    
    def parse_datetime(datetime_str):
        """Parse datetime string to timezone-aware datetime object"""
        if not datetime_str or str(datetime_str).strip() == '':
            return None
        import pytz
        utc = pytz.UTC
        
        # If already a datetime object, make it timezone-aware and return
        if isinstance(datetime_str, datetime):
            if timezone.is_aware(datetime_str):
                return datetime_str
            # For naive datetime, assume UTC to avoid timezone offset issues
            return utc.localize(datetime_str)
        # If date object, convert to datetime
        if isinstance(datetime_str, date):
            dt = datetime.combine(datetime_str, datetime.min.time())
            # Assume UTC for date objects
            return utc.localize(dt)
        datetime_str = str(datetime_str).strip()
        
        # Try parsing ISO format with timezone first (e.g., "2024-01-01T10:00:00Z" or "2024-01-01T10:00:00+00:00")
        try:
            from dateutil import parser
            parsed = parser.parse(datetime_str)
            if timezone.is_aware(parsed):
                return parsed
            # If parsed as naive, assume UTC
            return utc.localize(parsed)
        except (ValueError, ImportError):
            pass
        
        # Try common datetime formats (assume UTC for naive datetimes)
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%S.%f',
            '%Y-%m-%dT%H:%M',
            '%Y-%m-%d',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y',
        ]
        for fmt in formats:
            try:
                parsed = datetime.strptime(datetime_str, fmt)
                # Assume UTC instead of server local timezone to avoid offset issues
                return utc.localize(parsed)
            except ValueError:
                continue
        return None
    
    def phone_to_int(value):
        if not value:
            return None
        cleaned = ''.join(str(value).split())
        if not cleaned:
            return None
        try:
            return int(cleaned)
        except (ValueError, TypeError):
            return None
    
    # Pre-fetch all related objects in bulk to avoid N+1 queries
    status_ids = set()
    source_ids = set()
    teleoperator_ids = set()
    confirmateur_ids = set()
    platform_ids = set()
    emails_to_check = set()
    old_contact_ids_to_check = set()
    
    for contact_data in contacts_data:
        if contact_data.get('statusId'):
            status_ids.add(contact_data['statusId'])
        if contact_data.get('sourceId'):
            source_ids.add(contact_data['sourceId'])
        if contact_data.get('teleoperatorId'):
            # Normalize to string and strip whitespace for consistent lookup
            teleoperator_ids.add(str(contact_data['teleoperatorId']).strip())
        if contact_data.get('confirmateurId'):
            # Normalize to string and strip whitespace for consistent lookup
            confirmateur_ids.add(str(contact_data['confirmateurId']).strip())
        if contact_data.get('platformId'):
            platform_ids.add(contact_data['platformId'])
        email = contact_data.get('email', '').strip()
        if email:
            emails_to_check.add(email)
        # Collect old_contact_ids for update check
        old_contact_id = contact_data.get('oldContactId', '').strip() if contact_data.get('oldContactId') else ''
        if old_contact_id:
            old_contact_ids_to_check.add(old_contact_id)
    
    # Bulk fetch all related objects
    statuses_dict = {s.id: s for s in Status.objects.filter(id__in=status_ids)} if status_ids else {}
    sources_dict = {s.id: s for s in Source.objects.filter(id__in=source_ids)} if source_ids else {}
    platforms_dict = {p.id: p for p in Platform.objects.filter(id__in=platform_ids)} if platform_ids else {}
    
    # Fetch users (both Django User and UserDetails)
    # IMPORTANT: Frontend sends UserDetails.id (string), not Django User.id (integer)
    # We need to prioritize UserDetails lookup first, then fall back to Django User ID if numeric
    teleoperator_users_dict = {}
    confirmateur_users_dict = {}
    if teleoperator_ids or confirmateur_ids:
        from api.models import UserDetails
        
        # Step 1: Try UserDetails IDs first (strings) - this is what the frontend sends
        # Convert all IDs to strings for UserDetails lookup (UserDetails.id is CharField)
        user_details_id_list = [str(uid).strip() for uid in list(teleoperator_ids) + list(confirmateur_ids)]
        user_details_list = UserDetails.objects.filter(id__in=user_details_id_list).select_related('django_user')
        
        # Track which IDs we found via UserDetails
        found_via_userdetails = set()
        
        for ud in user_details_list:
            if not ud.django_user:
                continue
                
            ud_id_str = str(ud.id).strip()
            ud_id_original = str(ud.id)
            django_user = ud.django_user
            
            # Store in teleoperator dict if this UserDetails ID was requested for teleoperator
            for uid in teleoperator_ids:
                uid_str = str(uid).strip()
                uid_original = str(uid)
                
                # Exact match: UserDetails ID matches the requested ID
                if uid_str == ud_id_str or uid_original == ud_id_original or uid == ud.id:
                    # Store with all possible key formats for reliable lookup
                    teleoperator_users_dict[uid_str] = django_user
                    teleoperator_users_dict[ud_id_str] = django_user
                    if uid_original != uid_str:
                        teleoperator_users_dict[uid_original] = django_user
                    if ud_id_original != ud_id_str:
                        teleoperator_users_dict[ud_id_original] = django_user
                    found_via_userdetails.add(uid_str)
                    found_via_userdetails.add(uid_original)
                    break
            
            # Store in confirmateur dict if this UserDetails ID was requested for confirmateur
            for uid in confirmateur_ids:
                uid_str = str(uid).strip()
                uid_original = str(uid)
                
                # Exact match: UserDetails ID matches the requested ID
                if uid_str == ud_id_str or uid_original == ud_id_original or uid == ud.id:
                    # Store with all possible key formats for reliable lookup
                    confirmateur_users_dict[uid_str] = django_user
                    confirmateur_users_dict[ud_id_str] = django_user
                    if uid_original != uid_str:
                        confirmateur_users_dict[uid_original] = django_user
                    if ud_id_original != ud_id_str:
                        confirmateur_users_dict[ud_id_original] = django_user
                    found_via_userdetails.add(uid_str)
                    found_via_userdetails.add(uid_original)
                    break
        
        # Step 2: Fallback to Django User IDs only for IDs not found via UserDetails
        # This handles cases where someone might send Django User ID directly
        django_user_ids_to_try = []
        for uid in list(teleoperator_ids) + list(confirmateur_ids):
            uid_str = str(uid).strip()
            # Only try Django User ID lookup if:
            # 1. Not already found via UserDetails
            # 2. Can be converted to integer (Django User IDs are integers)
            if uid_str not in found_via_userdetails:
                try:
                    django_user_ids_to_try.append(int(uid))
                except (ValueError, TypeError):
                    pass
        
        if django_user_ids_to_try:
            django_users = {u.id: u for u in DjangoUser.objects.filter(id__in=django_user_ids_to_try)}
            
            # Add Django users to dictionaries for IDs not found via UserDetails
            for uid in teleoperator_ids:
                uid_str = str(uid).strip()
                if uid_str not in found_via_userdetails:
                    try:
                        int_uid = int(uid)
                        if int_uid in django_users:
                            teleoperator_users_dict[uid_str] = django_users[int_uid]
                            teleoperator_users_dict[int_uid] = django_users[int_uid]
                    except (ValueError, TypeError):
                        pass
            
            for uid in confirmateur_ids:
                uid_str = str(uid).strip()
                if uid_str not in found_via_userdetails:
                    try:
                        int_uid = int(uid)
                        if int_uid in django_users:
                            confirmateur_users_dict[uid_str] = django_users[int_uid]
                            confirmateur_users_dict[int_uid] = django_users[int_uid]
                    except (ValueError, TypeError):
                        pass
    
    # Pre-fetch existing contacts by old_contact_id for updates
    # NOTE: We ONLY match by old_contact_id, NOT by email (duplicate emails are allowed)
    existing_contacts_by_old_id = {}
    if old_contact_ids_to_check:
        existing_contacts = Contact.objects.filter(old_contact_id__in=old_contact_ids_to_check).exclude(old_contact_id__isnull=True).exclude(old_contact_id='')
        for contact in existing_contacts:
            if contact.old_contact_id:
                existing_contacts_by_old_id[contact.old_contact_id.strip()] = contact
    
    # Track emails seen in this batch to detect CSV duplicates
    emails_seen_in_batch = {}
    
    # Prepare contacts for bulk creation and updates
    contacts_to_create = []
    contacts_to_update = []
    results = []
    from django.utils import timezone
    
    for idx, contact_data in enumerate(contacts_data):
        try:
            # Validate required fields - only statusId is required
            if not contact_data.get('statusId'):
                results.append({'row': idx, 'success': False, 'error': 'Le statut est requis'})
                continue
            
            # Handle old_contact_id and check for existing contact
            old_contact_id_value = contact_data.get('oldContactId', '') or ''
            old_contact_id_clean = old_contact_id_value.strip() if old_contact_id_value.strip() else None
            
            # Check if contact with this old_contact_id already exists
            # ONLY match by old_contact_id, NOT by email (duplicate emails are allowed)
            existing_contact = None
            if old_contact_id_clean:
                existing_contact = existing_contacts_by_old_id.get(old_contact_id_clean)
            
            # Check email only for CSV duplicate detection (not for matching existing contacts)
            email = contact_data.get('email', '').strip()
            is_duplicate = False
            duplicate_reason = None
            
            if email:
                email_lower = email.lower()
                # Check if this email appears multiple times in the CSV batch
                if email_lower in emails_seen_in_batch:
                    is_duplicate = True
                    duplicate_reason = f'L\'email {email} apparaît plusieurs fois dans le CSV (première occurrence à l\'index {emails_seen_in_batch[email_lower]})'
                else:
                    # Mark this email as seen in this batch
                    emails_seen_in_batch[email_lower] = idx
            
            if existing_contact:
                # Update existing contact instead of creating new one
                # Update all fields from contact_data
                existing_contact.civility = contact_data.get('civility', '') or ''
                existing_contact.fname = contact_data.get('firstName', '') or ''
                existing_contact.lname = contact_data.get('lastName', '') or ''
                existing_contact.phone = phone_to_int(contact_data.get('phone', ''))
                existing_contact.mobile = phone_to_int(contact_data.get('mobile', ''))
                existing_contact.email = email
                existing_contact.birth_date = parse_date(contact_data.get('birthDate'))
                existing_contact.birth_place = contact_data.get('birthPlace', '') or ''
                existing_contact.address = contact_data.get('address', '') or ''
                existing_contact.address_complement = contact_data.get('addressComplement', '') or ''
                existing_contact.postal_code = contact_data.get('postalCode', '') or ''
                existing_contact.city = contact_data.get('city', '') or ''
                existing_contact.nationality = contact_data.get('nationality', '') or ''
                existing_contact.campaign = contact_data.get('campaign', '') or ''
                
                # Update old_contact_id if provided (useful for migration scenarios)
                if old_contact_id_clean:
                    existing_contact.old_contact_id = old_contact_id_clean
                
                # Update status
                status_id = contact_data.get('statusId')
                if status_id and status_id in statuses_dict:
                    # Get old status type directly from database to ensure we have the correct value
                    old_status_type = None
                    if existing_contact.status_id:
                        old_status = Status.objects.filter(id=existing_contact.status_id).values_list('type', flat=True).first()
                        if old_status:
                            old_status_type = old_status
                    
                    # Get new status type directly from database
                    new_status_type = Status.objects.filter(id=status_id).values_list('type', flat=True).first()
                    
                    # If moving from lead to client, set date_lead_to_client (only if not already set)
                    if old_status_type == 'lead' and new_status_type == 'client':
                        if not existing_contact.date_lead_to_client:
                            from django.utils import timezone as tz
                            existing_contact.date_lead_to_client = tz.now()
                    
                    existing_contact.status = statuses_dict[status_id]
                elif 'statusId' in contact_data and not status_id:
                    existing_contact.status = None
                
                # Update source
                source_id = contact_data.get('sourceId')
                if source_id and source_id in sources_dict:
                    existing_contact.source = sources_dict[source_id]
                elif 'sourceId' in contact_data and not source_id:
                    existing_contact.source = None
                
                # Update teleoperator - normalize ID for lookup
                teleoperator_id_raw = contact_data.get('teleoperatorId')
                if teleoperator_id_raw:
                    teleoperator_id = str(teleoperator_id_raw).strip()
                    # Try exact match first
                    if teleoperator_id in teleoperator_users_dict:
                        existing_contact.teleoperator = teleoperator_users_dict[teleoperator_id]
                        if not existing_contact.assigned_at:
                            existing_contact.assigned_at = timezone.now()
                    else:
                        # Try as integer if it's numeric
                        try:
                            int_id = int(teleoperator_id)
                            if int_id in teleoperator_users_dict:
                                existing_contact.teleoperator = teleoperator_users_dict[int_id]
                                if not existing_contact.assigned_at:
                                    existing_contact.assigned_at = timezone.now()
                        except (ValueError, TypeError):
                            pass
                elif 'teleoperatorId' in contact_data:
                    # Explicitly set to None if field is present but empty
                    existing_contact.teleoperator = None
                
                # Update confirmateur - normalize ID for lookup
                confirmateur_id_raw = contact_data.get('confirmateurId')
                if confirmateur_id_raw:
                    confirmateur_id = str(confirmateur_id_raw).strip()
                    confirmateur_found = False
                    
                    # Try exact match first (normalized string) - this should be the primary lookup
                    if confirmateur_id in confirmateur_users_dict:
                        existing_contact.confirmateur = confirmateur_users_dict[confirmateur_id]
                        confirmateur_found = True
                    else:
                        # Try original format (without strip) before trying integer
                        confirmateur_id_original = str(confirmateur_id_raw).strip()
                        if confirmateur_id_original != confirmateur_id and confirmateur_id_original in confirmateur_users_dict:
                            existing_contact.confirmateur = confirmateur_users_dict[confirmateur_id_original]
                            confirmateur_found = True
                        
                        # Try as integer if it's numeric and not found yet
                        if not confirmateur_found:
                            try:
                                int_id = int(confirmateur_id)
                                if int_id in confirmateur_users_dict:
                                    existing_contact.confirmateur = confirmateur_users_dict[int_id]
                                    confirmateur_found = True
                            except (ValueError, TypeError):
                                pass
                        
                        # Last resort: try without strip
                        if not confirmateur_found:
                            confirmateur_id_no_strip = str(confirmateur_id_raw)
                            if confirmateur_id_no_strip != confirmateur_id and confirmateur_id_no_strip in confirmateur_users_dict:
                                existing_contact.confirmateur = confirmateur_users_dict[confirmateur_id_no_strip]
                                confirmateur_found = True
                elif 'confirmateurId' in contact_data:
                    # Explicitly set to None if field is present but empty
                    existing_contact.confirmateur = None
                
                # Update platform
                platform_id = contact_data.get('platformId')
                if platform_id and platform_id in platforms_dict:
                    existing_contact.platform = platforms_dict[platform_id]
                elif 'platformId' in contact_data and not platform_id:
                    existing_contact.platform = None
                
                # Update custom fields
                if 'montantEncaisse' in contact_data:
                    existing_contact.montant_encaisse = contact_data.get('montantEncaisse') if contact_data.get('montantEncaisse') else None
                if 'bonus' in contact_data:
                    existing_contact.bonus = contact_data.get('bonus') if contact_data.get('bonus') else None
                if 'paiement' in contact_data:
                    existing_contact.paiement = contact_data.get('paiement', '') or ''
                if 'contrat' in contact_data:
                    existing_contact.contrat = contact_data.get('contrat', '') or ''
                if 'nomDeScene' in contact_data:
                    existing_contact.nom_de_scene = contact_data.get('nomDeScene', '') or ''
                if 'dateProTr' in contact_data:
                    existing_contact.date_pro_tr = str(contact_data.get('dateProTr')).strip() if contact_data.get('dateProTr') else ''
                if 'potentiel' in contact_data:
                    existing_contact.potentiel = contact_data.get('potentiel', '') or ''
                if 'produit' in contact_data:
                    existing_contact.produit = contact_data.get('produit', '') or ''
                
                # Update confirmateur email and telephone
                if 'confirmateurEmail' in contact_data:
                    existing_contact.confirmateur_email = contact_data.get('confirmateurEmail', '').strip()
                if 'confirmateurTelephone' in contact_data:
                    existing_contact.confirmateur_telephone = contact_data.get('confirmateurTelephone', '').strip()
                
                # Update datetime fields
                if contact_data.get('createdAt'):
                    parsed_dt = parse_datetime(contact_data.get('createdAt'))
                    if parsed_dt:
                        existing_contact.created_at = parsed_dt
                if contact_data.get('updatedAt'):
                    parsed_dt = parse_datetime(contact_data.get('updatedAt'))
                    if parsed_dt:
                        existing_contact.updated_at = parsed_dt
                if contact_data.get('assignedAt'):
                    parsed_dt = parse_datetime(contact_data.get('assignedAt'))
                    if parsed_dt:
                        existing_contact.assigned_at = parsed_dt
                
                # Track which fields were updated
                updated_fields = []
                # Track all fields that are being updated (presence in contact_data means update)
                if 'civility' in contact_data:
                    updated_fields.append('civility')
                if 'firstName' in contact_data:
                    updated_fields.append('firstName')
                if 'lastName' in contact_data:
                    updated_fields.append('lastName')
                if 'email' in contact_data:
                    updated_fields.append('email')
                if 'phone' in contact_data:
                    updated_fields.append('phone')
                if 'mobile' in contact_data:
                    updated_fields.append('mobile')
                if 'birthDate' in contact_data:
                    updated_fields.append('birthDate')
                if 'birthPlace' in contact_data:
                    updated_fields.append('birthPlace')
                if 'address' in contact_data:
                    updated_fields.append('address')
                if 'addressComplement' in contact_data:
                    updated_fields.append('addressComplement')
                if 'postalCode' in contact_data:
                    updated_fields.append('postalCode')
                if 'city' in contact_data:
                    updated_fields.append('city')
                if 'nationality' in contact_data:
                    updated_fields.append('nationality')
                if 'campaign' in contact_data:
                    updated_fields.append('campaign')
                if 'statusId' in contact_data:
                    updated_fields.append('statusId')
                if 'sourceId' in contact_data:
                    updated_fields.append('sourceId')
                if 'teleoperatorId' in contact_data:
                    updated_fields.append('teleoperatorId')
                if 'confirmateurId' in contact_data:
                    updated_fields.append('confirmateurId')
                if 'platformId' in contact_data:
                    updated_fields.append('platformId')
                if 'createdAt' in contact_data:
                    updated_fields.append('createdAt')
                if 'updatedAt' in contact_data:
                    updated_fields.append('updatedAt')
                if 'assignedAt' in contact_data:
                    updated_fields.append('assignedAt')
                if 'montantEncaisse' in contact_data:
                    updated_fields.append('montantEncaisse')
                if 'bonus' in contact_data:
                    updated_fields.append('bonus')
                if 'paiement' in contact_data:
                    updated_fields.append('paiement')
                if 'contrat' in contact_data:
                    updated_fields.append('contrat')
                if 'nomDeScene' in contact_data:
                    updated_fields.append('nomDeScene')
                if 'dateProTr' in contact_data:
                    updated_fields.append('dateProTr')
                if 'potentiel' in contact_data:
                    updated_fields.append('potentiel')
                if 'produit' in contact_data:
                    updated_fields.append('produit')
                if 'confirmateurEmail' in contact_data:
                    updated_fields.append('confirmateurEmail')
                if 'confirmateurTelephone' in contact_data:
                    updated_fields.append('confirmateurTelephone')
                
                contacts_to_update.append(existing_contact)
                result_entry = {
                    'row': idx, 
                    'success': True, 
                    'contactId': existing_contact.id,
                    'updated': True,
                    'updatedFields': updated_fields,
                    'contactName': f"{existing_contact.fname} {existing_contact.lname}".strip() or 'N/A',
                    'contactEmail': existing_contact.email or 'N/A',
                    'oldContactId': old_contact_id_clean
                }
                if is_duplicate:
                    result_entry['duplicate'] = True
                    result_entry['duplicateReason'] = duplicate_reason
                results.append(result_entry)
                continue  # Skip creation logic for this contact
            
            # No existing contact found, proceed with creation
            # Generate contact ID
            contact_id = uuid.uuid4().hex[:12]
            while Contact.objects.filter(id=contact_id).exists() or any(c[0].id == contact_id for c in contacts_to_create):
                contact_id = uuid.uuid4().hex[:12]
            
            # Build contact data
            contact_obj_data = {
                'id': contact_id,
                'civility': contact_data.get('civility', '') or '',
                'fname': contact_data.get('firstName', '') or '',
                'lname': contact_data.get('lastName', '') or '',
                'phone': phone_to_int(contact_data.get('phone', '')),
                'mobile': phone_to_int(contact_data.get('mobile', '')),
                'email': email,
                'birth_date': parse_date(contact_data.get('birthDate')),
                'birth_place': contact_data.get('birthPlace', '') or '',
                'address': contact_data.get('address', '') or '',
                'address_complement': contact_data.get('addressComplement', '') or '',
                'postal_code': contact_data.get('postalCode', '') or '',
                'city': contact_data.get('city', '') or '',
                'nationality': contact_data.get('nationality', '') or '',
                'campaign': contact_data.get('campaign', '') or '',
                'creator': request.user,
                'old_contact_id': old_contact_id_clean,
            }
            
            # Set status
            status_id = contact_data.get('statusId')
            if status_id and status_id in statuses_dict:
                contact_obj_data['status'] = statuses_dict[status_id]
            
            # Set source
            source_id = contact_data.get('sourceId')
            if source_id and source_id in sources_dict:
                contact_obj_data['source'] = sources_dict[source_id]
            
            # Handle datetime fields FIRST (before setting teleoperator assigned_at)
            # This ensures CSV values take precedence
            if contact_data.get('createdAt'):
                parsed_dt = parse_datetime(contact_data.get('createdAt'))
                if parsed_dt:
                    contact_obj_data['created_at'] = parsed_dt
            if contact_data.get('updatedAt'):
                parsed_dt = parse_datetime(contact_data.get('updatedAt'))
                if parsed_dt:
                    contact_obj_data['updated_at'] = parsed_dt
            if contact_data.get('assignedAt'):
                parsed_dt = parse_datetime(contact_data.get('assignedAt'))
                if parsed_dt:
                    contact_obj_data['assigned_at'] = parsed_dt
            
            # Set teleoperator - normalize ID for lookup
            teleoperator_id_raw = contact_data.get('teleoperatorId')
            if teleoperator_id_raw:
                teleoperator_id = str(teleoperator_id_raw).strip()
                # Try exact match first
                if teleoperator_id in teleoperator_users_dict:
                    contact_obj_data['teleoperator'] = teleoperator_users_dict[teleoperator_id]
                    # Only set assigned_at to now() if not already set from CSV
                    if 'assigned_at' not in contact_obj_data:
                        contact_obj_data['assigned_at'] = timezone.now()
                else:
                    # Try as integer if it's numeric
                    try:
                        int_id = int(teleoperator_id)
                        if int_id in teleoperator_users_dict:
                            contact_obj_data['teleoperator'] = teleoperator_users_dict[int_id]
                            # Only set assigned_at to now() if not already set from CSV
                            if 'assigned_at' not in contact_obj_data:
                                contact_obj_data['assigned_at'] = timezone.now()
                    except (ValueError, TypeError):
                        pass
            
            # Set confirmateur - normalize ID for lookup
            confirmateur_id_raw = contact_data.get('confirmateurId')
            if confirmateur_id_raw:
                confirmateur_id = str(confirmateur_id_raw).strip()
                confirmateur_found = False
                
                # Try exact match first (normalized string) - this should be the primary lookup
                if confirmateur_id in confirmateur_users_dict:
                    contact_obj_data['confirmateur'] = confirmateur_users_dict[confirmateur_id]
                    confirmateur_found = True
                else:
                    # Try original format (without strip) before trying integer
                    confirmateur_id_original = str(confirmateur_id_raw).strip()
                    if confirmateur_id_original != confirmateur_id and confirmateur_id_original in confirmateur_users_dict:
                        contact_obj_data['confirmateur'] = confirmateur_users_dict[confirmateur_id_original]
                        confirmateur_found = True
                    
                    # Try as integer if it's numeric and not found yet
                    if not confirmateur_found:
                        try:
                            int_id = int(confirmateur_id)
                            if int_id in confirmateur_users_dict:
                                contact_obj_data['confirmateur'] = confirmateur_users_dict[int_id]
                                confirmateur_found = True
                        except (ValueError, TypeError):
                            pass
                    
                    # Last resort: try without strip
                    if not confirmateur_found:
                        confirmateur_id_no_strip = str(confirmateur_id_raw)
                        if confirmateur_id_no_strip != confirmateur_id and confirmateur_id_no_strip in confirmateur_users_dict:
                            contact_obj_data['confirmateur'] = confirmateur_users_dict[confirmateur_id_no_strip]
                            confirmateur_found = True
            
            # Set platform
            platform_id = contact_data.get('platformId')
            if platform_id and platform_id in platforms_dict:
                contact_obj_data['platform'] = platforms_dict[platform_id]
            
            # Handle custom fields
            if contact_data.get('montantEncaisse'):
                contact_obj_data['montant_encaisse'] = contact_data.get('montantEncaisse')
            if contact_data.get('bonus'):
                contact_obj_data['bonus'] = contact_data.get('bonus')
            if contact_data.get('paiement'):
                contact_obj_data['paiement'] = contact_data.get('paiement')
            if contact_data.get('contrat'):
                contact_obj_data['contrat'] = contact_data.get('contrat')
            if contact_data.get('nomDeScene'):
                contact_obj_data['nom_de_scene'] = contact_data.get('nomDeScene')
            if contact_data.get('dateProTr'):
                # date_pro_tr is a CharField, so store as string
                contact_obj_data['date_pro_tr'] = str(contact_data.get('dateProTr')).strip() if contact_data.get('dateProTr') else ''
            if contact_data.get('potentiel'):
                contact_obj_data['potentiel'] = contact_data.get('potentiel')
            if contact_data.get('produit'):
                contact_obj_data['produit'] = contact_data.get('produit')
            # Handle confirmateur email and telephone
            if contact_data.get('confirmateurEmail'):
                contact_obj_data['confirmateur_email'] = contact_data.get('confirmateurEmail', '').strip()
            if contact_data.get('confirmateurTelephone'):
                contact_obj_data['confirmateur_telephone'] = contact_data.get('confirmateurTelephone', '').strip()
            
            contacts_to_create.append((Contact(**contact_obj_data), contact_data))
            result_entry = {'row': idx, 'success': True, 'contactId': contact_id}
            if is_duplicate:
                result_entry['duplicate'] = True
                result_entry['duplicateReason'] = duplicate_reason
            results.append(result_entry)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error preparing contact {idx}: {error_details}")
            results.append({'row': idx, 'success': False, 'error': str(e)})
    
    # Bulk update existing contacts
    if contacts_to_update:
        try:
            from django.db import transaction
            UPDATE_BATCH_SIZE = 50  # Smaller batch size to avoid timeouts
            
            # Process updates in batches within a transaction
            with transaction.atomic():
                for i in range(0, len(contacts_to_update), UPDATE_BATCH_SIZE):
                    batch = contacts_to_update[i:i + UPDATE_BATCH_SIZE]
                    Contact.objects.bulk_update(
                        batch,
                        [
                            'civility', 'fname', 'lname', 'phone', 'mobile', 'email',
                            'birth_date', 'birth_place', 'address', 'address_complement',
                            'postal_code', 'city', 'nationality', 'campaign',
                            'status', 'source', 'teleoperator', 'confirmateur', 'platform',
                            'montant_encaisse', 'bonus', 'paiement', 'contrat',
                            'nom_de_scene', 'date_pro_tr', 'potentiel', 'produit',
                            'confirmateur_email', 'confirmateur_telephone',
                            'created_at', 'updated_at', 'assigned_at'
                        ],
                        batch_size=UPDATE_BATCH_SIZE
                    )
            
            # Note: Log entries are skipped for bulk operations to improve performance
            # If logging is needed, it can be added as a background task
            
        except Exception as e:
            import traceback
            from django.db import OperationalError
            error_details = traceback.format_exc()
            print(f"Error bulk updating contacts: {error_details}")
            
            # Handle database connection errors
            if isinstance(e, OperationalError) or 'timeout' in str(e).lower() or 'connection' in str(e).lower():
                return Response({
                    'error': 'Database connection timeout. Please try again in a moment.',
                    'results': results,
                    'total': len(contacts_data),
                    'success': sum(1 for r in results if r.get('success')),
                    'failed': len(contacts_data)
                }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            
            # Mark update results as failed
            for result in results:
                if result.get('updated') and result.get('success'):
                    result['success'] = False
                    result['error'] = f'Bulk update failed: {str(e)}'
                    result.pop('updated', None)
    
    # Bulk create contacts in smaller batches to avoid timeouts
    if contacts_to_create:
        try:
            from django.db import transaction
            BATCH_SIZE = 50  # Smaller batch size to avoid timeouts
            
            contacts_objects = [c[0] for c in contacts_to_create]
            
            # Process in batches within a transaction
            with transaction.atomic():
                for i in range(0, len(contacts_objects), BATCH_SIZE):
                    batch = contacts_objects[i:i + BATCH_SIZE]
                    Contact.objects.bulk_create(batch, batch_size=BATCH_SIZE)
            
            # Note: Log entries are skipped for bulk operations to improve performance
            # If logging is needed, it can be added as a background task
            
        except Exception as e:
            import traceback
            from django.db import OperationalError
            error_details = traceback.format_exc()
            print(f"Error bulk creating contacts: {error_details}")
            
            # Handle database connection errors
            if isinstance(e, OperationalError) or 'timeout' in str(e).lower() or 'connection' in str(e).lower():
                return Response({
                    'error': 'Database connection timeout. Please try again in a moment.',
                    'results': results,
                    'total': len(contacts_data),
                    'success': sum(1 for r in results if r.get('success')),
                    'failed': len(contacts_data)
                }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            
            # Mark all as failed
            for result in results:
                if result.get('success'):
                    result['success'] = False
                    result['error'] = f'Bulk creation failed: {str(e)}'
    
    # Count updated vs created contacts
    updated_count = sum(1 for r in results if r.get('success') and r.get('updated'))
    created_count = sum(1 for r in results if r.get('success') and not r.get('updated'))
    
    return Response({
        'results': results,
        'total': len(contacts_data),
        'success': sum(1 for r in results if r.get('success')),
        'failed': sum(1 for r in results if not r.get('success')),
        'created': created_count,
        'updated': updated_count
    }, status=status.HTTP_201_CREATED)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def csv_import_preview(request):
    """Preview CSV or Excel file and return headers and sample rows"""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    file_name = file.name.lower()
    
    # Check file extension
    is_csv = file_name.endswith('.csv')
    is_xlsx = file_name.endswith('.xlsx')
    is_xls = file_name.endswith('.xls')
    is_excel = is_xlsx or is_xls
    
    if not is_csv and not is_excel:
        return Response({'error': 'File must be a CSV or Excel file (.csv, .xlsx)'}, status=status.HTTP_400_BAD_REQUEST)
    
    # .xls files (old Excel format) are not supported - only .xlsx
    if is_xls:
        return Response({
            'error': 'Old Excel format (.xls) is not supported. Please save your file as .xlsx format and try again.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        if is_excel:
            # Handle Excel file
            from openpyxl import load_workbook
            
            # Reset file pointer and read into BytesIO buffer
            # CRITICAL: Must read in binary mode - Django UploadedFile.read() returns bytes by default
            # but we need to ensure we're getting raw binary data, not decoded text
            file.seek(0)
            
            # Django's UploadedFile.read() should return bytes, but let's be explicit
            # Check if file has chunks() method (more reliable for binary)
            if hasattr(file, 'chunks'):
                # Read file in chunks to ensure binary mode
                file_content = b''.join(file.chunks())
            else:
                # Fallback: read directly
                file_content = file.read()
            
            # Ensure we have bytes, not a string
            if isinstance(file_content, str):
                # If somehow we got a string, this is wrong - log it
                logger.error(f"ERROR: File content is string, not bytes! First 50 chars: {file_content[:50]}")
                # Try to get binary content properly
                file.seek(0)
                if hasattr(file, 'open'):
                    try:
                        with file.open('rb') as f:
                            file_content = f.read()
                    except:
                        # Last resort: encode the string (but this is wrong)
                        file_content = file_content.encode('latin-1')
                else:
                    file_content = file_content.encode('latin-1')
            
            # Validate file content
            if not file_content:
                return Response({'error': 'Excel file is empty'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Ensure file_content is bytes (Django UploadedFile should return bytes, but check to be safe)
            if isinstance(file_content, str):
                file_content = file_content.encode('latin-1')  # Preserve binary data
            
            # Validate that we have actual bytes
            if not isinstance(file_content, bytes):
                return Response({
                    'error': f'Invalid file content type: {type(file_content)}. Expected bytes.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check file size
            if len(file_content) < 100:  # Excel files should be at least a few KB
                return Response({
                    'error': f'File is too small ({len(file_content)} bytes). Excel files should be larger.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Check magic bytes for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Excel import: file size={len(file_content)}, first 10 bytes={file_content[:10]}")
            
            file_buffer = BytesIO(file_content)
            file_buffer.seek(0)  # Ensure buffer is at the start
            
            # Try to load the workbook - let openpyxl handle validation
            try:
                workbook = load_workbook(file_buffer, read_only=True, data_only=True)
            except Exception as e:
                # If it fails, provide a helpful error message
                error_msg = str(e)
                if 'not a zip file' in error_msg.lower() or 'bad zipfile' in error_msg.lower():
                    return Response({
                        'error': 'File is not a valid Excel file. Please ensure you are uploading a .xlsx file (not .xls or another format). The file may be corrupted or in an unsupported format.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({
                        'error': f'Failed to read Excel file: {error_msg}. Please ensure the file is a valid .xlsx file and not corrupted.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Use the first sheet
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            if sheet.max_row > 0:
                first_row = sheet[1]
                for i, cell in enumerate(first_row):
                    value = cell.value
                    if value and str(value).strip():
                        headers.append(str(value).strip())
                    else:
                        headers.append(f'Colonne_{i+1}')
            
            # Get first 5 rows as preview
            preview_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=False), start=2):
                if row_idx > 6:
                    break
                row_dict = {}
                for i, cell in enumerate(row):
                    header = headers[i] if i < len(headers) else f'Colonne_{i+1}'
                    value = cell.value
                    row_dict[header] = str(value) if value is not None else ''
                preview_rows.append(row_dict)
            
            # Count total rows (excluding header)
            total_rows = sheet.max_row - 1 if sheet.max_row > 1 else 0
            
            workbook.close()
            
            return Response({
                'headers': headers,
                'preview': preview_rows,
                'totalRows': total_rows
            }, status=status.HTTP_200_OK)
        else:
            # Read CSV file - accept any CSV format
            file_content = file.read().decode('utf-8-sig')  # Handle BOM
            csv_reader = csv.DictReader(io.StringIO(file_content))
            
            # Get headers - accept any header names
            headers = csv_reader.fieldnames or []
            
            # Clean headers (remove whitespace, handle empty headers)
            cleaned_headers = []
            for i, header in enumerate(headers):
                if header and header.strip():
                    cleaned_headers.append(header.strip())
                else:
                    # If header is empty, create a placeholder
                    cleaned_headers.append(f'Colonne_{i+1}')
            
            # Get first 5 rows as preview
            preview_rows = []
            csv_reader_preview = csv.DictReader(io.StringIO(file_content))
            for i, row in enumerate(csv_reader_preview):
                if i >= 5:
                    break
                preview_rows.append(row)
            
            # Count total rows (excluding header)
            file_content_for_count = io.StringIO(file_content)
            total_rows = sum(1 for _ in csv.DictReader(file_content_for_count))
            
            return Response({
                'headers': cleaned_headers if cleaned_headers else headers,
                'preview': preview_rows,
                'totalRows': total_rows
            }, status=status.HTTP_200_OK)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def csv_import_contacts(request):
    """Import contacts from CSV or Excel with column mapping - optimized for large imports"""
    import logging
    logger = logging.getLogger(__name__)
    
    if 'file' not in request.FILES:
        logger.warning("CSV import: No file provided")
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    file_name = file.name.lower()
    logger.info(f"CSV import: Processing file '{file.name}'")
    
    # Check file extension
    is_csv = file_name.endswith('.csv')
    is_xlsx = file_name.endswith('.xlsx')
    is_xls = file_name.endswith('.xls')
    is_excel = is_xlsx or is_xls
    
    if not is_csv and not is_excel:
        logger.warning(f"CSV import: Invalid file type '{file_name}'")
        return Response({'error': 'File must be a CSV or Excel file (.csv, .xlsx)'}, status=status.HTTP_400_BAD_REQUEST)
    
    # .xls files (old Excel format) are not supported - only .xlsx
    if is_xls:
        logger.warning(f"CSV import: Old Excel format (.xls) not supported for '{file_name}'")
        return Response({
            'error': 'Old Excel format (.xls) is not supported. Please save your file as .xlsx format and try again.'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Parse column mapping JSON string if it's a string
    column_mapping_str = request.data.get('columnMapping', '{}')
    if isinstance(column_mapping_str, str):
        import json
        try:
            column_mapping = json.loads(column_mapping_str)
        except json.JSONDecodeError:
            column_mapping = {}
    else:
        column_mapping = column_mapping_str or {}
    
    default_status_id = request.data.get('defaultStatusId')
    default_source_id = request.data.get('defaultSourceId')
    default_teleoperator_id = request.data.get('defaultTeleoperatorId')
    include_first_row = request.data.get('includeFirstRow', 'false').lower() == 'true'
    
    logger.info(f"CSV import: default_status_id='{default_status_id}', include_first_row={include_first_row}")
    
    # Clean and validate IDs (strip whitespace, handle empty strings)
    if default_status_id:
        default_status_id = str(default_status_id).strip()
        if not default_status_id:
            default_status_id = None
    
    if default_source_id:
        default_source_id = str(default_source_id).strip()
        if not default_source_id:
            default_source_id = None
    
    if default_teleoperator_id:
        default_teleoperator_id = str(default_teleoperator_id).strip()
        if not default_teleoperator_id:
            default_teleoperator_id = None
    
    # Validate required mappings - only statusId is required (via default_status_id)
    # lastName is no longer required
    
    if not default_status_id:
        logger.warning(f"CSV import: Default status is required but not provided")
        return Response({'error': 'Default status is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    workbook = None
    try:
        # Read file content first to detect actual file type (not just extension)
        file.seek(0)
        
        # Django's UploadedFile.read() should return bytes, but let's be explicit
        # Check if file has chunks() method (more reliable for binary)
        if hasattr(file, 'chunks'):
            # Read file in chunks to ensure binary mode
            file_content = b''.join(file.chunks())
        else:
            # Fallback: read directly
            file_content = file.read()
        
        # Validate file content
        if not file_content:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Ensure we have bytes, not a string
        if isinstance(file_content, str):
            # If somehow we got a string, this is wrong - log it
            logger.error(f"ERROR: File content is string, not bytes! First 50 chars: {file_content[:50]}")
            # Try to get binary content properly
            file.seek(0)
            if hasattr(file, 'open'):
                try:
                    with file.open('rb') as f:
                        file_content = f.read()
                except:
                    # Last resort: encode the string (but this is wrong)
                    file_content = file_content.encode('latin-1')
            else:
                file_content = file_content.encode('latin-1')
        
        # Validate that we have actual bytes
        if not isinstance(file_content, bytes):
            return Response({
                'error': f'Invalid file content type: {type(file_content)}. Expected bytes.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check magic bytes to detect actual file type (not just extension)
        # Excel files (.xlsx) start with PK (ZIP signature: 0x50 0x4B)
        actual_is_excel = len(file_content) >= 2 and file_content[:2] == b'PK'
        actual_is_csv = False
        
        if not actual_is_excel:
            # Check if it looks like CSV text
            try:
                first_line = file_content[:100].decode('utf-8', errors='ignore')
                if ',' in first_line or ';' in first_line:
                    logger.warning(f"File appears to be CSV, not Excel. First line: {first_line[:100]}")
                    logger.info(f"File detected as CSV content (not Excel), despite .xlsx extension. Treating as CSV.")
                    actual_is_csv = True
                    is_excel = False
                    is_csv = True
            except:
                pass
        
        # Read file based on ACTUAL type (not just extension)
        if is_excel and actual_is_excel:
            from openpyxl import load_workbook
            
            # Check file size
            if len(file_content) < 100:  # Excel files should be at least a few KB
                return Response({
                    'error': f'File is too small ({len(file_content)} bytes). Excel files should be larger.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"Excel import: file size={len(file_content)}, first 10 bytes={file_content[:10]}")
            
            file_buffer = BytesIO(file_content)
            file_buffer.seek(0)  # Ensure buffer is at the start
            
            # Try to load the workbook - let openpyxl handle validation
            try:
                workbook = load_workbook(file_buffer, read_only=True, data_only=True)
            except Exception as e:
                # Log the full error for debugging
                import traceback
                error_details = traceback.format_exc()
                logger.error(f"Failed to load Excel workbook: {error_details}")
                
                # If it fails, provide a helpful error message
                error_msg = str(e)
                error_type = type(e).__name__
                
                # Include the actual error type and message in the response for debugging
                if 'not a zip file' in error_msg.lower() or 'bad zipfile' in error_msg.lower() or 'bad magic number' in error_msg.lower() or 'zipfile' in error_type.lower():
                    return Response({
                        'error': f'File is not a valid Excel file ({error_type}: {error_msg}). Please ensure you are uploading a .xlsx file (not .xls or another format). The file may be corrupted or in an unsupported format.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({
                        'error': f'Failed to read Excel file ({error_type}: {error_msg}). Please ensure the file is a valid .xlsx file and not corrupted.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            sheet = workbook.active
            
            # Get headers
            headers = []
            # Safely check max_row - handle None case
            max_row = sheet.max_row if sheet.max_row is not None else 0
            if max_row > 0:
                first_row = sheet[1]
                if include_first_row:
                    # Generate generic column headers - must match frontend format (Column1, Column2, etc.)
                    num_columns = len(first_row)
                    headers = [f'Column{i+1}' for i in range(num_columns)]
                else:
                    # Extract headers from first row
                    for i, cell in enumerate(first_row):
                        value = cell.value
                        if value and str(value).strip():
                            headers.append(str(value).strip())
                        else:
                            # Use Column1, Column2 format to match frontend
                            headers.append(f'Column{i+1}')
            
            # Convert Excel rows to dict-like format compatible with CSV reader
            # Use iter_rows() instead of indexing because read_only=True doesn't support indexing
            class ExcelDictReader:
                def __init__(self, sheet, headers, include_first_row=False):
                    self.sheet = sheet
                    self.headers = headers
                    # Start from row 1 if include_first_row is True, otherwise start from row 2 (skip header)
                    start_row = 1 if include_first_row else 2
                    # Handle None max_row - use sheet.max_row if available, otherwise let iter_rows find the end
                    self.max_row = sheet.max_row if sheet.max_row is not None else None
                    
                    # Create iterator using iter_rows() which works with read_only=True
                    # iter_rows() returns rows with cell objects, not values
                    # If max_row is None, don't pass it and let iter_rows iterate until the end
                    if self.max_row is not None:
                        self.row_iterator = sheet.iter_rows(
                            min_row=start_row,
                            max_row=self.max_row,
                            values_only=False
                        )
                    else:
                        self.row_iterator = sheet.iter_rows(
                            min_row=start_row,
                            values_only=False
                        )
                    self.iterator = iter(self.row_iterator)
                
                def __iter__(self):
                    return self
                
                def __next__(self):
                    try:
                        row = next(self.iterator)
                        row_dict = {}
                        for i, cell in enumerate(row):
                            header = self.headers[i] if i < len(self.headers) else f'Column{i+1}'
                            value = cell.value
                            row_dict[header] = str(value) if value is not None else ''
                        return row_dict
                    except StopIteration:
                        raise
                    except Exception as e:
                        # If we can't read the row, stop iteration
                        raise StopIteration
            
            file_reader = ExcelDictReader(sheet, headers, include_first_row)
        else:
            # Read CSV file - use already-read file_content if available, otherwise read from file
            if 'file_content' not in locals() or file_content is None:
                file.seek(0)
                file_content_bytes = file.read()
            else:
                file_content_bytes = file_content
            
            # Decode CSV content from bytes to string
            try:
                file_content = file_content_bytes.decode('utf-8-sig')  # Handle BOM
            except UnicodeDecodeError:
                # Try other encodings if UTF-8 fails
                try:
                    file_content = file_content_bytes.decode('latin-1')
                except:
                    file_content = file_content_bytes.decode('utf-8', errors='ignore')
            
            # For CSV files, handle include_first_row setting
            # If include_first_row is True, we need to generate headers and include first row as data
            # If include_first_row is False, treat first row as headers
            csv_lines = file_content.strip().split('\n')
            if not csv_lines:
                return Response({'error': 'CSV file is empty'}, status=status.HTTP_400_BAD_REQUEST)
            
            if include_first_row:
                # Generate generic headers and include first row as data
                first_row_values = csv_lines[0].split(',')
                num_columns = len(first_row_values)
                headers = [f'Column{i+1}' for i in range(num_columns)]
                # Create a reader that includes the first row
                all_lines = csv_lines  # Include first row
                logger.info(f"CSV import: Including first row as data, generated {num_columns} column headers")
            else:
                # Use first row as headers
                first_row = csv_lines[0]
                headers = [h.strip().strip('"') for h in first_row.split(',')]
                all_lines = csv_lines[1:]  # Skip first row
                logger.info(f"CSV import: Using first row as headers: {headers[:5]}...")
            
            # Create CSV reader with appropriate data
            csv_content = '\n'.join(all_lines)
            file_reader = csv.DictReader(io.StringIO(csv_content), fieldnames=headers if include_first_row else None)
        
        # Get status and source objects
        status_obj = Status.objects.filter(id=default_status_id).first()
        if not status_obj:
            return Response({'error': 'Invalid status ID'}, status=status.HTTP_400_BAD_REQUEST)
        
        source_obj = None
        if default_source_id:
            source_obj = Source.objects.filter(id=default_source_id).first()
            if not source_obj:
                # Log warning but don't fail - source is optional
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Source ID '{default_source_id}' not found during import")
        
        # Handle teleoperator - optional field
        # Only set if explicitly provided by user - do NOT auto-assign even if user is a teleoperateur
        teleoperator_obj = None
        if default_teleoperator_id:
            try:
                teleoperator_user = None
                # Prioritize UserDetails ID lookup first (since frontend sends UserDetails IDs)
                user_details = UserDetails.objects.filter(id=str(default_teleoperator_id)).select_related('django_user').only('id', 'django_user').first()
                if user_details and user_details.django_user:
                    teleoperator_user = user_details.django_user
                
                # Fallback: Try Django User ID lookup only if UserDetails lookup failed
                if not teleoperator_user:
                    try:
                        int_id = int(default_teleoperator_id)
                        teleoperator_user = DjangoUser.objects.filter(id=int_id).only('id').first()
                    except (ValueError, TypeError):
                        pass
                
                if teleoperator_user:
                    teleoperator_obj = teleoperator_user
                else:
                    # Log warning but don't fail - teleoperator is optional
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Teleoperator ID '{default_teleoperator_id}' not found during import")
            except Exception as e:
                # Log error but don't fail - teleoperator is optional
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Error setting teleoperator during CSV import: {e}")
        # Removed auto-assignment logic - if user removes teleoperator selection, respect that choice
        
        # Field mapping from frontend names to model field names
        field_mapping = {
            'civility': 'civility',
            'firstName': 'fname',
            'lastName': 'lname',
            'phone': 'phone',
            'mobile': 'mobile',
            'email': 'email',
            'birthDate': 'birth_date',
            'birthPlace': 'birth_place',
            'address': 'address',
            'addressComplement': 'address_complement',
            'postalCode': 'postal_code',
            'city': 'city',
            'nationality': 'nationality',
            'autreInformations': 'autre_informations',
            'dateInscription': 'date_d_inscription',
            'campaign': 'campaign',
            'oldContactId': 'old_contact_id',
            'createdAt': 'created_at',
            'updatedAt': 'updated_at',
            'assignedAt': 'assigned_at',
        }
        
        results = {
            'success': [],
            'errors': [],
            'duplicates': [],
            'total': 0,
            'imported': 0,
            'failed': 0
        }
        
        # Helper function to parse date
        def parse_date(date_str):
            if not date_str or date_str.strip() == '':
                return None
            date_str = date_str.strip()
            # Try common date formats
            formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None
        
        # Helper function to parse datetime
        def parse_datetime(datetime_str):
            if not datetime_str or datetime_str.strip() == '':
                return None
            datetime_str = datetime_str.strip()
            # Try common datetime formats
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%dT%H:%M:%S.%f',
                '%Y-%m-%dT%H:%M',
                '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y %H:%M',
                '%m/%d/%Y',
            ]
            for fmt in formats:
                try:
                    parsed = datetime.strptime(datetime_str, fmt)
                    from django.utils import timezone
                    return timezone.make_aware(parsed)
                except ValueError:
                    continue
            return None
        
        # Batch processing configuration
        BATCH_SIZE = 1000  # Process contacts in batches of 1000
        contacts_to_create = []
        row_data_map = {}  # Map contact_id to row number and name for results
        
        # First pass: Parse all rows and collect valid contacts
        row_count = 0
        
        # Debug: Log file reader info
        import logging
        logger = logging.getLogger(__name__)
        if is_excel:
            logger.info(f"Excel import: sheet.max_row={sheet.max_row}, include_first_row={include_first_row}, file_reader.max_row={file_reader.max_row if hasattr(file_reader, 'max_row') else 'N/A'}, file_reader.current_row={file_reader.current_row if hasattr(file_reader, 'current_row') else 'N/A'}")
        else:
            logger.info(f"CSV import: include_first_row={include_first_row}")
        
        # Try to iterate and log first few rows
        rows_processed = 0
        for row in file_reader:
            rows_processed += 1
            row_count += 1
            # Adjust row number: if include_first_row is False, first data row is row 2 (row 1 is header)
            # If include_first_row is True, first data row is row 1
            row_num = row_count + 1 if not include_first_row else row_count
            results['total'] += 1
            
            # Log first row for debugging
            if rows_processed == 1:
                logger.info(f"First row processed: row_num={row_num}, row_keys={list(row.keys())[:5] if row else 'empty'}")
                logger.info(f"Column mapping: {column_mapping}")
                logger.info(f"Row data sample: {dict(list(row.items())[:3])}")
            
            try:
                # Build contact data from CSV row
                contact_data = {}
                
                # Check if column_mapping is empty
                if not column_mapping:
                    if rows_processed == 1:
                        logger.error("Column mapping is empty! No data will be imported.")
                    results['errors'].append({
                        'row': row_num,
                        'error': 'Column mapping is empty. Please configure column mapping in the import interface.'
                    })
                    results['failed'] += 1
                    continue
                
                # Map CSV columns to contact fields
                mapped_fields_count = 0
                for frontend_field, csv_column in column_mapping.items():
                    if not csv_column:
                        continue
                    
                    # Special handling for autreInformations - can be a list of columns
                    if frontend_field == 'autreInformations' and isinstance(csv_column, list):
                        # Concatenate multiple columns with line breaks
                        values = []
                        for col in csv_column:
                            if not col:
                                continue
                            
                            # Normalize column names for matching
                            normalized_csv_column = col.strip().replace('"', '').replace("'", '')
                            
                            # Try exact match first, then case-insensitive match, then normalized match
                            csv_value = None
                            if col in row:
                                csv_value = row[col]
                            else:
                                # Try case-insensitive match
                                for key in row.keys():
                                    if not key:
                                        continue
                                    normalized_key = key.strip().replace('"', '').replace("'", '')
                                    # Try exact match first
                                    if key.strip() == col.strip():
                                        csv_value = row[key]
                                        break
                                    # Try normalized match
                                    elif normalized_key.lower() == normalized_csv_column.lower():
                                        csv_value = row[key]
                                        break
                                    # Try case-insensitive match on original values
                                    elif key.strip().lower() == col.strip().lower():
                                        csv_value = row[key]
                                        break
                            
                            if csv_value is not None:
                                # Convert to string and strip
                                str_value = str(csv_value).strip() if csv_value else ''
                                
                                # Only add non-empty values
                                if str_value:
                                    values.append(str_value)
                        
                        # Concatenate all values with line breaks
                        if values:
                            model_field = field_mapping.get(frontend_field)
                            if model_field:
                                contact_data[model_field] = '\n'.join(values)
                                mapped_fields_count += 1
                        continue
                    
                    # Handle single column mapping (existing logic)
                    # Normalize column names for matching (remove extra spaces, quotes, etc.)
                    normalized_csv_column = csv_column.strip().replace('"', '').replace("'", '')
                    
                    # Try exact match first, then case-insensitive match, then normalized match
                    csv_value = None
                    if csv_column in row:
                        csv_value = row[csv_column]
                    else:
                        # Try case-insensitive match
                        for key in row.keys():
                            if not key:
                                continue
                            normalized_key = key.strip().replace('"', '').replace("'", '')
                            # Try exact match first
                            if key.strip() == csv_column.strip():
                                csv_value = row[key]
                                break
                            # Try normalized match
                            elif normalized_key.lower() == normalized_csv_column.lower():
                                csv_value = row[key]
                                break
                            # Try case-insensitive match on original values
                            elif key.strip().lower() == csv_column.strip().lower():
                                csv_value = row[key]
                                break
                    
                    if csv_value is None:
                        # Log missing column for debugging
                        if rows_processed == 1:
                            logger.warning(f"Column '{csv_column}' not found in row. Available columns: {list(row.keys())[:10]}")
                        continue
                    
                    # Convert to string and strip - handle None, empty strings, and whitespace
                    if csv_value is None:
                        value = ''
                    else:
                        value = str(csv_value).strip() if csv_value else ''
                    
                    # Log phone field mapping for debugging
                    if rows_processed == 1 and frontend_field in ['phone', 'mobile']:
                        logger.info(f"Phone field mapping: frontend_field='{frontend_field}', csv_column='{csv_column}', csv_value='{csv_value}', value='{value}'")
                    
                    # Skip empty values for most fields (except where empty is meaningful)
                    # But still process the field to count it as mapped
                    
                    # Map to model field name
                    if frontend_field in field_mapping:
                        model_field = field_mapping[frontend_field]
                        
                        # Handle date field
                        if model_field == 'birth_date':
                            parsed_date = parse_date(value)
                            if parsed_date:
                                contact_data[model_field] = parsed_date
                        # Handle datetime fields - store separately for later update
                        elif model_field in ['created_at', 'updated_at', 'assigned_at']:
                            parsed_dt = parse_datetime(value)
                            if parsed_dt:
                                contact_data[model_field] = parsed_dt
                        # Handle phone number fields - convert to integer
                        elif model_field in ['phone', 'mobile']:
                            if value:
                                try:
                                    # Remove all whitespace, dashes, parentheses, and other non-digit characters
                                    cleaned = ''.join(c for c in value if c.isdigit())
                                    if cleaned:
                                        contact_data[model_field] = int(cleaned)
                                        # Log for debugging first row
                                        if rows_processed == 1:
                                            logger.info(f"Phone field '{model_field}' mapped: original='{value}', cleaned='{cleaned}', final={contact_data[model_field]}")
                                    else:
                                        if rows_processed == 1:
                                            logger.warning(f"Phone field '{model_field}' value '{value}' resulted in empty string after cleaning")
                                except (ValueError, TypeError) as e:
                                    # Log error but don't fail - phone is optional
                                    if rows_processed == 1:
                                        logger.warning(f"Error converting phone field '{model_field}' value '{value}' to int: {e}")
                                    pass  # Skip invalid phone numbers
                            else:
                                if rows_processed == 1:
                                    logger.info(f"Phone field '{model_field}' has empty value, skipping")
                        # Handle old_contact_id - convert empty strings to None
                        elif model_field == 'old_contact_id':
                            if value:
                                contact_data[model_field] = value
                        # For other string fields, only set if value is not empty
                        else:
                            if value:
                                contact_data[model_field] = value
                        
                        mapped_fields_count += 1
                
                # Log mapping results for first row
                if rows_processed == 1:
                    logger.info(f"Mapped {mapped_fields_count} fields. Contact data keys: {list(contact_data.keys())}")
                    logger.info(f"Sample contact_data: {dict(list(contact_data.items())[:5])}")
                
                # Validate required fields - lastName is no longer required
                # Only statusId is required (handled via default_status_id)
                
                # Store email for bulk duplicate check
                email = contact_data.get('email', '').strip()
                
                # Verify email domain (DNS MX record check) - with short timeout to avoid blocking import
                email_verification_status = 'not_verified'
                if email:
                    try:
                        from api.utils.email_verification import verify_email_domain
                        # Use short timeout (2 seconds) to avoid blocking bulk imports
                        email_verification_status, _ = verify_email_domain(email, timeout=2)
                    except Exception as e:
                        # If verification fails for any reason, default to not_verified
                        # Don't let email verification block contact creation
                        email_verification_status = 'not_verified'
                
                contact_data['email_verification_status'] = email_verification_status
                
                # Generate contact ID (UUID collisions are extremely rare, so we'll handle them during bulk_create if needed)
                contact_id = uuid.uuid4().hex[:12]
                # Check uniqueness only against pending contacts in this batch
                existing_ids = {c.id for c in contacts_to_create}
                while contact_id in existing_ids:
                    contact_id = uuid.uuid4().hex[:12]
                
                contact_data['id'] = contact_id
                contact_data['creator'] = request.user
                contact_data['status'] = status_obj
                # Always set source if source_obj exists (even if None, to ensure it's saved)
                if source_obj is not None:
                    contact_data['source'] = source_obj
                # Teleoperator is optional
                if teleoperator_obj is not None:
                    contact_data['teleoperator'] = teleoperator_obj
                    # Set assigned_at only if not provided in CSV (when importing contact with teleoperator)
                    if 'assigned_at' not in contact_data:
                        from django.utils import timezone
                        contact_data['assigned_at'] = timezone.now()
                
                # Store custom timestamps separately for later update (since auto_now_add/auto_now may override)
                custom_created_at = contact_data.pop('created_at', None)
                custom_updated_at = contact_data.pop('updated_at', None)
                custom_assigned_at = contact_data.pop('assigned_at', None)
                
                # Store row data for results
                contact_name = contact_data.get('fname', '')
                if contact_data.get('lname'):
                    contact_name = f"{contact_data.get('fname')} {contact_data.get('lname')}"
                
                row_data_map[contact_id] = {
                    'row': row_num,
                    'name': contact_name,
                    'email': email,
                    'created_at': custom_created_at,
                    'updated_at': custom_updated_at,
                    'assigned_at': custom_assigned_at,
                }
                
                # Create Contact instance (not saved yet)
                contact = Contact(**contact_data)
                contacts_to_create.append(contact)
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                results['errors'].append({
                    'row': row_num,
                    'error': str(e),
                    'details': error_details
                })
                results['failed'] += 1
        
        # Bulk check for duplicate emails - track but don't remove
        emails_to_check = [row_data['email'] for row_data in row_data_map.values() if row_data['email']]
        emails_seen_in_batch = {}  # Track emails within the CSV batch
        if emails_to_check:
            # Get existing emails and normalize to lowercase for comparison
            existing_emails_lower = set()
            existing_contacts_query = Contact.objects.filter(email__in=emails_to_check).exclude(email__isnull=True).exclude(email='')
            for contact in existing_contacts_query:
                if contact.email:
                    existing_emails_lower.add(contact.email.lower())
            
            # Track duplicates but don't remove them - they will still be inserted
            for contact in contacts_to_create:
                email = row_data_map[contact.id]['email']
                if email:
                    email_lower = email.lower()
                    is_duplicate = False
                    duplicate_reason = None
                    
                    # Check if email exists in database (case-insensitive)
                    if email_lower in existing_emails_lower:
                        is_duplicate = True
                        duplicate_reason = f'L\'email {email} existe déjà dans la base de données'
                    
                    # Check if email appears multiple times in this CSV batch
                    if email_lower in emails_seen_in_batch:
                        is_duplicate = True
                        if duplicate_reason:
                            duplicate_reason += f' et apparaît plusieurs fois dans le CSV (ligne {emails_seen_in_batch[email_lower]})'
                        else:
                            duplicate_reason = f'L\'email {email} apparaît plusieurs fois dans le CSV (première occurrence à la ligne {emails_seen_in_batch[email_lower]})'
                    else:
                        emails_seen_in_batch[email_lower] = row_data_map[contact.id]['row']
                    
                    if is_duplicate:
                        row_num = row_data_map[contact.id]['row']
                        results['duplicates'].append({
                            'row': row_num,
                            'email': email,
                            'reason': duplicate_reason,
                            'data': {'firstName': contact.fname, 'lastName': contact.lname}
                        })
        
        # Bulk create contacts in batches
        from django.db import transaction, IntegrityError
        
        with transaction.atomic():
            for i in range(0, len(contacts_to_create), BATCH_SIZE):
                batch = contacts_to_create[i:i + BATCH_SIZE]
                try:
                    # Ensure all contacts have email_verification_status set
                    for contact in batch:
                        if not hasattr(contact, 'email_verification_status') or contact.email_verification_status is None:
                            contact.email_verification_status = 'not_verified'
                    
                    Contact.objects.bulk_create(batch, batch_size=BATCH_SIZE)
                    
                    # Update custom timestamps directly in database for contacts that have them from CSV
                    for contact in batch:
                        row_data = row_data_map[contact.id]
                        custom_created_at = row_data.get('created_at')
                        custom_updated_at = row_data.get('updated_at')
                        custom_assigned_at = row_data.get('assigned_at')
                        
                        # Update timestamps directly in database if provided in CSV
                        update_fields = {}
                        if custom_created_at:
                            update_fields['created_at'] = custom_created_at
                        if custom_updated_at:
                            update_fields['updated_at'] = custom_updated_at
                        if custom_assigned_at:
                            update_fields['assigned_at'] = custom_assigned_at
                        
                        if update_fields:
                            Contact.objects.filter(id=contact.id).update(**update_fields)
                    
                    # Add to success results
                    for contact in batch:
                        row_data = row_data_map[contact.id]
                        results['success'].append({
                            'row': row_data['row'],
                            'contactId': contact.id,
                            'name': row_data['name']
                        })
                        results['imported'] += 1
                except Exception as e:
                    # Log the error for debugging
                    import logging
                    import traceback
                    logger = logging.getLogger(__name__)
                    error_details = traceback.format_exc()
                    logger.error(f"Error bulk creating contacts batch: {str(e)}\n{error_details}")
                    
                    # Handle potential ID collisions or other errors by falling back to individual creates for this batch
                    for contact in batch:
                        try:
                            row_data = row_data_map[contact.id]
                            custom_created_at = row_data.get('created_at')
                            custom_updated_at = row_data.get('updated_at')
                            custom_assigned_at = row_data.get('assigned_at')
                            
                            # Ensure email_verification_status is set
                            if not hasattr(contact, 'email_verification_status') or contact.email_verification_status is None:
                                contact.email_verification_status = 'not_verified'
                            
                            # Save contact first (without custom timestamps)
                            contact.save()
                            
                            # Update custom timestamps directly in database if provided
                            update_fields = {}
                            if custom_created_at:
                                update_fields['created_at'] = custom_created_at
                            if custom_updated_at:
                                update_fields['updated_at'] = custom_updated_at
                            if custom_assigned_at:
                                update_fields['assigned_at'] = custom_assigned_at
                            
                            if update_fields:
                                Contact.objects.filter(id=contact.id).update(**update_fields)
                            
                            results['success'].append({
                                'row': row_data['row'],
                                'contactId': contact.id,
                                'name': row_data['name']
                            })
                            results['imported'] += 1
                        except IntegrityError:
                            # ID collision - regenerate and try once more
                            old_id = contact.id
                            contact.id = uuid.uuid4().hex[:12]
                            # Update row_data_map with new ID
                            if old_id in row_data_map:
                                row_data_map[contact.id] = row_data_map.pop(old_id)
                            
                            try:
                                row_data = row_data_map[contact.id]
                                custom_created_at = row_data.get('created_at')
                                custom_updated_at = row_data.get('updated_at')
                                custom_assigned_at = row_data.get('assigned_at')
                                
                                # Save contact first (without custom timestamps)
                                contact.save()
                                
                                # Update custom timestamps directly in database if provided
                                update_fields = {}
                                if custom_created_at:
                                    update_fields['created_at'] = custom_created_at
                                if custom_updated_at:
                                    update_fields['updated_at'] = custom_updated_at
                                if custom_assigned_at:
                                    update_fields['assigned_at'] = custom_assigned_at
                                
                                if update_fields:
                                    Contact.objects.filter(id=contact.id).update(**update_fields)
                                
                                results['success'].append({
                                    'row': row_data['row'],
                                    'contactId': contact.id,
                                    'name': row_data['name']
                                })
                                results['imported'] += 1
                            except Exception as e:
                                row_data = row_data_map.get(contact.id, {})
                                results['errors'].append({
                                    'row': row_data.get('row', 'unknown'),
                                    'error': f'Failed to create contact: {str(e)}'
                                })
                                results['failed'] += 1
        
        # Create a single bulk log entry for the import (more efficient than individual logs)
        # This logs the import action itself rather than each individual contact
        if results['imported'] > 0:
            try:
                bulk_log_details = {
                    'ip_address': get_client_ip(request),
                    'browser': get_browser_info(request),
                    'imported_count': results['imported'],
                    'total_rows': results['total'],
                    'failed_count': results['failed'],
                }
                
                log_id = uuid.uuid4().hex[:12]
                while Log.objects.filter(id=log_id).exists():
                    log_id = uuid.uuid4().hex[:12]
                
                Log.objects.create(
                    id=log_id,
                    event_type='bulkImportContacts',
                    user_id=request.user if request.user.is_authenticated else None,
                    contact_id=None,  # Bulk import doesn't have a single contact
                    details=bulk_log_details,
                    old_value={},
                    new_value={'imported': results['imported'], 'total': results['total']}
                )
            except Exception as e:
                # Don't fail the import if logging fails
                pass
        
        # Close workbook if Excel file
        if workbook is not None:
            workbook.close()
        
        # Check if no rows were processed - this helps debug import issues
        if results['total'] == 0:
            return Response({
                'error': 'No data rows were processed. Please check:\n1. Your file contains data rows\n2. Column mapping is correct\n3. File format is valid (CSV or Excel)\n4. If using Excel, ensure data starts from the correct row',
                'results': results,
                'total': 0,
                'imported': 0,
                'failed': 0
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(results, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"CSV import error: {str(e)}")
        logger.error(f"CSV import error details: {error_details}")
        
        # Close workbook if Excel file (even on error)
        if workbook is not None:
            try:
                workbook.close()
            except:
                pass
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def contacts_integration_update(request):
    """
    Update existing contacts with timestamp fields from CSV using old_contact_id mapping.
    
    IMPORTANT: This function ONLY matches contacts by old_contact_id, NEVER by email.
    Each CSV row must have a valid old_contact_id that exists in the database.
    If a contact is not found by old_contact_id, the row will be skipped with an error.
    """
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    csv_file = request.FILES['file']
    
    # Parse column mapping JSON string if it's a string
    column_mapping_str = request.data.get('columnMapping', '{}')
    if isinstance(column_mapping_str, str):
        import json
        try:
            column_mapping = json.loads(column_mapping_str)
        except json.JSONDecodeError:
            column_mapping = {}
    else:
        column_mapping = column_mapping_str or {}
    
    # Parse value mappings (teleoperatorMapping, confirmateurMapping, sourceMapping)
    teleoperator_mapping_str = request.data.get('teleoperatorMapping', '{}')
    if isinstance(teleoperator_mapping_str, str):
        import json
        try:
            teleoperator_mapping = json.loads(teleoperator_mapping_str)
        except json.JSONDecodeError:
            teleoperator_mapping = {}
    else:
        teleoperator_mapping = teleoperator_mapping_str or {}
    
    confirmateur_mapping_str = request.data.get('confirmateurMapping', '{}')
    if isinstance(confirmateur_mapping_str, str):
        import json
        try:
            confirmateur_mapping = json.loads(confirmateur_mapping_str)
        except json.JSONDecodeError:
            confirmateur_mapping = {}
    else:
        confirmateur_mapping = confirmateur_mapping_str or {}
    
    source_mapping_str = request.data.get('sourceMapping', '{}')
    if isinstance(source_mapping_str, str):
        import json
        try:
            source_mapping = json.loads(source_mapping_str)
        except json.JSONDecodeError:
            source_mapping = {}
    else:
        source_mapping = source_mapping_str or {}
    
    # Validate required mappings - oldContactId is required
    if 'oldContactId' not in column_mapping or not column_mapping['oldContactId']:
        return Response({
            'error': 'oldContactId column mapping is required'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        import csv
        import io
        from datetime import datetime
        from django.utils import timezone
        import pytz
        
        # Read CSV file
        csv_content = csv_file.read().decode('utf-8-sig')  # Handle BOM
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        # Helper function to parse datetime
        def parse_datetime(datetime_str):
            """Parse datetime string to timezone-aware datetime object"""
            if not datetime_str or str(datetime_str).strip() == '':
                return None
            utc = pytz.UTC
            
            # If already a datetime object, make it timezone-aware and return
            if isinstance(datetime_str, datetime):
                if timezone.is_aware(datetime_str):
                    return datetime_str
                return utc.localize(datetime_str)
            datetime_str = str(datetime_str).strip()
            
            # Handle invalid dates like "0000-00-00 00:00:00" - return None silently
            if datetime_str.startswith('0000-00-00') or datetime_str.startswith('0000/00/00'):
                return None
            
            # Try parsing ISO format with timezone first
            try:
                from dateutil import parser
                parsed = parser.parse(datetime_str)
                # Check if parsed date is invalid (year 0 or similar)
                if parsed.year == 0 or parsed.year < 1900:
                    return None
                if timezone.is_aware(parsed):
                    return parsed
                return utc.localize(parsed)
            except (ValueError, ImportError):
                pass
            
            # Try common datetime formats (assume UTC for naive datetimes)
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%dT%H:%M:%S',
                '%Y-%m-%dT%H:%M:%S.%f',
                '%Y-%m-%dT%H:%M',
                '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y %H:%M',
                '%m/%d/%Y',
            ]
            for fmt in formats:
                try:
                    parsed = datetime.strptime(datetime_str, fmt)
                    # Check if parsed date is invalid (year 0 or similar)
                    if parsed.year == 0 or parsed.year < 1900:
                        return None
                    return utc.localize(parsed)
                except ValueError:
                    continue
            return None
        
        # Pre-load all contacts with old_contact_id into a dictionary for fast lookup
        # IMPORTANT: Only match by old_contact_id, NEVER by email
        contacts_by_old_id = {}
        contacts_with_old_id = Contact.objects.filter(old_contact_id__isnull=False).exclude(old_contact_id='')
        duplicate_old_ids = set()
        for contact in contacts_with_old_id:
            old_id_key = str(contact.old_contact_id).strip()
            if old_id_key in contacts_by_old_id:
                # Track duplicate old_contact_id values (shouldn't happen but handle gracefully)
                duplicate_old_ids.add(old_id_key)
            contacts_by_old_id[old_id_key] = contact
        
        # Log warning if duplicates found
        if duplicate_old_ids:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Found {len(duplicate_old_ids)} duplicate old_contact_id values: {list(duplicate_old_ids)[:10]}")
        
        # Pre-load users and sources for efficient lookup
        from api.models import UserDetails, Source
        users_by_userdetails_id = {}
        users_by_django_id = {}
        sources_by_id = {}
        
        # Load all UserDetails with their Django users
        user_details_list = UserDetails.objects.select_related('django_user').all()
        for user_detail in user_details_list:
            if user_detail.id:
                users_by_userdetails_id[str(user_detail.id).strip()] = user_detail.django_user
            if user_detail.django_user_id:
                users_by_django_id[user_detail.django_user_id] = user_detail.django_user
        
        # Load all sources
        sources_list = Source.objects.all()
        for source in sources_list:
            if source.id:
                sources_by_id[str(source.id).strip()] = source
        
        results = {
            'success': [],
            'errors': [],
            'total': 0,
            'updated': 0,
            'failed': 0
        }
        
        # Process CSV rows
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (row 1 is header)
            results['total'] += 1
            try:
                # Get old_contact_id from CSV
                old_contact_id_col = column_mapping.get('oldContactId', '')
                if not old_contact_id_col:
                    results['errors'].append({
                        'row': row_num,
                        'error': 'oldContactId column not mapped'
                    })
                    results['failed'] += 1
                    continue
                
                # Try exact match first, then case-insensitive match
                old_contact_id_value = None
                if old_contact_id_col in row:
                    old_contact_id_value = row[old_contact_id_col]
                else:
                    # Try case-insensitive match
                    for key in row.keys():
                        if key and key.strip().lower() == old_contact_id_col.strip().lower():
                            old_contact_id_value = row[key]
                            break
                
                if not old_contact_id_value or not old_contact_id_value.strip():
                    results['errors'].append({
                        'row': row_num,
                        'error': 'oldContactId is empty'
                    })
                    results['failed'] += 1
                    continue
                
                old_contact_id_value = str(old_contact_id_value).strip()
                
                # Find contact by old_contact_id ONLY - NEVER use email for matching
                # This ensures we update the correct contact based on old_contact_id
                # If duplicate old_contact_id values exist, reject the update to prevent wrong contact updates
                if old_contact_id_value in duplicate_old_ids:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Multiple contacts found with old_contact_id: {old_contact_id_value}. Please ensure old_contact_id values are unique in the database before importing.'
                    })
                    results['failed'] += 1
                    continue
                
                contact = contacts_by_old_id.get(old_contact_id_value)
                if not contact:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Contact not found with old_contact_id: {old_contact_id_value}'
                    })
                    results['failed'] += 1
                    continue
                
                # Helper function to get CSV value
                def get_csv_value(column_name):
                    if not column_name:
                        return None
                    if column_name in row:
                        return row[column_name]
                    # Try case-insensitive match
                    for key in row.keys():
                        if key and key.strip().lower() == column_name.strip().lower():
                            return row[key]
                    return None
                
                # Parse timestamp fields and other fields
                update_fields = {}
                contact_updates = {}  # For ForeignKey fields that need object assignment
                
                # Handle created_at
                if 'createdAt' in column_mapping and column_mapping['createdAt']:
                    created_at_value = get_csv_value(column_mapping['createdAt'])
                    if created_at_value and str(created_at_value).strip():
                        parsed_dt = parse_datetime(created_at_value)
                        if parsed_dt:
                            update_fields['created_at'] = parsed_dt
                        # If parsing fails, silently skip (set to None/null) - don't log error
                
                # Handle updated_at
                if 'updatedAt' in column_mapping and column_mapping['updatedAt']:
                    updated_at_value = get_csv_value(column_mapping['updatedAt'])
                    if updated_at_value and str(updated_at_value).strip():
                        parsed_dt = parse_datetime(updated_at_value)
                        if parsed_dt:
                            update_fields['updated_at'] = parsed_dt
                        # If parsing fails, silently skip (set to None/null) - don't log error
                
                # Handle assigned_at
                if 'assignedAt' in column_mapping and column_mapping['assignedAt']:
                    assigned_at_value = get_csv_value(column_mapping['assignedAt'])
                    if assigned_at_value and str(assigned_at_value).strip():
                        parsed_dt = parse_datetime(assigned_at_value)
                        if parsed_dt:
                            update_fields['assigned_at'] = parsed_dt
                        # If parsing fails, silently skip (set to None/null) - don't log error
                
                # Handle teleoperatorId
                if 'teleoperatorId' in column_mapping and column_mapping['teleoperatorId']:
                    teleoperator_value = get_csv_value(column_mapping['teleoperatorId'])
                    if teleoperator_value and teleoperator_value.strip():
                        teleoperator_value_str = str(teleoperator_value).strip()
                        
                        # Check if there's a mapping for this CSV value
                        mapped_teleoperator_id = teleoperator_mapping.get(teleoperator_value_str)
                        if mapped_teleoperator_id:
                            teleoperator_id = str(mapped_teleoperator_id).strip()
                        else:
                            # Use CSV value directly (might be an ID)
                            teleoperator_id = teleoperator_value_str
                        
                        # Try UserDetails ID first (this is what frontend sends), then Django User ID as fallback
                        # Normalize the ID for lookup (strip whitespace)
                        teleoperator_id_normalized = teleoperator_id.strip()
                        teleoperator_user = users_by_userdetails_id.get(teleoperator_id_normalized)
                        if not teleoperator_user:
                            # Try original format (without strip) in case it was stored differently
                            teleoperator_user = users_by_userdetails_id.get(teleoperator_id)
                        if not teleoperator_user:
                            # Fallback: try Django User ID if it's numeric
                            try:
                                django_id = int(teleoperator_id_normalized)
                                teleoperator_user = users_by_django_id.get(django_id)
                            except (ValueError, TypeError):
                                pass
                        
                        if teleoperator_user:
                            contact_updates['teleoperator'] = teleoperator_user
                        else:
                            # Log warning but don't fail - allow other fields to update
                            results['errors'].append({
                                'row': row_num,
                                'error': f'Teleoperator not found with ID: {teleoperator_id} (field skipped)'
                            })
                    elif teleoperator_value is not None and teleoperator_value.strip() == '':
                        # Empty string means clear the teleoperator
                        contact_updates['teleoperator'] = None
                
                # Handle confirmateurId
                if 'confirmateurId' in column_mapping and column_mapping['confirmateurId']:
                    confirmateur_value = get_csv_value(column_mapping['confirmateurId'])
                    if confirmateur_value and confirmateur_value.strip():
                        confirmateur_value_str = str(confirmateur_value).strip()
                        
                        # Check if there's a mapping for this CSV value
                        mapped_confirmateur_id = confirmateur_mapping.get(confirmateur_value_str)
                        if mapped_confirmateur_id:
                            confirmateur_id = str(mapped_confirmateur_id).strip()
                        else:
                            # Use CSV value directly (might be an ID)
                            confirmateur_id = confirmateur_value_str
                        
                        # Try UserDetails ID first (this is what frontend sends), then Django User ID as fallback
                        # Normalize the ID for lookup (strip whitespace)
                        confirmateur_id_normalized = confirmateur_id.strip()
                        confirmateur_user = users_by_userdetails_id.get(confirmateur_id_normalized)
                        if not confirmateur_user:
                            # Try original format (without strip) in case it was stored differently
                            confirmateur_user = users_by_userdetails_id.get(confirmateur_id)
                        if not confirmateur_user:
                            # Fallback: try Django User ID if it's numeric
                            try:
                                django_id = int(confirmateur_id_normalized)
                                confirmateur_user = users_by_django_id.get(django_id)
                            except (ValueError, TypeError):
                                pass
                        
                        if confirmateur_user:
                            contact_updates['confirmateur'] = confirmateur_user
                        else:
                            # Log warning but don't fail - allow other fields to update
                            results['errors'].append({
                                'row': row_num,
                                'error': f'Confirmateur not found with ID: {confirmateur_id} (field skipped)'
                            })
                    elif confirmateur_value is not None and confirmateur_value.strip() == '':
                        # Empty string means clear the confirmateur
                        contact_updates['confirmateur'] = None
                
                # Handle sourceId
                if 'sourceId' in column_mapping and column_mapping['sourceId']:
                    source_value = get_csv_value(column_mapping['sourceId'])
                    if source_value and source_value.strip():
                        source_value_str = str(source_value).strip()
                        
                        # Check if there's a mapping for this CSV value
                        mapped_source_id = source_mapping.get(source_value_str)
                        if mapped_source_id:
                            source_id = str(mapped_source_id).strip()
                        else:
                            # Use CSV value directly (might be an ID)
                            source_id = source_value_str
                        
                        source_obj = sources_by_id.get(source_id)
                        
                        if source_obj:
                            contact_updates['source'] = source_obj
                        else:
                            # Log warning but don't fail - allow other fields to update
                            results['errors'].append({
                                'row': row_num,
                                'error': f'Source not found with ID: {source_id} (field skipped)'
                            })
                    elif source_value is not None and source_value.strip() == '':
                        # Empty string means clear the source
                        contact_updates['source'] = None
                
                # Update contact if we have fields to update
                if update_fields or contact_updates:
                    # Prepare all fields for database update
                    # We need to update ForeignKey fields by their ID, not the object
                    db_update_fields = {}
                    
                    # Add timestamp fields
                    if update_fields:
                        db_update_fields.update(update_fields)
                    
                    # Convert ForeignKey objects to IDs for database update
                    if contact_updates:
                        if 'teleoperator' in contact_updates:
                            # teleoperator is a DjangoUser object, so .id gives us the Django User ID (integer)
                            db_update_fields['teleoperator_id'] = contact_updates['teleoperator'].id if contact_updates['teleoperator'] else None
                        if 'confirmateur' in contact_updates:
                            # confirmateur is a DjangoUser object, so .id gives us the Django User ID (integer)
                            db_update_fields['confirmateur_id'] = contact_updates['confirmateur'].id if contact_updates['confirmateur'] else None
                        if 'source' in contact_updates:
                            # source is a Source object, so .id gives us the Source ID (string)
                            db_update_fields['source_id'] = contact_updates['source'].id if contact_updates['source'] else None
                    
                    # Use direct database update to bypass auto_now and auto_now_add
                    # This ensures our CSV values are preserved
                    # Note: .update() bypasses Django's save() method, so auto_now and auto_now_add are ignored
                    if db_update_fields:
                        try:
                            # Use update() which bypasses auto_now and auto_now_add
                            rows_updated = Contact.objects.filter(id=contact.id).update(**db_update_fields)
                            if rows_updated == 0:
                                results['errors'].append({
                                    'row': row_num,
                                    'error': f'Contact update failed - no rows updated'
                                })
                                results['failed'] += 1
                                continue
                            contact.refresh_from_db()
                        except Exception as update_error:
                            import traceback
                            error_details = traceback.format_exc()
                            results['errors'].append({
                                'row': row_num,
                                'error': f'Error updating contact: {str(update_error)}',
                                'details': error_details
                            })
                            results['failed'] += 1
                            continue
                    
                    updated_field_names = list(update_fields.keys()) + list(contact_updates.keys())
                    results['success'].append({
                        'row': row_num,
                        'contactId': contact.id,
                        'oldContactId': old_contact_id_value,
                        'updatedFields': updated_field_names
                    })
                    results['updated'] += 1
                else:
                    results['errors'].append({
                        'row': row_num,
                        'error': 'No fields provided to update'
                    })
                    results['failed'] += 1
                    
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                results['errors'].append({
                    'row': row_num,
                    'error': str(e),
                    'details': error_details
                })
                results['failed'] += 1
        
        return Response(results, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def contacts_by_old_ids(request):
    """
    Get contacts by a list of oldContactIds for migration purposes.
    Accepts a list of oldContactIds and returns matching contacts with their IDs and teleoperatorIds.
    Bypasses pagination limits for migration efficiency.
    """
    try:
        old_contact_ids = request.data.get('oldContactIds', [])
        
        if not old_contact_ids or not isinstance(old_contact_ids, list):
            return Response({'error': 'oldContactIds must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Normalize oldContactIds (trim whitespace, convert to string)
        normalized_ids = [str(id).strip() for id in old_contact_ids if id is not None and str(id).strip()]
        
        if not normalized_ids:
            return Response({'contacts': []}, status=status.HTTP_200_OK)
        
        # Build query conditions for exact matches (try both string and numeric versions)
        from django.db.models import Q
        q_conditions = Q()
        
        for old_id in normalized_ids:
            # Exact string match
            q_conditions |= Q(old_contact_id=old_id)
            # Also try numeric match if applicable (for cases where DB has "123" and CSV has "123")
            try:
                # If it's numeric, also try matching the numeric string version
                numeric_id = str(int(old_id))
                if numeric_id != old_id:
                    q_conditions |= Q(old_contact_id=numeric_id)
            except (ValueError, TypeError):
                pass
        
        # Query contacts with matching old_contact_ids
        # Use select_related to optimize teleoperator access
        contacts = Contact.objects.filter(
            old_contact_id__isnull=False
        ).filter(q_conditions).select_related(
            'teleoperator',
            'teleoperator__user_details'
        ).distinct()
        
        # Build response with contactId and teleoperatorId
        result = []
        for contact in contacts:
            # Normalize old_contact_id for matching
            contact_old_id = str(contact.old_contact_id).strip() if contact.old_contact_id else None
            
            # Check if this contact matches any of the requested IDs
            matches = False
            for requested_id in normalized_ids:
                # Exact match
                if contact_old_id == requested_id:
                    matches = True
                    break
                # Numeric match (for cases where DB has "123" and CSV has "123")
                try:
                    numeric_requested = str(int(requested_id))
                    if contact_old_id == numeric_requested:
                        matches = True
                        break
                except (ValueError, TypeError):
                    pass
            
            if matches:
                # Get teleoperatorId (UserDetails ID, not Django User ID)
                teleoperator_id = None
                if contact.teleoperator and contact.teleoperator.user_details:
                    teleoperator_id = contact.teleoperator.user_details.id
                
                result.append({
                    'oldContactId': contact.old_contact_id,
                    'contactId': contact.id,
                    'teleoperatorId': teleoperator_id
                })
        
        return Response({'contacts': result}, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error in contacts_by_old_ids: {error_details}")
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def contacts_migration_missing(request):
    """Check CSV file for old IDs not in database and return missing rows as CSV"""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    csv_file = request.FILES['file']
    old_id_column = request.data.get('oldIdColumn', None)
    
    try:
        # Read CSV file
        csv_content = csv_file.read().decode('utf-8-sig')  # Handle BOM
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        rows = list(csv_reader)
        
        if not rows:
            return Response({'error': 'CSV file is empty'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get column names
        fieldnames = csv_reader.fieldnames
        if not fieldnames:
            return Response({'error': 'CSV file has no headers'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Find old ID column
        if not old_id_column:
            # Try common variations
            possible_names = [
                'old id', 'old_id', 'old_contact_id', 'oldContactId',
                'old contact id', 'old-contact-id', 'oldcontactid',
                'OLD_ID', 'OLD_CONTACT_ID', 'Old ID', 'Old Contact ID'
            ]
            old_id_column = None
            for name in possible_names:
                if name in fieldnames:
                    old_id_column = name
                    break
            
            if not old_id_column:
                # Try case-insensitive search
                fieldnames_lower = {f.lower().replace('_', ' ').replace('-', ' '): f for f in fieldnames}
                for name in possible_names:
                    name_lower = name.lower().replace('_', ' ').replace('-', ' ')
                    if name_lower in fieldnames_lower:
                        old_id_column = fieldnames_lower[name_lower]
                        break
        
        if not old_id_column:
            return Response({
                'error': 'Could not find old ID column',
                'availableColumns': list(fieldnames),
                'message': 'Please specify the column name using oldIdColumn parameter'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Extract all old IDs from CSV
        old_ids_in_csv = []
        for i, row in enumerate(rows):
            old_id = str(row.get(old_id_column, '')).strip()
            if old_id:
                old_ids_in_csv.append((i, old_id, row))
        
        if not old_ids_in_csv:
            return Response({
                'error': 'No old IDs found in CSV file',
                'detectedColumn': old_id_column
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Query database for existing old IDs in batches
        # Normalize both CSV values and database values to handle whitespace differences
        existing_old_ids = set()
        old_id_values = [old_id for _, old_id, _ in old_ids_in_csv]
        # Create a set of normalized CSV values for efficient lookup
        normalized_csv_ids = {old_id.strip() for old_id in old_id_values}
        batch_size = 1000
        
        for i in range(0, len(old_id_values), batch_size):
            batch = old_id_values[i:i + batch_size]
            # Build a Q object to query contacts where old_contact_id could match
            # This includes exact matches and potential matches with whitespace
            # We use Q objects to check if old_contact_id contains any batch value
            # (to catch cases like " 123 " matching "123")
            q_conditions = Q()
            for batch_id in batch:
                batch_id_str = str(batch_id).strip()
                # Check for exact match
                q_conditions |= Q(old_contact_id=batch_id)
                # Also check if old_contact_id contains the batch value (to catch whitespace variations)
                # This will match " 123 ", "123 ", " 123", etc.
                # We'll filter more precisely in Python to avoid false positives
                if batch_id_str:
                    q_conditions |= Q(old_contact_id__contains=batch_id_str)
            
            # Query contacts that could potentially match
            existing = Contact.objects.filter(
                old_contact_id__isnull=False
            ).filter(q_conditions).values_list('old_contact_id', flat=True).distinct()
            
            # Normalize database values and check if they match any normalized CSV value
            # This ensures we only match values that, when stripped, exactly match the CSV value
            for db_id in existing:
                if db_id:
                    normalized_db_id = str(db_id).strip()
                    # Only add if normalized database ID exactly matches a normalized CSV ID
                    # This prevents false positives from __contains (e.g., "123" matching "1234")
                    if normalized_db_id in normalized_csv_ids:
                        existing_old_ids.add(normalized_db_id)
        
        # Find rows with old IDs NOT in database
        missing_rows = []
        for row_index, old_id, row in old_ids_in_csv:
            old_id_stripped = old_id.strip()
            if old_id_stripped not in existing_old_ids:
                missing_rows.append(row)
        
        # Create CSV content with missing rows
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(missing_rows)
        csv_output = output.getvalue()
        
        # Return JSON response with statistics and CSV content
        return Response({
            'success': True,
            'statistics': {
                'totalRows': len(rows),
                'rowsWithOldIds': len(old_ids_in_csv),
                'rowsInDatabase': len(existing_old_ids),
                'rowsMissing': len(missing_rows),
            },
            'detectedColumn': old_id_column,
            'csvContent': csv_output,
            'filename': f'missing_contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def csv_import_notes(request):
    """Import notes from CSV with column mapping - optimized for large imports"""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    csv_file = request.FILES['file']
    
    # Parse column mapping JSON string if it's a string
    column_mapping_str = request.data.get('columnMapping', '{}')
    if isinstance(column_mapping_str, str):
        import json
        try:
            column_mapping = json.loads(column_mapping_str)
        except json.JSONDecodeError:
            column_mapping = {}
    else:
        column_mapping = column_mapping_str or {}
    
    # Validate required mappings - text is required
    required_fields = ['text']
    missing_fields = [field for field in required_fields if field not in column_mapping or not column_mapping[field]]
    if missing_fields:
        return Response({
            'error': f'Missing required column mappings: {", ".join(missing_fields)}'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Always use current user for notes
    default_user_obj = request.user
    
    # Get default category if specified
    default_category_id = request.data.get('defaultCategoryId')
    default_category_obj = None
    if default_category_id:
        default_category_id = str(default_category_id).strip()
        if default_category_id:
            try:
                default_category_obj = NoteCategory.objects.get(id=default_category_id)
            except NoteCategory.DoesNotExist:
                return Response({'error': 'Default category not found'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        import csv
        import io
        from datetime import datetime
        
        # Read CSV file
        csv_content = csv_file.read().decode('utf-8-sig')  # Handle BOM
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        # Field mapping from frontend names to model field names
        field_mapping = {
            'id': 'id',
            'text': 'text',
            'oldContactId': 'old_contact_id',  # For mapping by old contact ID
            'createdAt': 'created_at',
        }
        
        results = {
            'success': [],
            'errors': [],
            'total': 0,
            'imported': 0,
            'failed': 0
        }
        
        # Helper function to parse date
        def parse_datetime(datetime_str):
            if not datetime_str or datetime_str.strip() == '':
                return None
            datetime_str = datetime_str.strip()
            # Try common datetime formats
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y %H:%M',
                '%m/%d/%Y',
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(datetime_str, fmt)
                except ValueError:
                    continue
            return None
        
        # Batch processing configuration
        BATCH_SIZE = 1000  # Process notes in batches of 1000
        notes_to_create = []
        row_data_map = {}  # Map note_id to row number and details for results
        
        # Pre-load all contacts with old_contact_id into a dictionary for fast lookup
        # This avoids N+1 query problem when processing CSV rows
        contacts_by_old_id = {}
        contacts_with_old_id = Contact.objects.filter(old_contact_id__isnull=False).exclude(old_contact_id='')
        for contact in contacts_with_old_id:
            contacts_by_old_id[contact.old_contact_id] = contact
        
        # Pre-load existing note IDs to check for duplicates
        # This avoids N+1 query problem when checking if note IDs already exist
        existing_note_ids = set(Note.objects.values_list('id', flat=True))
        
        # First pass: Parse all rows and collect valid notes
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (row 1 is header)
            results['total'] += 1
            try:
                # Build note data from CSV row
                note_data = {}
                old_contact_id = None
                
                # Map CSV columns to note fields
                for frontend_field, csv_column in column_mapping.items():
                    if not csv_column:
                        continue
                    
                    # Try exact match first, then case-insensitive match
                    csv_value = None
                    if csv_column in row:
                        csv_value = row[csv_column]
                    else:
                        # Try case-insensitive match
                        for key in row.keys():
                            if key and key.strip().lower() == csv_column.strip().lower():
                                csv_value = row[key]
                                break
                    
                    if csv_value is None:
                        continue
                    
                    value = csv_value.strip() if csv_value else ''
                    
                    # Map to model field name
                    if frontend_field in field_mapping:
                        model_field = field_mapping[frontend_field]
                        
                        # Handle datetime field
                        if model_field == 'created_at':
                            parsed_dt = parse_datetime(value)
                            if parsed_dt:
                                from django.utils import timezone
                                note_data[model_field] = timezone.make_aware(parsed_dt)
                            else:
                                note_data[model_field] = None
                        # Handle old contact ID (store separately for lookup)
                        elif model_field == 'old_contact_id':
                            old_contact_id = value if value else None
                        else:
                            note_data[model_field] = value
                
                # Validate required fields - text is required
                if not note_data.get('text'):
                    results['errors'].append({
                        'row': row_num,
                        'error': 'Text is required'
                    })
                    results['failed'] += 1
                    continue
                
                # Handle contact ID mapping by old_contact_id only
                # Only import notes where the old contact ID was found in the contact table
                contact_obj = None
                if old_contact_id:
                    # Use pre-loaded dictionary for fast lookup (O(1) instead of database query)
                    contact_obj = contacts_by_old_id.get(old_contact_id)
                    if not contact_obj:
                        # Contact not found - skip this note
                        results['errors'].append({
                            'row': row_num,
                            'error': f'Contact not found with old_contact_id: {old_contact_id}'
                        })
                        results['failed'] += 1
                        continue
                else:
                    # No old_contact_id provided - skip this note
                    results['errors'].append({
                        'row': row_num,
                        'error': 'old_contact_id is required for note import'
                    })
                    results['failed'] += 1
                    continue
                
                # Always use current user
                user_obj = default_user_obj
                
                # Use default category if provided
                category_obj = default_category_obj
                
                # Handle note ID - use from CSV if provided, otherwise generate
                note_id = None
                if 'id' in note_data and note_data['id']:
                    # Use ID from CSV
                    note_id = str(note_data['id']).strip()
                    if len(note_id) > 12:
                        note_id = note_id[:12]  # Truncate to 12 characters if longer
                    
                    # Check if ID already exists in database
                    if note_id in existing_note_ids:
                        results['errors'].append({
                            'row': row_num,
                            'error': f'Note with ID {note_id} already exists in database'
                        })
                        results['failed'] += 1
                        continue
                    
                    # Check uniqueness against pending notes in this batch
                    existing_ids = {n.id for n in notes_to_create}
                    if note_id in existing_ids:
                        results['errors'].append({
                            'row': row_num,
                            'error': f'Duplicate ID {note_id} found in CSV'
                        })
                        results['failed'] += 1
                        continue
                else:
                    # Generate new note ID
                    note_id = uuid.uuid4().hex[:12]
                    # Check uniqueness against pending notes in this batch
                    existing_ids = {n.id for n in notes_to_create}
                    while note_id in existing_ids or note_id in existing_note_ids:
                        note_id = uuid.uuid4().hex[:12]
                
                note_data['id'] = note_id
                note_data['userId'] = user_obj
                note_data['contactId'] = contact_obj
                note_data['categ_id'] = category_obj
                
                # Add ID to existing_note_ids to prevent duplicates in the same batch
                existing_note_ids.add(note_id)
                
                # Remove old_contact_id from note_data (it's not a Note field)
                note_data.pop('old_contact_id', None)
                
                # Store created_at separately if provided (needs special handling)
                custom_created_at = note_data.pop('created_at', None)
                
                # Store row data for results
                note_text_preview = note_data.get('text', '')[:50] + ('...' if len(note_data.get('text', '')) > 50 else '')
                
                row_data_map[note_id] = {
                    'row': row_num,
                    'text': note_text_preview,
                    'created_at': custom_created_at,  # Store for later update
                }
                
                # Create Note instance (not saved yet)
                note = Note(**note_data)
                # Set created_at if provided (this will override auto_now_add)
                if custom_created_at:
                    note.created_at = custom_created_at
                notes_to_create.append(note)
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                results['errors'].append({
                    'row': row_num,
                    'error': str(e),
                    'details': error_details
                })
                results['failed'] += 1
        
        # Bulk create notes in batches
        from django.db import transaction, IntegrityError
        
        with transaction.atomic():
            for i in range(0, len(notes_to_create), BATCH_SIZE):
                batch = notes_to_create[i:i + BATCH_SIZE]
                try:
                    Note.objects.bulk_create(batch, batch_size=BATCH_SIZE)
                    
                    # Update created_at for notes that have custom timestamps from CSV
                    # bulk_create may override created_at due to auto_now_add, so we update it afterward
                    notes_to_update = []
                    for note in batch:
                        row_data = row_data_map[note.id]
                        custom_created_at = row_data.get('created_at')
                        if custom_created_at:
                            note.created_at = custom_created_at
                            notes_to_update.append(note)
                    
                    # Bulk update created_at for notes with custom timestamps
                    if notes_to_update:
                        Note.objects.bulk_update(notes_to_update, ['created_at'], batch_size=BATCH_SIZE)
                    
                    # Add to success results
                    for note in batch:
                        row_data = row_data_map[note.id]
                        results['success'].append({
                            'row': row_data['row'],
                            'noteId': note.id,
                            'text': row_data['text']
                        })
                        results['imported'] += 1
                except IntegrityError as e:
                    # Handle potential ID collisions by falling back to individual creates for this batch
                    for note in batch:
                        try:
                            note.save()
                            row_data = row_data_map[note.id]
                            results['success'].append({
                                'row': row_data['row'],
                                'noteId': note.id,
                                'text': row_data['text']
                            })
                            results['imported'] += 1
                        except IntegrityError:
                            # ID collision - regenerate and try once more
                            note.id = uuid.uuid4().hex[:12]
                            try:
                                note.save()
                                row_data = row_data_map[note.id]
                                results['success'].append({
                                    'row': row_data['row'],
                                    'noteId': note.id,
                                    'text': row_data['text']
                                })
                                results['imported'] += 1
                            except Exception as e:
                                row_data = row_data_map.get(note.id, {})
                                results['errors'].append({
                                    'row': row_data.get('row', 'unknown'),
                                    'error': f'Failed to create note: {str(e)}'
                                })
                                results['failed'] += 1
                        except Exception as e:
                            row_data = row_data_map.get(note.id, {})
                            results['errors'].append({
                                'row': row_data.get('row', 'unknown'),
                                'error': f'Failed to create note: {str(e)}'
                            })
                            results['failed'] += 1
        
        # Create a single bulk log entry for the import
        if results['imported'] > 0:
            try:
                bulk_log_details = {
                    'ip_address': get_client_ip(request),
                    'browser': get_browser_info(request),
                    'imported_count': results['imported'],
                    'total_rows': results['total'],
                    'failed_count': results['failed'],
                }
                
                log_id = uuid.uuid4().hex[:12]
                while Log.objects.filter(id=log_id).exists():
                    log_id = uuid.uuid4().hex[:12]
                
                Log.objects.create(
                    id=log_id,
                    event_type='bulkImportNotes',
                    user_id=request.user if request.user.is_authenticated else None,
                    contact_id=None,  # Bulk import doesn't have a single contact
                    details=bulk_log_details,
                    old_value={},
                    new_value={'imported': results['imported'], 'total': results['total']}
                )
            except Exception as e:
                # Don't fail the import if logging fails
                pass
        
        return Response(results, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def log_event_types(request):
    """Get all event types used in the system (hardcoded list)"""
    # Hardcoded list of all event types used in the codebase
    event_types = [
        'addContact',
        'bulkImportContacts',
        'bulkImportNotes',
        'createEvent',
        'createTeam',
        'createUser',
        'deleteEvent',
        'deleteTeam',
        'deleteUser',
        'editContact',
        'editEvent',
        'editTeam',
        'editUser',
        'resetPassword',
    ]
    return Response(event_types, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def csv_import_logs(request):
    """Import logs from CSV with column mapping - optimized for large imports"""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    csv_file = request.FILES['file']
    
    # Parse column mapping JSON string if it's a string
    column_mapping_str = request.data.get('columnMapping', '{}')
    if isinstance(column_mapping_str, str):
        import json
        try:
            column_mapping = json.loads(column_mapping_str)
        except json.JSONDecodeError:
            column_mapping = {}
    else:
        column_mapping = column_mapping_str or {}
    
    # Validate required mappings - event_type is required
    required_fields = ['eventType']
    missing_fields = [field for field in required_fields if field not in column_mapping or not column_mapping[field]]
    if missing_fields:
        return Response({
            'error': f'Missing required column mappings: {", ".join(missing_fields)}'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    # Parse event type mapping JSON string if it's a string
    event_type_mapping_str = request.data.get('eventTypeMapping', '{}')
    if isinstance(event_type_mapping_str, str):
        import json
        try:
            event_type_mapping = json.loads(event_type_mapping_str)
        except json.JSONDecodeError:
            event_type_mapping = {}
    else:
        event_type_mapping = event_type_mapping_str or {}
    
    # Parse user ID mapping JSON string if it's a string
    user_id_mapping_str = request.data.get('userIdMapping', '{}')
    if isinstance(user_id_mapping_str, list) and len(user_id_mapping_str) > 0:
        user_id_mapping_str = user_id_mapping_str[0]
    if isinstance(user_id_mapping_str, str):
        import json
        try:
            user_id_mapping = json.loads(user_id_mapping_str)
        except json.JSONDecodeError as e:
            user_id_mapping = {}
    else:
        user_id_mapping = user_id_mapping_str or {}
    
    # Get default user ID if provided
    # FormData sends values as strings, so we need to handle that
    default_user_id = request.data.get('defaultUserId', '')
    if isinstance(default_user_id, list) and len(default_user_id) > 0:
        default_user_id = default_user_id[0]
    default_user_id = str(default_user_id).strip() if default_user_id else ''
    
    default_user_obj = None
    if default_user_id:
        try:
            # DjangoUser is already imported at the top of the file
            # Try both string and integer formats
            try:
                default_user_obj = DjangoUser.objects.get(id=default_user_id)
            except (DjangoUser.DoesNotExist, ValueError, TypeError):
                try:
                    if default_user_id.isdigit():
                        default_user_obj = DjangoUser.objects.get(id=int(default_user_id))
                    else:
                        # Try UUID format if applicable
                        default_user_obj = DjangoUser.objects.get(id=default_user_id)
                except (DjangoUser.DoesNotExist, ValueError, TypeError):
                    default_user_obj = None
        except Exception as e:
            # Handle any other exception (invalid ID format, etc.)
            default_user_obj = None
    
    try:
        import csv
        import io
        from datetime import datetime
        
        # Read CSV file
        csv_content = csv_file.read().decode('utf-8-sig')  # Handle BOM
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        # Field mapping from frontend names to model field names
        field_mapping = {
            'id': 'id',
            'eventType': 'event_type',
            'oldContactId': 'old_contact_id',  # For mapping by old contact ID
            'userId': 'user_id',
            'creatorId': 'creator_id',
            'createdAt': 'created_at',
            'details': 'details',
            'oldValue': 'old_value',
            'newValue': 'new_value',
            'oldLogs': 'old_logs',  # Text field for old logs that are not in JSON format
        }
        
        results = {
            'success': [],
            'errors': [],
            'total': 0,
            'imported': 0,
            'failed': 0
        }
        
        # Helper function to parse date
        def parse_datetime(datetime_str):
            if not datetime_str or datetime_str.strip() == '':
                return None
            datetime_str = datetime_str.strip()
            # Try common datetime formats
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d %H:%M',
                '%Y-%m-%d',
                '%d/%m/%Y %H:%M:%S',
                '%d/%m/%Y %H:%M',
                '%d/%m/%Y',
                '%m/%d/%Y %H:%M:%S',
                '%m/%d/%Y %H:%M',
                '%m/%d/%Y',
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(datetime_str, fmt)
                except ValueError:
                    continue
            return None
        
        # Helper function to parse JSON
        def parse_json(json_str):
            if not json_str or json_str.strip() == '':
                return {}
            try:
                import json
                return json.loads(json_str)
            except (json.JSONDecodeError, ValueError):
                return {}
        
        # Batch processing configuration
        BATCH_SIZE = 1000  # Process logs in batches of 1000
        logs_to_create = []
        row_data_map = {}  # Map log_id to row number and details for results
        
        # Pre-load all contacts with old_contact_id into a dictionary for fast lookup
        # This avoids N+1 query problem when processing CSV rows
        contacts_by_old_id = {}
        contacts_with_old_id = Contact.objects.filter(old_contact_id__isnull=False).exclude(old_contact_id='')
        for contact in contacts_with_old_id:
            # Use stripped value as key to ensure matching
            old_id_key = str(contact.old_contact_id).strip()
            if old_id_key:
                contacts_by_old_id[old_id_key] = contact
        
        # Pre-load existing log IDs to check for duplicates
        # This avoids N+1 query problem when checking if log IDs already exist
        existing_log_ids = set(Log.objects.values_list('id', flat=True))
        
        # Pre-load existing logs to check for duplicates based on content
        # Create a set of tuples (event_type, contact_id, user_id, created_at) for fast lookup
        # This allows us to skip logs that are exact duplicates even if they have different IDs
        existing_logs_signatures = set()
        existing_logs_query = Log.objects.select_related('contact_id', 'user_id').values(
            'event_type', 'contact_id', 'user_id', 'created_at'
        )
        for log in existing_logs_query:
            # Create a signature tuple for duplicate detection
            # Normalize created_at to minute precision to handle slight time differences
            created_at = log['created_at']
            if created_at:
                # Round to minute precision to handle slight time differences from CSV imports
                if timezone.is_aware(created_at):
                    created_at_normalized = created_at.replace(second=0, microsecond=0)
                else:
                    created_at_normalized = timezone.make_aware(created_at).replace(second=0, microsecond=0)
            else:
                created_at_normalized = None
            
            signature = (
                log['event_type'] or '',
                log['contact_id'] or None,
                log['user_id'] or None,
                created_at_normalized
            )
            existing_logs_signatures.add(signature)
        
        # Pre-load all users for fast lookup
        # Use both string and integer keys to handle different ID formats
        users_by_id = {}
        all_users = DjangoUser.objects.all()
        for user in all_users:
            user_id_str = str(user.id)
            user_id_int = user.id
            # Store with both string and integer keys for flexible lookup
            users_by_id[user_id_str] = user
            if isinstance(user_id_int, int):
                users_by_id[user_id_int] = user
        
        # First pass: Parse all rows and collect valid logs
        for row_num, row in enumerate(csv_reader, start=2):  # Start at 2 (row 1 is header)
            results['total'] += 1
            try:
                # Build log data from CSV row
                log_data = {}
                old_contact_id = None
                
                # Map CSV columns to log fields
                for frontend_field, csv_column in column_mapping.items():
                    if not csv_column:
                        continue
                    
                    # Try exact match first, then case-insensitive match
                    csv_value = None
                    if csv_column in row:
                        csv_value = row[csv_column]
                    else:
                        # Try case-insensitive match
                        for key in row.keys():
                            if key and key.strip().lower() == csv_column.strip().lower():
                                csv_value = row[key]
                                break
                    
                    if csv_value is None:
                        continue
                    
                    value = csv_value.strip() if csv_value else ''
                    
                    # Map to model field name
                    if frontend_field in field_mapping:
                        model_field = field_mapping[frontend_field]
                        
                        # Handle datetime field
                        if model_field == 'created_at':
                            parsed_dt = parse_datetime(value)
                            if parsed_dt:
                                from django.utils import timezone
                                log_data[model_field] = timezone.make_aware(parsed_dt)
                            else:
                                log_data[model_field] = None
                        # Handle JSON fields
                        elif model_field in ['details', 'old_value', 'new_value']:
                            log_data[model_field] = parse_json(value)
                        # Handle old_logs as text field (not JSON)
                        elif model_field == 'old_logs':
                            log_data[model_field] = value.strip() if value else None
                        # Handle old contact ID (store separately for lookup)
                        elif model_field == 'old_contact_id':
                            old_contact_id = value.strip() if value and value.strip() else None
                        # Handle event_type with mapping
                        elif model_field == 'event_type':
                            event_type_value = value if value else ''
                            # Apply event type mapping if provided
                            if event_type_value in event_type_mapping and event_type_mapping[event_type_value]:
                                log_data[model_field] = str(event_type_mapping[event_type_value]).strip()
                            else:
                                log_data[model_field] = event_type_value.strip()
                        else:
                            log_data[model_field] = value
                
                # Validate required fields - event_type is required
                if not log_data.get('event_type'):
                    results['errors'].append({
                        'row': row_num,
                        'error': 'event_type is required'
                    })
                    results['failed'] += 1
                    continue
                
                # Handle contact ID mapping by old_contact_id
                # If old_contact_id is provided in CSV, contact must be found, otherwise skip this log
                contact_obj = None
                if old_contact_id:
                    # Strip whitespace and use pre-loaded dictionary for fast lookup (O(1) instead of database query)
                    old_contact_id_clean = str(old_contact_id).strip()
                    contact_obj = contacts_by_old_id.get(old_contact_id_clean)
                    if not contact_obj:
                        # Contact not found - skip this log since old_contact_id was provided
                        results['errors'].append({
                            'row': row_num,
                            'error': f'Contact not found with old_contact_id: {old_contact_id_clean}'
                        })
                        results['failed'] += 1
                        continue
                
                # Handle user_id - use from CSV if provided, otherwise use default or None
                # If userId column is mapped, user_id should not be null
                user_obj = None
                userId_column_mapped = 'userId' in column_mapping and column_mapping.get('userId')
                
                if 'user_id' in log_data and log_data['user_id']:
                    user_id_str = str(log_data['user_id']).strip()
                    # Apply user ID mapping if provided
                    if user_id_str in user_id_mapping and user_id_mapping[user_id_str]:
                        mapped_user_id = str(user_id_mapping[user_id_str]).strip()
                        # Try to find user with mapped ID (try both string and original format)
                        user_obj = users_by_id.get(mapped_user_id) or users_by_id.get(user_id_mapping[user_id_str])
                        # If mapped user not found, try default user
                        if not user_obj:
                            user_obj = default_user_obj
                    else:
                        # No mapping found for this CSV value, check if CSV value itself is a valid user ID
                        user_obj = users_by_id.get(user_id_str)
                        # Also try integer version if user_id_str is numeric
                        if not user_obj and user_id_str.isdigit():
                            try:
                                user_obj = users_by_id.get(int(user_id_str))
                            except ValueError:
                                pass
                        # If CSV value is not a valid user ID, use default user
                        if not user_obj:
                            user_obj = default_user_obj
                elif userId_column_mapped:
                    # Column is mapped but value is empty/null, use default user
                    user_obj = default_user_obj
                else:
                    # No userId column mapped, use default user if available (optional)
                    user_obj = default_user_obj
                
                # If userId column is mapped but user_obj is still None, skip this log
                if userId_column_mapped and not user_obj:
                    csv_val = log_data.get('user_id', 'N/A')
                    mapping_val = user_id_mapping.get(str(csv_val), 'N/A')
                    results['errors'].append({
                        'row': row_num,
                        'error': f'User ID column is mapped but no valid user found. CSV value: "{csv_val}", Mapping: "{mapping_val}", Default user ID: "{default_user_id}"'
                    })
                    results['failed'] += 1
                    continue
                
                # Handle log ID - use from CSV if provided, otherwise generate
                log_id = None
                if 'id' in log_data and log_data['id']:
                    # Use ID from CSV
                    log_id = str(log_data['id']).strip()
                    if len(log_id) > 12:
                        log_id = log_id[:12]  # Truncate to 12 characters if longer
                    
                    # Check if ID already exists in database
                    if log_id in existing_log_ids:
                        results['errors'].append({
                            'row': row_num,
                            'error': f'Log with ID {log_id} already exists in database'
                        })
                        results['failed'] += 1
                        continue
                    
                    # Check uniqueness against pending logs in this batch
                    existing_ids = {l.id for l in logs_to_create}
                    if log_id in existing_ids:
                        results['errors'].append({
                            'row': row_num,
                            'error': f'Duplicate ID {log_id} found in CSV'
                        })
                        results['failed'] += 1
                        continue
                else:
                    # Generate new log ID
                    log_id = uuid.uuid4().hex[:12]
                    # Check uniqueness against pending logs in this batch
                    existing_ids = {l.id for l in logs_to_create}
                    while log_id in existing_ids or log_id in existing_log_ids:
                        log_id = uuid.uuid4().hex[:12]
                
                log_data['id'] = log_id
                log_data['user_id'] = user_obj
                
                # Remove old_contact_id from log_data (it's not a Log field)
                # Remove creator_id from log_data (it's not a Log field anymore)
                log_data.pop('creator_id', None)
                log_data.pop('old_contact_id', None)
                
                # Store created_at separately if provided (needs special handling)
                custom_created_at = log_data.pop('created_at', None)
                
                # Check if log already exists in database based on content (not just ID)
                # Create signature for duplicate detection
                if custom_created_at:
                    # Normalize created_at to minute precision to handle slight time differences
                    if timezone.is_aware(custom_created_at):
                        created_at_normalized = custom_created_at.replace(second=0, microsecond=0)
                    else:
                        created_at_normalized = timezone.make_aware(custom_created_at).replace(second=0, microsecond=0)
                else:
                    created_at_normalized = None
                
                log_signature = (
                    log_data.get('event_type', '') or '',
                    contact_obj.id if contact_obj else None,
                    user_obj.id if user_obj else None,
                    created_at_normalized
                )
                
                # Check if this log already exists
                if log_signature in existing_logs_signatures:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Log already exists in database (duplicate: event_type={log_data.get("event_type")}, contact_id={contact_obj.id if contact_obj else None}, user_id={user_obj.id if user_obj else None}, created_at={created_at_normalized})'
                    })
                    results['failed'] += 1
                    continue
                
                # Also check against logs in current batch to avoid duplicates within CSV
                batch_signature_exists = False
                for pending_log in logs_to_create:
                    pending_created_at = row_data_map.get(pending_log.id, {}).get('created_at')
                    if pending_created_at:
                        if timezone.is_aware(pending_created_at):
                            pending_created_at_normalized = pending_created_at.replace(second=0, microsecond=0)
                        else:
                            pending_created_at_normalized = timezone.make_aware(pending_created_at).replace(second=0, microsecond=0)
                    else:
                        pending_created_at_normalized = None
                    
                    pending_signature = (
                        pending_log.event_type or '',
                        pending_log.contact_id.id if pending_log.contact_id else None,
                        pending_log.user_id.id if pending_log.user_id else None,
                        pending_created_at_normalized
                    )
                    
                    if pending_signature == log_signature:
                        batch_signature_exists = True
                        break
                
                if batch_signature_exists:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Duplicate log found in CSV (same event_type, contact_id, user_id, and created_at)'
                    })
                    results['failed'] += 1
                    continue
                
                # Add to existing signatures set to track duplicates within batch
                existing_logs_signatures.add(log_signature)
                
                # Store row data for results
                event_type_preview = log_data.get('event_type', '')[:50] + ('...' if len(log_data.get('event_type', '')) > 50 else '')
                
                row_data_map[log_id] = {
                    'row': row_num,
                    'event_type': event_type_preview,
                    'created_at': custom_created_at,  # Store for later update
                }
                
                # Create Log instance (not saved yet)
                log = Log(**log_data)
                # Explicitly set contact_id and user_id to ensure they're assigned correctly
                log.contact_id = contact_obj
                log.user_id = user_obj
                # Set created_at if provided (this will override auto_now_add)
                if custom_created_at:
                    log.created_at = custom_created_at
                logs_to_create.append(log)
                
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                results['errors'].append({
                    'row': row_num,
                    'error': str(e),
                    'details': error_details
                })
                results['failed'] += 1
        
        # Bulk create logs in batches
        from django.db import transaction, IntegrityError
        
        with transaction.atomic():
            for i in range(0, len(logs_to_create), BATCH_SIZE):
                batch = logs_to_create[i:i + BATCH_SIZE]
                try:
                    Log.objects.bulk_create(batch, batch_size=BATCH_SIZE)
                    
                    # Update created_at for logs that have custom timestamps from CSV
                    # bulk_create may override created_at due to auto_now_add, so we update it afterward
                    logs_to_update = []
                    for log in batch:
                        row_data = row_data_map[log.id]
                        custom_created_at = row_data.get('created_at')
                        if custom_created_at:
                            log.created_at = custom_created_at
                            logs_to_update.append(log)
                    
                    # Bulk update created_at for logs with custom timestamps
                    if logs_to_update:
                        Log.objects.bulk_update(logs_to_update, ['created_at'], batch_size=BATCH_SIZE)
                    
                    # Add to success results
                    for log in batch:
                        row_data = row_data_map[log.id]
                        results['success'].append({
                            'row': row_data['row'],
                            'logId': log.id,
                            'event_type': row_data['event_type']
                        })
                        results['imported'] += 1
                except IntegrityError as e:
                    # Handle potential ID collisions by falling back to individual creates for this batch
                    for log in batch:
                        try:
                            log.save()
                            row_data = row_data_map[log.id]
                            results['success'].append({
                                'row': row_data['row'],
                                'logId': log.id,
                                'event_type': row_data['event_type']
                            })
                            results['imported'] += 1
                        except IntegrityError:
                            # ID collision - regenerate and try once more
                            log.id = uuid.uuid4().hex[:12]
                            try:
                                log.save()
                                row_data = row_data_map[log.id]
                                results['success'].append({
                                    'row': row_data['row'],
                                    'logId': log.id,
                                    'event_type': row_data['event_type']
                                })
                                results['imported'] += 1
                            except Exception as e:
                                row_data = row_data_map.get(log.id, {})
                                results['errors'].append({
                                    'row': row_data.get('row', 'unknown'),
                                    'error': f'Failed to create log: {str(e)}'
                                })
                                results['failed'] += 1
                        except Exception as e:
                            row_data = row_data_map.get(log.id, {})
                            results['errors'].append({
                                'row': row_data.get('row', 'unknown'),
                                'error': f'Failed to create log: {str(e)}'
                            })
                            results['failed'] += 1
        
        return Response(results, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def contact_detail(request, contact_id):
    """
    Get or update a contact, respecting data_access restrictions.
    Users with own_only:
        - If teleoperateur: Can only access contacts where they are teleoperator
        - If confirmateur: Can only access contacts where they are confirmateur
        - Otherwise: Can access contacts where they are teleoperator, confirmateur, or creator
    Users with team_only can access contacts from their team.
    Users with all can access any contact.
    """
    # Optimize query with select_related and prefetch_related to avoid N+1 queries
    try:
        from django.db.models import Prefetch
        contact = Contact.objects.select_related(
            'status',
            'source',
            'platform',
            'teleoperator',
            'confirmateur',
            'creator'
        ).prefetch_related(
            Prefetch(
                'teleoperator__user_details__team_memberships',
                queryset=TeamMember.objects.select_related('team')
            ),
            Prefetch(
                'confirmateur__user_details__team_memberships',
                queryset=TeamMember.objects.select_related('team')
            ),
            Prefetch(
                'creator__user_details__team_memberships',
                queryset=TeamMember.objects.select_related('team')
            )
        ).get(id=contact_id)
    except Contact.DoesNotExist:
        return Response(
            {'error': 'Contact non trouvé'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    user = request.user
    
    # Check if contact is a fosse contact (unassigned: teleoperator is null AND confirmateur is null)
    is_fosse_contact = contact.teleoperator is None and contact.confirmateur is None
    
    # Use module-level import (UserDetails is imported at top of file)
    # Alias it locally to avoid any scoping conflicts
    UserDetailsModel = UserDetails
    
    # Check data access restrictions
    try:
        user_details = UserDetailsModel.objects.get(django_user=user)
        
        # Special case: If contact is in fosse, check if user has fosse view permission
        has_fosse_permission = False
        if is_fosse_contact and user_details.role:
            from api.models import Permission, PermissionRole
            has_fosse_permission = PermissionRole.objects.filter(
                role=user_details.role,
                permission__component='fosse',
                permission__action='view'
            ).exists()
            
            if has_fosse_permission and request.method == 'GET':
                # User has fosse view permission, allow access for GET requests
                serializer = ContactSerializer(contact, context={'request': request})
                return Response({'contact': serializer.data})
        
        if user_details.role:
            data_access = user_details.role.data_access
            
            # Skip data_access restrictions if user has fosse permission for fosse contacts
            if has_fosse_permission and is_fosse_contact:
                # User has fosse permission, skip data_access checks
                pass
            else:
                # Special case for PATCH: Allow if user is setting themselves as teleoperator/confirmateur
                allow_patch_for_self_assignment = False
                if request.method == 'PATCH':
                    is_teleoperateur = user_details.role.is_teleoperateur
                    is_confirmateur = user_details.role.is_confirmateur
                    
                    # Check if user is trying to assign themselves as teleoperator
                    if is_teleoperateur and 'teleoperatorId' in request.data:
                        teleoperator_id = request.data.get('teleoperatorId')
                        if teleoperator_id and str(teleoperator_id) == str(user.id):
                            allow_patch_for_self_assignment = True
                    
                    # Check if user is trying to assign themselves as confirmateur
                    if is_confirmateur and 'confirmateurId' in request.data:
                        confirmateur_id = request.data.get('confirmateurId')
                        if confirmateur_id and str(confirmateur_id) == str(user.id):
                            allow_patch_for_self_assignment = True
                
                if data_access == 'own_only' and not allow_patch_for_self_assignment:
                    # Check if user is teleoperateur or confirmateur
                    is_teleoperateur = user_details.role.is_teleoperateur
                    is_confirmateur = user_details.role.is_confirmateur
                    
                    # For GET requests, also check if user has field-level permissions that might allow access
                    # This handles cases where user is viewing a contact they might assign themselves to
                    has_field_permission = False
                    if request.method == 'GET' and user_details.role:
                        # Check if user's role has fiche_contact edit permission for teleoperatorId or confirmateurId
                        from api.models import Permission, PermissionRole
                        if is_teleoperateur:
                            has_field_permission = PermissionRole.objects.filter(
                                role=user_details.role,
                                permission__component='fiche_contact',
                                permission__action='edit',
                                permission__field_name='teleoperatorId'
                            ).exists()
                        if is_confirmateur and not has_field_permission:
                            has_field_permission = PermissionRole.objects.filter(
                                role=user_details.role,
                                permission__component='fiche_contact',
                                permission__action='edit',
                                permission__field_name='confirmateurId'
                            ).exists()
                    
                    if is_teleoperateur and is_confirmateur:
                        # User is both: allow if user is teleoperator OR confirmateur OR has field permission
                        if contact.teleoperator != user and contact.confirmateur != user and not has_field_permission:
                            return Response(
                                {'error': 'Vous n\'avez pas accès à ce contact'},
                                status=status.HTTP_403_FORBIDDEN
                            )
                    elif is_teleoperateur:
                        # Teleoperateur with own_only: allow if user is teleoperator OR has field permission
                        if contact.teleoperator != user and not has_field_permission:
                            return Response(
                                {'error': 'Vous n\'avez pas accès à ce contact'},
                                status=status.HTTP_403_FORBIDDEN
                            )
                    elif is_confirmateur:
                        # Confirmateur with own_only: allow if user is confirmateur OR has field permission
                        if contact.confirmateur != user and not has_field_permission:
                            return Response(
                                {'error': 'Vous n\'avez pas accès à ce contact'},
                                status=status.HTTP_403_FORBIDDEN
                            )
                    else:
                        # Default behavior: only allow if user is teleoperator, confirmateur, or creator
                        if contact.teleoperator != user and contact.confirmateur != user and contact.creator != user:
                            return Response(
                                {'error': 'Vous n\'avez pas accès à ce contact'},
                                status=status.HTTP_403_FORBIDDEN
                            )
                elif data_access == 'team_only':
                    # Check if user has access (either assigned to them or from their team)
                    team_member = user_details.team_memberships.first()
                    if team_member:
                        team = team_member.team
                        team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                        # Allow if user is assigned OR if contact's assignees are in the same team
                        if (contact.teleoperator != user and contact.confirmateur != user and contact.creator != user and
                            (not contact.teleoperator or contact.teleoperator.id not in team_user_ids) and
                            (not contact.confirmateur or contact.confirmateur.id not in team_user_ids) and
                            (not contact.creator or contact.creator.id not in team_user_ids)):
                            return Response(
                                {'error': 'Vous n\'avez pas accès à ce contact'},
                                status=status.HTTP_403_FORBIDDEN
                            )
                    else:
                        # User has no team, fall back to own_only behavior
                        if contact.teleoperator != user and contact.confirmateur != user and contact.creator != user:
                            return Response(
                                {'error': 'Vous n\'avez pas accès à ce contact'},
                                status=status.HTTP_403_FORBIDDEN
                            )
                # If data_access is 'all', allow access (no check needed)
    except UserDetailsModel.DoesNotExist:
        # If user has no UserDetails, deny access (safety default)
        return Response(
            {'error': 'Vous n\'avez pas accès à ce contact'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    if request.method == 'GET':
        serializer = ContactSerializer(contact, context={'request': request})
        return Response({'contact': serializer.data})
    
    if request.method == 'PATCH':
        # Track if we need to send client notification
        should_send_client_notification = False
        # Track if we need to send confirmateur assignment notification
        should_send_confirmateur_notification = False
        new_confirmateur_user = None
        
        try:
            # Get old value BEFORE any modifications
            try:
                old_serializer = ContactSerializer(contact, context={'request': request})
                old_value_raw = old_serializer.data
                old_value = clean_contact_data_for_log(old_value_raw, include_created_at=False)
            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"Error getting old value for contact: {error_details}")
                return Response(
                    {'error': f'Erreur lors de la récupération des données: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Helper functions
            def get_date(value):
                if not value:
                    return None
                try:
                    from datetime import datetime
                    # Handle both YYYY-MM-DD and DD/MM/YYYY formats
                    if '/' in str(value):
                        parts = str(value).split('/')
                        if len(parts) == 3:
                            day, month, year = parts
                            return datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d").date()
                    return datetime.strptime(str(value), "%Y-%m-%d").date()
                except:
                    return None
            
            # Update personal information fields
            if 'civility' in request.data:
                contact.civility = request.data.get('civility', '') or ''
            if 'firstName' in request.data:
                contact.fname = request.data.get('firstName', '') or ''
            if 'lastName' in request.data:
                contact.lname = request.data.get('lastName', '') or ''
            if 'phone' in request.data:
                phone_value = request.data.get('phone')
                # Handle None, empty string, or whitespace-only strings
                # Convert empty strings and whitespace to None explicitly
                if phone_value is None:
                    contact.phone = None
                elif isinstance(phone_value, str) and not phone_value.strip():
                    # Explicitly handle empty string or whitespace-only string
                    contact.phone = None
                else:
                    try:
                        # Remove spaces and convert to int
                        cleaned = ''.join(str(phone_value).split())
                        if cleaned:
                            contact.phone = int(cleaned)
                        else:
                            contact.phone = None
                    except (ValueError, TypeError):
                        contact.phone = None
            if 'mobile' in request.data:
                mobile_value = request.data.get('mobile')
                # Handle None, empty string, or whitespace-only strings
                # IMPORTANT: If mobile is required (NOT NULL constraint), we need to handle empty values differently
                # Check if the value is actually empty/null
                if mobile_value is None or (isinstance(mobile_value, str) and not mobile_value.strip()):
                    # If mobile is empty/null, check if DB allows NULL
                    # If DB has NOT NULL constraint, we'll keep existing value (handled in final check before save)
                    # For now, set to None - will be corrected before save if DB doesn't allow it
                    contact.mobile = None
                else:
                    try:
                        # Remove spaces and convert to int
                        cleaned = ''.join(str(mobile_value).split())
                        if cleaned:
                            contact.mobile = int(cleaned)
                        else:
                            # Empty after cleaning - set to None (will be corrected before save if needed)
                            contact.mobile = None
                    except (ValueError, TypeError):
                        # Invalid value - set to None (will be corrected before save if needed)
                        contact.mobile = None
            if 'email' in request.data:
                contact.email = request.data.get('email', '') or ''
            if 'birthDate' in request.data:
                contact.birth_date = get_date(request.data.get('birthDate'))
            if 'birthPlace' in request.data:
                contact.birth_place = request.data.get('birthPlace', '') or ''
            if 'address' in request.data:
                contact.address = request.data.get('address', '') or ''
            if 'postalCode' in request.data:
                contact.postal_code = request.data.get('postalCode', '') or ''
            if 'city' in request.data:
                contact.city = request.data.get('city', '') or ''
            if 'nationality' in request.data:
                contact.nationality = request.data.get('nationality', '') or ''
            if 'autreInformations' in request.data:
                contact.autre_informations = request.data.get('autreInformations', '') or ''
            if 'dateInscription' in request.data:
                contact.date_d_inscription = request.data.get('dateInscription', '') or ''
            
            # Update status if provided
            if 'statusId' in request.data:
                status_id = request.data.get('statusId')
                if status_id:
                    try:
                        status_obj = Status.objects.filter(id=status_id).first()
                        if status_obj:
                            # Get old status type directly from database to ensure we have the correct value
                            old_status_type = None
                            print(f"[DEBUG] Checking old status - contact.status_id={contact.status_id}, contact.status={contact.status}")
                            if contact.status_id:
                                old_status = Status.objects.filter(id=contact.status_id).values_list('type', flat=True).first()
                                print(f"[DEBUG] Queried old status type from DB: {old_status}")
                                if old_status:
                                    old_status_type = old_status
                            else:
                                print(f"[DEBUG] No old status_id, contact has no current status")
                            
                            # Get new status type directly from database
                            new_status_type = Status.objects.filter(id=status_id).values_list('type', flat=True).first()
                            
                            # Debug logging with print statements for visibility
                            print(f"[DEBUG] Status change for contact {contact.id}:")
                            print(f"  - old_status_id={contact.status_id}")
                            print(f"  - old_type={old_status_type}")
                            print(f"  - new_status_id={status_id}")
                            print(f"  - new_type={new_status_type}")
                            print(f"  - current date_lead_to_client={contact.date_lead_to_client}")
                            
                            # If moving from lead to client, set date_lead_to_client (only if not already set)
                            if old_status_type == 'lead' and new_status_type == 'client':
                                print(f"[DEBUG] Condition met: lead -> client transition")
                                if not contact.date_lead_to_client:
                                    from django.utils import timezone as tz
                                    contact.date_lead_to_client = tz.now()
                                    print(f"[DEBUG] Set date_lead_to_client={contact.date_lead_to_client} for contact {contact.id}")
                                else:
                                    print(f"[DEBUG] date_lead_to_client already set to {contact.date_lead_to_client}, skipping")
                            else:
                                print(f"[DEBUG] Condition NOT met: old_type={old_status_type}, new_type={new_status_type}")
                            
                            # Check if new status is client_default=True
                            old_status_obj = None
                            if contact.status_id:
                                old_status_obj = Status.objects.filter(id=contact.status_id).first()
                            
                            # Check if new status has client_default=True
                            is_new_client_default = status_obj.client_default if status_obj else False
                            is_old_client_default = old_status_obj.client_default if old_status_obj else False
                            
                            # Track if we need to send notification (will send after contact is saved)
                            if is_new_client_default and not is_old_client_default:
                                should_send_client_notification = True
                            
                            contact.status = status_obj
                    except Exception as e:
                        # Log the exception instead of silently passing
                        import logging
                        import traceback
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error updating contact status: {str(e)}\n{traceback.format_exc()}")
                else:
                    contact.status = None
            
            # Update source if provided
            if 'sourceId' in request.data:
                source_id = request.data.get('sourceId')
                if source_id:
                    try:
                        source_obj = Source.objects.filter(id=source_id).first()
                        if source_obj:
                            contact.source = source_obj
                    except Exception:
                        pass
                else:
                    contact.source = None
            
            # Update platform if provided
            if 'platformId' in request.data:
                platform_id = request.data.get('platformId')
                if platform_id:
                    try:
                        platform_obj = Platform.objects.filter(id=platform_id).first()
                        if platform_obj:
                            contact.platform = platform_obj
                    except Exception:
                        pass
                else:
                    contact.platform = None
            
            # Update confirmateur fields if provided
            if 'montantEncaisse' in request.data:
                montant_encaisse = request.data.get('montantEncaisse')
                if montant_encaisse:
                    try:
                        contact.montant_encaisse = float(montant_encaisse)
                    except (ValueError, TypeError):
                        contact.montant_encaisse = None
                else:
                    contact.montant_encaisse = None
            
            if 'bonus' in request.data:
                bonus = request.data.get('bonus')
                if bonus:
                    try:
                        contact.bonus = float(bonus)
                    except (ValueError, TypeError):
                        contact.bonus = None
                else:
                    contact.bonus = None
            
            if 'paiement' in request.data:
                contact.paiement = request.data.get('paiement', '') or ''
            
            if 'contrat' in request.data:
                contact.contrat = request.data.get('contrat', '') or ''
            
            if 'nomDeScene' in request.data:
                contact.nom_de_scene = request.data.get('nomDeScene', '') or ''
            
            if 'dateProTr' in request.data:
                contact.date_pro_tr = request.data.get('dateProTr', '') or ''
            
            if 'potentiel' in request.data:
                contact.potentiel = request.data.get('potentiel', '') or ''
            
            if 'produit' in request.data:
                contact.produit = request.data.get('produit', '') or ''
            
            # Initialize change tracking variables
            teleoperator_changed = False
            confirmateur_changed = False
            
            # Update teleoperator if provided
            if 'teleoperatorId' in request.data:
                teleoperator_id_raw = request.data.get('teleoperatorId')
                
                # Normalize empty string, 'none', or None to None
                if teleoperator_id_raw is None or teleoperator_id_raw == '' or teleoperator_id_raw == 'none':
                    new_teleoperator_id = None
                else:
                    # Keep as string - frontend sends UserDetails IDs
                    new_teleoperator_id = str(teleoperator_id_raw).strip()
                
                # Get old teleoperator Django User ID (stored as integer in database)
                old_teleoperator_django_id = contact.teleoperator_id
                
                # Get old teleoperator UserDetails ID for comparison
                old_teleoperator_userdetails_id = None
                if old_teleoperator_django_id:
                    try:
                        # Use UserDetailsModel alias to avoid scoping issues
                        old_user_details = UserDetailsModel.objects.filter(django_user_id=old_teleoperator_django_id).first()
                        if old_user_details:
                            old_teleoperator_userdetails_id = old_user_details.id
                    except Exception:
                        pass
                
                # Teleoperator changed if:
                # 1. Old is None and new is not None (assigning)
                # 2. Old is not None and new is None (clearing)
                # 3. Old and new are both not None but different (reassigning)
                # Compare UserDetails IDs (strings)
                if old_teleoperator_userdetails_id is None and new_teleoperator_id is not None:
                    teleoperator_changed = True
                    change_type = "assigning (None -> value)"
                elif old_teleoperator_userdetails_id is not None and new_teleoperator_id is None:
                    teleoperator_changed = True
                    change_type = "clearing (value -> None)"
                elif old_teleoperator_userdetails_id is not None and new_teleoperator_id is not None:
                    # Both are not None, compare as strings (UserDetails IDs)
                    teleoperator_changed = (str(old_teleoperator_userdetails_id) != str(new_teleoperator_id))
                    change_type = f"reassigning ({old_teleoperator_userdetails_id} -> {new_teleoperator_id})" if teleoperator_changed else f"same value ({old_teleoperator_userdetails_id})"
                else:
                    # Both are None
                    teleoperator_changed = False
                
                if new_teleoperator_id is not None:
                    try:
                        # Prioritize UserDetails ID lookup first (since frontend sends UserDetails IDs)
                        teleoperator_user = None
                        
                        # First try as UserDetails ID (string) - this is what frontend sends
                        try:
                            user_details = UserDetailsModel.objects.filter(id=str(new_teleoperator_id)).first()
                            if user_details and user_details.django_user:
                                teleoperator_user = user_details.django_user
                        except Exception:
                            pass
                        
                        # Fallback: Try Django User ID lookup only if UserDetails lookup failed
                        if not teleoperator_user:
                            try:
                                # Try as Django User ID (if it's numeric)
                                int_id = int(new_teleoperator_id)
                                teleoperator_user = DjangoUser.objects.filter(id=int_id).first()
                            except (ValueError, TypeError):
                                pass
                        
                        if teleoperator_user:
                            contact.teleoperator = teleoperator_user
                            # Update assigned_at when teleoperator changes (assigns or reassigns)
                            if teleoperator_changed:
                                from django.utils import timezone
                                contact.assigned_at = timezone.now()
                                contact._assigned_at_was_set = True  # Mark that we set assigned_at
                            # Clear any cached relationship
                            if hasattr(contact, '_teleoperator_cache'):
                                delattr(contact, '_teleoperator_cache')
                        else:
                            # Don't clear the teleoperator if user not found - return error instead
                            return Response(
                                {'error': f'Utilisateur téléopérateur avec l\'ID {new_teleoperator_id} non trouvé'},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                else:
                    # Clearing teleoperator
                    contact.teleoperator = None
                    # Clear assigned_at when teleoperator is removed
                    if teleoperator_changed:
                        contact.assigned_at = None
                        contact._assigned_at_was_set = True  # Mark that we modified assigned_at
                    # Clear any cached relationship
                    if hasattr(contact, '_teleoperator_cache'):
                        delattr(contact, '_teleoperator_cache')
            
            # Update confirmateur if provided
            if 'confirmateurId' in request.data:
                confirmateur_id_raw = request.data.get('confirmateurId')
                
                # Normalize empty string, 'none', or None to None
                if confirmateur_id_raw is None or confirmateur_id_raw == '' or confirmateur_id_raw == 'none':
                    new_confirmateur_id = None
                else:
                    # Keep as string - frontend sends UserDetails IDs
                    new_confirmateur_id = str(confirmateur_id_raw).strip()
                
                # Get old confirmateur Django User ID (stored as integer in database)
                old_confirmateur_django_id = contact.confirmateur_id
                
                # Get old confirmateur UserDetails ID for comparison
                old_confirmateur_userdetails_id = None
                if old_confirmateur_django_id:
                    try:
                        # Use UserDetailsModel alias to avoid scoping issues
                        old_user_details = UserDetailsModel.objects.filter(django_user_id=old_confirmateur_django_id).first()
                        if old_user_details:
                            old_confirmateur_userdetails_id = old_user_details.id
                    except Exception:
                        pass
                
                # Confirmateur changed if:
                # 1. Old is None and new is not None (assigning)
                # 2. Old is not None and new is None (clearing)
                # 3. Old and new are both not None but different (reassigning)
                # Compare UserDetails IDs (strings)
                if old_confirmateur_userdetails_id is None and new_confirmateur_id is not None:
                    confirmateur_changed = True
                elif old_confirmateur_userdetails_id is not None and new_confirmateur_id is None:
                    confirmateur_changed = True
                elif old_confirmateur_userdetails_id is not None and new_confirmateur_id is not None:
                    # Both are not None, compare as strings (UserDetails IDs)
                    confirmateur_changed = (str(old_confirmateur_userdetails_id) != str(new_confirmateur_id))
                else:
                    # Both are None
                    confirmateur_changed = False
                
                if new_confirmateur_id is not None:
                    try:
                        # Prioritize UserDetails ID lookup first (since frontend sends UserDetails IDs)
                        confirmateur_user = None
                        
                        # First try as UserDetails ID (string) - this is what frontend sends
                        try:
                            user_details = UserDetailsModel.objects.filter(id=str(new_confirmateur_id)).first()
                            if user_details and user_details.django_user:
                                confirmateur_user = user_details.django_user
                        except Exception:
                            pass
                        
                        # Fallback: Try Django User ID lookup only if UserDetails lookup failed
                        if not confirmateur_user:
                            try:
                                # Try as Django User ID (if it's numeric)
                                int_id = int(new_confirmateur_id)
                                confirmateur_user = DjangoUser.objects.filter(id=int_id).first()
                            except (ValueError, TypeError):
                                pass
                        
                        if confirmateur_user:
                            contact.confirmateur = confirmateur_user
                            # Clear any cached relationship
                            if hasattr(contact, '_confirmateur_cache'):
                                delattr(contact, '_confirmateur_cache')
                            
                            # Check if confirmateur was assigned (from null) and contact has client_default status
                            if old_confirmateur_django_id is None and contact.status and contact.status.client_default:
                                should_send_confirmateur_notification = True
                                new_confirmateur_user = confirmateur_user
                            
                            # Automation: If confirmateur was null and we're assigning one,
                            # assign all upcoming events to this confirmateur
                            # Check old_confirmateur_django_id directly (database value) to ensure it was null
                            if old_confirmateur_django_id is None:
                                try:
                                    # Import timezone locally to avoid scoping conflict with line 5884
                                    from django.utils import timezone as tz
                                    now = tz.now()
                                    # Get all upcoming events for this contact using explicit ID lookup
                                    # Use datetime__gt to match event_list behavior (future events)
                                    upcoming_events = Event.objects.filter(
                                        contactId__id=contact.id,
                                        datetime__gt=now
                                    )
                                    
                                    # Update all upcoming events to assign them to the confirmateur
                                    upcoming_events.update(userId=confirmateur_user)
                                except Exception:
                                    # Silently fail - don't break the contact update
                                    pass
                        else:
                            # Don't clear the confirmateur if user not found - return error instead
                            return Response(
                                {'error': f'Utilisateur confirmateur avec l\'ID {new_confirmateur_id} non trouvé'},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                    except Exception:
                        import traceback
                        traceback.print_exc()
                else:
                    # Clearing confirmateur
                    contact.confirmateur = None
                    # Clear any cached relationship
                    if hasattr(contact, '_confirmateur_cache'):
                        delattr(contact, '_confirmateur_cache')
            
            # Update campaign if provided
            if 'campaign' in request.data:
                contact.campaign = request.data.get('campaign', '') or ''
            
            # Update addressComplement if provided
            if 'addressComplement' in request.data:
                contact.address_complement = request.data.get('addressComplement', '') or ''
            
            # Update confirmateur email and telephone if provided
            if 'confirmateur_email' in request.data:
                confirmateur_email_value = request.data.get('confirmateur_email', '')
                contact.confirmateur_email = confirmateur_email_value.strip() if confirmateur_email_value else ''
            elif 'confirmateurEmail' in request.data:
                # Handle camelCase version from frontend
                confirmateur_email_value = request.data.get('confirmateurEmail', '')
                contact.confirmateur_email = confirmateur_email_value.strip() if confirmateur_email_value else ''
            
            if 'confirmateur_telephone' in request.data:
                confirmateur_telephone_value = request.data.get('confirmateur_telephone')
                # Handle None, empty string, or whitespace-only strings
                # Note: CharField without null=True cannot store None, so use empty string instead
                if confirmateur_telephone_value is None or (isinstance(confirmateur_telephone_value, str) and not confirmateur_telephone_value.strip()):
                    contact.confirmateur_telephone = ''
                else:
                    try:
                        # Remove spaces and convert to int (if it's a phone number)
                        cleaned = ''.join(str(confirmateur_telephone_value).split())
                        if cleaned:
                            contact.confirmateur_telephone = cleaned  # Keep as string for phone numbers
                        else:
                            contact.confirmateur_telephone = ''
                    except (ValueError, TypeError):
                        # If conversion fails, keep as string
                        contact.confirmateur_telephone = str(confirmateur_telephone_value).strip() if confirmateur_telephone_value else ''
            elif 'confirmateurTelephone' in request.data:
                # Handle camelCase version from frontend
                confirmateur_telephone_value = request.data.get('confirmateurTelephone')
                # Handle None, empty string, or whitespace-only strings
                # Note: CharField without null=True cannot store None, so use empty string instead
                if confirmateur_telephone_value is None or (isinstance(confirmateur_telephone_value, str) and not confirmateur_telephone_value.strip()):
                    contact.confirmateur_telephone = ''
                else:
                    try:
                        # Remove spaces and convert to int (if it's a phone number)
                        cleaned = ''.join(str(confirmateur_telephone_value).split())
                        if cleaned:
                            contact.confirmateur_telephone = cleaned  # Keep as string for phone numbers
                        else:
                            contact.confirmateur_telephone = ''
                    except (ValueError, TypeError):
                        # If conversion fails, keep as string
                        contact.confirmateur_telephone = str(confirmateur_telephone_value).strip() if confirmateur_telephone_value else ''
            
            # CRITICAL: Ensure phone and mobile are None (not empty strings) before saving
            # This prevents ValueError when saving to BigIntegerField
            # Check both the attribute value and ensure it's not an empty string
            # NOTE: Only update if the field was actually provided in the request
            # This prevents overwriting existing values with None when the field wasn't in the update
            if hasattr(contact, 'phone'):
                if contact.phone == '' or (isinstance(contact.phone, str) and not contact.phone.strip()):
                    # Only set to None if phone was in the request (meaning user wants to clear it)
                    if 'phone' in request.data:
                        contact.phone = None
            if hasattr(contact, 'mobile'):
                if contact.mobile == '' or (isinstance(contact.mobile, str) and not contact.mobile.strip()):
                    # Only set to None if mobile was in the request (meaning user wants to clear it)
                    # But check if the database allows NULL - if not, we might need to keep existing value
                    if 'mobile' in request.data:
                        # Try to set to None, but if database constraint prevents it, keep existing value
                        # The actual None assignment happens above in the mobile update section
                        pass
            
            # CRITICAL: Final check before saving - ensure mobile is not None if DB requires it
            # If mobile is None and DB has NOT NULL constraint, keep existing value
            if hasattr(contact, 'mobile') and contact.mobile is None:
                # Check if mobile was in request.data - if not, don't update it
                if 'mobile' in request.data:
                    # Mobile was provided but is None - check if we need to keep existing value
                    # Reload the contact from DB to get current mobile value
                    try:
                        # Use select_for_update to avoid race conditions, but don't lock if not needed
                        current_contact = Contact.objects.select_for_update(nowait=True).get(id=contact.id)
                        if current_contact.mobile is not None:
                            # Keep existing value if DB doesn't allow NULL
                            contact.mobile = current_contact.mobile
                        else:
                            # Current value is also None - DB constraint will fail, but we'll catch it
                            pass
                    except Contact.DoesNotExist:
                        pass
                    except Exception:
                        # If we can't check (e.g., lock timeout), try to save anyway and let the exception handler catch it
                        pass
            
            # Check if teleoperator or confirmateur was changed and if both are now null
            # If both are null, set status to default fosse status from user's role settings
            # Note: teleoperator_changed and confirmateur_changed are already set above in the update blocks
            
            # Use _id fields to check actual database values (avoids Django ORM caching issues)
            teleoperator_id = getattr(contact, 'teleoperator_id', None)
            confirmateur_id = getattr(contact, 'confirmateur_id', None)
            
            # Only check for default fosse status if we're clearing assignments (both become None)
            # This prevents running the logic when assigning users
            if (teleoperator_changed or confirmateur_changed):
                # Check if both teleoperator and confirmateur are null/empty AFTER the update
                # We check the actual contact object values after they've been updated above
                # Use _id fields to avoid Django ORM caching issues
                # Only set default status if BOTH fields are None (meaning both were cleared)
                if teleoperator_id is None and confirmateur_id is None:
                    # Both are None - check for previous status first, then default fosse status
                    # Note: UserDetailsModel, FosseSettings, Status, and Log are already imported at the top of the file
                    try:
                        # First, try to get previous status from logs
                        previous_status_name = None
                        logs = Log.objects.filter(
                            contact_id=contact,
                            event_type='editContact'
                        ).order_by('-created_at')
                        
                        current_status_name = contact.status.name if contact.status else ''
                        for log in logs:
                            if log.old_value and log.new_value:
                                old_status = log.old_value.get('statusName', '')
                                new_status = log.new_value.get('statusName', '')
                                if old_status and old_status != new_status and new_status == current_status_name:
                                    previous_status_name = old_status
                                    break
                        
                        status_to_use = None
                        
                        # If previous_status exists, use it
                        if previous_status_name:
                            try:
                                status_to_use = Status.objects.get(name=previous_status_name)
                            except Status.DoesNotExist:
                                status_to_use = None
                        
                        # If previous_status is empty, use default status from FosseSettings
                        if not status_to_use:
                            user_details = UserDetailsModel.objects.filter(django_user=request.user).first()
                            if user_details and user_details.role_id:
                                # Use select_related to avoid N+1 query
                                fosse_setting = FosseSettings.objects.select_related('default_status').filter(role_id=user_details.role_id).first()
                                if fosse_setting:
                                    default_status_id = getattr(fosse_setting, 'default_status_id', None)
                                    
                                    if default_status_id:
                                        # Use the status configured in FosseSettings
                                        try:
                                            status_to_use = Status.objects.get(id=default_status_id)
                                        except Status.DoesNotExist:
                                            status_to_use = None
                                    
                                    # Fallback: if no status from FosseSettings, use the status with is_fosse_default=True
                                    if not status_to_use:
                                        status_to_use = Status.objects.filter(is_fosse_default=True).first()
                        
                        # Apply the status if we found one and statusId wasn't explicitly set
                        if status_to_use:
                            if 'statusId' not in request.data:
                                try:
                                    contact.status = status_to_use
                                except Exception:
                                    import traceback
                                    traceback.print_exc()
                    except Exception:
                        # If there's an error getting the status, log it but don't fail the update
                        import traceback
                        traceback.print_exc()
            
            # Save the contact with all modifications
            try:
                # If assigned_at was modified, update it directly in the database first
                # This ensures it's persisted regardless of Django's change tracking
                if hasattr(contact, '_assigned_at_was_set'):
                    assigned_at_value = contact.assigned_at
                    # Direct database update to ensure assigned_at is saved
                    Contact.objects.filter(id=contact.id).update(assigned_at=assigned_at_value)
                    # Refresh the contact object so Django knows about the DB change
                    contact.refresh_from_db(fields=['assigned_at'])
                    delattr(contact, '_assigned_at_was_set')
                
                # Save all other modified fields normally
                contact.save()
                
                # Verify assigned_at was saved
                contact.refresh_from_db()
                
                # Create transaction with type 'Ouverture' when fiche client is filled
                # Check if montantEncaisse was provided in the request
                montant_encaisse_provided = 'montantEncaisse' in request.data
                montant_encaisse_value = None
                
                if montant_encaisse_provided:
                    montant_encaisse_raw = request.data.get('montantEncaisse')
                    
                    # Handle None, empty string, or numeric values
                    if montant_encaisse_raw is not None and montant_encaisse_raw != '':
                        try:
                            montant_encaisse_value = float(montant_encaisse_raw)
                        except (ValueError, TypeError):
                            montant_encaisse_value = None
                
                # Check if contact has client status
                # Priority: Check statusId from request first (the status being set), then check saved contact status
                is_client_status = False
                status_id_from_request = request.data.get('statusId')
                
                if status_id_from_request:
                    # Check the status that was just set in the request
                    new_status_type = Status.objects.filter(id=status_id_from_request).values_list('type', flat=True).first()
                    is_client_status = new_status_type == 'client'
                elif contact.status_id:
                    # Fallback: check the contact's current status after save
                    status_type = Status.objects.filter(id=contact.status_id).values_list('type', flat=True).first()
                    is_client_status = status_type == 'client'
                
                # Create or update transaction with type 'Ouverture' when fiche client is filled
                # A contact can only have one transaction with type 'Ouverture'
                # Condition: montantEncaisse must be provided (can be 0), and contact must have client status
                if montant_encaisse_provided and montant_encaisse_value is not None and is_client_status:
                    try:
                        # Check if a transaction with type 'Ouverture' already exists for this contact
                        existing_ouverture_transaction = Transaction.objects.filter(
                            contact=contact,
                            type='Ouverture'
                        ).first()
                        
                        # Get bonus value if provided
                        bonus_value = None
                        if 'bonus' in request.data:
                            bonus_raw = request.data.get('bonus')
                            if bonus_raw:
                                try:
                                    bonus_value = float(bonus_raw)
                                except (ValueError, TypeError):
                                    bonus_value = None
                        
                        # Get payment type if provided
                        payment_type = request.data.get('paiement', '') or ''
                        
                        # Build comment with montant encaisse and bonus if provided
                        comment_parts = [f'Montant encaissé: {montant_encaisse_value}€']
                        if bonus_value and bonus_value > 0:
                            comment_parts.append(f'Bonus: {bonus_value}€')
                        comment = ' - '.join(comment_parts)
                        
                        if existing_ouverture_transaction:
                            # Update existing transaction (contact already has one 'Ouverture' transaction)
                            # Update: montant encaisse (amount), mode de paiement (payment_type), comment, bonus, and updated_at (automatic)
                            existing_ouverture_transaction.amount = montant_encaisse_value
                            existing_ouverture_transaction.payment_type = payment_type
                            existing_ouverture_transaction.comment = comment
                            existing_ouverture_transaction.bonus = True if bonus_value and bonus_value > 0 else False
                            # Note: updated_at is automatically updated by Django's auto_now=True
                            existing_ouverture_transaction.save()
                            
                            # Send notification to all users who have access to the contact
                            send_transaction_update_notification(contact, existing_ouverture_transaction)
                        else:
                            # No transaction found - create a new one
                            from api.signals import generate_unique_id
                            from django.utils import timezone as tz
                            transaction_id = generate_unique_id(Transaction)
                            
                            transaction = Transaction.objects.create(
                                id=transaction_id,
                                contact=contact,
                                type='Ouverture',
                                status='to_verify',
                                payment_type=payment_type,
                                amount=montant_encaisse_value,
                                date=tz.now(),
                                comment=comment,
                                created_by=request.user if request.user.is_authenticated else None,
                                bonus=True if bonus_value and bonus_value > 0 else False
                            )
                    except Exception as e:
                        # Log error but don't fail the contact update
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error creating/updating transaction with type 'Ouverture' for contact {contact.id}: {str(e)}")
                
                # Send notification if contact moved to client_default status
                if should_send_client_notification:
                    print(f"[DEBUG] Contact {contact.id} moved to client_default status, sending notification")
                    send_contact_notification(contact, notification_type='nouveau_client')
                
                # Send notification if confirmateur was assigned to contact with client_default status
                if should_send_confirmateur_notification and new_confirmateur_user:
                    print(f"[DEBUG] Confirmateur {new_confirmateur_user.id} assigned to contact {contact.id} with client_default status, sending notification")
                    send_confirmateur_assignment_notification(contact, new_confirmateur_user)
                
                # Django automatically handles transaction commits - no manual commit needed
            except ValueError as e:
                # Handle ValueError specifically (e.g., invalid phone/mobile format)
                import traceback
                error_details = traceback.format_exc()
                print(f"Error saving contact: {error_details}")
                return Response(
                    {'error': f'Erreur lors de la sauvegarde: {str(e)}', 'details': error_details},
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                # Handle any other errors (including database constraint violations)
                import traceback
                error_details = traceback.format_exc()
                print(f"Unexpected error saving contact: {error_details}")
                # Check if it's a NOT NULL constraint violation
                error_str = str(e).lower()
                if 'not-null constraint' in error_str or 'null value' in error_str:
                    return Response(
                        {'error': 'Le champ mobile est requis et ne peut pas être vide. Veuillez fournir un numéro de téléphone mobile.', 'details': error_details},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                return Response(
                    {'error': f'Erreur inattendue: {str(e)}', 'details': error_details},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Reload contact from database with all relationships to ensure confirmateur/teleoperator are properly loaded
            # This ensures we get the latest committed data with all relationships
            try:
                contact_id = contact.id
                # Use a fresh query to get the updated contact with all relationships
                contact = Contact.objects.select_related(
                    'status',
                    'source',
                    'teleoperator',
                    'confirmateur',
                    'creator'
                ).prefetch_related(
                    Prefetch(
                        'teleoperator__user_details__team_memberships',
                        queryset=TeamMember.objects.select_related('team')
                    ),
                    Prefetch(
                        'confirmateur__user_details__team_memberships',
                        queryset=TeamMember.objects.select_related('team')
                    ),
                    Prefetch(
                        'creator__user_details__team_memberships',
                        queryset=TeamMember.objects.select_related('team')
                    )
                ).get(id=contact_id)
                
                # Force Django to load the confirmateur relationship
                # Access the confirmateur to ensure it's loaded from DB
                if hasattr(contact, 'confirmateur_id') and contact.confirmateur_id:
                    contact.confirmateur  # This forces Django to load the relationship
            except Contact.DoesNotExist:
                # Contact was deleted or doesn't exist - this shouldn't happen but handle it gracefully
                # Use the contact object we just saved (it should still be valid)
                pass
            except Exception:
                # If reload fails, continue with the contact object we saved
                pass
            
            # Get new value after saving
            serializer = ContactSerializer(contact, context={'request': request})
            new_value_raw = serializer.data
            new_value = clean_contact_data_for_log(new_value_raw, include_created_at=False)
            
            # Compute only changed fields
            changed_fields = compute_changed_fields(old_value, new_value)
            
            # Only create log entry if there are actual changes
            if changed_fields:
                # Store only the changed fields in old_value and new_value
                # old_value will contain the old values of changed fields
                # new_value will contain the new values of changed fields
                old_value_changes = {field: change['old'] for field, change in changed_fields.items()}
                new_value_changes = {field: change['new'] for field, change in changed_fields.items()}
                
                create_log_entry(
                    event_type='editContact',
                    user_id=request.user if request.user.is_authenticated else None,
                    request=request,
                    old_value=old_value_changes,
                    new_value=new_value_changes,
                    contact_id=contact
                )
            
            return Response({'contact': serializer.data})
        except Exception as e:
            # Catch any other unexpected errors during PATCH processing
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in PATCH contact_detail: {error_details}")
            return Response(
                {'error': f'Erreur lors de la mise à jour: {str(e)}', 'details': error_details},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def contact_delete(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    contact.delete()
    return Response({'message': 'Contact supprimé avec succès'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_current_user(request):
    django_user = request.user
    try:
        # Optimize query with prefetch_related to load permissions efficiently
        # This loads all permissions in one query instead of N+1 queries
        # Check if user is deleted
        user_details = UserDetails.objects.select_related(
            'role_id'
        ).prefetch_related(
            'team_memberships__team',  # For teamId
            'role_id__permission_roles__permission__status'
        ).get(django_user=django_user, deleted_at__isnull=True)
        # Use UserDetailsSerializer to ensure consistent format with other endpoints
        serializer = UserDetailsSerializer(user_details)
        return Response(serializer.data)
    except UserDetails.DoesNotExist:
        # Check if user was deleted (exists but deleted_at is not null)
        try:
            deleted_user = UserDetails.objects.get(django_user=django_user)
            if deleted_user.deleted_at:
                # User is deleted, return 403 Forbidden
                return Response(
                    {'detail': 'This account has been deleted.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except UserDetails.DoesNotExist:
            pass
        
        # If custom user doesn't exist, return Django user data with default role
        # Still include first_name and last_name from Django Auth
        return Response({
            'id': str(django_user.id),
            'djangoUserId': django_user.id,  # Add djangoUserId for consistency
            'username': django_user.username,
            'email': django_user.email or '',
            'firstName': django_user.first_name or '',
            'lastName': django_user.last_name or '',
            'role': None,
            'roleName': None,
            'phone': '',
            'active': True,
            'permissions': [],  # No permissions if no role
        })

# Teams endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def team_list(request):
    teams = Team.objects.all()
    serializer = TeamSerializer(teams, many=True)
    return Response({'teams': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def team_create(request):
    serializer = TeamSerializer(data=request.data)
    if serializer.is_valid():
        # Generate team ID
        team_id = uuid.uuid4().hex[:12]
        while Team.objects.filter(id=team_id).exists():
            team_id = uuid.uuid4().hex[:12]
        team = serializer.save(id=team_id, created_by=request.user if request.user.is_authenticated else None)
        
        # Create log entry
        new_value = get_team_data_for_log(team)
        create_log_entry(
            event_type='createTeam',
            user_id=request.user,
            request=request,
            old_value={},  # No old value for creation
            new_value=new_value
        )
        
        return Response(TeamSerializer(team).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def team_delete(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    
    # Get old value before deletion for logging
    old_value = get_team_data_for_log(team)
    
    # Delete the team
    team.delete()
    
    # Create log entry
    create_log_entry(
        event_type='deleteTeam',
        user_id=request.user,
        request=request,
        old_value=old_value,
        new_value={}  # No new value for deletion
    )
    
    return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    
    if request.method == 'PATCH':
        # Get old value before update for logging
        old_value = get_team_data_for_log(team)
        
        # Update team name
        if 'name' in request.data:
            team.name = request.data['name']
            team.save()
        
        # Refresh team to get updated timestamp
        team.refresh_from_db()
        
        # Get new value after update for logging
        new_value = get_team_data_for_log(team)
        
        # Create log entry
        create_log_entry(
            event_type='editTeam',
            user_id=request.user,
            request=request,
            old_value=old_value,
            new_value=new_value
        )
        
        serializer = TeamSerializer(team)
        return Response(serializer.data)
    
    serializer = TeamDetailSerializer({
        'team': team,
        'team_members': team.team_members.all()
    })
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_list(request):
    # Optimize queries with select_related and prefetch_related to avoid N+1 queries
    # Filter out deleted users (only show users where deleted_at is null)
    users = UserDetails.objects.select_related(
        'django_user',  # For firstName, lastName, username, email
        'role_id'  # For role data (model field is role_id)
    ).prefetch_related(
        'team_memberships__team',  # For teamId
        'role_id__permission_roles__permission__status'  # For permissions
    ).filter(deleted_at__isnull=True)
    serializer = UserDetailsSerializer(users, many=True)
    return Response({'users': serializer.data})



@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def user_delete(request, user_id):
    user_details = get_object_or_404(UserDetails, id=user_id, deleted_at__isnull=True)
    
    # Get old value before deletion for logging
    old_value = {}
    if user_details.django_user:
        old_value = get_user_data_for_log(user_details.django_user, user_details)
    
    # Soft delete: set deleted_at instead of actually deleting
    user_details.deleted_at = timezone.now()
    user_details.save()
    
    # Create log entry
    create_log_entry(
        event_type='deleteUser',
        user_id=request.user,
        request=request,
        old_value=old_value,
        new_value={}  # No new value for deletion
    )
    
    return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_toggle_active(request, user_id):
    user_details = get_object_or_404(UserDetails, id=user_id, deleted_at__isnull=True)
    # Toggle the active status
    user_details.active = not user_details.active
    user_details.save()
    return Response({'active': user_details.active})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_toggle_otp(request, user_id):
    user_details = get_object_or_404(UserDetails, id=user_id, deleted_at__isnull=True)
    # Toggle the require_otp status
    user_details.require_otp = not user_details.require_otp
    user_details.save()
    return Response({'require_otp': user_details.require_otp})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_reset_password(request, user_id):
    """Reset password for a user"""
    try:
        user_details = get_object_or_404(UserDetails, id=user_id, deleted_at__isnull=True)
        django_user = user_details.django_user
        
        if not django_user:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Debug: Log received data
        print(f"Reset password request for user {user_id}")
        print(f"Request data: {request.data}")
        print(f"Request data type: {type(request.data)}")
        
        # Get new password from request, or use default
        new_password = request.data.get('password')
        
        # If password is not provided, use default
        if not new_password:
            new_password = 'Access@123'
        
        # Validate password length
        if len(new_password) < 6:
            return Response({'error': 'Password must be at least 6 characters long'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get old value for logging (without password for security)
        old_value = get_user_data_for_log(django_user, user_details)
        
        # Reset password using Django's set_password which properly hashes it
        django_user.set_password(new_password)
        django_user.save()
        
        # Get new value for logging (password is not included in user data)
        new_value = get_user_data_for_log(django_user, user_details)
        # Add indicator that password was reset
        new_value['password_reset'] = True
        
        # Create log entry
        create_log_entry(
            event_type='resetPassword',
            user_id=request.user,
            request=request,
            old_value=old_value,
            new_value=new_value
        )
        
        return Response({'message': 'Password reset successfully'}, status=status.HTTP_200_OK)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error resetting password: {error_details}")
        return Response({'error': f'Failed to reset password: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def user_update(request, user_id):
    user_details = get_object_or_404(UserDetails, id=user_id, deleted_at__isnull=True)
    django_user = user_details.django_user
    
    if not django_user:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Get old value before update for logging
    old_value = get_user_data_for_log(django_user, user_details)
    
    # Update Django User fields
    if 'first_name' in request.data:
        django_user.first_name = request.data['first_name']
    if 'last_name' in request.data:
        django_user.last_name = request.data['last_name']
    if 'username' in request.data:
        django_user.username = request.data['username']
    if 'email' in request.data:
        django_user.email = request.data['email']
        # When email is updated, also update username to match the email
        django_user.username = request.data['email']
    django_user.save()
    
    # Update UserDetails fields
    if 'roleId' in request.data or 'role' in request.data:
        role_id = request.data.get('roleId') or request.data.get('role')
        if role_id:
            try:
                role = Role.objects.get(id=role_id)
                user_details.role = role
            except Role.DoesNotExist:
                return Response({'error': 'Role not found'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            user_details.role = None
    if 'phone' in request.data:
        phone_value = request.data.get('phone', '') or ''
        if phone_value:
            try:
                # Remove spaces and convert to int
                cleaned = ''.join(str(phone_value).split())
                user_details.phone = int(cleaned) if cleaned else None
            except (ValueError, TypeError):
                user_details.phone = None
        else:
            user_details.phone = None
    if 'hrex' in request.data:
        user_details.hrex = request.data.get('hrex', '').strip() or ''
    
    # Update IP whitelist first, then enable/disable flag (order matters)
    if 'ipWhitelist' in request.data:
        ip_whitelist = request.data.get('ipWhitelist', [])
        # Ensure it's a list and filter out empty strings
        if isinstance(ip_whitelist, list):
            cleaned_whitelist = [ip.strip() for ip in ip_whitelist if ip and ip.strip()]
            user_details.ip_whitelist = cleaned_whitelist
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Updated IP whitelist for user {django_user.email}: {cleaned_whitelist}")
        else:
            user_details.ip_whitelist = []
    
    if 'ipWhitelistEnabled' in request.data:
        ip_whitelist_enabled = bool(request.data.get('ipWhitelistEnabled', False))
        user_details.ip_whitelist_enabled = ip_whitelist_enabled
        # If enabling but no IPs provided, log a warning
        if ip_whitelist_enabled and (not user_details.ip_whitelist or len(user_details.ip_whitelist) == 0):
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"IP whitelist enabled for user {django_user.email} but no IPs in whitelist - access will be denied")
    
    # Update team membership using TeamMember table
    if 'teamId' in request.data:
        team_id = request.data['teamId']
        # Remove user from all teams first (using TeamMember relationship)
        TeamMember.objects.filter(user=user_details).delete()
        
        # If a team is specified, create a new TeamMember relationship
        if team_id:
            try:
                team = Team.objects.get(id=team_id)
                # Generate TeamMember ID
                team_member_id = uuid.uuid4().hex[:12]
                while TeamMember.objects.filter(id=team_member_id).exists():
                    team_member_id = uuid.uuid4().hex[:12]
                TeamMember.objects.create(
                    id=team_member_id,
                    user=user_details,
                    team=team
                )
            except Team.DoesNotExist:
                return Response({'error': 'Team not found'}, status=status.HTTP_400_BAD_REQUEST)
        # If team_id is None or empty, user is removed from all teams (already done above)
    
    user_details.save()
    
    # Refresh user_details to get updated team membership and verify IP whitelist was saved
    user_details.refresh_from_db()
    
    # Log IP whitelist settings after save for debugging
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"User {django_user.email} updated - IP whitelist enabled: {user_details.ip_whitelist_enabled}, IP whitelist: {user_details.ip_whitelist}")
    
    # Get new value after update for logging
    new_value = get_user_data_for_log(django_user, user_details)
    
    # Create log entry
    create_log_entry(
        event_type='editUser',
        user_id=request.user,
        request=request,
        old_value=old_value,
        new_value=new_value
    )
    
    # Return updated user data
    serializer = UserDetailsSerializer(user_details)
    return Response(serializer.data)

# Events endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def event_list(request):
    user = request.user
    
    # Allow filtering by contactId if provided as query parameter
    contact_id = request.query_params.get('contactId', None)
    
    # Allow filtering by future/past events
    future_only = request.query_params.get('future_only', 'false').lower() == 'true'
    past_only = request.query_params.get('past_only', 'false').lower() == 'true'
    
    # Allow bypassing permission filtering for admin views (all_events=true)
    all_events = request.query_params.get('all_events', 'false').lower() == 'true'
    
    # Check if pagination is requested (for contactId filtering)
    requested_page = request.query_params.get('page')
    requested_page_size = request.query_params.get('page_size')
    
    if contact_id:
        # Return events for this contact with pagination support
        events = Event.objects.filter(contactId=contact_id).select_related('userId', 'contactId', 'contactId__status').order_by('-datetime')  # Most recent first
        
        # Apply pagination if requested
        if requested_page or requested_page_size:
            from rest_framework.pagination import PageNumberPagination
            
            page_size = int(requested_page_size) if requested_page_size else 20  # Default 20 events per page
            page = int(requested_page) if requested_page else 1
            
            # Ensure reasonable page size (max 100 per page)
            if page_size > 100:
                page_size = 100
            if page_size < 1:
                page_size = 20
            
            # Create pagination class with proper page_size
            # Use a closure to capture page_size value
            def create_event_pagination(page_size_val):
                class EventPagination(PageNumberPagination):
                    page_size = page_size_val
                    page_size_query_param = 'page_size'
                    max_page_size = 100
                return EventPagination
            
            EventPagination = create_event_pagination(page_size)
            paginator = EventPagination()
            paginated_events = paginator.paginate_queryset(events, request)
            
            serializer = EventSerializer(paginated_events, many=True)
            
            # Return paginated response in consistent format
            return Response({
                'events': serializer.data,
                'total': paginator.page.paginator.count,
                'next': paginator.get_next_link(),
                'previous': paginator.get_previous_link(),
                'page': page,
                'page_size': page_size,
                'has_next': paginator.page.has_next(),
                'has_previous': paginator.page.has_previous()
            })
        else:
            # No pagination requested, return all events (backward compatibility)
            serializer = EventSerializer(events, many=True)
            return Response({'events': serializer.data})
    else:
        # If all_events=true, bypass permission filtering and return all events (admin view)
        if all_events:
            events = Event.objects.all().select_related('userId', 'contactId', 'contactId__status')
        else:
            # Filter events based on user's role data_access level
            # Events are filtered based on the contacts the user can access
            try:
                user_details = UserDetails.objects.get(django_user=user)
                if user_details.role:
                    data_access = user_details.role.data_access
                    
                    if data_access == 'all':
                        # User has access to all contacts, so show all events (including events without contacts)
                        events = Event.objects.all().select_related('userId', 'contactId', 'contactId__status')
                    elif data_access == 'team_only':
                        # Get user's team members
                        team_member = user_details.team_memberships.first()
                        if team_member:
                            team = team_member.team
                            # Get all users in the same team
                            team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                            # Get contacts accessible to the user or their team
                            accessible_contact_ids = Contact.objects.filter(
                                models.Q(teleoperator=user) |
                                models.Q(confirmateur=user) |
                                models.Q(creator=user) |
                                models.Q(teleoperator__id__in=team_user_ids) |
                                models.Q(confirmateur__id__in=team_user_ids) |
                                models.Q(creator__id__in=team_user_ids)
                            ).values_list('id', flat=True)
                            # Return events for accessible contacts OR events created by team members (even without contactId)
                            events = Event.objects.filter(
                                models.Q(contactId__id__in=accessible_contact_ids) |
                                models.Q(contactId__isnull=True, userId__id__in=team_user_ids)
                            ).select_related('userId', 'contactId', 'contactId__status')
                        else:
                            # User has no team, fall back to own_only behavior
                            is_teleoperateur = user_details.role.is_teleoperateur
                            is_confirmateur = user_details.role.is_confirmateur
                            
                            if is_teleoperateur and is_confirmateur:
                                accessible_contact_ids = Contact.objects.filter(
                                    models.Q(teleoperator=user) |
                                    models.Q(confirmateur=user)
                                ).values_list('id', flat=True)
                            elif is_teleoperateur:
                                accessible_contact_ids = Contact.objects.filter(teleoperator=user).values_list('id', flat=True)
                            elif is_confirmateur:
                                accessible_contact_ids = Contact.objects.filter(confirmateur=user).values_list('id', flat=True)
                            else:
                                accessible_contact_ids = Contact.objects.filter(
                                    models.Q(teleoperator=user) |
                                    models.Q(confirmateur=user) |
                                    models.Q(creator=user)
                                ).values_list('id', flat=True)
                            # Return events for accessible contacts OR events created by user (even without contactId)
                            events = Event.objects.filter(
                                models.Q(contactId__id__in=accessible_contact_ids) |
                                models.Q(contactId__isnull=True, userId=user)
                            ).select_related('userId', 'contactId', 'contactId__status')
                    else:  # own_only
                        # Show events where BOTH conditions are true:
                        # 1. User is assigned to the event (userId = user)
                        # 2. AND the contact is assigned to the user (teleoperateur or confirmateur)
                        accessible_contact_ids = Contact.objects.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user)
                        ).values_list('id', flat=True)
                        
                        events = Event.objects.filter(
                            userId=user,
                            contactId__id__in=accessible_contact_ids
                        ).select_related('userId', 'contactId')
                else:
                    # User has no role, show no events (safety default)
                    events = Event.objects.none()
            except UserDetails.DoesNotExist:
                # If user has no UserDetails, show no events (safety default)
                events = Event.objects.none()
    
    # Filter by future/past if requested
    now = timezone.now()
    if future_only:
        events = events.filter(datetime__gt=now)
    elif past_only:
        events = events.filter(datetime__lte=now)
    
    # Order events: future events first (ascending), then past events (descending)
    # This ensures future events appear at the top when paginated
    # Past events should show most recent first
    # Use raw SQL to properly order: future events ASC, past events DESC
    # For past events, we'll use a large timestamp minus the difference to invert order
    from django.db import connection
    if connection.vendor == 'postgresql':
        # PostgreSQL: Use a computed sort_datetime
        # For future events: use datetime as-is (sorts ASC - earliest first)
        # For past events: use negative datetime to invert order (most recent = less negative = sorts first)
        # Convert to epoch seconds for consistent numeric sorting
        events = events.extra(
            select={
                'is_future': "CASE WHEN datetime > %s THEN 0 ELSE 1 END",
                'sort_datetime': "CASE WHEN datetime > %s THEN EXTRACT(EPOCH FROM datetime) ELSE -EXTRACT(EPOCH FROM datetime) END"
            },
            select_params=[now, now],
            order_by=['is_future', 'sort_datetime']
        )
    else:
        # Fallback for other databases: just order future first, then datetime descending
        # This means past events will be most recent first
        events = events.annotate(
            is_future=Case(
                When(datetime__gt=now, then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('is_future', Case(
            When(datetime__gt=now, then=F('datetime')),  # Future: ascending
            default=F('datetime'),  # Past: ascending (not ideal, but frontend will fix)
            output_field=models.DateTimeField()
        ))
    
    # Apply pagination if requested (for all events or filtered events)
    if requested_page or requested_page_size:
        from rest_framework.pagination import PageNumberPagination
        
        page_size = int(requested_page_size) if requested_page_size else 10  # Default 10 events per page
        page = int(requested_page) if requested_page else 1
        
        # Check if user has 'all' data access - allow larger page sizes
        max_page_size = 100
        try:
            user_details = UserDetails.objects.get(django_user=user)
            if user_details.role and user_details.role.data_access == 'all':
                max_page_size = 1000  # Allow up to 1000 events per page for admins
        except UserDetails.DoesNotExist:
            pass
        
        # Ensure reasonable page size
        if page_size > max_page_size:
            page_size = max_page_size
        if page_size < 1:
            page_size = 10
        
        # Create pagination class with proper page_size
        def create_event_pagination(page_size_val, max_page_size_val):
            class EventPagination(PageNumberPagination):
                page_size = page_size_val
                page_size_query_param = 'page_size'
                max_page_size = max_page_size_val
            return EventPagination
        
        EventPagination = create_event_pagination(page_size, max_page_size)
        paginator = EventPagination()
        paginated_events = paginator.paginate_queryset(events, request)
        
        serializer = EventSerializer(paginated_events, many=True)
        
        # Return paginated response in consistent format
        return Response({
            'events': serializer.data,
            'total': paginator.page.paginator.count,
            'next': paginator.get_next_link(),
            'previous': paginator.get_previous_link(),
            'page': page,
            'page_size': page_size,
            'has_next': paginator.page.has_next(),
            'has_previous': paginator.page.has_previous()
        })
    else:
        # No pagination requested, return all events (backward compatibility)
        serializer = EventSerializer(events, many=True)
        return Response({'events': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def event_create(request):
    serializer = EventSerializer(data=request.data)
    if serializer.is_valid():
        # Get contact if contactId provided
        contact = None
        contact_id = request.data.get('contactId')
        if contact_id:
            try:
                contact = Contact.objects.get(id=contact_id)
            except Contact.DoesNotExist:
                pass
        
        # Check for duplicate event: same date, same hour, same contact
        # Use validated_data to get properly parsed datetime
        validated_datetime = serializer.validated_data.get('datetime')
        if validated_datetime and contact:
            try:
                # Ensure datetime is timezone-aware
                if timezone.is_naive(validated_datetime):
                    validated_datetime = timezone.make_aware(validated_datetime, timezone.utc)
                
                # Extract date and hour
                event_date = validated_datetime.date()
                event_hour = validated_datetime.hour
                
                # Check if an event already exists with the same contact, date, and hour
                existing_event = Event.objects.filter(
                    contactId=contact,
                    datetime__date=event_date,
                    datetime__hour=event_hour
                ).first()
                
                if existing_event:
                    # Return existing event instead of creating a duplicate
                    return Response(EventSerializer(existing_event).data, status=status.HTTP_200_OK)
            except (ValueError, AttributeError, TypeError) as e:
                # If datetime parsing fails, continue with normal creation
                pass
        
        # Generate event ID
        event_id = uuid.uuid4().hex[:12]
        while Event.objects.filter(id=event_id).exists():
            event_id = uuid.uuid4().hex[:12]
        
        # Get user if userId provided, otherwise use current user
        # userId can be either UserDetails ID or DjangoUser ID (for backward compatibility)
        user = request.user
        user_id = request.data.get('userId')
        if user_id:
            try:
                # First try to get UserDetails by ID
                user_details = UserDetails.objects.get(id=user_id)
                user = user_details.django_user
            except UserDetails.DoesNotExist:
                try:
                    # Fallback: try to get DjangoUser by ID (for backward compatibility)
                    user = DjangoUser.objects.get(id=user_id)
                except DjangoUser.DoesNotExist:
                    pass  # Use current user as fallback
        
        event = serializer.save(
            id=event_id,
            userId=user,
            contactId=contact,
            created_by=request.user if request.user.is_authenticated else None
        )
        
        # Refresh from database to ensure we have the latest data
        event.refresh_from_db()
        
        # Send notification to assigned user
        # Commented out: "nouvel evenement assigne" notification
        # if event.userId:
        #     send_event_notification(event, notification_type='assigned')
        
        # Create log entry for event creation
        # Use event.contactId after saving to ensure we have the correct contact reference
        print(f"[EVENT LOG] Event created: id={event.id}, contactId={event.contactId.id if event.contactId else None}")
        if event.contactId:
            try:
                event_data = {
                    'eventId': event.id,
                    'datetime': event.datetime.isoformat() if event.datetime else None,
                    'comment': event.comment or '',
                    'userId': user.id if user else None,
                    'userName': f"{user.first_name} {user.last_name}".strip() if user and (user.first_name or user.last_name) else (user.username if user else None),
                    'createdAt': event.created_at.isoformat() if event.created_at else None,
                    'updatedAt': event.updated_at.isoformat() if event.updated_at else None,
                }
                print(f"[EVENT LOG] Creating log entry for createEvent, contact_id={event.contactId.id}")
                create_log_entry(
                    event_type='createEvent',
                    user_id=request.user if request.user.is_authenticated else None,
                    request=request,
                    old_value={},
                    new_value=event_data,
                    contact_id=event.contactId
                )
                print(f"[EVENT LOG] Log entry created successfully")
            except Exception as e:
                import traceback
                print(f"[EVENT LOG] Error creating log entry: {str(e)}")
                print(f"[EVENT LOG] Traceback: {traceback.format_exc()}")
        else:
            print(f"[EVENT LOG] No contactId for event {event.id}, skipping log creation")
        
        return Response(EventSerializer(event).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def event_update(request, event_id):
    # Allow updating if user owns the event OR if event is linked to a contact (for contact management)
    try:
        event = Event.objects.get(id=event_id)
        # Check if user owns the event OR if event has a contact (allows contact-related edits)
        if event.userId != request.user and not event.contactId:
            return Response({'detail': 'You do not have permission to update this event.'}, status=status.HTTP_403_FORBIDDEN)
    except Event.DoesNotExist:
        return Response({'detail': 'Event not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    # Store old values before update
    # Get UserDetails ID for userId field
    user_details_id = None
    if event.userId:
        try:
            user_details = event.userId.user_details
            user_details_id = user_details.id if user_details else None
        except UserDetails.DoesNotExist:
            pass
    
    old_event_data = {
        'eventId': event.id,
        'datetime': event.datetime.isoformat() if event.datetime else None,
        'comment': event.comment or '',
        'userId': user_details_id,  # Use UserDetails ID instead of DjangoUser ID
        'userName': f"{event.userId.first_name} {event.userId.last_name}".strip() if event.userId and (event.userId.first_name or event.userId.last_name) else (event.userId.username if event.userId else None),
        'createdAt': event.created_at.isoformat() if event.created_at else None,
        'updatedAt': event.updated_at.isoformat() if event.updated_at else None,
    }
    contact_before_update = event.contactId
    
    serializer = EventSerializer(event, data=request.data, partial=True)
    if serializer.is_valid():
        # Get contact if contactId provided
        contact = None
        contact_id = request.data.get('contactId')
        if contact_id:
            try:
                contact = Contact.objects.get(id=contact_id)
            except Contact.DoesNotExist:
                pass
        elif contact_id == '' or contact_id is None:
            contact = None
        
        # Get user if userId provided, otherwise keep existing user
        # userId can be either UserDetails ID or DjangoUser ID (for backward compatibility)
        user = event.userId
        user_id = request.data.get('userId')
        if user_id:
            try:
                # First try to get UserDetails by ID
                user_details = UserDetails.objects.get(id=user_id)
                user = user_details.django_user
            except UserDetails.DoesNotExist:
                try:
                    # Fallback: try to get DjangoUser by ID (for backward compatibility)
                    user = DjangoUser.objects.get(id=user_id)
                except DjangoUser.DoesNotExist:
                    pass  # Keep existing user as fallback
        
        # Check if user assignment changed (using UserDetails ID)
        old_user_details_id = None
        if event.userId:
            try:
                old_user_details = event.userId.user_details
                old_user_details_id = old_user_details.id if old_user_details else None
            except UserDetails.DoesNotExist:
                pass
        
        new_user_details_id = None
        if user:
            try:
                new_user_details = user.user_details
                new_user_details_id = new_user_details.id if new_user_details else None
            except UserDetails.DoesNotExist:
                pass
        
        # Update event with new data
        event = serializer.save(contactId=contact, userId=user)
        
        # Refresh from database to get updated timestamps
        event.refresh_from_db()
        
        # Send notification if user was assigned or changed
        # Commented out: "nouvel evenement assigne" notification
        # if event.userId and (old_user_details_id != new_user_details_id or old_user_details_id is None):
        #     send_event_notification(event, notification_type='assigned')
        
        # Create log entry for event update
        contact_for_log = contact or contact_before_update
        print(f"[EVENT LOG] Event updated: id={event.id}, contact_for_log={contact_for_log.id if contact_for_log else None}")
        if contact_for_log:
            try:
                new_event_data = {
                    'eventId': event.id,
                    'datetime': event.datetime.isoformat() if event.datetime else None,
                    'comment': event.comment or '',
                    'userId': user.id if user else None,
                    'userName': f"{user.first_name} {user.last_name}".strip() if user and (user.first_name or user.last_name) else (user.username if user else None),
                    'createdAt': event.created_at.isoformat() if event.created_at else None,
                    'updatedAt': event.updated_at.isoformat() if event.updated_at else None,
                }
                print(f"[EVENT LOG] Creating log entry for editEvent, contact_id={contact_for_log.id}")
                create_log_entry(
                    event_type='editEvent',
                    user_id=request.user if request.user.is_authenticated else None,
                    request=request,
                    old_value=old_event_data,
                    new_value=new_event_data,
                    contact_id=contact_for_log
                )
                print(f"[EVENT LOG] Log entry created successfully")
            except Exception as e:
                import traceback
                print(f"[EVENT LOG] Error creating log entry: {str(e)}")
                print(f"[EVENT LOG] Traceback: {traceback.format_exc()}")
        else:
            print(f"[EVENT LOG] No contact_for_log for event {event.id}, skipping log creation")
        
        return Response(EventSerializer(event).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def event_delete(request, event_id):
    # Allow deleting if user owns the event OR if event is linked to a contact (for contact management)
    try:
        event = Event.objects.get(id=event_id)
        # Check if user owns the event OR if event has a contact (allows contact-related deletes)
        if event.userId != request.user and not event.contactId:
            return Response({'detail': 'You do not have permission to delete this event.'}, status=status.HTTP_403_FORBIDDEN)
    except Event.DoesNotExist:
        return Response({'detail': 'Event not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    # Store event data before deletion
    contact = event.contactId
    event_data = {
        'eventId': event.id,
        'datetime': event.datetime.isoformat() if event.datetime else None,
        'comment': event.comment or '',
        'userId': event.userId.id if event.userId else None,
        'userName': f"{event.userId.first_name} {event.userId.last_name}".strip() if event.userId and (event.userId.first_name or event.userId.last_name) else (event.userId.username if event.userId else None),
        'createdAt': event.created_at.isoformat() if event.created_at else None,
        'updatedAt': event.updated_at.isoformat() if event.updated_at else None,
    }
    
    print(f"[EVENT LOG] Event deleted: id={event.id}, contact={contact.id if contact else None}")
    
    event.delete()
    
    # Create log entry for event deletion
    if contact:
        try:
            print(f"[EVENT LOG] Creating log entry for deleteEvent, contact_id={contact.id}")
            create_log_entry(
                event_type='deleteEvent',
                user_id=request.user if request.user.is_authenticated else None,
                request=request,
                old_value=event_data,
                new_value={},
                contact_id=contact
            )
            print(f"[EVENT LOG] Log entry created successfully")
        except Exception as e:
            import traceback
            print(f"[EVENT LOG] Error creating log entry: {str(e)}")
            print(f"[EVENT LOG] Traceback: {traceback.format_exc()}")
    else:
        print(f"[EVENT LOG] No contact for event {event.id}, skipping log creation")
    
    return Response(status=status.HTTP_204_NO_CONTENT)

# Transactions endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transaction_list(request):
    user = request.user
    
    # Allow filtering by contactId if provided as query parameter
    contact_id = request.query_params.get('contactId', None)
    
    # Check if pagination is requested
    requested_page = request.query_params.get('page')
    requested_page_size = request.query_params.get('page_size')
    
    if contact_id:
        # Check if user has permission to access this contact
        try:
            contact = Contact.objects.get(id=contact_id)
        except Contact.DoesNotExist:
            return Response(
                {'error': 'Contact non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Verify user has access to this contact based on data_access permissions
        try:
            user_details = UserDetails.objects.get(django_user=user)
            if user_details.role:
                data_access = user_details.role.data_access
                
                has_access = False
                
                if data_access == 'all':
                    # User has access to all contacts
                    has_access = True
                elif data_access == 'team_only':
                    # Check if user has access (either assigned to them or from their team)
                    team_member = user_details.team_memberships.first()
                    if team_member:
                        team = team_member.team
                        team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                        # Allow if user is assigned OR if contact's assignees are in the same team
                        if (contact.teleoperator == user or contact.confirmateur == user or contact.creator == user or
                            (contact.teleoperator and contact.teleoperator.id in team_user_ids) or
                            (contact.confirmateur and contact.confirmateur.id in team_user_ids) or
                            (contact.creator and contact.creator.id in team_user_ids)):
                            has_access = True
                    else:
                        # User has no team, fall back to own_only behavior
                        if contact.teleoperator == user or contact.confirmateur == user or contact.creator == user:
                            has_access = True
                else:  # own_only
                    # Check if user is assigned to this contact
                    is_teleoperateur = user_details.role.is_teleoperateur
                    is_confirmateur = user_details.role.is_confirmateur
                    
                    if is_teleoperateur and is_confirmateur:
                        # User is both: allow if user is teleoperator OR confirmateur
                        if contact.teleoperator == user or contact.confirmateur == user:
                            has_access = True
                    elif is_teleoperateur:
                        # User is only teleoperateur: allow only if user is teleoperator
                        if contact.teleoperator == user:
                            has_access = True
                    elif is_confirmateur:
                        # User is only confirmateur: allow only if user is confirmateur
                        if contact.confirmateur == user:
                            has_access = True
                    else:
                        # Default behavior: allow if user is teleoperator, confirmateur, or creator
                        if contact.teleoperator == user or contact.confirmateur == user or contact.creator == user:
                            has_access = True
                
                if not has_access:
                    return Response(
                        {'error': 'Vous n\'avez pas accès à ce contact'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                # User has no role, deny access (safety default)
                return Response(
                    {'error': 'Vous n\'avez pas accès à ce contact'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except UserDetails.DoesNotExist:
            # If user has no UserDetails, deny access (safety default)
            return Response(
                {'error': 'Vous n\'avez pas accès à ce contact'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # User has permission, return transactions for this contact with pagination support
        # Note: Using contact_id=contact_id (not contact=contact_id) to filter by ForeignKey ID value directly
        # This is the correct Django ORM syntax when filtering by ID (string) rather than model instance
        transactions = Transaction.objects.filter(contact_id=contact_id).select_related('contact', 'contact__teleoperator', 'contact__confirmateur', 'created_by').order_by('-date', '-created_at')
        
        # Apply pagination if requested
        if requested_page or requested_page_size:
            from rest_framework.pagination import PageNumberPagination
            
            page_size = int(requested_page_size) if requested_page_size else 20
            page = int(requested_page) if requested_page else 1
            
            if page_size > 100:
                page_size = 100
            if page_size < 1:
                page_size = 20
            
            def create_transaction_pagination(page_size_val):
                class TransactionPagination(PageNumberPagination):
                    page_size = page_size_val
                    page_size_query_param = 'page_size'
                    max_page_size = 100
                return TransactionPagination
            
            TransactionPagination = create_transaction_pagination(page_size)
            paginator = TransactionPagination()
            paginated_transactions = paginator.paginate_queryset(transactions, request)
            
            serializer = TransactionSerializer(paginated_transactions, many=True)
            
            return Response({
                'transactions': serializer.data,
                'total': paginator.page.paginator.count,
                'next': paginator.get_next_link(),
                'previous': paginator.get_previous_link(),
                'page': page,
                'page_size': page_size,
                'has_next': paginator.page.has_next(),
                'has_previous': paginator.page.has_previous()
            })
        else:
            serializer = TransactionSerializer(transactions, many=True)
            return Response({'transactions': serializer.data})
    else:
        # Filter transactions based on user's role data_access level
        # Transactions are filtered based on the contacts the user can access
        try:
            user_details = UserDetails.objects.get(django_user=user)
            if user_details.role:
                data_access = user_details.role.data_access
                
                if data_access == 'all':
                    # User has access to all contacts, so show all transactions
                    transactions = Transaction.objects.all().select_related('contact', 'contact__teleoperator', 'contact__confirmateur', 'created_by')
                elif data_access == 'team_only':
                    # Get user's team members
                    team_member = user_details.team_memberships.first()
                    if team_member:
                        team = team_member.team
                        # Get all users in the same team
                        team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                        # Get contacts accessible to the user or their team
                        accessible_contact_ids = Contact.objects.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user) |
                            models.Q(teleoperator__id__in=team_user_ids) |
                            models.Q(confirmateur__id__in=team_user_ids) |
                            models.Q(creator__id__in=team_user_ids)
                        ).values_list('id', flat=True)
                        # Return transactions for accessible contacts
                        transactions = Transaction.objects.filter(
                            contact__id__in=accessible_contact_ids
                        ).select_related('contact', 'contact__teleoperator', 'contact__confirmateur', 'created_by')
                    else:
                        # User has no team, fall back to own_only behavior
                        is_teleoperateur = user_details.role.is_teleoperateur
                        is_confirmateur = user_details.role.is_confirmateur
                        
                        if is_teleoperateur and is_confirmateur:
                            accessible_contact_ids = Contact.objects.filter(
                                models.Q(teleoperator=user) |
                                models.Q(confirmateur=user)
                            ).values_list('id', flat=True)
                        elif is_teleoperateur:
                            accessible_contact_ids = Contact.objects.filter(teleoperator=user).values_list('id', flat=True)
                        elif is_confirmateur:
                            accessible_contact_ids = Contact.objects.filter(confirmateur=user).values_list('id', flat=True)
                        else:
                            accessible_contact_ids = Contact.objects.filter(
                                models.Q(teleoperator=user) |
                                models.Q(confirmateur=user) |
                                models.Q(creator=user)
                            ).values_list('id', flat=True)
                        transactions = Transaction.objects.filter(
                            contact__id__in=accessible_contact_ids
                        ).select_related('contact', 'contact__teleoperator', 'contact__confirmateur', 'created_by')
                else:  # own_only
                    # Show transactions where the contact is assigned to the user (teleoperateur or confirmateur)
                    accessible_contact_ids = Contact.objects.filter(
                        models.Q(teleoperator=user) |
                        models.Q(confirmateur=user)
                    ).values_list('id', flat=True)
                    
                    transactions = Transaction.objects.filter(
                        contact__id__in=accessible_contact_ids
                    ).select_related('contact', 'contact__teleoperator', 'contact__confirmateur', 'created_by')
            else:
                # User has no role, show no transactions (safety default)
                transactions = Transaction.objects.none()
        except UserDetails.DoesNotExist:
            # If user has no UserDetails, show no transactions (safety default)
            transactions = Transaction.objects.none()
    
    # Order by date descending (most recent first)
    transactions = transactions.order_by('-date', '-created_at')
    
    # Apply pagination if requested
    if requested_page or requested_page_size:
        from rest_framework.pagination import PageNumberPagination
        
        page_size = int(requested_page_size) if requested_page_size else 20
        page = int(requested_page) if requested_page else 1
        
        max_page_size = 100
        try:
            user_details = UserDetails.objects.get(django_user=user)
            if user_details.role and user_details.role.data_access == 'all':
                max_page_size = 1000
        except UserDetails.DoesNotExist:
            pass
        
        if page_size > max_page_size:
            page_size = max_page_size
        if page_size < 1:
            page_size = 20
        
        def create_transaction_pagination(page_size_val, max_size):
            class TransactionPagination(PageNumberPagination):
                page_size = page_size_val
                page_size_query_param = 'page_size'
                max_page_size = max_size
            return TransactionPagination
        
        TransactionPagination = create_transaction_pagination(page_size, max_page_size)
        paginator = TransactionPagination()
        paginated_transactions = paginator.paginate_queryset(transactions, request)
        
        serializer = TransactionSerializer(paginated_transactions, many=True)
        
        return Response({
            'transactions': serializer.data,
            'total': paginator.page.paginator.count,
            'next': paginator.get_next_link(),
            'previous': paginator.get_previous_link(),
            'page': page,
            'page_size': page_size,
            'has_next': paginator.page.has_next(),
            'has_previous': paginator.page.has_previous()
        })
    else:
        serializer = TransactionSerializer(transactions, many=True)
        return Response({'transactions': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transaction_create(request):
    user = request.user
    serializer = TransactionSerializer(data=request.data)
    
    if serializer.is_valid():
        # Generate unique ID if not provided
        transaction_id = serializer.validated_data.get('id')
        if not transaction_id:
            from api.signals import generate_unique_id
            transaction_id = generate_unique_id(Transaction)
            serializer.validated_data['id'] = transaction_id
        
        # Get contact if contactId provided
        contact = None
        contact_id = request.data.get('contactId')
        if contact_id:
            try:
                contact = Contact.objects.get(id=contact_id)
            except Contact.DoesNotExist:
                return Response({'error': 'Contact not found'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not contact:
            return Response({'error': 'contactId is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if user has access to this contact (for own_only users)
        try:
            user_details = UserDetails.objects.get(django_user=user)
            if user_details.role:
                data_access = user_details.role.data_access
                if data_access == 'own_only':
                    # Check if user is teleoperateur or confirmateur of this contact
                    if contact.teleoperator != user and contact.confirmateur != user:
                        return Response(
                            {'error': 'Vous n\'avez pas accès à ce contact'},
                            status=status.HTTP_403_FORBIDDEN
                        )
        except UserDetails.DoesNotExist:
            pass
        
        # Get RIB if ribId provided
        rib = None
        rib_id = request.data.get('ribId')
        if rib_id:
            try:
                rib = RIB.objects.get(id=rib_id)
            except RIB.DoesNotExist:
                pass
        
        # Create transaction
        transaction = serializer.save(contact=contact, rib=rib, created_by=user)
        
        return Response(TransactionSerializer(transaction).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def transaction_update(request, transaction_id):
    try:
        transaction = Transaction.objects.get(id=transaction_id)
    except Transaction.DoesNotExist:
        return Response({'detail': 'Transaction not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    # Check if user has access to this transaction's contact
    user = request.user
    try:
        user_details = UserDetails.objects.get(django_user=user)
        if user_details.role:
            data_access = user_details.role.data_access
            if data_access == 'own_only':
                # Check if user is teleoperateur or confirmateur of this contact
                if transaction.contact.teleoperator != user and transaction.contact.confirmateur != user:
                    return Response(
                        {'error': 'Vous n\'avez pas accès à cette transaction'},
                        status=status.HTTP_403_FORBIDDEN
                    )
    except UserDetails.DoesNotExist:
        pass
    
    serializer = TransactionSerializer(transaction, data=request.data, partial=True)
    if serializer.is_valid():
        # Get contact if contactId provided
        contact = transaction.contact
        contact_id = request.data.get('contactId')
        if contact_id:
            try:
                contact = Contact.objects.get(id=contact_id)
            except Contact.DoesNotExist:
                pass
        
        # Get RIB if ribId provided
        rib = transaction.rib
        rib_id = request.data.get('ribId')
        if rib_id:
            try:
                rib = RIB.objects.get(id=rib_id)
            except RIB.DoesNotExist:
                pass
        elif rib_id == '' or rib_id is None:
            # Explicitly set to None if empty string or None
            rib = None
        
        # Store old transaction type to check if it's an 'Ouverture' transaction
        old_transaction_type = transaction.type
        
        transaction = serializer.save(contact=contact, rib=rib)
        
        # Send notification if transaction type is 'Ouverture' (either was already 'Ouverture' or was updated to 'Ouverture')
        if transaction.type == 'Ouverture':
            send_transaction_update_notification(contact, transaction)
        
        return Response(TransactionSerializer(transaction).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def transaction_delete(request, transaction_id):
    try:
        transaction = Transaction.objects.get(id=transaction_id)
    except Transaction.DoesNotExist:
        return Response({'detail': 'Transaction not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    # Check if user has access to this transaction's contact
    user = request.user
    try:
        user_details = UserDetails.objects.get(django_user=user)
        if user_details.role:
            data_access = user_details.role.data_access
            if data_access == 'own_only':
                # Check if user is teleoperateur or confirmateur of this contact
                if transaction.contact.teleoperator != user and transaction.contact.confirmateur != user:
                    return Response(
                        {'error': 'Vous n\'avez pas accès à cette transaction'},
                        status=status.HTTP_403_FORBIDDEN
                    )
    except UserDetails.DoesNotExist:
        pass
    
    transaction.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# RIB endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rib_list(request):
    """List all RIBs"""
    ribs = RIB.objects.all().order_by('-created_at')
    serializer = RIBSerializer(ribs, many=True)
    return Response({'ribs': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rib_create(request):
    """Create a new RIB"""
    user = request.user
    serializer = RIBSerializer(data=request.data)
    
    if serializer.is_valid():
        # Generate unique ID if not provided
        rib_id = serializer.validated_data.get('id')
        if not rib_id:
            from api.signals import generate_unique_id
            rib_id = generate_unique_id(RIB)
            serializer.validated_data['id'] = rib_id
        
        # Create RIB
        rib = serializer.save(created_by=user)
        
        return Response(RIBSerializer(rib).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def rib_delete(request, rib_id):
    """Delete a RIB"""
    try:
        rib = RIB.objects.get(id=rib_id)
    except RIB.DoesNotExist:
        return Response({'detail': 'RIB not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    rib.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def team_add_member(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    user_id = request.data.get('userId')
    
    if not user_id:
        return Response({'error': 'userId is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user_details = UserDetails.objects.get(id=user_id)
        
        # Check if user is already in the team
        if TeamMember.objects.filter(user=user_details, team=team).exists():
            return Response({'error': 'User is already in this team'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate TeamMember ID
        import uuid
        team_member_id = uuid.uuid4().hex[:12]
        while TeamMember.objects.filter(id=team_member_id).exists():
            team_member_id = uuid.uuid4().hex[:12]
        
        # Create TeamMember relationship
        team_member = TeamMember.objects.create(
            id=team_member_id,
            user=user_details,
            team=team
        )
        
        
        serializer = TeamMemberSerializer(team_member)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    except UserDetails.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def team_remove_member(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    user_id = request.data.get('userId')
    
    if not user_id:
        return Response({'error': 'userId is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user_details = UserDetails.objects.get(id=user_id)
        team_member = TeamMember.objects.get(user=user_details, team=team)
        
        # If user is teamleader, change role to something else (e.g., 'gestionnaire')
        if user_details.role and user_details.role.name.lower() == 'teamleader':
            try:
                gestionnaire_role = Role.objects.get(name__iexact='gestionnaire')
                user_details.role = gestionnaire_role
                user_details.save()
            except Role.DoesNotExist:
                pass  # If gestionnaire role doesn't exist, just remove from team
        
        # Remove TeamMember relationship
        team_member.delete()
        
        
        return Response(status=status.HTTP_204_NO_CONTENT)
    except UserDetails.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
    except TeamMember.DoesNotExist:
        return Response({'error': 'User not found in this team'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def team_set_leader(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    user_id = request.data.get('userId')
    
    if not user_id:
        return Response({'error': 'userId is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Get team members through TeamMember
        team_members = TeamMember.objects.filter(team=team)
        user_ids_in_team = [tm.user.id for tm in team_members]
        
        # Remove leader status from all team members (change role from teamleader to gestionnaire)
        try:
            teamleader_role = Role.objects.get(name__iexact='teamleader')
            gestionnaire_role = Role.objects.get(name__iexact='gestionnaire')
            UserDetails.objects.filter(
                id__in=user_ids_in_team, 
                role=teamleader_role
            ).update(role=gestionnaire_role)
        except Role.DoesNotExist:
            pass  # If roles don't exist, skip
        
        # Set new leader (change role to teamleader)
        user_details = UserDetails.objects.get(id=user_id)
        
        # Verify user is in the team
        if not TeamMember.objects.filter(user=user_details, team=team).exists():
            return Response({'error': 'User not found in this team'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            teamleader_role = Role.objects.get(name__iexact='teamleader')
            user_details.role = teamleader_role
            user_details.save()
        except Role.DoesNotExist:
            return Response({'error': 'Teamleader role not found'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = UserDetailsSerializer(user_details)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except UserDetails.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

# Roles endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def role_list(request):
    """List all roles"""
    roles = Role.objects.all().order_by('name')
    serializer = RoleSerializer(roles, many=True)
    return Response({'roles': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def role_create(request):
    """Create a new role"""
    serializer = RoleSerializer(data=request.data)
    if serializer.is_valid():
        # Generate role ID
        role_id = uuid.uuid4().hex[:12]
        while Role.objects.filter(id=role_id).exists():
            role_id = uuid.uuid4().hex[:12]
        role = serializer.save(id=role_id, created_by=request.user if request.user.is_authenticated else None)
        return Response(RoleSerializer(role).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def role_update(request, role_id):
    """Update a role"""
    role = get_object_or_404(Role, id=role_id)
    serializer = RoleSerializer(role, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(RoleSerializer(role).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def role_delete(request, role_id):
    """Delete a role"""
    role = get_object_or_404(Role, id=role_id)
    role.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# Permissions endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def permission_list(request):
    """List all permissions"""
    permissions = Permission.objects.all().order_by('component', 'field_name')
    serializer = PermissionSerializer(permissions, many=True)
    return Response({'permissions': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def permission_create(request):
    """Create a new permission"""
    serializer = PermissionSerializer(data=request.data)
    if serializer.is_valid():
        # Generate permission ID
        permission_id = uuid.uuid4().hex[:12]
        while Permission.objects.filter(id=permission_id).exists():
            permission_id = uuid.uuid4().hex[:12]
        
        # Ensure action is provided, default to 'view'
        action = request.data.get('action', 'view')
        if action not in ['view', 'create', 'edit', 'delete']:
            action = 'view'
        
        # The serializer will handle statusId conversion to status
        permission = serializer.save(id=permission_id, action=action)
        return Response(PermissionSerializer(permission).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def permission_update(request, permission_id):
    """Update a permission"""
    permission = get_object_or_404(Permission, id=permission_id)
    
    # The serializer will handle statusId conversion to status
    serializer = PermissionSerializer(permission, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(PermissionSerializer(permission).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def permission_delete(request, permission_id):
    """Delete a permission"""
    permission = get_object_or_404(Permission, id=permission_id)
    permission.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# Permission-Role endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def permission_role_list(request):
    """List all permission-role relationships"""
    permission_roles = PermissionRole.objects.all().select_related('role', 'permission')
    serializer = PermissionRoleSerializer(permission_roles, many=True)
    return Response({'permissionRoles': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def permission_role_create(request):
    """Assign a permission to a role"""
    role_id = request.data.get('roleId')
    permission_id = request.data.get('permissionId')
    
    if not role_id or not permission_id:
        return Response({'error': 'roleId and permissionId are required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        role = Role.objects.get(id=role_id)
        permission = Permission.objects.get(id=permission_id)
        
        # Check if relationship already exists
        if PermissionRole.objects.filter(role=role, permission=permission).exists():
            return Response({'error': 'Permission already assigned to this role'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate permission role ID
        permission_role_id = uuid.uuid4().hex[:12]
        while PermissionRole.objects.filter(id=permission_role_id).exists():
            permission_role_id = uuid.uuid4().hex[:12]
        
        permission_role = PermissionRole.objects.create(
            id=permission_role_id,
            role=role,
            permission=permission
        )
        
        serializer = PermissionRoleSerializer(permission_role)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    except Role.DoesNotExist:
        return Response({'error': 'Role not found'}, status=status.HTTP_404_NOT_FOUND)
    except Permission.DoesNotExist:
        return Response({'error': 'Permission not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def permission_role_delete(request, permission_role_id):
    """Remove a permission from a role"""
    permission_role = get_object_or_404(PermissionRole, id=permission_role_id)
    permission_role.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# Statuses endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def status_list(request):
    """List all statuses"""
    status_type = request.query_params.get('type', None)
    statuses = Status.objects.all()
    if status_type:
        statuses = statuses.filter(type=status_type)
    statuses = statuses.order_by('order_index', 'name')
    serializer = StatusSerializer(statuses, many=True)
    return Response({'statuses': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def status_create(request):
    """Create a new status"""
    # Normalize the data - ensure type has a default value if not provided
    data = request.data.copy()
    if 'type' not in data or not data['type']:
        data['type'] = 'lead'
    
    # Ensure color is an empty string if not provided or None
    if 'color' not in data or data['color'] is None:
        data['color'] = ''
    
    serializer = StatusSerializer(data=data)
    if serializer.is_valid():
        try:
            # Generate status ID
            status_id = uuid.uuid4().hex[:12]
            while Status.objects.filter(id=status_id).exists():
                status_id = uuid.uuid4().hex[:12]
            
            # Auto-assign orderIndex: get the max orderIndex for the same type and add 1
            status_type = serializer.validated_data.get('type', 'lead')
            max_order = Status.objects.filter(type=status_type).aggregate(
                max_order=models.Max('order_index')
            )['max_order'] or -1
            order_index = max_order + 1
            
            status_obj = serializer.save(id=status_id, order_index=order_index, created_by=request.user if request.user.is_authenticated else None)
            return Response(StatusSerializer(status_obj).data, status=status.HTTP_201_CREATED)
        except IntegrityError as e:
            # Catch database-level unique constraint violations
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Status creation database integrity error: {str(e)}, data: {request.data}")
            
            # Extract name and type from the data
            name = data.get('name', '').strip()
            status_type = data.get('type', 'lead')
            
            # Return a user-friendly error message
            return Response({
                'non_field_errors': [f"A status with name '{name}' and type '{status_type}' already exists."]
            }, status=status.HTTP_400_BAD_REQUEST)
    
    # Log validation errors for debugging
    import logging
    logger = logging.getLogger(__name__)
    logger.error(f"Status creation validation failed: {serializer.errors}, data: {request.data}")
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def status_update(request, status_id):
    """Update a status"""
    import sys
    import logging
    logger = logging.getLogger(__name__)
    
    status_obj = get_object_or_404(Status, id=status_id)
    
    # Debug: log incoming data using both print and logger
    logger.info(f"[DEBUG] Status update request data: {request.data}")
    logger.info(f"[DEBUG] Status ID: {status_id}")
    logger.info(f"[DEBUG] Request method: {request.method}")
    print(f"[DEBUG] Status update request data: {request.data}", flush=True, file=sys.stderr)
    print(f"[DEBUG] Status ID: {status_id}", flush=True, file=sys.stderr)
    print(f"[DEBUG] Request method: {request.method}", flush=True, file=sys.stderr)
    sys.stderr.flush()
    
    serializer = StatusSerializer(status_obj, data=request.data, partial=True)
    if serializer.is_valid():
        logger.info(f"[DEBUG] Serializer is valid, saving...")
        print(f"[DEBUG] Serializer is valid, saving...", flush=True, file=sys.stderr)
        sys.stderr.flush()
        serializer.save()
        # Refresh from database to get updated values
        status_obj.refresh_from_db()
        sys.stderr.flush()
        return Response(StatusSerializer(status_obj).data, status=status.HTTP_200_OK)
    logger.error(f"[ERROR] Status update validation errors: {serializer.errors}")
    print(f"[ERROR] Status update validation errors: {serializer.errors}", flush=True, file=sys.stderr)
    sys.stderr.flush()
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def status_delete(request, status_id):
    """Delete a status"""
    status_obj = get_object_or_404(Status, id=status_id)
    status_obj.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# Sources endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def source_list(request):
    """List all sources"""
    sources = Source.objects.all().order_by('name')
    serializer = SourceSerializer(sources, many=True)
    return Response({'sources': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def source_create(request):
    """Create a new source"""
    source_id = uuid.uuid4().hex[:12]
    while Source.objects.filter(id=source_id).exists():
        source_id = uuid.uuid4().hex[:12]
    
    # Get name from request data
    name = request.data.get('name', '').strip()
    if not name:
        return Response({'error': 'Le nom de la source est requis'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if source with same name already exists
    if Source.objects.filter(name=name).exists():
        return Response({'error': 'Une source avec ce nom existe déjà'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Create source directly (bypass serializer for creation to avoid issues)
    try:
        source = Source.objects.create(id=source_id, name=name, created_by=request.user if request.user.is_authenticated else None)
        serializer = SourceSerializer(source)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error creating source: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def source_update(request, source_id):
    """Update a source"""
    source_obj = get_object_or_404(Source, id=source_id)
    serializer = SourceSerializer(source_obj, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(SourceSerializer(source_obj).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def source_delete(request, source_id):
    """Delete a source"""
    source_obj = get_object_or_404(Source, id=source_id)
    source_obj.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# Platforms endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def platform_list(request):
    """List all platforms"""
    platforms = Platform.objects.all().order_by('name')
    serializer = PlatformSerializer(platforms, many=True)
    return Response({'platforms': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def platform_create(request):
    """Create a new platform"""
    platform_id = uuid.uuid4().hex[:12]
    while Platform.objects.filter(id=platform_id).exists():
        platform_id = uuid.uuid4().hex[:12]
    
    # Get name from request data
    name = request.data.get('name', '').strip()
    if not name:
        return Response({'error': 'Le nom de la plateforme est requis'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if platform with same name already exists
    if Platform.objects.filter(name=name).exists():
        return Response({'error': 'Une plateforme avec ce nom existe déjà'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Create platform directly (bypass serializer for creation to avoid issues)
    try:
        platform = Platform.objects.create(id=platform_id, name=name, created_by=request.user if request.user.is_authenticated else None)
        serializer = PlatformSerializer(platform)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error creating platform: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def platform_update(request, platform_id):
    """Update a platform"""
    platform_obj = get_object_or_404(Platform, id=platform_id)
    serializer = PlatformSerializer(platform_obj, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(PlatformSerializer(platform_obj).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def platform_delete(request, platform_id):
    """Delete a platform"""
    platform_obj = get_object_or_404(Platform, id=platform_id)
    platform_obj.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

# Note Categories endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def note_category_list(request):
    """List all note categories"""
    categories = NoteCategory.objects.all().order_by('order_index', 'name')
    serializer = NoteCategorySerializer(categories, many=True)
    return Response({'categories': serializer.data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def note_category_create(request):
    """Create a new note category"""
    serializer = NoteCategorySerializer(data=request.data)
    if serializer.is_valid():
        # Generate category ID
        category_id = uuid.uuid4().hex[:12]
        while NoteCategory.objects.filter(id=category_id).exists():
            category_id = uuid.uuid4().hex[:12]
        
        # Auto-assign orderIndex: get the max orderIndex and add 1
        max_order = NoteCategory.objects.aggregate(
            max_order=models.Max('order_index')
        )['max_order'] or -1
        order_index = max_order + 1
        
        category = serializer.save(id=category_id, order_index=order_index, created_by=request.user if request.user.is_authenticated else None)
        return Response(NoteCategorySerializer(category).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def note_category_update(request, category_id):
    """Update a note category"""
    category = get_object_or_404(NoteCategory, id=category_id)
    serializer = NoteCategorySerializer(category, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(NoteCategorySerializer(category).data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def note_category_delete(request, category_id):
    """Delete a note category"""
    category = get_object_or_404(NoteCategory, id=category_id)
    category.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def note_category_reorder(request):
    """Update orderIndex for multiple note categories"""
    try:
        categories_data = request.data.get('categories', [])
        if not isinstance(categories_data, list):
            return Response(
                {'error': 'categories must be a list'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        for item in categories_data:
            category_id = item.get('id')
            order_index = item.get('orderIndex')
            
            if not category_id or order_index is None:
                continue
                
            try:
                category_obj = NoteCategory.objects.get(id=category_id)
                category_obj.order_index = order_index
                category_obj.save()
            except NoteCategory.DoesNotExist:
                continue
        
        return Response({'success': True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def contact_logs(request, contact_id):
    """Get all logs related to a contact"""
    try:
        # Verify contact exists
        contact = get_object_or_404(Contact, id=contact_id)
        
        # Get all logs for this contact with optimized queries
        # Use select_related to avoid N+1 queries when accessing user_id, contact_id
        logs = Log.objects.filter(
            contact_id=contact
        ).select_related(
            'user_id',  # For userId and userName (who performed the action) in serializer
            'contact_id'  # For contactId in serializer
        ).order_by('-created_at')
        
        # Debug: Print log types being returned
        log_types = logs.values_list('event_type', flat=True).distinct()
        print(f"[CONTACT LOGS] Contact {contact_id}: Found {logs.count()} logs with types: {list(log_types)}")
        
        serializer = LogSerializer(logs, many=True)
        return Response({'logs': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error in contact_logs: {error_details}")
        return Response({'error': str(e), 'details': error_details}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def contact_documents(request, contact_id):
    """Get all documents for a contact"""
    try:
        contact = get_object_or_404(Contact, id=contact_id)
        documents = Document.objects.filter(contact_id=contact).order_by('document_type')
        serializer = DocumentSerializer(documents, many=True)
        return Response({'documents': serializer.data}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def document_upload(request):
    """Upload a file to Impossible Cloud and return the URL"""
    try:
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        contact_id = request.data.get('contactId')
        document_type = request.data.get('documentType')
        
        if not contact_id or not document_type:
            return Response({'error': 'contactId and documentType are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate file size (max 10MB) - reject before starting upload
        MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
        if file.size > MAX_FILE_SIZE:
            file_size_mb = file.size / 1024 / 1024
            return Response({
                'error': f'File too large ({file_size_mb:.2f} MB). Maximum size: 10 MB'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if file is empty
        if file.size == 0:
            return Response({'error': 'File is empty'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Configure S3 client for Impossible Cloud with reasonable timeouts
        # Note: These timeouts apply per HTTP request, not the entire upload
        # For a 10MB file, the upload should complete within reasonable time
        from botocore.config import Config
        s3_config = Config(
            connect_timeout=10,  # 10 seconds to establish connection
            read_timeout=30,     # 30 seconds per read operation (allows for slower connections)
            retries={'max_attempts': 2}  # Allow 2 retries for transient errors
        )
        
        s3_client = boto3.client(
            's3',
            endpoint_url=os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net'),
            aws_access_key_id=os.getenv('IMPOSSIBLE_CLOUD_ACCESS_KEY'),
            aws_secret_access_key=os.getenv('IMPOSSIBLE_CLOUD_SECRET_KEY'),
            region_name=os.getenv('IMPOSSIBLE_CLOUD_REGION', 'eu-central-2'),
            config=s3_config
        )
        
        bucket_name = os.getenv('IMPOSSIBLE_CLOUD_BUCKET', 'leadflow-documents')
        
        # Generate unique file path
        file_extension = os.path.splitext(file.name)[1]
        file_path = f"contacts/{contact_id}/{document_type}/{uuid.uuid4().hex[:12]}{file_extension}"
        
        # Upload file to Impossible Cloud
        file.seek(0)  # Reset file pointer
        s3_client.upload_fileobj(
            file,
            bucket_name,
            file_path,
            ExtraArgs={'ContentType': file.content_type}
        )
        
        # Generate public URL (using the endpoint URL format)
        endpoint = os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net')
        # Remove trailing slash if present
        endpoint = endpoint.rstrip('/')
        file_url = f"{endpoint}/{bucket_name}/{file_path}"
        
        return Response({
            'fileUrl': file_url,
            'fileName': file.name,
            'filePath': file_path
        }, status=status.HTTP_200_OK)
        
    except ClientError as e:
        import traceback
        error_details = traceback.format_exc()
        error_code = e.response.get('Error', {}).get('Code', 'Unknown')
        error_message = e.response.get('Error', {}).get('Message', str(e))
        
        # Check for timeout errors
        if 'timeout' in str(e).lower() or 'timed out' in str(e).lower():
            return Response({
                'error': 'Upload timeout. File may be too large or connection too slow. Maximum file size: 10 MB'
            }, status=status.HTTP_408_REQUEST_TIMEOUT)
        
        print(f"Error uploading to Impossible Cloud: {error_details}")
        return Response({
            'error': f'Failed to upload file: {error_message}',
            'errorCode': error_code
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        
        # Check for timeout or connection errors
        error_str = str(e).lower()
        if 'timeout' in error_str or 'timed out' in error_str:
            return Response({
                'error': 'Upload timeout. File may be too large or connection too slow. Maximum file size: 10 MB'
            }, status=status.HTTP_408_REQUEST_TIMEOUT)
        
        print(f"Error uploading document: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def document_create(request):
    """Create or update a document for a contact"""
    try:
        contact_id = request.data.get('contactId')
        document_type = request.data.get('documentType')
        file_url = request.data.get('fileUrl', '')
        file_name = request.data.get('fileName', '')
        
        if not contact_id or not document_type:
            return Response({'error': 'contactId and documentType are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        contact = get_object_or_404(Contact, id=contact_id)
        
        # Check if document already exists
        document, created = Document.objects.get_or_create(
            contact_id=contact,
            document_type=document_type,
            defaults={
                'id': uuid.uuid4().hex[:12],
                'has_document': bool(file_url),
                'file_url': file_url,
                'file_name': file_name,
                'uploaded_by': request.user if request.user.is_authenticated else None
            }
        )
        
        if not created:
            # Update existing document
            document.has_document = bool(file_url)
            document.file_url = file_url
            document.file_name = file_name
            document.uploaded_by = request.user if request.user.is_authenticated else None
            document.save()
        
        serializer = DocumentSerializer(document)
        return Response(serializer.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error creating/updating document: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def document_update(request, document_id):
    """Update a document"""
    try:
        document = get_object_or_404(Document, id=document_id)
        
        if 'hasDocument' in request.data:
            document.has_document = request.data.get('hasDocument', False)
        if 'fileUrl' in request.data:
            document.file_url = request.data.get('fileUrl', '')
        if 'fileName' in request.data:
            document.file_name = request.data.get('fileName', '')
        
        document.uploaded_by = request.user if request.user.is_authenticated else None
        document.save()
        
        serializer = DocumentSerializer(document)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def document_delete(request, document_id):
    """Delete a document"""
    try:
        document = get_object_or_404(Document, id=document_id)
        document.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

def _get_s3_client_and_path(document):
    """Helper function to get S3 client and file path from document"""
    if not document.file_url:
        return None, None, None
    
    file_url = document.file_url
    endpoint = os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net').rstrip('/')
    
    if file_url.startswith(endpoint):
        path_part = file_url[len(endpoint):].lstrip('/')
        parts = path_part.split('/', 1)
        if len(parts) == 2:
            bucket_name = parts[0]
            file_path = parts[1]
            
            s3_client = boto3.client(
                's3',
                endpoint_url=endpoint,
                aws_access_key_id=os.getenv('IMPOSSIBLE_CLOUD_ACCESS_KEY'),
                aws_secret_access_key=os.getenv('IMPOSSIBLE_CLOUD_SECRET_KEY'),
                region_name=os.getenv('IMPOSSIBLE_CLOUD_REGION', 'eu-central-2')
            )
            return s3_client, bucket_name, file_path
    
    return None, None, None

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def document_download(request, document_id):
    """Proxy file download from Impossible Cloud"""
    try:
        document = get_object_or_404(Document, id=document_id)
        
        s3_client, bucket_name, file_path = _get_s3_client_and_path(document)
        if not s3_client:
            return Response({'error': 'Invalid file URL format'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get the object from S3
        s3_object = s3_client.get_object(Bucket=bucket_name, Key=file_path)
        
        # Create a streaming response directly from S3
        def file_iterator():
            chunk_size = 8192
            while True:
                chunk = s3_object['Body'].read(chunk_size)
                if not chunk:
                    break
                yield chunk
        
        # Create streaming response
        response = StreamingHttpResponse(
            file_iterator(),
            content_type=s3_object.get('ContentType', 'application/octet-stream')
        )
        
        # Set the filename for download
        file_name = document.file_name or 'document'
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        if 'ContentLength' in s3_object:
            response['Content-Length'] = str(s3_object['ContentLength'])
        
        return response
        
    except ClientError as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error downloading from S3: {error_details}")
        return Response({'error': f'Failed to download file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error downloading document: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def document_view_url(request, document_id):
    """Generate a presigned URL for viewing a document"""
    try:
        document = get_object_or_404(Document, id=document_id)
        
        s3_client, bucket_name, file_path = _get_s3_client_and_path(document)
        if not s3_client:
            return Response({'error': 'Invalid file URL format'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Generate presigned URL (valid for 1 hour)
        presigned_url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': file_path},
            ExpiresIn=3600
        )
        
        return Response({
            'viewUrl': presigned_url,
            'fileName': document.file_name
        }, status=status.HTTP_200_OK)
        
    except ClientError as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error generating presigned URL: {error_details}")
        return Response({'error': f'Failed to generate view URL: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error getting view URL: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def status_reorder(request):
    """Update orderIndex for multiple statuses"""
    try:
        statuses_data = request.data.get('statuses', [])
        if not isinstance(statuses_data, list):
            return Response(
                {'error': 'statuses must be a list'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        for item in statuses_data:
            status_id = item.get('id')
            order_index = item.get('orderIndex')
            
            if not status_id or order_index is None:
                continue
                
            try:
                status_obj = Status.objects.get(id=status_id)
                status_obj.order_index = order_index
                status_obj.save()
            except Status.DoesNotExist:
                continue
        
        return Response({'success': True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stats(request):
    """Get dashboard statistics"""
    try:
        user = request.user
        
        # Get filter parameters
        date_from = request.GET.get('dateFrom')
        date_to = request.GET.get('dateTo')
        team_id = request.GET.get('teamId')
        user_id = request.GET.get('userId')
        
        # Base querysets
        contacts_qs = Contact.objects.all()
        notes_qs = Note.objects.all()
        events_qs = Event.objects.all()
        users_qs = UserDetails.objects.filter(active=True, deleted_at__isnull=True)
        
        # Apply data_access filtering based on user's role
        try:
            user_details = UserDetails.objects.select_related('role_id').get(django_user=user)
            if user_details.role:
                data_access = user_details.role.data_access
                print(f"[DEBUG] User {user.username} has data_access: {data_access}")
                
                if data_access == 'all':
                    # User has access to all data, no filtering needed
                    print(f"[DEBUG] data_access is 'all', no filtering applied")
                    pass
                elif data_access == 'own_only':
                    # Show only data where user is assigned as teleoperator or confirmateur
                    is_teleoperateur = user_details.role.is_teleoperateur
                    is_confirmateur = user_details.role.is_confirmateur
                    
                    if is_teleoperateur and is_confirmateur:
                        # User is both: show contacts where user is teleoperator OR confirmateur
                        contacts_qs = contacts_qs.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user)
                        )
                        # For notes, filter by userId (user created the note)
                        notes_qs = notes_qs.filter(userId=user)
                        # For events, filter by contacts accessible to user
                        accessible_contact_ids = Contact.objects.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user)
                        ).values_list('id', flat=True)
                        events_qs = events_qs.filter(contactId__id__in=accessible_contact_ids)
                    elif is_teleoperateur:
                        # Teleoperateur with own_only: only show contacts where user is teleoperator
                        contacts_qs = contacts_qs.filter(teleoperator=user)
                        # For notes, filter by userId (user created the note)
                        notes_qs = notes_qs.filter(userId=user)
                        accessible_contact_ids = Contact.objects.filter(teleoperator=user).values_list('id', flat=True)
                        events_qs = events_qs.filter(contactId__id__in=accessible_contact_ids)
                    elif is_confirmateur:
                        # Confirmateur with own_only: only show contacts where user is confirmateur
                        contacts_qs = contacts_qs.filter(confirmateur=user)
                        # For notes, filter by userId (user created the note)
                        notes_qs = notes_qs.filter(userId=user)
                        accessible_contact_ids = Contact.objects.filter(confirmateur=user).values_list('id', flat=True)
                        events_qs = events_qs.filter(contactId__id__in=accessible_contact_ids)
                    else:
                        # Default behavior: show contacts where user is teleoperator, confirmateur, or creator
                        contacts_qs = contacts_qs.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user)
                        )
                        # For notes, filter by userId (user created the note)
                        notes_qs = notes_qs.filter(userId=user)
                        accessible_contact_ids = Contact.objects.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user)
                        ).values_list('id', flat=True)
                        events_qs = events_qs.filter(contactId__id__in=accessible_contact_ids)
                        
                elif data_access == 'team_only':
                    # Show data linked to teleoperateurs of the team the user is in
                    team_member = user_details.team_memberships.select_related('team').first()
                    if team_member:
                        team = team_member.team
                        # Get all users in the same team
                        team_user_ids = TeamMember.objects.filter(team=team).values_list('user__django_user__id', flat=True)
                        # Show contacts where:
                        # - User is teleoperator, confirmateur, or creator
                        # - OR contact's teleoperator/confirmateur/creator is in the same team
                        contacts_qs = contacts_qs.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user) |
                            models.Q(teleoperator__id__in=team_user_ids) |
                            models.Q(confirmateur__id__in=team_user_ids) |
                            models.Q(creator__id__in=team_user_ids)
                        )
                        # For notes, filter by userId (notes created by users in the same team)
                        notes_qs = notes_qs.filter(userId__id__in=team_user_ids)
                        # For events, filter by contacts accessible to user or team
                        accessible_contact_ids = Contact.objects.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user) |
                            models.Q(teleoperator__id__in=team_user_ids) |
                            models.Q(confirmateur__id__in=team_user_ids) |
                            models.Q(creator__id__in=team_user_ids)
                        ).values_list('id', flat=True)
                        events_qs = events_qs.filter(contactId__id__in=accessible_contact_ids)
                    else:
                        # User has no team, fall back to own_only behavior
                        contacts_qs = contacts_qs.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user)
                        )
                        # For notes, filter by userId (user created the note)
                        notes_qs = notes_qs.filter(userId=user)
                        accessible_contact_ids = Contact.objects.filter(
                            models.Q(teleoperator=user) |
                            models.Q(confirmateur=user) |
                            models.Q(creator=user)
                        ).values_list('id', flat=True)
                        events_qs = events_qs.filter(contactId__id__in=accessible_contact_ids)
                # If user_details.role is None or data_access is 'all', show all data (no filtering)
            else:
                print(f"[DEBUG] User {user.username} has no role, no filtering applied")
            # If user_details.role is None, no filtering is applied (show all data)
        except UserDetails.DoesNotExist:
            # If user has no UserDetails, show no data (safety default)
            print(f"[DEBUG] User {user.username} has no UserDetails, showing no data")
            contacts_qs = Contact.objects.none()
            notes_qs = Note.objects.none()
            events_qs = Event.objects.none()
        
        # Debug: Print queryset counts before date filters
        print(f"[DEBUG] After data_access filtering - Contacts: {contacts_qs.count()}, Notes: {notes_qs.count()}, Events: {events_qs.count()}")
        
        # Apply date filters
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                contacts_qs = contacts_qs.filter(created_at__date__gte=date_from_obj)
                notes_qs = notes_qs.filter(created_at__date__gte=date_from_obj)
                events_qs = events_qs.filter(created_at__date__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                # Use __lt with next day to ensure we only include items up to end of selected day
                date_to_next = date_to_obj + timedelta(days=1)
                contacts_qs = contacts_qs.filter(created_at__date__lt=date_to_next)
                notes_qs = notes_qs.filter(created_at__date__lt=date_to_next)
                events_qs = events_qs.filter(created_at__date__lt=date_to_next)
            except ValueError:
                pass
        
        # Apply team filter
        if team_id and team_id != 'all':
            try:
                team = Team.objects.get(id=team_id)
                team_members = TeamMember.objects.filter(team=team).values_list('user__django_user', flat=True)
                contacts_qs = contacts_qs.filter(
                    Q(creator__in=team_members) | 
                    Q(teleoperator__in=team_members) | 
                    Q(confirmateur__in=team_members)
                )
                notes_qs = notes_qs.filter(userId__in=team_members)
                events_qs = events_qs.filter(userId__in=team_members)
            except Team.DoesNotExist:
                pass
        
        # Apply user filter
        if user_id and user_id != 'all':
            try:
                user_filter = DjangoUser.objects.get(id=user_id)
                # Filter contacts where user is creator, teleoperator, or confirmateur
                contacts_qs = contacts_qs.filter(
                    Q(creator=user_filter) | 
                    Q(teleoperator=user_filter) | 
                    Q(confirmateur=user_filter)
                )
                # Filter notes and events by user
                notes_qs = notes_qs.filter(userId=user_filter)
                events_qs = events_qs.filter(userId=user_filter)
            except DjangoUser.DoesNotExist:
                pass
        
        # Calculate statistics
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())
        month_start = today_start.replace(day=1)
        
        # Total counts
        total_contacts = contacts_qs.count()
        total_notes = notes_qs.count()
        total_events = events_qs.count()
        total_users = users_qs.count()
        
        # Contacts by status type
        contacts_by_status_type = contacts_qs.values('status__type').annotate(count=Count('id'))
        leads_count = sum(item['count'] for item in contacts_by_status_type if item['status__type'] == 'lead')
        contacts_count = sum(item['count'] for item in contacts_by_status_type if item['status__type'] == 'contact')
        clients_count = sum(item['count'] for item in contacts_by_status_type if item['status__type'] == 'client')
        
        # Recent activity
        contacts_today = contacts_qs.filter(created_at__gte=today_start).count()
        contacts_this_week = contacts_qs.filter(created_at__gte=week_start).count()
        contacts_this_month = contacts_qs.filter(created_at__gte=month_start).count()
        
        notes_today = notes_qs.filter(created_at__gte=today_start).count()
        events_today = events_qs.filter(created_at__gte=today_start).count()
        
        # Contacts by source
        contacts_by_source = contacts_qs.values('source__name').annotate(count=Count('id')).order_by('-count')[:5]
        top_sources = [{'name': item['source__name'] or 'Non défini', 'count': item['count']} for item in contacts_by_source]
        
        # Contacts by teleoperator
        contacts_by_teleoperator = contacts_qs.filter(teleoperator__isnull=False).values(
            'teleoperator__first_name', 
            'teleoperator__last_name'
        ).annotate(count=Count('id')).order_by('-count')[:5]
        top_teleoperators = [
            {
                'name': f"{item['teleoperator__first_name'] or ''} {item['teleoperator__last_name'] or ''}".strip() or 'Non défini',
                'count': item['count']
            } 
            for item in contacts_by_teleoperator
        ]
        
        # Notes by user (only for admins - data_access == 'all')
        notes_by_user = []
        try:
            user_details = UserDetails.objects.select_related('role_id').get(django_user=user)
            if user_details.role and user_details.role.data_access == 'all':
                # Admin: show notes count for each user
                # Get all notes (no filtering for admin)
                all_notes_qs = Note.objects.all()
                
                # Apply date filters if provided
                if date_from:
                    try:
                        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                        all_notes_qs = all_notes_qs.filter(created_at__date__gte=date_from_obj)
                    except ValueError:
                        pass
                
                if date_to:
                    try:
                        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                        date_to_next = date_to_obj + timedelta(days=1)
                        all_notes_qs = all_notes_qs.filter(created_at__date__lt=date_to_next)
                    except ValueError:
                        pass
                
                # Apply team filter if provided
                if team_id and team_id != 'all':
                    try:
                        team = Team.objects.get(id=team_id)
                        team_members = TeamMember.objects.filter(team=team).values_list('user__django_user', flat=True)
                        all_notes_qs = all_notes_qs.filter(userId__in=team_members)
                    except Team.DoesNotExist:
                        pass
                
                # Apply user filter if provided
                if user_id and user_id != 'all':
                    try:
                        user_filter = DjangoUser.objects.get(id=user_id)
                        all_notes_qs = all_notes_qs.filter(userId=user_filter)
                    except DjangoUser.DoesNotExist:
                        pass
                
                # Count notes by user
                notes_by_user_data = all_notes_qs.values(
                    'userId__id',
                    'userId__first_name',
                    'userId__last_name',
                    'userId__username'
                ).annotate(count=Count('id')).order_by('-count')
                
                notes_by_user = [
                    {
                        'userId': item['userId__id'],
                        'name': f"{item['userId__first_name'] or ''} {item['userId__last_name'] or ''}".strip() or item['userId__username'] or 'Non défini',
                        'count': item['count']
                    }
                    for item in notes_by_user_data
                ]
        except UserDetails.DoesNotExist:
            pass
        
        # Upcoming events (next 7 days)
        upcoming_events = events_qs.filter(
            datetime__gte=now,
            datetime__lte=now + timedelta(days=7)
        ).order_by('datetime')[:10]
        
        upcoming_events_data = []
        for event in upcoming_events:
            upcoming_events_data.append({
                'id': event.id,
                'datetime': event.datetime.isoformat(),
                'contactId': event.contactId.id if event.contactId else None,
                'contactName': f"{event.contactId.fname} {event.contactId.lname}".strip() if event.contactId else None,
                'comment': event.comment,
                'userId': event.userId.id,
                'userName': f"{event.userId.first_name} {event.userId.last_name}".strip() or event.userId.username
            })
        
        # Recent contacts (last 10)
        recent_contacts = contacts_qs.order_by('-created_at')[:10]
        recent_contacts_data = []
        for contact in recent_contacts:
            recent_contacts_data.append({
                'id': contact.id,
                'name': f"{contact.fname} {contact.lname}".strip(),
                'status': contact.status.name if contact.status else None,
                'source': contact.source.name if contact.source else None,
                'createdAt': contact.created_at.isoformat()
            })
        
        response_data = {
            'totalContacts': total_contacts,
            'totalLeads': leads_count,
            'totalContactsCount': contacts_count,
            'totalClients': clients_count,
            'totalNotes': total_notes,
            'totalEvents': total_events,
            'totalUsers': total_users,
            'contactsToday': contacts_today,
            'contactsThisWeek': contacts_this_week,
            'contactsThisMonth': contacts_this_month,
            'notesToday': notes_today,
            'eventsToday': events_today,
            'topSources': top_sources,
            'topTeleoperators': top_teleoperators,
            'upcomingEvents': upcoming_events_data,
            'recentContacts': recent_contacts_data
        }
        
        # Add notesByUser only for admins
        if notes_by_user:
            response_data['notesByUser'] = notes_by_user
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error getting stats: {error_details}")
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# ==================== EMAIL ENDPOINTS ====================

@api_view(['GET', 'POST', 'PUT'])
@permission_classes([IsAuthenticated])
def smtp_config(request):
    """Get, create, or update SMTP configuration for current user"""
    user = request.user
    
    if request.method == 'GET':
        try:
            config = SMTPConfig.objects.filter(user=user).first()
            if config:
                serializer = SMTPConfigSerializer(config)
                return Response({'config': serializer.data})
            return Response({'config': None})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'POST':
        # Create new config
        try:
            # Check if config already exists
            existing = SMTPConfig.objects.filter(user=user).first()
            if existing:
                return Response({'error': 'SMTP configuration already exists. Use PUT to update.'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            
            config_id = uuid.uuid4().hex[:12]
            while SMTPConfig.objects.filter(id=config_id).exists():
                config_id = uuid.uuid4().hex[:12]
            
            config = SMTPConfig.objects.create(
                id=config_id,
                user=user,
                email_address=request.data.get('emailAddress', ''),
                smtp_server=request.data.get('smtpServer', ''),
                smtp_port=request.data.get('smtpPort', 587),
                smtp_use_tls=request.data.get('smtpUseTls', True),
                smtp_username=request.data.get('smtpUsername', ''),
                smtp_password=request.data.get('smtpPassword', ''),
                imap_server=request.data.get('imapServer', ''),
                imap_port=request.data.get('imapPort', 993),
                imap_use_ssl=request.data.get('imapUseSsl', True),
                is_active=request.data.get('isActive', True)
            )
            
            serializer = SMTPConfigSerializer(config)
            return Response({'config': serializer.data}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'PUT':
        # Update existing config
        try:
            config = SMTPConfig.objects.filter(user=user).first()
            if not config:
                return Response({'error': 'SMTP configuration not found. Use POST to create.'}, 
                              status=status.HTTP_404_NOT_FOUND)
            
            # Update fields
            if 'emailAddress' in request.data:
                config.email_address = request.data['emailAddress']
            if 'smtpServer' in request.data:
                config.smtp_server = request.data['smtpServer']
            if 'smtpPort' in request.data:
                config.smtp_port = request.data['smtpPort']
            if 'smtpUseTls' in request.data:
                config.smtp_use_tls = request.data['smtpUseTls']
            if 'smtpUsername' in request.data:
                config.smtp_username = request.data['smtpUsername']
            if 'smtpPassword' in request.data:
                config.smtp_password = request.data['smtpPassword']
            if 'imapServer' in request.data:
                config.imap_server = request.data.get('imapServer', '')
            if 'imapPort' in request.data:
                config.imap_port = request.data.get('imapPort', 993)
            if 'imapUseSsl' in request.data:
                config.imap_use_ssl = request.data.get('imapUseSsl', True)
            if 'isActive' in request.data:
                config.is_active = request.data['isActive']
            
            config.save()
            serializer = SMTPConfigSerializer(config)
            return Response({'config': serializer.data})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_smtp_connection(request):
    """Test SMTP connection with current configuration"""
    user = request.user
    try:
        config = SMTPConfig.objects.filter(user=user, is_active=True).first()
        if not config:
            return Response({'error': 'No active SMTP configuration found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Test SMTP connection with timeout
        server = None
        try:
            # Port 465 SSL connections may need slightly more time for SSL handshake
            timeout = 45 if config.smtp_port == 465 else 30  # 45 seconds for SSL, 30 for others
            # Port 465 requires SSL from the start, not STARTTLS
            # Port 587 typically uses STARTTLS
            if config.smtp_port == 465:
                # Port 465 always uses SSL/TLS from the start
                server = smtplib.SMTP_SSL(config.smtp_server, config.smtp_port, timeout=timeout)
            elif config.smtp_use_tls:
                # Port 587 or other ports with TLS enabled - use STARTTLS
                server = smtplib.SMTP(config.smtp_server, config.smtp_port, timeout=timeout)
                server.starttls()
            else:
                # No encryption (not recommended but supported)
                server = smtplib.SMTP(config.smtp_server, config.smtp_port, timeout=timeout)
            
            server.login(config.smtp_username, config.smtp_password)
            server.quit()
            server = None  # Mark as closed
            
            return Response({'success': True, 'message': 'SMTP connection successful'})
        except smtplib.SMTPConnectError as e:
            return Response({'success': False, 'error': f'Could not connect to SMTP server: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPAuthenticationError as e:
            return Response({'success': False, 'error': f'Authentication failed: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPException as e:
            return Response({'success': False, 'error': f'SMTP error: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'success': False, 'error': f'Connection error: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        finally:
            if server:
                try:
                    server.quit()
                except:
                    pass
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_email(request):
    """Send an email using user's SMTP configuration"""
    user = request.user
    try:
        config = SMTPConfig.objects.filter(user=user, is_active=True).first()
        if not config:
            return Response({'error': 'No active SMTP configuration found. Please configure SMTP settings first.'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Get email data
        to_emails = request.data.get('toEmails', [])
        if not to_emails or not isinstance(to_emails, list):
            return Response({'error': 'toEmails is required and must be a list'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        subject = request.data.get('subject', '')
        body_text = request.data.get('bodyText', '')
        body_html = request.data.get('bodyHtml', '')
        cc_emails = request.data.get('ccEmails', [])
        bcc_emails = request.data.get('bccEmails', [])
        contact_id = request.data.get('contactId', None)
        
        # Convert logo URLs in body_html to base64 embedded images (publicly accessible)
        if body_html:
            import re
            import base64
            endpoint = os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net').rstrip('/')
            bucket_name = os.getenv('IMPOSSIBLE_CLOUD_BUCKET', 'leadflow-documents')
            
            # Find all img tags with src pointing to Impossible Cloud
            def replace_logo_url(match):
                img_tag = match.group(0)
                src_match = re.search(r'src="([^"]+)"', img_tag)
                if src_match:
                    logo_url = src_match.group(1)
                    # Check if it's an Impossible Cloud URL
                    if logo_url.startswith(endpoint):
                        # Extract file path
                        path_part = logo_url[len(endpoint):].lstrip('/')
                        if path_part.startswith(f'{bucket_name}/'):
                            file_path = path_part[len(bucket_name) + 1:]  # Remove bucket name and leading slash
                            
                            # Check if it's a signature logo
                            if file_path.startswith('email-signatures/'):
                                # Download image and convert to base64
                                try:
                                    s3_client = boto3.client(
                                        's3',
                                        endpoint_url=endpoint,
                                        aws_access_key_id=os.getenv('IMPOSSIBLE_CLOUD_ACCESS_KEY'),
                                        aws_secret_access_key=os.getenv('IMPOSSIBLE_CLOUD_SECRET_KEY'),
                                        region_name=os.getenv('IMPOSSIBLE_CLOUD_REGION', 'eu-central-2')
                                    )
                                    
                                    # Get the image from S3
                                    s3_object = s3_client.get_object(Bucket=bucket_name, Key=file_path)
                                    image_data = s3_object['Body'].read()
                                    content_type = s3_object.get('ContentType', 'image/png')
                                    
                                    # Convert to base64
                                    base64_data = base64.b64encode(image_data).decode('utf-8')
                                    data_uri = f"data:{content_type};base64,{base64_data}"
                                    
                                    print(f"Embedded logo as base64: {file_path} ({len(image_data)} bytes)")
                                    
                                    # Replace URL in img tag with data URI
                                    return img_tag.replace(logo_url, data_uri)
                                except Exception as e:
                                    import traceback
                                    error_details = traceback.format_exc()
                                    print(f"Error embedding logo as base64 for {file_path}: {error_details}")
                                    return img_tag  # Return original if error
                
                return img_tag  # Return original if not a signature logo
            
            # Replace all img src URLs with base64 data URIs
            body_html = re.sub(r'<img[^>]+src="[^"]+"[^>]*>', replace_logo_url, body_html)
        
        # Create email message
        msg = MIMEMultipart('alternative')
        msg['From'] = config.email_address
        msg['To'] = ', '.join(to_emails)
        if cc_emails:
            msg['Cc'] = ', '.join(cc_emails)
        msg['Subject'] = subject
        
        # Add body
        if body_html:
            msg.attach(MIMEText(body_html, 'html'))
        if body_text:
            msg.attach(MIMEText(body_text, 'plain'))
        
        # Handle attachments (if any)
        attachments = request.data.get('attachments', [])
        # Note: In a real implementation, you'd need to handle file uploads
        
        # Send email with timeout handling
        server = None
        try:
            # Port 465 SSL connections may need slightly more time for SSL handshake
            timeout = 45 if config.smtp_port == 465 else 30  # 45 seconds for SSL, 30 for others
            # Port 465 requires SSL from the start, not STARTTLS
            # Port 587 typically uses STARTTLS
            if config.smtp_port == 465:
                # Port 465 always uses SSL/TLS from the start
                server = smtplib.SMTP_SSL(config.smtp_server, config.smtp_port, timeout=timeout)
            elif config.smtp_use_tls:
                # Port 587 or other ports with TLS enabled - use STARTTLS
                server = smtplib.SMTP(config.smtp_server, config.smtp_port, timeout=timeout)
                server.starttls()
            else:
                # No encryption (not recommended but supported)
                server = smtplib.SMTP(config.smtp_server, config.smtp_port, timeout=timeout)
            
            server.login(config.smtp_username, config.smtp_password)
            
            # Combine all recipients
            recipients = to_emails + (cc_emails if cc_emails else []) + (bcc_emails if bcc_emails else [])
            server.sendmail(config.email_address, recipients, msg.as_string())
            server.quit()
            server = None  # Mark as closed
            
            # Save email to database
            contact = None
            if contact_id:
                contact = Contact.objects.filter(id=contact_id).first()
            
            email_id = uuid.uuid4().hex[:12]
            while Email.objects.filter(id=email_id).exists():
                email_id = uuid.uuid4().hex[:12]
            
            email_obj = Email.objects.create(
                id=email_id,
                user=user,
                email_type='sent',
                subject=subject,
                from_email=config.email_address,
                to_emails=to_emails,
                cc_emails=cc_emails or [],
                bcc_emails=bcc_emails or [],
                body_text=body_text,
                body_html=body_html,
                attachments=attachments or [],
                contact=contact,
                sent_at=timezone.now(),
                is_read=True
            )
            
            serializer = EmailSerializer(email_obj)
            return Response({'email': serializer.data, 'message': 'Email sent successfully'}, 
                          status=status.HTTP_201_CREATED)
        except smtplib.SMTPConnectError as e:
            return Response({'error': f'Could not connect to SMTP server. Please check your server address and port: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPAuthenticationError as e:
            return Response({'error': f'Authentication failed. Please check your username and password: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPRecipientsRefused as e:
            return Response({'error': f'One or more recipients were refused: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPSenderRefused as e:
            return Response({'error': f'Sender address was refused: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPDataError as e:
            return Response({'error': f'SMTP server refused the email data: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except smtplib.SMTPException as e:
            return Response({'error': f'SMTP error occurred: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        except TimeoutError as e:
            return Response({'error': f'Connection timeout. Please check your network connection and SMTP settings: {str(e)}'}, 
                          status=status.HTTP_408_REQUEST_TIMEOUT)
        except Exception as e:
            error_msg = str(e)
            if 'timeout' in error_msg.lower() or 'timed out' in error_msg.lower():
                return Response({'error': f'Request timeout. Please check your SMTP server settings and network connection: {error_msg}'}, 
                              status=status.HTTP_408_REQUEST_TIMEOUT)
            return Response({'error': f'Failed to send email: {error_msg}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        finally:
            # Ensure server connection is closed
            if server:
                try:
                    server.quit()
                except:
                    try:
                        server.close()
                    except:
                        pass
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def email_list(request):
    """Get list of emails for current user with pagination"""
    user = request.user
    try:
        email_type = request.query_params.get('type', 'all')  # all, sent, received, draft
        is_read = request.query_params.get('isRead', None)
        is_starred = request.query_params.get('isStarred', None)
        contact_id = request.query_params.get('contactId', None)
        
        # Pagination parameters - define outside try block to avoid NameError in except block
        try:
            page = int(request.query_params.get('page', 1))
        except (ValueError, TypeError):
            page = 1
        try:
            limit = int(request.query_params.get('limit', 50))
        except (ValueError, TypeError):
            limit = 50
        offset = (page - 1) * limit
        
        emails = Email.objects.filter(user=user)
        
        if email_type != 'all':
            emails = emails.filter(email_type=email_type)
        
        if is_read is not None:
            emails = emails.filter(is_read=is_read.lower() == 'true')
        
        if is_starred is not None:
            emails = emails.filter(is_starred=is_starred.lower() == 'true')
        
        if contact_id:
            emails = emails.filter(contact_id=contact_id)
        
        # Get total count before pagination
        total_count = emails.count()
        
        # Order by sent_at (most recent first)
        emails = emails.order_by('-sent_at', '-created_at')
        
        # Apply pagination
        emails = emails[offset:offset + limit]
        
        serializer = EmailSerializer(emails, many=True)
        return Response({
            'emails': serializer.data,
            'total': total_count,
            'page': page,
            'limit': limit
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def email_detail(request, email_id):
    """Get email details"""
    user = request.user
    try:
        email_obj = Email.objects.filter(id=email_id, user=user).first()
        if not email_obj:
            return Response({'error': 'Email not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Mark as read if it's a received email
        if email_obj.email_type == 'received' and not email_obj.is_read:
            email_obj.is_read = True
            email_obj.save()
        
        serializer = EmailSerializer(email_obj)
        return Response({'email': serializer.data})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def fetch_emails(request):
    """Fetch emails from IMAP server"""
    user = request.user
    try:
        config = SMTPConfig.objects.filter(user=user, is_active=True).first()
        if not config or not config.imap_server:
            return Response({'error': 'No active IMAP configuration found'}, 
                          status=status.HTTP_404_NOT_FOUND)
        
        # Connect to IMAP server
        try:
            if config.imap_use_ssl:
                mail = imaplib.IMAP4_SSL(config.imap_server, config.imap_port)
            else:
                mail = imaplib.IMAP4(config.imap_server, config.imap_port)
            
            mail.login(config.smtp_username, config.smtp_password)
            mail.select('INBOX')
            
            # Search for unread emails
            status, messages = mail.search(None, 'UNSEEN')
            email_ids = messages[0].split()
            
            fetched_count = 0
            for email_id_bytes in email_ids:
                try:
                    # Fetch email
                    status, msg_data = mail.fetch(email_id_bytes, '(RFC822)')
                    email_body = msg_data[0][1]
                    email_message = email.message_from_bytes(email_body)
                    
                    # Parse email
                    subject = decode_header(email_message['Subject'])[0][0]
                    if isinstance(subject, bytes):
                        subject = subject.decode()
                    
                    from_email = email_message['From']
                    to_email = email_message['To']
                    
                    # Extract email addresses
                    from_match = re.search(r'[\w\.-]+@[\w\.-]+', from_email)
                    from_addr = from_match.group(0) if from_match else from_email
                    
                    # Get body
                    body_text = ''
                    body_html = ''
                    if email_message.is_multipart():
                        for part in email_message.walk():
                            content_type = part.get_content_type()
                            if content_type == 'text/plain':
                                body_text = part.get_payload(decode=True).decode()
                            elif content_type == 'text/html':
                                body_html = part.get_payload(decode=True).decode()
                    else:
                        body_text = email_message.get_payload(decode=True).decode()
                    
                    # Check if email already exists
                    message_id = email_message.get('Message-ID', '')
                    existing = Email.objects.filter(user=user, message_id=message_id).first()
                    if existing:
                        continue
                    
                    # Save email
                    new_email_id = uuid.uuid4().hex[:12]
                    while Email.objects.filter(id=new_email_id).exists():
                        new_email_id = uuid.uuid4().hex[:12]
                    
                    Email.objects.create(
                        id=new_email_id,
                        user=user,
                        email_type='received',
                        subject=subject or '(No Subject)',
                        from_email=from_addr,
                        to_emails=[config.email_address],
                        body_text=body_text,
                        body_html=body_html,
                        message_id=message_id,
                        in_reply_to=email_message.get('In-Reply-To', ''),
                        references=email_message.get('References', ''),
                        sent_at=email.utils.parsedate_to_datetime(email_message['Date']) if email_message.get('Date') else timezone.now(),
                        is_read=False
                    )
                    fetched_count += 1
                except Exception as e:
                    # Continue with next email if one fails
                    continue
            
            mail.close()
            mail.logout()
            
            return Response({'message': f'Fetched {fetched_count} new email(s)'})
        except Exception as e:
            return Response({'error': f'Failed to fetch emails: {str(e)}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def email_update(request, email_id):
    """Update email (mark as read/unread, star/unstar, etc.)"""
    user = request.user
    try:
        email_obj = Email.objects.filter(id=email_id, user=user).first()
        if not email_obj:
            return Response({'error': 'Email not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if 'isRead' in request.data:
            email_obj.is_read = request.data['isRead']
        if 'isStarred' in request.data:
            email_obj.is_starred = request.data['isStarred']
        
        email_obj.save()
        serializer = EmailSerializer(email_obj)
        return Response({'email': serializer.data})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def email_delete(request, email_id):
    """Delete an email"""
    user = request.user
    try:
        email_obj = Email.objects.filter(id=email_id, user=user).first()
        if not email_obj:
            return Response({'error': 'Email not found'}, status=status.HTTP_404_NOT_FOUND)
        
        email_obj.delete()
        return Response({'message': 'Email deleted successfully'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==================== EMAIL SIGNATURE ENDPOINTS ====================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def email_signatures(request):
    """Get list of email signatures or create a new one"""
    user = request.user
    
    if request.method == 'GET':
        try:
            signatures = EmailSignature.objects.filter(user=user).order_by('-is_default', '-created_at')
            serializer = EmailSignatureSerializer(signatures, many=True)
            return Response({'signatures': serializer.data})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'POST':
        # Create new signature
        try:
            name = request.data.get('name', '')
            if not name:
                return Response({'error': 'Signature name is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            content_html = request.data.get('contentHtml', '')
            content_text = request.data.get('contentText', '')
            logo_url = request.data.get('logoUrl', '')
            logo_position = request.data.get('logoPosition', 'left')
            is_default = request.data.get('isDefault', False)
            
            # If setting as default, unset other defaults
            if is_default:
                EmailSignature.objects.filter(user=user, is_default=True).update(is_default=False)
            
            signature_id = uuid.uuid4().hex[:12]
            while EmailSignature.objects.filter(id=signature_id).exists():
                signature_id = uuid.uuid4().hex[:12]
            
            signature = EmailSignature.objects.create(
                id=signature_id,
                user=user,
                name=name,
                content_html=content_html,
                content_text=content_text,
                logo_url=logo_url,
                logo_position=logo_position,
                is_default=is_default,
                created_by=request.user if request.user.is_authenticated else None
            )
            
            serializer = EmailSignatureSerializer(signature)
            return Response({'signature': serializer.data}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def email_signature_detail(request, signature_id):
    """Get, update, or delete a specific email signature"""
    user = request.user
    
    try:
        signature = EmailSignature.objects.filter(id=signature_id, user=user).first()
        if not signature:
            return Response({'error': 'Signature not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            serializer = EmailSignatureSerializer(signature)
            return Response({'signature': serializer.data})
        
        elif request.method == 'PUT':
            # Update signature
            if 'name' in request.data:
                signature.name = request.data['name']
            if 'contentHtml' in request.data:
                signature.content_html = request.data['contentHtml']
            if 'contentText' in request.data:
                signature.content_text = request.data['contentText']
            if 'logoUrl' in request.data:
                signature.logo_url = request.data['logoUrl']
            if 'logoPosition' in request.data:
                signature.logo_position = request.data['logoPosition']
            if 'isDefault' in request.data:
                is_default = request.data['isDefault']
                # If setting as default, unset other defaults
                if is_default:
                    EmailSignature.objects.filter(user=user, is_default=True).exclude(id=signature_id).update(is_default=False)
                signature.is_default = is_default
            
            signature.save()
            serializer = EmailSignatureSerializer(signature)
            return Response({'signature': serializer.data})
        
        elif request.method == 'DELETE':
            signature.delete()
            return Response({'message': 'Signature deleted successfully'}, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def email_signature_logo_upload(request):
    """Upload a logo image for email signature to Impossible Cloud and return the URL"""
    try:
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        user = request.user
        
        # Validate file type (only images)
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
        if file.content_type not in allowed_types:
            return Response({'error': 'Only image files are allowed (JPEG, PNG, GIF, WebP)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Configure S3 client for Impossible Cloud (same as document_upload)
        s3_client = boto3.client(
            's3',
            endpoint_url=os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net'),
            aws_access_key_id=os.getenv('IMPOSSIBLE_CLOUD_ACCESS_KEY'),
            aws_secret_access_key=os.getenv('IMPOSSIBLE_CLOUD_SECRET_KEY'),
            region_name=os.getenv('IMPOSSIBLE_CLOUD_REGION', 'eu-central-2')
        )
        
        bucket_name = os.getenv('IMPOSSIBLE_CLOUD_BUCKET', 'leadflow-documents')
        
        # Generate unique file path for signature logos
        file_extension = os.path.splitext(file.name)[1]
        file_path = f"email-signatures/{user.id}/{uuid.uuid4().hex[:12]}{file_extension}"
        
        # Upload file to Impossible Cloud (same logic as document_upload)
        file.seek(0)  # Reset file pointer
        s3_client.upload_fileobj(
            file,
            bucket_name,
            file_path,
            ExtraArgs={'ContentType': file.content_type}
        )
        
        # Generate public URL (using the endpoint URL format)
        endpoint = os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net')
        # Remove trailing slash if present
        endpoint = endpoint.rstrip('/')
        file_url = f"{endpoint}/{bucket_name}/{file_path}"
        
        # Return proxy URL instead of direct S3 URL to avoid CORS issues
        # The frontend will use /api/emails/signatures/logo-proxy/<signature_id>/ to load the image
        return Response({
            'logoUrl': file_url,  # Keep original URL for reference
            'logoProxyUrl': f'/api/emails/signatures/logo-proxy/{file_path}',  # Proxy URL for frontend
            'fileName': file.name,
            'filePath': file_path
        }, status=status.HTTP_200_OK)
        
    except ClientError as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error uploading signature logo to Impossible Cloud: {error_details}")
        return Response({'error': f'Failed to upload logo: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error uploading signature logo: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def email_signature_logo_presigned_url(request):
    """Generate presigned URL for signature logo to use in emails (publicly accessible)"""
    try:
        user = request.user
        file_path = request.GET.get('filePath')
        
        if not file_path:
            return Response({'error': 'filePath parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate that the file path belongs to the current user
        if not file_path.startswith(f'email-signatures/{user.id}/'):
            return Response({'error': 'Unauthorized access'}, status=status.HTTP_403_FORBIDDEN)
        
        # Configure S3 client
        s3_client = boto3.client(
            's3',
            endpoint_url=os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net'),
            aws_access_key_id=os.getenv('IMPOSSIBLE_CLOUD_ACCESS_KEY'),
            aws_secret_access_key=os.getenv('IMPOSSIBLE_CLOUD_SECRET_KEY'),
            region_name=os.getenv('IMPOSSIBLE_CLOUD_REGION', 'eu-central-2')
        )
        
        bucket_name = os.getenv('IMPOSSIBLE_CLOUD_BUCKET', 'leadflow-documents')
        
        # Generate presigned URL (valid for 7 days - emails may be read later)
        presigned_url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': bucket_name, 'Key': file_path},
            ExpiresIn=604800  # 7 days
        )
        
        return Response({
            'presignedUrl': presigned_url
        }, status=status.HTTP_200_OK)
        
    except ClientError as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error generating presigned URL for signature logo: {error_details}")
        return Response({'error': f'Failed to generate presigned URL: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Unexpected error generating presigned URL: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def email_signature_logo_proxy(request, file_path):
    """
    Proxy signature logo images from Impossible Cloud to avoid CORS issues.
    file_path format: email-signatures/{user_id}/{filename}
    Note: file_path may contain slashes, so we use path converter in URL
    """
    try:
        user = request.user
        
        # Validate that the file path belongs to the current user
        # file_path should be like "email-signatures/1/filename.png"
        if not file_path.startswith(f'email-signatures/{user.id}/'):
            # Also check if it's a different user's signature (for admin access, you might want to allow)
            # For now, only allow own signatures
            return Response({'error': 'Unauthorized access'}, status=status.HTTP_403_FORBIDDEN)
        
        # Configure S3 client
        s3_client = boto3.client(
            's3',
            endpoint_url=os.getenv('IMPOSSIBLE_CLOUD_ENDPOINT', 'https://eu-central-2.storage.impossibleapi.net'),
            aws_access_key_id=os.getenv('IMPOSSIBLE_CLOUD_ACCESS_KEY'),
            aws_secret_access_key=os.getenv('IMPOSSIBLE_CLOUD_SECRET_KEY'),
            region_name=os.getenv('IMPOSSIBLE_CLOUD_REGION', 'eu-central-2')
        )
        
        bucket_name = os.getenv('IMPOSSIBLE_CLOUD_BUCKET', 'leadflow-documents')
        
        # Get the object from S3
        s3_object = s3_client.get_object(Bucket=bucket_name, Key=file_path)
        
        # Create a streaming response
        def file_iterator():
            chunk_size = 8192
            while True:
                chunk = s3_object['Body'].read(chunk_size)
                if not chunk:
                    break
                yield chunk
        
        # Determine content type
        content_type = s3_object.get('ContentType', 'image/png')
        
        # Create streaming response with CORS headers
        response = StreamingHttpResponse(
            file_iterator(),
            content_type=content_type
        )
        
        # Add CORS headers
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Allow-Methods'] = 'GET'
        response['Access-Control-Allow-Headers'] = '*'
        response['Cache-Control'] = 'public, max-age=31536000'  # Cache for 1 year
        
        if 'ContentLength' in s3_object:
            response['Content-Length'] = str(s3_object['ContentLength'])
        
        return response
        
    except ClientError as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error proxying signature logo from S3: {error_details}")
        return Response({'error': f'Failed to load image: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error proxying signature logo: {error_details}")
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Chat endpoints
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def chat_rooms(request):
    """List all chat rooms for the current user or create a new one"""
    if request.method == 'GET':
        # Get pagination parameters - define outside try block to avoid NameError in except
        try:
            limit = int(request.query_params.get('limit', 15))  # Default to 15 conversations
        except (ValueError, TypeError):
            limit = 15
        
        try:
            offset = int(request.query_params.get('offset', 0))  # Default to 0 (start from most recent)
        except (ValueError, TypeError):
            offset = 0
        
        try:
            # Get total count for pagination info (optimized query)
            # Use values('id') to avoid selecting name field which may not exist yet
            total_count = ChatRoom.objects.filter(participants=request.user).values('id').distinct().count()
            
            # Get chat rooms ordered by updated_at descending (most recent first)
            # Use prefetch_related to avoid N+1 queries for participants and messages
            from django.db.models import Prefetch
            from django.db import ProgrammingError
            
            # Prefetch messages with sender to avoid N+1 queries
            # Order by created_at descending for efficient lastMessage retrieval
            messages_prefetch = Prefetch(
                'messages',
                queryset=Message.objects.select_related('sender').order_by('-created_at')
            )
            
            # Get chat rooms - defer name field to avoid selecting it if column doesn't exist
            # The serializer handles missing name column gracefully via SerializerMethodField
            chat_rooms = ChatRoom.objects.filter(
                participants=request.user
            ).distinct().defer('name').prefetch_related(
                'participants',
                messages_prefetch
            ).order_by('-updated_at')[offset:offset + limit]
            
            serializer = ChatRoomSerializer(chat_rooms, many=True, context={'request': request})
            
            # Return chat rooms with pagination metadata
            # Catch ProgrammingError in case name column doesn't exist and defer didn't work
            try:
                serializer_data = serializer.data
            except ProgrammingError as pe:
                if 'name' in str(pe).lower() or 'api_chatroom.name' in str(pe):
                    # Name column doesn't exist - recreate queryset without it
                    chat_rooms = ChatRoom.objects.filter(
                        participants=request.user
                    ).distinct().only('id', 'created_at', 'updated_at').prefetch_related(
                        'participants',
                        messages_prefetch
                    ).order_by('-updated_at')[offset:offset + limit]
                    serializer = ChatRoomSerializer(chat_rooms, many=True, context={'request': request})
                    serializer_data = serializer.data
                else:
                    raise
            
            return Response({
                'chatRooms': serializer_data,
                'hasMore': offset + limit < total_count,
                'total': total_count,
                'offset': offset,
                'limit': limit
            })
        except Exception as e:
            # Handle database connection errors gracefully
            from django.db import OperationalError
            import traceback
            
            error_details = traceback.format_exc()
            print(f"Error in chat_rooms: {error_details}")
            
            if isinstance(e, OperationalError) or 'timeout' in str(e).lower() or 'connection' in str(e).lower():
                return Response({
                    'error': 'Database connection timeout. Please try again in a moment.',
                    'chatRooms': [],
                    'hasMore': False,
                    'total': 0,
                    'offset': offset,
                    'limit': limit
                }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            
            # For other errors, return generic error
            return Response({
                'error': 'An error occurred while loading chat rooms',
                'chatRooms': [],
                'hasMore': False,
                'total': 0,
                'offset': offset,
                'limit': limit
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    elif request.method == 'POST':
        # Create a new chat room
        participant_ids = request.data.get('participants', [])
        name = request.data.get('name', '').strip()  # Optional name for group chats
        
        # Ensure current user is included
        if request.user.id not in participant_ids:
            participant_ids.append(request.user.id)
        
        # Remove duplicates
        participant_ids = list(set(participant_ids))
        
        # Only check for existing rooms if it's a 1-on-1 chat (2 participants)
        # Group chats can have multiple rooms with same participants
        if len(participant_ids) == 2:
            from django.db import ProgrammingError
            try:
                existing_rooms = ChatRoom.objects.filter(participants__in=[request.user.id]).distinct()
            except ProgrammingError:
                # Fallback if name column doesn't exist - use values to exclude it
                existing_room_ids = ChatRoom.objects.filter(participants__in=[request.user.id]).values_list('id', flat=True).distinct()
                existing_rooms = ChatRoom.objects.filter(id__in=existing_room_ids)
            
            for room in existing_rooms:
                room_participant_ids = set(room.participants.values_list('id', flat=True))
                if room_participant_ids == set(participant_ids):
                    # Room already exists, return it
                    serializer = ChatRoomSerializer(room, context={'request': request})
                    return Response(serializer.data, status=status.HTTP_200_OK)
        
        # Create new chat room
        chat_room_id = uuid.uuid4().hex[:12]
        while ChatRoom.objects.filter(id=chat_room_id).exists():
            chat_room_id = uuid.uuid4().hex[:12]
        
        # Create chat room with optional name
        chat_room = ChatRoom.objects.create(id=chat_room_id, name=name if name else None)
        
        # Add participants
        for participant_id in participant_ids:
            try:
                participant = DjangoUser.objects.get(id=participant_id)
                chat_room.participants.add(participant)
            except DjangoUser.DoesNotExist:
                continue
        
        serializer = ChatRoomSerializer(chat_room, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chat_room_detail(request, chat_room_id):
    """Get details of a specific chat room"""
    try:
        chat_room = ChatRoom.objects.get(id=chat_room_id)
        
        # Check if user is a participant
        if request.user not in chat_room.participants.all():
            return Response({'error': 'Unauthorized access'}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = ChatRoomSerializer(chat_room, context={'request': request})
        return Response(serializer.data)
    except ChatRoom.DoesNotExist:
        return Response({'error': 'Chat room not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def chat_messages(request, chat_room_id):
    """Get messages for a chat room or send a new message"""
    try:
        chat_room = ChatRoom.objects.get(id=chat_room_id)
        
        # Check if user is a participant
        if request.user not in chat_room.participants.all():
            return Response({'error': 'Unauthorized access'}, status=status.HTTP_403_FORBIDDEN)
        
        if request.method == 'GET':
            # Get pagination parameters - ALWAYS limit to 15 messages per page
            # Define outside try block to avoid NameError in except block
            try:
                limit = min(int(request.query_params.get('limit', 15)), 15)  # Max 15 messages per page
            except (ValueError, TypeError):
                limit = 15
            try:
                offset = int(request.query_params.get('offset', 0))  # Default to 0 (start from most recent)
            except (ValueError, TypeError):
                offset = 0
            
            # Get total count for pagination info
            total_count = Message.objects.filter(chat_room=chat_room).count()
            
            # Get messages ordered by created_at descending (most recent first)
            # Offset 0 = most recent messages, higher offset = older messages
            messages = Message.objects.filter(chat_room=chat_room).order_by('-created_at')[offset:offset + limit]
            
            # Reverse to get chronological order (oldest to newest) for display
            messages_list = list(messages)
            messages_list.reverse()
            
            serializer = MessageSerializer(messages_list, many=True)
            
            # Return messages with pagination metadata
            # hasMore = True if there are more messages to load (older messages when scrolling up)
            return Response({
                'messages': serializer.data,
                'hasMore': offset + limit < total_count,
                'total': total_count,
                'offset': offset,
                'limit': limit
            })
        
        elif request.method == 'POST':
            # Send a new message
            content = request.data.get('content', '').strip()
            if not content:
                return Response({'error': 'Message content is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            message_id = uuid.uuid4().hex[:12]
            while Message.objects.filter(id=message_id).exists():
                message_id = uuid.uuid4().hex[:12]
            
            message = Message.objects.create(
                id=message_id,
                chat_room=chat_room,
                sender=request.user,
                content=content
            )
            
            # Update chat room's updated_at timestamp
            chat_room.save()
            
            # Send message via WebSocket to chat room
            # Wrap in try-except to prevent message send failure if WebSocket fails
            try:
                channel_layer = get_channel_layer()
                if channel_layer:
                    message_data = {
                        'id': message.id,
                        'chatRoomId': message.chat_room.id,
                        'senderId': message.sender.id,
                        'senderName': f"{message.sender.first_name} {message.sender.last_name}".strip() or message.sender.username,
                        'content': message.content,
                        'isRead': message.is_read,
                        'createdAt': message.created_at.isoformat(),
                    }
                    
                    async_to_sync(channel_layer.group_send)(
                        f'chat_{chat_room.id}',
                        {
                            'type': 'chat_message',
                            'message': message_data,
                            'sender_id': request.user.id
                        }
                    )
                    
                    # Send message notification via WebSocket (no database notification for messages)
                    # This allows real-time popup without cluttering notifications
                    participants = chat_room.participants.exclude(id=request.user.id)
                    
                    for participant in participants:
                        try:
                            # Check if this is the first message the recipient receives in this room
                            # (excluding messages they sent themselves)
                            recipient_message_count = Message.objects.filter(
                                chat_room=chat_room
                            ).exclude(sender=participant).count()
                            is_first_message_for_recipient = recipient_message_count == 1
                            
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.info(f"[chat_messages] Sending message to participant {participant.id}. "
                                      f"Recipient message count: {recipient_message_count}, "
                                      f"Is first message: {is_first_message_for_recipient}, "
                                      f"Chat room ID: {chat_room.id}")
                            
                            async_to_sync(channel_layer.group_send)(
                                f'chat_message_{participant.id}',
                                {
                                    'type': 'new_message',
                                    'message': {
                                        'id': message.id,
                                        'chatRoomId': chat_room.id,
                                        'senderId': message.sender.id,
                                        'senderName': f"{message.sender.first_name} {message.sender.last_name}".strip() or message.sender.username,
                                        'content': message.content,
                                        'createdAt': message.created_at.isoformat(),
                                    },
                                    'chat_room_id': chat_room.id,
                                    'is_new_chat_room': is_first_message_for_recipient,
                                }
                            )
                            
                            # If this is the first message for the recipient, also send a new_chat_room event
                            # This ensures the chat appears in their list
                            if is_first_message_for_recipient:
                                from .serializer import ChatRoomSerializer
                                # Create a mock request object for serializer context
                                class MockRequest:
                                    def __init__(self, user):
                                        self.user = user
                                
                                mock_request = MockRequest(participant)
                                chat_room_data = ChatRoomSerializer(chat_room, context={'request': mock_request}).data
                                
                                async_to_sync(channel_layer.group_send)(
                                    f'chat_message_{participant.id}',
                                    {
                                        'type': 'new_chat_room',
                                        'chat_room': chat_room_data,
                                    }
                                )
                        except Exception as ws_error:
                            # Log but don't fail - message is already saved
                            import traceback
                            print(f"Error sending WebSocket notification to participant {participant.id}: {str(ws_error)}")
                            print(traceback.format_exc())
            except Exception as e:
                # Log but don't fail - message is already saved
                import traceback
                print(f"Error sending message via WebSocket: {str(e)}")
                print(traceback.format_exc())
            
            serializer = MessageSerializer(message)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
    except ChatRoom.DoesNotExist:
        return Response({'error': 'Chat room not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_chat_room(request, chat_room_id):
    """Update chat room name or add/remove participants"""
    try:
        chat_room = ChatRoom.objects.get(id=chat_room_id)
        
        # Check if user is a participant
        if request.user not in chat_room.participants.all():
            return Response({'error': 'Unauthorized access'}, status=status.HTTP_403_FORBIDDEN)
        
        # Update name if provided
        if 'name' in request.data:
            chat_room.name = request.data.get('name', '').strip() or None
            chat_room.save()
        
        # Add participants if provided
        if 'add_participants' in request.data:
            participant_ids = request.data.get('add_participants', [])
            for participant_id in participant_ids:
                try:
                    participant = DjangoUser.objects.get(id=participant_id)
                    if participant not in chat_room.participants.all():
                        chat_room.participants.add(participant)
                except DjangoUser.DoesNotExist:
                    continue
            chat_room.save()  # Update timestamp
        
        # Remove participants if provided
        if 'remove_participants' in request.data:
            participant_ids = request.data.get('remove_participants', [])
            # Get initial participant count before removals
            initial_participant_count = chat_room.participants.count()
            
            for participant_id in participant_ids:
                try:
                    participant = DjangoUser.objects.get(id=participant_id)
                    # Get current count before this removal
                    current_count = chat_room.participants.count()
                    
                    # Don't allow removing if it would leave less than 1 participant
                    if current_count <= 1:
                        continue
                    
                    # Allow removing other participants if there are more than 1 total
                    if participant.id != request.user.id:
                        chat_room.participants.remove(participant)
                    # Allow current user to leave if there are other participants
                    elif participant.id == request.user.id and current_count > 1:
                        chat_room.participants.remove(participant)
                except DjangoUser.DoesNotExist:
                    continue
            chat_room.save()  # Update timestamp
        
        serializer = ChatRoomSerializer(chat_room, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
        
    except ChatRoom.DoesNotExist:
        return Response({'error': 'Chat room not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_messages_read(request, chat_room_id):
    """Mark all messages in a chat room as read for the current user"""
    try:
        chat_room = ChatRoom.objects.get(id=chat_room_id)
        
        # Check if user is a participant
        if request.user not in chat_room.participants.all():
            return Response({'error': 'Unauthorized access'}, status=status.HTTP_403_FORBIDDEN)
        
        # Mark all unread messages (except those sent by the user) as read
        Message.objects.filter(
            chat_room=chat_room,
            is_read=False
        ).exclude(sender=request.user).update(is_read=True)
        
        return Response({'success': True})
    except ChatRoom.DoesNotExist:
        return Response({'error': 'Chat room not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chat_users(request):
    """Get list of users that can be chatted with"""
    # Filter out deleted users
    users = DjangoUser.objects.filter(is_active=True).exclude(id=request.user.id)
    user_list = []
    for user in users:
        # Check if user is deleted
        try:
            user_details = UserDetails.objects.get(django_user=user)
            if user_details.deleted_at:
                continue  # Skip deleted users
        except UserDetails.DoesNotExist:
            pass  # UserDetails doesn't exist, include user
        
        first_name = user.first_name or ''
        last_name = user.last_name or ''
        name = f"{first_name} {last_name}".strip() if (first_name or last_name) else user.username
        user_list.append({
            'id': user.id,
            'username': user.username,
            'name': name,
            'email': user.email or ''
        })
    return Response(user_list)

# Notification views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_list(request):
    """Get all notifications for the current user"""
    # Exclude message notifications - they are handled separately via chat popup
    # Order by unread status first (unread first), then by created_at (newest first)
    notifications = Notification.objects.filter(user=request.user).exclude(type='message').order_by('is_read', '-created_at')
    
    # Pagination - increase limit to ensure all unread notifications are included
    limit = int(request.query_params.get('limit', 200))  # Increased to ensure unread notifications aren't paginated out
    offset = int(request.query_params.get('offset', 0))
    
    notifications = notifications[offset:offset + limit]
    
    serializer = NotificationSerializer(notifications, many=True)
    
    # Calculate unread count for debugging
    unread_count = Notification.objects.filter(user=request.user, is_read=False).exclude(type='message').count()
    unread_in_response = sum(1 for n in serializer.data if not n.get('is_read', True))
    
    return Response({
        'notifications': serializer.data,
        'total': Notification.objects.filter(user=request.user).exclude(type='message').count(),
        'unread_count': unread_count,  # Include unread count in response for debugging
        'unread_in_response': unread_in_response  # How many unread notifications are in this response
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_unread_count(request):
    """Get unread notifications count for the current user (excluding message notifications)"""
    # Exclude message notifications - they are handled separately via chat popup
    count = Notification.objects.filter(user=request.user, is_read=False).exclude(type='message').count()
    return Response({'unread_count': count})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def notification_mark_read(request, notification_id):
    """Mark a notification as read"""
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        
        # Send websocket update to notify frontend immediately
        channel_layer = get_channel_layer()
        if channel_layer:
            notification_data = {
                'id': notification.id,
                'type': notification.type,
                'title': notification.title,
                'message': notification.message,
                'message_id': notification.message_id if notification.message_id else None,
                'email_id': notification.email_id if notification.email_id else None,
                'contact_id': notification.contact_id if notification.contact_id else None,
                'event_id': notification.event_id if notification.event_id else None,
                'data': notification.data if notification.data else {},
                'is_read': notification.is_read,
                'created_at': notification.created_at.isoformat(),
            }
            
            # Get updated unread count (excluding message notifications)
            unread_count = Notification.objects.filter(user=request.user, is_read=False).exclude(type='message').count()
            
            # Send via WebSocket
            async_to_sync(channel_layer.group_send)(
                f'notifications_{request.user.id}',
                {
                    'type': 'notification_updated',
                    'notification': notification_data,
                    'unread_count': unread_count,
                }
            )
        
        serializer = NotificationSerializer(notification)
        return Response(serializer.data)
    except Notification.DoesNotExist:
        return Response({'error': 'Notification not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def notification_mark_all_read(request):
    """Mark all notifications as read for the current user (excluding message notifications)"""
    # Only mark non-message notifications as read - message notifications are handled separately
    Notification.objects.filter(user=request.user, is_read=False).exclude(type='message').update(is_read=True)
    
    # Send websocket update to notify frontend immediately
    channel_layer = get_channel_layer()
    if channel_layer:
        # Get updated unread count (excluding message notifications, should be 0)
        unread_count = Notification.objects.filter(user=request.user, is_read=False).exclude(type='message').count()
        
        # Send via WebSocket
        async_to_sync(channel_layer.group_send)(
            f'notifications_{request.user.id}',
            {
                'type': 'unread_count_updated',
                'unread_count': unread_count,
            }
        )
    
    return Response({'success': True})

# Notification Preferences endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_preference_list(request):
    """Get all notification preferences for all roles"""
    preferences = NotificationPreference.objects.all().select_related('role').order_by('role__name')
    serializer = NotificationPreferenceSerializer(preferences, many=True)
    return Response({'preferences': serializer.data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_preference_detail(request, role_id):
    """Get notification preference for a specific role"""
    try:
        role = Role.objects.get(id=role_id)
        preference, created = NotificationPreference.objects.get_or_create(
            role=role,
            defaults={
                'id': uuid.uuid4().hex[:12],
                'notify_message_received': True,
                'notify_sensitive_contact_modification': True,
                'notify_contact_edit': True
            }
        )
        # Ensure ID is set if it was just created
        if created:
            pref_id = uuid.uuid4().hex[:12]
            while NotificationPreference.objects.filter(id=pref_id).exists():
                pref_id = uuid.uuid4().hex[:12]
            preference.id = pref_id
            preference.save()
        
        serializer = NotificationPreferenceSerializer(preference)
        return Response(serializer.data)
    except Role.DoesNotExist:
        return Response({'error': 'Role not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def notification_preference_update(request, role_id):
    """Update notification preference for a specific role"""
    try:
        role = Role.objects.get(id=role_id)
        preference, created = NotificationPreference.objects.get_or_create(
            role=role,
            defaults={
                'id': uuid.uuid4().hex[:12],
                'notify_message_received': True,
                'notify_sensitive_contact_modification': True,
                'notify_contact_edit': True
            }
        )
        
        # Ensure ID is set if it was just created
        if created:
            pref_id = uuid.uuid4().hex[:12]
            while NotificationPreference.objects.filter(id=pref_id).exists():
                pref_id = uuid.uuid4().hex[:12]
            preference.id = pref_id
            preference.save()
        
        # Update fields from request data
        if 'notifyMessageReceived' in request.data:
            preference.notify_message_received = request.data['notifyMessageReceived']
        if 'notifySensitiveContactModification' in request.data:
            preference.notify_sensitive_contact_modification = request.data['notifySensitiveContactModification']
        if 'notifyContactEdit' in request.data:
            preference.notify_contact_edit = request.data['notifyContactEdit']
        
        preference.save()
        serializer = NotificationPreferenceSerializer(preference)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Role.DoesNotExist:
        return Response({'error': 'Role not found'}, status=status.HTTP_404_NOT_FOUND)

# Fosse Settings endpoints
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def fosse_settings_list(request):
    """Get all Fosse settings for all roles"""
    settings_list = FosseSettings.objects.all().select_related('role').order_by('role__name')
    serializer = FosseSettingsSerializer(settings_list, many=True)
    return Response({'settings': serializer.data})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def fosse_settings_detail(request, role_id):
    """Get Fosse settings for a specific role"""
    try:
        role = Role.objects.get(id=role_id)
        fosse_setting, created = FosseSettings.objects.get_or_create(
            role=role,
            defaults={
                'id': uuid.uuid4().hex[:12],
                'forced_columns': [],
                'forced_filters': {},
                'default_order': 'default'
            }
        )
        # Ensure ID is set if it was just created
        if created:
            setting_id = uuid.uuid4().hex[:12]
            while FosseSettings.objects.filter(id=setting_id).exists():
                setting_id = uuid.uuid4().hex[:12]
            fosse_setting.id = setting_id
            fosse_setting.save()
        
        serializer = FosseSettingsSerializer(fosse_setting)
        return Response(serializer.data)
    except Role.DoesNotExist:
        return Response({'error': 'Role not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def fosse_settings_update(request, role_id):
    """Update Fosse settings for a specific role"""
    # Import status module locally to avoid scoping conflicts
    from rest_framework import status as http_status
    try:
        role = Role.objects.get(id=role_id)
        fosse_setting, created = FosseSettings.objects.get_or_create(
            role=role,
            defaults={
                'id': uuid.uuid4().hex[:12],
                'forced_columns': [],
                'forced_filters': {},
                'default_order': 'default'
            }
        )
        
        # Ensure ID is set if it was just created
        if created:
            setting_id = uuid.uuid4().hex[:12]
            while FosseSettings.objects.filter(id=setting_id).exists():
                setting_id = uuid.uuid4().hex[:12]
            fosse_setting.id = setting_id
            fosse_setting.save()
        
        # Use serializer to validate and update
        serializer = FosseSettingsSerializer(fosse_setting, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=http_status.HTTP_200_OK)
        else:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"FosseSettings validation errors: {serializer.errors}")
            logger.error(f"Request data: {request.data}")
            return Response({
                'error': 'Validation failed',
                'details': serializer.errors
            }, status=http_status.HTTP_400_BAD_REQUEST)
    except Role.DoesNotExist:
        return Response({'error': 'Role not found'}, status=http_status.HTTP_404_NOT_FOUND)
    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error updating Fosse settings: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'error': 'Internal server error',
            'detail': str(e),
            'traceback': traceback.format_exc()
        }, status=http_status.HTTP_500_INTERNAL_SERVER_ERROR)

# TEMPORARY ENDPOINT - DELETE ALL CONTACTS
# WARNING: This is a dangerous operation that will delete ALL contacts from the database
# Remove this endpoint after use
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def delete_all_contacts(request):
    """
    TEMPORARY ENDPOINT: Delete all contacts from the database.
    WARNING: This operation cannot be undone!
    
    Requires confirmation parameter: {'confirm': 'DELETE_ALL'}
    """
    try:
        # Require explicit confirmation to prevent accidental deletion
        confirmation = request.data.get('confirm', '')
        if confirmation != 'DELETE_ALL':
            return Response({
                'error': 'Confirmation required. Send {"confirm": "DELETE_ALL"} to proceed.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get count before deletion
        total_count = Contact.objects.count()
        
        # Delete all contacts
        deleted_count, _ = Contact.objects.all().delete()
        
        return Response({
            'success': True,
            'message': f'Successfully deleted {deleted_count} contact(s)',
            'deleted_count': deleted_count,
            'total_before_deletion': total_count
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return Response({
            'error': str(e),
            'details': error_details
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def send_otp(request):
    """Send OTP code to user's email using Resend API after password verification"""
    # Close old connections at the start
    from django.db import close_old_connections, OperationalError
    close_old_connections()
    
    try:
        email = request.data.get('email', '').strip().lower()
        password = request.data.get('password', '')
        
        if not email:
            return Response({'error': 'Email is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not password:
            return Response({'error': 'Password is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate email format
        from django.core.validators import validate_email
        from django.core.exceptions import ValidationError
        try:
            validate_email(email)
        except ValidationError:
            return Response({'error': 'Invalid email format'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if user exists with this email
        try:
            # Close old connections before database query
            from django.db import close_old_connections
            close_old_connections()
            
            user = DjangoUser.objects.get(email=email)
            
            # Close connections after query
            close_old_connections()
        except DjangoUser.DoesNotExist:
            # Close connections even on error
            from django.db import close_old_connections
            close_old_connections()
            return Response({'error': 'Invalid email or password'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            # Close connections on any error
            from django.db import close_old_connections
            close_old_connections()
            # Re-raise if it's a connection error, otherwise handle normally
            from django.db import OperationalError
            if isinstance(e, OperationalError):
                return Response({'error': 'Database connection error. Please try again.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            raise
        
        # Verify password
        from django.contrib.auth import authenticate
        authenticated_user = authenticate(username=user.username, password=password)
        if not authenticated_user:
            return Response({'error': 'Invalid email or password'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # Check if user requires OTP
        user_details = None
        try:
            user_details = UserDetails.objects.get(django_user=user)
            require_otp = user_details.require_otp
            # Check if user is deleted
            if user_details.deleted_at:
                return Response({'error': 'This account has been deleted.'}, status=status.HTTP_403_FORBIDDEN)
            # Check if user is active
            if not user_details.active:
                return Response({'error': 'This account has been disabled. Please contact your administrator.'}, status=status.HTTP_403_FORBIDDEN)
            # Check IP whitelist if enabled
            if user_details.ip_whitelist_enabled:
                client_ip = get_client_ip(request)
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"IP whitelist check for user {user.email}: enabled={user_details.ip_whitelist_enabled}, whitelist={user_details.ip_whitelist}, client_ip={client_ip}")
                is_allowed, error_msg = is_ip_allowed(client_ip, user_details.ip_whitelist)
                if not is_allowed:
                    error_message = f"Access denied. Your IP address ({client_ip}) is not authorized."
                    if error_msg:
                        error_message += f" {error_msg}"
                    logger.warning(f"IP whitelist check FAILED for user {user.email}: client_ip={client_ip} not in whitelist={user_details.ip_whitelist}")
                    return Response({'error': error_message}, status=status.HTTP_403_FORBIDDEN)
                logger.info(f"IP whitelist check PASSED for user {user.email}: client_ip={client_ip}")
        except UserDetails.DoesNotExist:
            # If UserDetails doesn't exist, default to require_otp = False
            require_otp = False
        
        # If user doesn't require OTP, return tokens directly
        if not require_otp:
            # Generate JWT tokens
            from rest_framework_simplejwt.tokens import RefreshToken
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'success': True,
                'message': 'Login successful',
                'access': str(refresh.access_token),
                'refresh': str(refresh)
            }, status=status.HTTP_200_OK)
        
        # User requires OTP - proceed with OTP generation and sending
        # Generate 6-digit OTP code
        import random
        otp_code = str(random.randint(100000, 999999))
        
        # Set expiration time (10 minutes from now)
        from django.utils import timezone
        expires_at = timezone.now() + timedelta(minutes=10)
        
        # Invalidate any existing unverified OTPs for this email
        OTP.objects.filter(email=email, is_verified=False).update(is_verified=True)
        
        # Create new OTP record
        otp_id = uuid.uuid4().hex[:12]
        while OTP.objects.filter(id=otp_id).exists():
            otp_id = uuid.uuid4().hex[:12]
        
        otp = OTP.objects.create(
            id=otp_id,
            email=email,
            code=otp_code,
            expires_at=expires_at
        )
        
        # Send email using Resend API
        import os
        resend_api_key = os.getenv('RESEND_API_KEY')
        if not resend_api_key:
            return Response({'error': 'Resend API key not configured'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        try:
            try:
                import resend
            except ImportError as import_err:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to import resend module: {str(import_err)}")
                return Response({
                    'error': 'Resend module not installed',
                    'detail': f'Please install resend: pip install resend==2.19.0. Error: {str(import_err)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            resend.api_key = resend_api_key
            
            # Use noreply@crm-prospection.online for OTP emails
            from_email = 'noreply@crm-prospection.online'
            
            # Send email
            params = {
                "from": from_email,
                "to": [email],
                "subject": "Votre code OTP",
                "html": f"""
                <html>
                <body style="font-family: Arial, sans-serif; padding: 20px;">
                    <h2>Votre code OTP</h2>
                    <p>Bonjour,</p>
                    <p>Votre code de mot de passe à usage unique (OTP) est :</p>
                    <div style="background-color: #f4f4f4; padding: 15px; border-radius: 5px; text-align: center; margin: 20px 0;">
                        <h1 style="margin: 0; color: #333; letter-spacing: 5px;">{otp_code}</h1>
                    </div>
                    <p>Ce code expirera dans 10 minutes.</p>
                    <p>Si vous n'avez pas demandé ce code, veuillez ignorer cet email.</p>
                    <p>Cordialement,<br>L'équipe LeadFlow</p>
                </body>
                </html>
                """,
            }
            
            email_response = resend.Emails.send(params)
            
            return Response({
                'success': True,
                'message': 'OTP code sent successfully',
                'otp_id': otp_id
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            # Close connections on error
            close_old_connections()
            
            # Delete OTP record if email sending fails
            try:
                otp.delete()
            except:
                pass
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send OTP email: {str(e)}")
            return Response({
                'error': 'Failed to send OTP email',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except OperationalError as e:
        # Close connections on database error
        close_old_connections()
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Database connection error in send_otp: {str(e)}")
        return Response({
            'error': 'Database connection error. Please try again in a moment.',
            'detail': 'The database is temporarily unavailable. Please wait a few seconds and try again.'
        }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
    except Exception as e:
        # Close connections on any error
        close_old_connections()
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error in send_otp: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'error': 'Internal server error',
            'detail': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    finally:
        # Always close connections at the end
        close_old_connections()

@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    """Verify OTP code and return JWT tokens if valid"""
    try:
        email = request.data.get('email', '').strip().lower()
        otp_code = request.data.get('otp', '').strip()
        
        if not email or not otp_code:
            return Response({'error': 'Email and OTP code are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Find the most recent unverified OTP for this email
        try:
            otp = OTP.objects.filter(email=email, is_verified=False).order_by('-created_at').first()
            
            if not otp:
                return Response({'error': 'No OTP found for this email. Please request a new one.'}, status=status.HTTP_404_NOT_FOUND)
            
            # Check if OTP is expired
            if otp.is_expired():
                return Response({'error': 'OTP code has expired. Please request a new one.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Verify OTP code
            if otp.code != otp_code:
                return Response({'error': 'Invalid OTP code'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Mark OTP as verified
            otp.is_verified = True
            otp.save()
            
            # Get user by email
            try:
                user = DjangoUser.objects.get(email=email)
            except DjangoUser.DoesNotExist:
                return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Check if user is deleted or disabled
            try:
                user_details = UserDetails.objects.get(django_user=user)
                if user_details.deleted_at:
                    return Response({'error': 'This account has been deleted.'}, status=status.HTTP_403_FORBIDDEN)
                if not user_details.active:
                    return Response({'error': 'This account has been disabled. Please contact your administrator.'}, status=status.HTTP_403_FORBIDDEN)
                # Check IP whitelist if enabled
                if user_details.ip_whitelist_enabled:
                    client_ip = get_client_ip(request)
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.info(f"IP whitelist check for user {user.email} (verify_otp): enabled={user_details.ip_whitelist_enabled}, whitelist={user_details.ip_whitelist}, client_ip={client_ip}")
                    is_allowed, error_msg = is_ip_allowed(client_ip, user_details.ip_whitelist)
                    if not is_allowed:
                        error_message = f"Access denied. Your IP address ({client_ip}) is not authorized."
                        if error_msg:
                            error_message += f" {error_msg}"
                        logger.warning(f"IP whitelist check FAILED for user {user.email} (verify_otp): client_ip={client_ip} not in whitelist={user_details.ip_whitelist}")
                        return Response({'error': error_message}, status=status.HTTP_403_FORBIDDEN)
                    logger.info(f"IP whitelist check PASSED for user {user.email} (verify_otp): client_ip={client_ip}")
            except UserDetails.DoesNotExist:
                # UserDetails doesn't exist, allow login (might be a new user)
                pass
            
            # Generate JWT tokens
            from rest_framework_simplejwt.tokens import RefreshToken
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'success': True,
                'message': 'OTP verified successfully',
                'access': str(refresh.access_token),
                'refresh': str(refresh)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error verifying OTP: {str(e)}")
            return Response({
                'error': 'Failed to verify OTP',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error in verify_otp: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'error': 'Internal server error',
            'detail': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Health check endpoint
@api_view(['GET'])
@permission_classes([AllowAny])
def health_check(request):
    """Health check endpoint for API."""
    return Response({"status": "healthy", "service": "backend"}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def api_root(request):
    """API root endpoint - returns API information."""
    return Response({
        "message": "LeadFlow API is running",
        "status": "ok",
        "version": "1.0",
        "endpoints": {
            "health": "/api/health/",
            "token": "/api/token/",
            "contacts": "/api/contacts/",
            "users": "/api/users/",
            "teams": "/api/teams/",
            "events": "/api/events/",
            "notes": "/api/notes/",
            "logs": "/api/logs/",
            "documents": "/api/documents/",
            "chat": "/api/chat/",
            "notifications": "/api/notifications/",
        }
    }, status=status.HTTP_200_OK)

# Custom token serializer to check if user is deleted or disabled
class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        data = super().validate(attrs)
        # Check if user is deleted or disabled
        try:
            user_details = UserDetails.objects.get(django_user=self.user)
            if user_details.deleted_at:
                from rest_framework_simplejwt.exceptions import AuthenticationFailed
                raise AuthenticationFailed(
                    'This account has been deleted.',
                    code='user_deleted'
                )
            if not user_details.active:
                from rest_framework_simplejwt.exceptions import AuthenticationFailed
                raise AuthenticationFailed(
                    'This account has been disabled. Please contact your administrator.',
                    code='user_disabled'
                )
            # Check IP whitelist if enabled
            if user_details.ip_whitelist_enabled:
                # Get request from context
                request = self.context.get('request')
                if request:
                    client_ip = get_client_ip(request)
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.info(f"IP whitelist check for user {self.user.email} (token): enabled={user_details.ip_whitelist_enabled}, whitelist={user_details.ip_whitelist}, client_ip={client_ip}")
                    is_allowed, error_msg = is_ip_allowed(client_ip, user_details.ip_whitelist)
                    if not is_allowed:
                        error_message = f"Access denied. Your IP address ({client_ip}) is not authorized."
                        if error_msg:
                            error_message += f" {error_msg}"
                        logger.warning(f"IP whitelist check FAILED for user {self.user.email} (token): client_ip={client_ip} not in whitelist={user_details.ip_whitelist}")
                        from rest_framework_simplejwt.exceptions import AuthenticationFailed
                        raise AuthenticationFailed(
                            error_message,
                            code='ip_not_authorized'
                        )
                    logger.info(f"IP whitelist check PASSED for user {self.user.email} (token): client_ip={client_ip}")
                else:
                    # If no request context, deny access when IP whitelist is enabled
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"IP whitelist check FAILED for user {self.user.email} (token): No request context available")
                    from rest_framework_simplejwt.exceptions import AuthenticationFailed
                    raise AuthenticationFailed(
                        'Access denied. Your IP address is not authorized.',
                        code='ip_not_authorized'
                    )
        except UserDetails.DoesNotExist:
            # UserDetails doesn't exist, allow login (might be a new user)
            pass
        return data

# Custom token view that uses the custom serializer
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

# Import contact view endpoints
from .contact_views_endpoints import get_contact_views, create_contact_view, update_contact_view, delete_contact_view

